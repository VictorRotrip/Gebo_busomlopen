#!/usr/bin/env python3
"""
create_additional_inputs.py

Creates additional_inputs.xlsx — the consolidated input file for the bus roster
optimizer. Contains all financial data, bus specifications, and station data
in one place.

Sheets:
  1. Tarieven         - Revenue rates per bus type
  2. Chauffeurkosten  - Driver costs (CAO BB)
  3. Buskosten        - Bus operating costs, fuel consumption, tank capacity, ZE specs
  4. Duurzaamheid     - Sustainability KPIs and incentives
  5. Brandstofprijzen - Fuel/electricity prices (auto-fetched)
  6. Tanklocaties     - Fuel stations per bus station (from tanklocaties.json)
  7. Laadstations     - Charging stations per bus station (from tanklocaties.json)

Usage:
  python create_additional_inputs.py                    # Create with defaults
  python create_additional_inputs.py --tanklocaties tanklocaties.json  # Include station data
  python create_additional_inputs.py -o my_inputs.xlsx  # Custom output name
"""

import argparse
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# Styling constants
BOLD = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
AUTO_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
MANUAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        if fill and c == 2:
            cell.fill = fill


def add_section_header(ws, row, title, cols):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=11, color="2F5496")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    return row


def write_var(ws, row, name, value, unit, description, fill=None):
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=value)
    ws.cell(row=row, column=3, value=unit)
    ws.cell(row=row, column=4, value=description)
    style_data(ws, row, 4, fill=fill)
    return row + 1


# ---------------------------------------------------------------------------
# Sheet 1: Tarieven
# ---------------------------------------------------------------------------

def create_tarieven_sheet(wb):
    """Revenue rates per bus type."""
    ws = wb.create_sheet("Tarieven")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2
    r = add_section_header(ws, r, "Uurtarieven per bustype (omzet = inzet-uren x tarief)", 4) + 1
    rates = [
        ("rate_dubbeldekker_eur_per_hour", 116.37, "EUR/uur", "Uurtarief Dubbeldekker"),
        ("rate_touringcar_eur_per_hour", 80.455, "EUR/uur", "Uurtarief Touringcar"),
        ("rate_lagevloer_gelede_eur_per_hour", 80.445, "EUR/uur", "Uurtarief Lagevloer/Gelede"),
        ("rate_midibus_eur_per_hour", 74.85, "EUR/uur", "Uurtarief Midi bus"),
        ("rate_taxibus_eur_per_hour", 50.455, "EUR/uur", "Uurtarief Taxibus"),
    ]
    for name, val, unit, desc in rates:
        r = write_var(ws, r, name, val, unit, desc)

    for c, w in enumerate([42, 14, 12, 60], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


# ---------------------------------------------------------------------------
# Sheet 2: Chauffeurkosten
# ---------------------------------------------------------------------------

def create_chauffeurkosten_sheet(wb):
    """Driver costs from CAO BB."""
    ws = wb.create_sheet("Chauffeurkosten")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2

    # Pauzestaffel
    r = add_section_header(ws, r, "Pauzestaffel (pauze-aftrek op betaalde uren)", 4) + 1
    staffel = [
        ("Pauzestaffel_1_max_diensttijd_uren", 4.5, "uren", "Bracket 1: diensttijd <= 4.5h"),
        ("Pauzestaffel_1_pauze_min", 0, "minuten", "Pauze-aftrek: 0 min"),
        ("Pauzestaffel_2_max_diensttijd_uren", 7.5, "uren", "Bracket 2: diensttijd <= 7.5h"),
        ("Pauzestaffel_2_pauze_min", 30, "minuten", "Pauze-aftrek: 30 min"),
        ("Pauzestaffel_3_max_diensttijd_uren", 10.5, "uren", "Bracket 3: diensttijd <= 10.5h"),
        ("Pauzestaffel_3_pauze_min", 60, "minuten", "Pauze-aftrek: 60 min"),
        ("Pauzestaffel_4_max_diensttijd_uren", 13.5, "uren", "Bracket 4: diensttijd <= 13.5h"),
        ("Pauzestaffel_4_pauze_min", 90, "minuten", "Pauze-aftrek: 90 min"),
        ("Pauzestaffel_5_max_diensttijd_uren", 16.5, "uren", "Bracket 5: diensttijd <= 16.5h"),
        ("Pauzestaffel_5_pauze_min", 120, "minuten", "Pauze-aftrek: 120 min"),
        ("Pauzestaffel_6_min_diensttijd_uren", 16.5, "uren", "Bracket 6: diensttijd > 16.5h"),
        ("Pauzestaffel_6_pauze_min", 150, "minuten", "Pauze-aftrek: 150 min"),
    ]
    for name, val, unit, desc in staffel:
        r = write_var(ws, r, name, val, unit, desc)

    # Nachtwerk
    r = add_section_header(ws, r, "Nachtwerk", 4) + 1
    nacht = [
        ("Nachtwerk_window_start_min", 60, "min na middernacht", "Nachtwerk start (01:00)"),
        ("Nachtwerk_window_end_min", 300, "min na middernacht", "Nachtwerk einde (05:00)"),
        ("Max_arbeidstijd_per_24h_bij_nachtwerk_uren", 12, "uren", "Max arbeidstijd bij nachtwerk"),
    ]
    for name, val, unit, desc in nacht:
        r = write_var(ws, r, name, val, unit, desc)

    # Basis uurloon
    r = add_section_header(ws, r, "Basis uurloon chauffeur", 4) + 1
    r = write_var(ws, r, "basis_uurloon_chauffeur_eur", 18.50, "EUR/uur",
                  "Bruto basisuurloon chauffeur", fill=MANUAL_FILL)
    r = write_var(ws, r, "werkgever_opslag_factor", 1.35, "factor",
                  "Werkgeverslasten factor", fill=MANUAL_FILL)

    for c, w in enumerate([55, 14, 18, 55], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


# ---------------------------------------------------------------------------
# Sheet 3: Buskosten (with tank capacity and refuel times)
# ---------------------------------------------------------------------------

def create_buskosten_sheet(wb):
    """Bus operating costs including fuel consumption, tank capacity, ZE specs."""
    ws = wb.create_sheet("Buskosten")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2

    # --- Diesel fuel consumption ---
    r = add_section_header(ws, r, "Brandstofverbruik diesel per bustype", 4) + 1
    consumption = [
        ("verbruik_dubbeldekker_diesel_l_per_100km", 45, "L/100km", "Dieselverbruik Dubbeldekker"),
        ("verbruik_touringcar_diesel_l_per_100km", 32, "L/100km", "Dieselverbruik Touringcar"),
        ("verbruik_lagevloer_diesel_l_per_100km", 38, "L/100km", "Dieselverbruik Lagevloer"),
        ("verbruik_midibus_diesel_l_per_100km", 25, "L/100km", "Dieselverbruik Midi bus"),
        ("verbruik_taxibus_diesel_l_per_100km", 12, "L/100km", "Dieselverbruik Taxibus"),
    ]
    for name, val, unit, desc in consumption:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    # --- Tank capacity (NEW) ---
    r = add_section_header(ws, r, "Tankinhoud per bustype (diesel)", 4) + 1
    tank_capacity = [
        ("tankcapaciteit_dubbeldekker_liter", 400, "liter", "Tankinhoud Dubbeldekker"),
        ("tankcapaciteit_touringcar_liter", 500, "liter", "Tankinhoud Touringcar"),
        ("tankcapaciteit_lagevloer_liter", 300, "liter", "Tankinhoud Lagevloer"),
        ("tankcapaciteit_midibus_liter", 200, "liter", "Tankinhoud Midi bus"),
        ("tankcapaciteit_taxibus_liter", 80, "liter", "Tankinhoud Taxibus"),
    ]
    for name, val, unit, desc in tank_capacity:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    # --- Calculated diesel range (formula in description) ---
    r = add_section_header(ws, r, "Berekende actieradius diesel (tank / verbruik x 100)", 4) + 1
    diesel_range = [
        ("actieradius_dubbeldekker_diesel_km", 889, "km", "400L / 45L × 100 = 889 km"),
        ("actieradius_touringcar_diesel_km", 1562, "km", "500L / 32L × 100 = 1562 km"),
        ("actieradius_lagevloer_diesel_km", 789, "km", "300L / 38L × 100 = 789 km"),
        ("actieradius_midibus_diesel_km", 800, "km", "200L / 25L × 100 = 800 km"),
        ("actieradius_taxibus_diesel_km", 667, "km", "80L / 12L × 100 = 667 km"),
    ]
    for name, val, unit, desc in diesel_range:
        r = write_var(ws, r, name, val, unit, desc)

    # --- Refuel time (NEW) ---
    r = add_section_header(ws, r, "Tanktijd en rijtijd naar tankstation", 4) + 1
    refuel = [
        ("tanktijd_diesel_min", 15, "minuten", "Tijd om diesel te tanken"),
        ("tanktijd_buffer_min", 5, "minuten", "Extra tijd (afrekenen, etc.)"),
        ("avg_snelheid_naar_tankstation_kmh", 30, "km/h", "Gemiddelde snelheid naar tankstation"),
    ]
    for name, val, unit, desc in refuel:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    # --- ZE consumption ---
    r = add_section_header(ws, r, "Elektrisch verbruik ZE bussen", 4) + 1
    electric = [
        ("verbruik_dubbeldekker_ze_kwh_per_100km", 180, "kWh/100km", "Verbruik Dubbeldekker ZE"),
        ("verbruik_touringcar_ze_kwh_per_100km", 130, "kWh/100km", "Verbruik Touringcar ZE"),
        ("verbruik_lagevloer_ze_kwh_per_100km", 150, "kWh/100km", "Verbruik Lagevloer ZE"),
        ("verbruik_midibus_ze_kwh_per_100km", 100, "kWh/100km", "Verbruik Midi bus ZE"),
        ("verbruik_taxibus_ze_kwh_per_100km", 50, "kWh/100km", "Verbruik Taxibus ZE"),
    ]
    for name, val, unit, desc in electric:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    # --- ZE battery capacity (NEW) ---
    r = add_section_header(ws, r, "Accucapaciteit ZE bussen", 4) + 1
    battery = [
        ("accucapaciteit_dubbeldekker_ze_kwh", 450, "kWh", "Accucapaciteit Dubbeldekker ZE"),
        ("accucapaciteit_touringcar_ze_kwh", 390, "kWh", "Accucapaciteit Touringcar ZE"),
        ("accucapaciteit_lagevloer_ze_kwh", 420, "kWh", "Accucapaciteit Lagevloer ZE"),
        ("accucapaciteit_midibus_ze_kwh", 350, "kWh", "Accucapaciteit Midi bus ZE"),
        ("accucapaciteit_taxibus_ze_kwh", 200, "kWh", "Accucapaciteit Taxibus ZE"),
    ]
    for name, val, unit, desc in battery:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    # --- ZE range ---
    r = add_section_header(ws, r, "Actieradius ZE bussen", 4) + 1
    ze_range = [
        ("actieradius_dubbeldekker_ze_km", 250, "km", "Actieradius Dubbeldekker ZE"),
        ("actieradius_touringcar_ze_km", 300, "km", "Actieradius Touringcar ZE"),
        ("actieradius_lagevloer_ze_km", 280, "km", "Actieradius Lagevloer ZE"),
        ("actieradius_midibus_ze_km", 350, "km", "Actieradius Midi bus ZE"),
        ("actieradius_taxibus_ze_km", 400, "km", "Actieradius Taxibus ZE"),
    ]
    for name, val, unit, desc in ze_range:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    # --- ZE charging times ---
    r = add_section_header(ws, r, "Laadtijden ZE bussen", 4) + 1
    charging = [
        ("laadtijd_ze_snelladen_min", 30, "minuten", "Snelladen (150kW) voor ~100km bereik"),
        ("laadtijd_ze_normaal_uren", 6, "uren", "Normaal laden (22kW AC) voor vol"),
        ("min_laadtijd_nuttig_min", 10, "minuten", "Minimale laadtijd om te stoppen"),
    ]
    for name, val, unit, desc in charging:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    for c, w in enumerate([48, 14, 16, 55], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


# ---------------------------------------------------------------------------
# Sheet 4: Duurzaamheid
# ---------------------------------------------------------------------------

def create_duurzaamheid_sheet(wb):
    """Sustainability KPIs and incentives."""
    ws = wb.create_sheet("Duurzaamheid")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2
    r = add_section_header(ws, r, "Zero-emissie en HVO100 vergoedingen", 4) + 1
    incentives = [
        ("zero_emissie_stimulans_eur_per_km", 0.12, "EUR/km", "ZE stimulans per km"),
        ("hvo_stimulans_eur_per_liter", 0.05, "EUR/liter", "HVO100 stimulans per liter"),
        ("hvo_max_total_eur_per_liter", 0.40, "EUR/liter", "Max HVO vergoeding"),
    ]
    for name, val, unit, desc in incentives:
        r = write_var(ws, r, name, val, unit, desc)

    r = add_section_header(ws, r, "Duurzame brandstof KPI bodems", 4) + 1
    kpi = [
        ("KPI_DuurzameBrandstof_bodem_jaar1_pct", 35, "%", "Minimum % duurzaam jaar 1"),
        ("KPI_DuurzameBrandstof_bodem_jaar4_pct", 50, "%", "Minimum % duurzaam jaar 4"),
        ("KPI_DuurzameBrandstof_bodem_jaar8_pct", 75, "%", "Minimum % duurzaam jaar 8"),
    ]
    for name, val, unit, desc in kpi:
        r = write_var(ws, r, name, val, unit, desc)

    for c, w in enumerate([48, 14, 12, 50], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


# ---------------------------------------------------------------------------
# Sheet 5: Brandstofprijzen
# ---------------------------------------------------------------------------

def create_brandstofprijzen_sheet(wb):
    """Fuel and electricity prices (placeholder for auto-fetch)."""
    ws = wb.create_sheet("Brandstofprijzen")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2
    r = add_section_header(ws, r, "Brandstofprijzen (automatisch bij te werken)", 4) + 1
    prices = [
        ("diesel_b7_pompprijs_eur_per_liter", None, "EUR/liter", "Diesel B7 pompprijs"),
        ("hvo100_prijs_eur_per_liter", None, "EUR/liter", "HVO100 prijs"),
        ("elektriciteit_snelladen_eur_per_kwh", None, "EUR/kWh", "Elektriciteit snelladen"),
        ("elektriciteit_normaal_eur_per_kwh", None, "EUR/kWh", "Elektriciteit normaal laden"),
    ]
    for name, val, unit, desc in prices:
        r = write_var(ws, r, name, val, unit, desc, fill=AUTO_FILL)

    r += 1
    r = add_section_header(ws, r, "Handmatige prijzen (indien API niet beschikbaar)", 4) + 1
    manual = [
        ("diesel_b7_handmatig_eur_per_liter", 1.65, "EUR/liter", "Handmatige diesel prijs"),
        ("hvo100_handmatig_eur_per_liter", 1.95, "EUR/liter", "Handmatige HVO100 prijs"),
        ("elektriciteit_handmatig_eur_per_kwh", 0.35, "EUR/kWh", "Handmatige stroomprijs"),
    ]
    for name, val, unit, desc in manual:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    for c, w in enumerate([48, 14, 12, 50], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


# ---------------------------------------------------------------------------
# Sheet 6 & 7: Tanklocaties & Laadstations (from JSON)
# ---------------------------------------------------------------------------

def create_tanklocaties_sheet(wb, tanklocaties_data: dict):
    """Fuel stations per bus station."""
    ws = wb.create_sheet("Tanklocaties")
    headers = ["Busstation", "Tankstation", "Merk", "Afstand (km)",
               "Diesel", "HVO100", "Adres"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    r = 2
    for station_name, data in sorted(tanklocaties_data.get("stations", {}).items()):
        fuel_stations = data.get("fuel_stations", [])
        if not fuel_stations:
            ws.cell(row=r, column=1, value=station_name)
            ws.cell(row=r, column=2, value="(geen tankstations gevonden)")
            style_data(ws, r, len(headers))
            r += 1
            continue

        for fs in fuel_stations[:5]:  # Top 5 nearest
            ws.cell(row=r, column=1, value=station_name)
            ws.cell(row=r, column=2, value=fs.get("name", ""))
            ws.cell(row=r, column=3, value=fs.get("brand", ""))
            ws.cell(row=r, column=4, value=fs.get("distance_km", ""))
            ws.cell(row=r, column=5, value="Ja" if fs.get("has_diesel") else "")
            ws.cell(row=r, column=6, value="Ja" if fs.get("has_hvo100") else "")
            ws.cell(row=r, column=7, value=fs.get("address", ""))
            style_data(ws, r, len(headers))
            r += 1

    for c, w in enumerate([20, 30, 15, 12, 8, 8, 40], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


def create_laadstations_sheet(wb, tanklocaties_data: dict):
    """Charging stations per bus station."""
    ws = wb.create_sheet("Laadstations")
    headers = ["Busstation", "Laadstation", "Operator", "Afstand (km)",
               "Vermogen (kW)", "Categorie", "Aantal punten"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, len(headers))

    r = 2
    for station_name, data in sorted(tanklocaties_data.get("stations", {}).items()):
        charging_stations = data.get("charging_stations", [])
        if not charging_stations:
            ws.cell(row=r, column=1, value=station_name)
            ws.cell(row=r, column=2, value="(geen laadstations gevonden)")
            style_data(ws, r, len(headers))
            r += 1
            continue

        for cs in charging_stations[:5]:  # Top 5 nearest
            ws.cell(row=r, column=1, value=station_name)
            ws.cell(row=r, column=2, value=cs.get("name", ""))
            ws.cell(row=r, column=3, value=cs.get("operator", ""))
            ws.cell(row=r, column=4, value=cs.get("distance_km", ""))
            ws.cell(row=r, column=5, value=cs.get("max_power_kw", ""))
            ws.cell(row=r, column=6, value=cs.get("category", ""))
            ws.cell(row=r, column=7, value=cs.get("num_points", ""))
            style_data(ws, r, len(headers))
            r += 1

    for c, w in enumerate([20, 35, 20, 12, 14, 12, 14], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Create additional_inputs.xlsx — consolidated input file for bus roster optimizer"
    )
    parser.add_argument("-o", "--output", default="additional_inputs.xlsx",
                        help="Output filename (default: additional_inputs.xlsx)")
    parser.add_argument("--tanklocaties", default=None,
                        help="JSON file with fuel/charging stations (from fetch_tanklocaties.py)")
    args = parser.parse_args()

    print("=" * 60)
    print("Creating additional_inputs.xlsx")
    print("=" * 60)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create core sheets
    print("  [1/5] Tarieven...")
    create_tarieven_sheet(wb)

    print("  [2/5] Chauffeurkosten...")
    create_chauffeurkosten_sheet(wb)

    print("  [3/5] Buskosten (incl. tankinhoud, laadtijden)...")
    create_buskosten_sheet(wb)

    print("  [4/5] Duurzaamheid...")
    create_duurzaamheid_sheet(wb)

    print("  [5/5] Brandstofprijzen...")
    create_brandstofprijzen_sheet(wb)

    # Optional: load tanklocaties.json
    tanklocaties_data = {}
    if args.tanklocaties:
        tanklocaties_path = Path(args.tanklocaties)
        if tanklocaties_path.exists():
            print(f"\n  Loading {args.tanklocaties}...")
            with open(tanklocaties_path, "r") as f:
                tanklocaties_data = json.load(f)
            print(f"    {len(tanklocaties_data.get('stations', {}))} bus stations found")
        else:
            print(f"\n  Warning: {args.tanklocaties} not found, skipping station sheets")
    else:
        # Try default location
        default_path = Path("tanklocaties.json")
        if default_path.exists():
            print(f"\n  Found tanklocaties.json, loading...")
            with open(default_path, "r") as f:
                tanklocaties_data = json.load(f)
            print(f"    {len(tanklocaties_data.get('stations', {}))} bus stations found")

    if tanklocaties_data:
        print("  [+] Tanklocaties...")
        create_tanklocaties_sheet(wb, tanklocaties_data)
        print("  [+] Laadstations...")
        create_laadstations_sheet(wb, tanklocaties_data)
    else:
        print("\n  No tanklocaties.json found - run fetch_tanklocaties.py first")
        print("  to add fuel/charging station data to the Excel file.")

    # Save
    wb.save(args.output)
    print(f"\n  Saved: {args.output}")
    print("\nYellow cells = manual input (edit as needed)")
    print("Green cells = auto-updated by update script")
    print("=" * 60)


if __name__ == "__main__":
    main()
