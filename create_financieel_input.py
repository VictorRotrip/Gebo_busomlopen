#!/usr/bin/env python3
"""
create_financieel_input.py

Creates financieel_input.xlsx — the focused financial input file for
bus roster versions 6-9. Contains only the essential variables needed
for profit calculation, plus a sheet for auto-fetched fuel prices.

Run once to create the template; thereafter use update_financieel_input.py
to fetch latest fuel/electricity prices from APIs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BOLD = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
AUTO_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
MANUAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        if fill and c == 2:  # value column
            cell.fill = fill


def add_section_header(ws, row, title, cols):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, size=11, color="2F5496")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).fill = PatternFill(
            start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
        )
    return row


def write_var(ws, row, name, value, unit, description, fill=None):
    ws.cell(row=row, column=1, value=name)
    cell_val = ws.cell(row=row, column=2, value=value)
    ws.cell(row=row, column=3, value=unit)
    ws.cell(row=row, column=4, value=description)
    style_data(ws, row, 4, fill=fill)
    return row + 1


def create_tarieven_sheet(wb):
    """Sheet 1: Revenue rates per bus type (from Prijzenblad)."""
    ws = wb.create_sheet("Tarieven")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2
    r = add_section_header(ws, r, "Uurtarieven per bustype (omzet = inzet-uren × tarief)", 4) + 1
    r = write_var(ws, r, "rate_dubbeldekker_eur_per_hour", 116.37, "EUR/uur",
                  "Uurtarief Dubbeldekker (incl. km + chauffeur)")
    r = write_var(ws, r, "rate_touringcar_eur_per_hour", 80.455, "EUR/uur",
                  "Uurtarief Touringcar")
    r = write_var(ws, r, "rate_lagevloer_gelede_eur_per_hour", 80.445, "EUR/uur",
                  "Uurtarief Lagevloer / Gelede bus")
    r = write_var(ws, r, "rate_midibus_eur_per_hour", 74.85, "EUR/uur",
                  "Uurtarief Midi bus")
    r = write_var(ws, r, "rate_taxibus_eur_per_hour", 50.455, "EUR/uur",
                  "Uurtarief Taxibus")

    for c in [1, 2, 3, 4]:
        ws.column_dimensions[get_column_letter(c)].width = [42, 14, 12, 60][c - 1]
    return ws


def create_chauffeurkosten_sheet(wb):
    """Sheet 2: Driver costs from CAO BB."""
    ws = wb.create_sheet("Chauffeurkosten")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2

    # --- Pauzestaffel ---
    r = add_section_header(ws, r, "Pauzestaffel (pauze-aftrek op betaalde uren per diensttijd)", 4) + 1
    staffel = [
        ("Pauzestaffel_1_max_diensttijd_uren", 4.5, "uren", "Bracket 1: diensttijd <= 4.5h"),
        ("Pauzestaffel_1_pauze_min", 0, "minuten", "Pauze-aftrek: 0 min"),
        ("Pauzestaffel_2_max_diensttijd_uren", 7.5, "uren", "Bracket 2: diensttijd <= 7.5h"),
        ("Pauzestaffel_2_pauze_min", 30, "minuten", "Pauze-aftrek: 30 min"),
        ("Pauzestaffel_3_max_diensttijd_uren", 10.5, "uren", "Bracket 3: diensttijd <= 10.5h"),
        ("Pauzestaffel_3_pauze_min", 60, "minuten", "Pauze-aftrek: 60 min"),
        ("Pauzestaffel_4_max_diensttijd_uren", 13.5, "uren", "Bracket 4: diensttijd <= 13.5h"),
        ("Pauzestaffel_4_pauze_min", 90, "minuten", "Pauze-aftrek: 90 min"),
        ("Pauzestaffel_5_max_diensttijd_uren", 16.5, "uren", "Bracket 5: diensttijd <= 16.5h"),
        ("Pauzestaffel_5_pauze_min", 120, "minuten", "Pauze-aftrek: 120 min"),
        ("Pauzestaffel_6_min_diensttijd_uren", 16.5, "uren", "Bracket 6: diensttijd >= 16.5h"),
        ("Pauzestaffel_6_pauze_min", 150, "minuten", "Pauze-aftrek: 150 min"),
    ]
    for name, val, unit, desc in staffel:
        r = write_var(ws, r, name, val, unit, desc)

    # --- OV onbetaalde onderbreking ---
    r = add_section_header(ws, r, "Onbetaalde onderbreking (OV-regel)", 4) + 1
    r = write_var(ws, r, "OV_onbetaalde_onderbreking_max_uren", 1, "uren",
                  "Max 1h onbetaalde onderbreking per dienst (OV)")
    r = write_var(ws, r, "OV_onbetaalde_onderbreking_max_aantal_per_dienst", 1, "getal",
                  "Max 1x per dienst")

    # --- Nachtwerk ---
    r = add_section_header(ws, r, "Nachtwerk", 4) + 1
    r = write_var(ws, r, "Nachtwerk_window_start_min", 60, "min na middernacht",
                  "Nachtwerk-window start (01:00)")
    r = write_var(ws, r, "Nachtwerk_window_end_min", 300, "min na middernacht",
                  "Nachtwerk-window einde (05:00)")
    r = write_var(ws, r, "Max_arbeidstijd_per_24h_bij_nachtwerk_uren", 12, "uren",
                  "Max 12h arbeidstijd als dienst nachtwerk-window raakt")

    # --- ORT windows ---
    r = add_section_header(ws, r, "ORT tijdvensters", 4) + 1
    r = write_var(ws, r, "ORT_OV_ma_vr_window_start_min", 1140, "min na middernacht",
                  "Ma-Vr ORT window start (19:00)")
    r = write_var(ws, r, "ORT_OV_ma_vr_window_end_min", 450, "min na middernacht",
                  "Ma-Vr ORT window einde (07:30, over middernacht)")
    r = write_var(ws, r, "ORT_00_06_window_start_min", 0, "min na middernacht",
                  "Nacht ORT window start (00:00)")
    r = write_var(ws, r, "ORT_00_06_window_end_min", 360, "min na middernacht",
                  "Nacht ORT window einde (06:00)")
    r = write_var(ws, r, "ORT_touringcar_nacht_doordeweeks_window_start_min", 0, "min na middernacht",
                  "Touringcar nacht ORT start (00:00)")
    r = write_var(ws, r, "ORT_touringcar_nacht_doordeweeks_window_end_min", 360, "min na middernacht",
                  "Touringcar nacht ORT einde (06:00)")

    # --- ORT rates OV ---
    r = add_section_header(ws, r, "ORT tarieven OV (niet-touringcar)", 4) + 1
    ort_ov = [
        ("ORT_OV_ma_vr_19_00_07_30_EUR_per_uur_20250101", 4.80, "Ma-Vr 19:00-07:30 (v.a. 2025-01-01)"),
        ("ORT_OV_ma_vr_19_00_07_30_EUR_per_uur_20250701", 4.99, "Ma-Vr 19:00-07:30 (v.a. 2025-07-01)"),
        ("ORT_OV_ma_vr_19_00_07_30_EUR_per_uur_20260101", 5.19, "Ma-Vr 19:00-07:30 (v.a. 2026-01-01)"),
        ("ORT_OV_zaterdag_EUR_per_uur_20250101", 4.45, "Zaterdag (v.a. 2025-01-01)"),
        ("ORT_OV_zaterdag_EUR_per_uur_20250701", 4.63, "Zaterdag (v.a. 2025-07-01)"),
        ("ORT_OV_zaterdag_EUR_per_uur_20260101", 4.81, "Zaterdag (v.a. 2026-01-01)"),
        ("ORT_OV_zon_feest_EUR_per_uur_20250101", 6.18, "Zon/feestdag (v.a. 2025-01-01)"),
        ("ORT_OV_zon_feest_EUR_per_uur_20250701", 6.43, "Zon/feestdag (v.a. 2025-07-01)"),
        ("ORT_OV_zon_feest_EUR_per_uur_20260101", 6.68, "Zon/feestdag (v.a. 2026-01-01)"),
    ]
    for name, val, desc in ort_ov:
        r = write_var(ws, r, name, val, "EUR/uur", desc)

    # --- ORT rates Touringcar ---
    r = add_section_header(ws, r, "ORT tarieven Touringcar", 4) + 1
    ort_tc = [
        ("ORT_touringcar_zaterdag_EUR_per_uur_20250101", 3.71, "TC zaterdag (v.a. 2025-01-01)"),
        ("ORT_touringcar_zaterdag_EUR_per_uur_20250701", 3.86, "TC zaterdag (v.a. 2025-07-01)"),
        ("ORT_touringcar_zaterdag_EUR_per_uur_20260101", 4.01, "TC zaterdag (v.a. 2026-01-01)"),
        ("ORT_touringcar_zon_feest_EUR_per_uur_20250101", 5.58, "TC zon/feestdag (v.a. 2025-01-01)"),
        ("ORT_touringcar_zon_feest_EUR_per_uur_20250701", 5.80, "TC zon/feestdag (v.a. 2025-07-01)"),
        ("ORT_touringcar_zon_feest_EUR_per_uur_20260101", 6.04, "TC zon/feestdag (v.a. 2026-01-01)"),
        ("ORT_touringcar_nacht_doordeweeks_00_00_06_00_EUR_per_uur_20250101", 3.71,
         "TC doordeweeks 00:00-06:00 (v.a. 2025-01-01)"),
        ("ORT_touringcar_nacht_doordeweeks_00_00_06_00_EUR_per_uur_20250701", 3.86,
         "TC doordeweeks 00:00-06:00 (v.a. 2025-07-01)"),
        ("ORT_touringcar_nacht_doordeweeks_00_00_06_00_EUR_per_uur_20260101", 4.01,
         "TC doordeweeks 00:00-06:00 (v.a. 2026-01-01)"),
    ]
    for name, val, desc in ort_tc:
        r = write_var(ws, r, name, val, "EUR/uur", desc)

    # --- Overuren ---
    r = add_section_header(ws, r, "Overuren", 4) + 1
    overuren = [
        ("Jaarurenregeling_overuren_drempel_maand_uren", 173.33, "uren",
         "Overurendrempel per maand"),
        ("Jaarurenregeling_overuren_drempel_4weken_uren", 160, "uren",
         "Overurendrempel per 4-weken periode"),
        ("Jaarurenregeling_overuren_toeslag_pct", 0.35, "fractie",
         "Overurentoeslag (35%)"),
        ("Overuren_toeslag_rijdend_alles_pct", 0.35, "fractie",
         "Toeslag rijdende overuren (35%)"),
        ("Overuren_toeslag_niet_rijdend_werkdag_pct", 0.30, "fractie",
         "Toeslag niet-rijdende overuren werkdag (30%)"),
        ("Overuren_toeslag_niet_rijdend_zaterdag_pct", 0.50, "fractie",
         "Toeslag niet-rijdende overuren zaterdag (50%)"),
        ("Overuren_toeslag_niet_rijdend_zon_feest_pct", 1.00, "fractie",
         "Toeslag niet-rijdende overuren zon/feest (100%)"),
    ]
    for name, val, unit, desc in overuren:
        r = write_var(ws, r, name, val, unit, desc)

    # --- Onderbrekingstoeslag ---
    r = add_section_header(ws, r, "Onderbrekingstoeslag (gebroken dienst)", 4) + 1
    onderbreking = [
        ("Onderbrekingstoeslag_per_dienst_EUR_20250101", 14.72, "EUR/dienst",
         "Toeslag gebroken dienst (v.a. 2025-01-01)"),
        ("Onderbrekingstoeslag_per_dienst_EUR_20250701", 15.31, "EUR/dienst",
         "Toeslag gebroken dienst (v.a. 2025-07-01)"),
        ("Onderbrekingstoeslag_per_dienst_EUR_20260101", 15.92, "EUR/dienst",
         "Toeslag gebroken dienst (v.a. 2026-01-01)"),
        ("Onderbrekingstoeslag_bus_mee_naar_huis_EUR_20250101", 9.32, "EUR/dienst",
         "Gereduceerd tarief bus mee naar huis (v.a. 2025-01-01)"),
        ("Onderbrekingstoeslag_bus_mee_naar_huis_EUR_20250701", 9.69, "EUR/dienst",
         "Gereduceerd tarief bus mee naar huis (v.a. 2025-07-01)"),
        ("Onderbrekingstoeslag_bus_mee_naar_huis_EUR_20260101", 10.08, "EUR/dienst",
         "Gereduceerd tarief bus mee naar huis (v.a. 2026-01-01)"),
        ("Onderbrekingstoeslag_drempel_onderbreking_uren", 1, "uren",
         "Onderbreking moet >= 1h zijn om te tellen"),
        ("Onderbrekingstoeslag_geen_toeslag_bij_onderbreking_ge_uren", 8, "uren",
         "Geen toeslag als onderbreking >= 8h"),
        ("Onderbrekingstoeslag_diensttijd_grens_uren", 12, "uren",
         "Grens voor max aantal onderbrekingstoeslagen"),
        ("Onderbrekingstoeslag_max_aantal_bij_diensttijd_leq_12", 1, "getal",
         "Max 1 toeslag als diensttijd <= 12h"),
        ("Onderbrekingstoeslag_max_aantal_bij_diensttijd_gt_12", 2, "getal",
         "Max 2 toeslagen als diensttijd > 12h"),
    ]
    for name, val, unit, desc in onderbreking:
        r = write_var(ws, r, name, val, unit, desc)

    # --- Maaltijdvergoeding ---
    r = add_section_header(ws, r, "Maaltijdvergoeding", 4) + 1
    maaltijd = [
        ("Maaltijdvergoeding_drempel_diensttijd_ge_11u_uren", 11, "uren",
         "Maaltijdvergoeding vanaf >= 11h diensttijd"),
        ("Maaltijdvergoeding_drempel_diensttijd_ge_14u_uren", 14, "uren",
         "Hogere vergoeding vanaf >= 14h"),
        ("Maaltijdvergoeding_max_bij_diensttijd_ge_11u_EUR_20250101", 22.50, "EUR",
         "Max maaltijdvergoeding >= 11h (2025)"),
        ("Maaltijdvergoeding_max_bij_diensttijd_ge_11u_EUR_20260101", 22.95, "EUR",
         "Max maaltijdvergoeding >= 11h (2026)"),
        ("Maaltijdvergoeding_max_bij_diensttijd_ge_14u_EUR_20250101", 36.00, "EUR",
         "Max maaltijdvergoeding >= 14h (2025)"),
        ("Maaltijdvergoeding_max_bij_diensttijd_ge_14u_EUR_20260101", 36.72, "EUR",
         "Max maaltijdvergoeding >= 14h (2026)"),
    ]
    for name, val, unit, desc in maaltijd:
        r = write_var(ws, r, name, val, unit, desc)

    # --- Vakantietoeslag ---
    r = add_section_header(ws, r, "Vakantietoeslag", 4) + 1
    r = write_var(ws, r, "Vakantietoeslag_pct", 0.08, "fractie",
                  "8% vakantietoeslag over bruto loonkosten")

    # --- Basis uurloon (manual input) ---
    r = add_section_header(ws, r, "Basis uurloon chauffeur (handmatig in te vullen)", 4) + 1
    r = write_var(ws, r, "basis_uurloon_chauffeur_eur", 18.50, "EUR/uur",
                  "Bruto basisuurloon chauffeur (uit CAO loonschalen)", fill=MANUAL_FILL)
    r = write_var(ws, r, "werkgever_opslag_factor", 1.35, "factor",
                  "Werkgeverslasten bovenop bruto loon (pensioen, soc. premies)", fill=MANUAL_FILL)

    for c in [1, 2, 3, 4]:
        ws.column_dimensions[get_column_letter(c)].width = [55, 14, 18, 55][c - 1]
    return ws


def create_buskosten_sheet(wb):
    """Sheet 3: Bus operating costs (fuel consumption, manual input)."""
    ws = wb.create_sheet("Buskosten")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2
    r = add_section_header(ws, r, "Brandstofverbruik per bustype (handmatig in te vullen)", 4) + 1

    consumption = [
        ("verbruik_dubbeldekker_diesel_l_per_100km", 45, "L/100km",
         "Dieselverbruik Dubbeldekker"),
        ("verbruik_touringcar_diesel_l_per_100km", 32, "L/100km",
         "Dieselverbruik Touringcar"),
        ("verbruik_lagevloer_diesel_l_per_100km", 38, "L/100km",
         "Dieselverbruik Lagevloer/Gelede"),
        ("verbruik_midibus_diesel_l_per_100km", 25, "L/100km",
         "Dieselverbruik Midi bus"),
        ("verbruik_taxibus_diesel_l_per_100km", 12, "L/100km",
         "Dieselverbruik Taxibus"),
    ]
    for name, val, unit, desc in consumption:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    r = add_section_header(ws, r, "Elektrisch verbruik per bustype (voor ZE variant)", 4) + 1
    electric = [
        ("verbruik_dubbeldekker_ze_kwh_per_100km", 180, "kWh/100km",
         "Elektriciteitsverbruik Dubbeldekker ZE"),
        ("verbruik_touringcar_ze_kwh_per_100km", 130, "kWh/100km",
         "Elektriciteitsverbruik Touringcar ZE"),
        ("verbruik_lagevloer_ze_kwh_per_100km", 150, "kWh/100km",
         "Elektriciteitsverbruik Lagevloer ZE"),
        ("verbruik_midibus_ze_kwh_per_100km", 100, "kWh/100km",
         "Elektriciteitsverbruik Midi bus ZE"),
        ("verbruik_taxibus_ze_kwh_per_100km", 50, "kWh/100km",
         "Elektriciteitsverbruik Taxibus ZE"),
    ]
    for name, val, unit, desc in electric:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    r = add_section_header(ws, r, "Actieradius ZE bussen (voor range-constraint)", 4) + 1
    range_data = [
        ("actieradius_dubbeldekker_ze_km", 250, "km", "Actieradius Dubbeldekker ZE"),
        ("actieradius_touringcar_ze_km", 300, "km", "Actieradius Touringcar ZE"),
        ("actieradius_lagevloer_ze_km", 280, "km", "Actieradius Lagevloer ZE"),
        ("actieradius_midibus_ze_km", 350, "km", "Actieradius Midi bus ZE"),
        ("actieradius_taxibus_ze_km", 400, "km", "Actieradius Taxibus ZE"),
    ]
    for name, val, unit, desc in range_data:
        r = write_var(ws, r, name, val, unit, desc, fill=MANUAL_FILL)

    r = add_section_header(ws, r, "Laadtijd ZE bussen", 4) + 1
    r = write_var(ws, r, "laadtijd_ze_uren_snelladen", 1.0, "uren",
                  "Snellaadtijd (DC, ~150kW) voor 80% opladen", fill=MANUAL_FILL)
    r = write_var(ws, r, "laadtijd_ze_uren_normaal", 6.0, "uren",
                  "Normale laadtijd (AC, ~22kW) voor vol opladen", fill=MANUAL_FILL)

    for c in [1, 2, 3, 4]:
        ws.column_dimensions[get_column_letter(c)].width = [48, 14, 16, 50][c - 1]
    return ws


def create_duurzaamheid_sheet(wb):
    """Sheet 4: Sustainability incentives, KPIs, and malus."""
    ws = wb.create_sheet("Duurzaamheid")
    headers = ["Variabele", "Waarde", "Eenheid", "Toelichting"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 4)

    r = 2

    # --- ZE / HVO incentives ---
    r = add_section_header(ws, r, "Zero-emissie en HVO100 vergoedingen", 4) + 1
    incentives = [
        ("zero_emissie_stimulans_eur_per_km", 0.12, "EUR/km",
         "Bonus per km gereden met zero-emissie voertuig"),
        ("hvo_price_diff_threshold_eur_per_liter", 0, "EUR/L",
         "Drempel: als HVO-B7 verschil <= 0, geen vergoeding"),
        ("hvo_price_diff_max_eur_per_liter", 0.35, "EUR/L",
         "Cap op prijsverschilvergoeding HVO100 vs B7"),
        ("hvo_stimulans_eur_per_liter", 0.05, "EUR/L",
         "Extra stimulans per liter HVO100 (als verschil > 0)"),
        ("hvo_max_total_eur_per_liter", 0.40, "EUR/L",
         "Max totale HVO vergoeding (verschil + stimulans)"),
    ]
    for name, val, unit, desc in incentives:
        r = write_var(ws, r, name, val, unit, desc)

    # --- Sustainability fuel KPI targets ---
    r = add_section_header(ws, r, "Duurzame brandstof KPI bodems per contractjaar", 4) + 1
    fuel_kpi = [
        ("KPI_DuurzameBrandstof_bodem_jaar1_pct", 35, "%",
         "Min aandeel duurzame brandstof jaar 1 (2027)"),
        ("KPI_DuurzameBrandstof_bodem_jaar2_pct", 40, "%",
         "Min aandeel duurzame brandstof jaar 2 (2028)"),
        ("KPI_DuurzameBrandstof_bodem_jaar3_pct", 45, "%",
         "Min aandeel duurzame brandstof jaar 3 (2029)"),
        ("KPI_DuurzameBrandstof_bodem_jaar4_pct", 50, "%",
         "Min aandeel duurzame brandstof jaar 4 (2030)"),
        ("KPI_DuurzameBrandstof_bodem_jaar8_pct", 75, "%",
         "Min aandeel duurzame brandstof jaar 8 (2034)"),
        ("KPI_DuurzameBrandstof_step_pctpunt", 0.1, "%-punt",
         "Malusstap per 0.1%-punt onder bodem"),
        ("KPI_DuurzameBrandstof_malus_eur_per_0_1pctpunt", 1000, "EUR",
         "Malus EUR 1.000 per 0.1%-punt onder bodem"),
    ]
    for name, val, unit, desc in fuel_kpi:
        r = write_var(ws, r, name, val, unit, desc)

    # --- Emission norm ---
    r = add_section_header(ws, r, "Emissienorm KPI", 4) + 1
    emission = [
        ("KPI_Emissienorm_bodem_pct", 100, "%",
         "100% van km moet Euro VI of ZE zijn"),
        ("KPI_Emissienorm_step_pctpunt", 0.1, "%-punt",
         "Malusstap per 0.1%-punt onder 100%"),
        ("KPI_Emissienorm_malus_eur_per_0_1pctpunt", 1000, "EUR",
         "Malus EUR 1.000 per 0.1%-punt onder 100%"),
    ]
    for name, val, unit, desc in emission:
        r = write_var(ws, r, name, val, unit, desc)

    # --- Malus/Bonus caps ---
    r = add_section_header(ws, r, "Malus/Bonus plafonds", 4) + 1
    r = write_var(ws, r, "Plafond_malus_pct_jaaromzet", 0.01, "fractie",
                  "Max totale malus: 1% van jaaromzet")
    r = write_var(ws, r, "Plafond_bonus_pct_jaaromzet", 0.005, "fractie",
                  "Max totale bonus: 0.5% van jaaromzet")

    # --- Sustainability roadmap ---
    r = add_section_header(ws, r, "Duurzaamheid routekaart (fossiel/hernieuwbaar)", 4) + 1
    roadmap = [
        ("duurzaamheid_emissienorm_km_fractie", 1.0, "fractie",
         "100% van km met Euro VI of ZE"),
        ("duurzaamheid_emissienorm_min_euro_norm", 6, "EURO-norm",
         "Minimum Euro 6 emissieklasse"),
        ("duurzaamheid_fossiel_max_fractie_contract_start", 0.65, "fractie",
         "Max 65% fossiel bij contractstart"),
        ("duurzaamheid_fossiel_reductie_min_fractie_per_jaar_tm_2030", 0.05, "fractie/jaar",
         "Min 5pp fossiel reductie per jaar t/m 2030"),
        ("duurzaamheid_fossiel_reductie_min_fractie_per_jaar_2031_tm_2034", 0.063, "fractie/jaar",
         "Min 6.3pp fossiel reductie per jaar 2031-2034"),
        ("duurzaamheid_fossiel_max_fractie_2034", 0.25, "fractie",
         "Max 25% fossiel in 2034"),
        ("duurzaamheid_hernieuwbaar_ze_min_fractie_2027", 0.35, "fractie",
         "Min 35% hernieuwbaar/ZE in 2027"),
        ("duurzaamheid_hernieuwbaar_ze_min_fractie_2030", 0.50, "fractie",
         "Min 50% hernieuwbaar/ZE in 2030"),
        ("duurzaamheid_hernieuwbaar_ze_min_fractie_2034", 0.75, "fractie",
         "Min 75% hernieuwbaar/ZE in 2034"),
        ("duurzaamheid_hernieuwbaar_ze_groei_min_fractie_per_jaar_tm_2030", 0.05, "fractie/jaar",
         "Min 5pp groei hernieuwbaar/ZE per jaar t/m 2030"),
        ("duurzaamheid_hernieuwbaar_ze_groei_min_fractie_per_jaar_2031_tm_2034", 0.063, "fractie/jaar",
         "Min 6.3pp groei hernieuwbaar/ZE per jaar 2031-2034"),
    ]
    for name, val, unit, desc in roadmap:
        r = write_var(ws, r, name, val, unit, desc)

    for c in [1, 2, 3, 4]:
        ws.column_dimensions[get_column_letter(c)].width = [55, 14, 14, 55][c - 1]
    return ws


def create_brandstofprijzen_sheet(wb):
    """Sheet 5: Fuel prices — auto-updated by update_financieel_input.py."""
    ws = wb.create_sheet("Brandstofprijzen")
    headers = ["Variabele", "Waarde", "Eenheid", "Bron", "Laatst bijgewerkt"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, 1, 5)

    r = 2
    r = add_section_header(ws, r, "Brandstofprijzen (automatisch opgehaald)", 5) + 1

    now_str = "Nog niet opgehaald"

    prices = [
        ("diesel_b7_pompprijs_eur_per_liter", None, "EUR/L",
         "CBS OData (80416ned)", now_str),
        ("hvo100_adviesprijs_fieten_eur_per_liter", None, "EUR/L",
         "Fieten Olie adviesprijzen", now_str),
        ("hvo100_adviesprijs_pk_energy_eur_per_liter", None, "EUR/L",
         "PK Energy adviesprijzen", now_str),
        ("hvo100_adviesprijs_bp_eur_per_liter", None, "EUR/L",
         "BP Nederland (handmatig)", now_str),
        ("diesel_b7_adviesprijs_fieten_eur_per_liter", None, "EUR/L",
         "Fieten Olie adviesprijzen", now_str),
        ("diesel_b7_adviesprijs_pk_energy_eur_per_liter", None, "EUR/L",
         "PK Energy adviesprijzen", now_str),
        ("diesel_b7_adviesprijs_bp_eur_per_liter", None, "EUR/L",
         "BP Nederland adviesprijzen", now_str),
    ]
    for name, val, unit, source, updated in prices:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=source)
        ws.cell(row=r, column=5, value=updated)
        style_data(ws, r, 5, fill=AUTO_FILL)
        r += 1

    r = add_section_header(ws, r, "Berekend HVO100 prijsverschil (contractformule)", 5) + 1
    calc_rows = [
        ("hvo100_b7_verschil_gemiddeld_eur_per_liter", None, "EUR/L",
         "Gem. (Fieten + PK + BP)", now_str),
        ("hvo100_vergoeding_per_liter", None, "EUR/L",
         "min(verschil, 0.35) + 0.05 of 0", now_str),
    ]
    for name, val, unit, source, updated in calc_rows:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=source)
        ws.cell(row=r, column=5, value=updated)
        style_data(ws, r, 5, fill=AUTO_FILL)
        r += 1

    r = add_section_header(ws, r, "Elektriciteitsprijs (automatisch opgehaald)", 5) + 1
    elec = [
        ("elektriciteit_prijs_eur_per_kwh", None, "EUR/kWh",
         "EnergyZero API (EPEX spot incl. BTW)", now_str),
        ("elektriciteit_prijs_gemiddeld_24h_eur_per_kwh", None, "EUR/kWh",
         "EnergyZero API (24h gemiddelde)", now_str),
    ]
    for name, val, unit, source, updated in elec:
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=unit)
        ws.cell(row=r, column=4, value=source)
        ws.cell(row=r, column=5, value=updated)
        style_data(ws, r, 5, fill=AUTO_FILL)
        r += 1

    for c in [1, 2, 3, 4, 5]:
        ws.column_dimensions[get_column_letter(c)].width = [50, 14, 12, 40, 22][c - 1]
    return ws


def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    create_tarieven_sheet(wb)
    create_chauffeurkosten_sheet(wb)
    create_buskosten_sheet(wb)
    create_duurzaamheid_sheet(wb)
    create_brandstofprijzen_sheet(wb)

    output_path = "financieel_input.xlsx"
    wb.save(output_path)
    print(f"Created {output_path} with sheets:")
    for name in wb.sheetnames:
        print(f"  - {name}")
    print()
    print("Legend:")
    print("  Yellow cells  = manual input needed (estimates/fleet data)")
    print("  Green cells   = auto-updated by update_financieel_input.py")
    print()
    print("Next step: run 'python update_financieel_input.py' to fetch latest fuel prices.")


if __name__ == "__main__":
    main()
