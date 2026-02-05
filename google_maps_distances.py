"""
Google Maps Distance Matrix - Deadhead & Trip Validation
=========================================================
Fetches driving distances/times between all bus station locations using the
Google Maps Distance Matrix API. Produces:

1. A deadhead matrix (travel time in minutes between all station pairs)
2. A validation report comparing scheduled trip durations to Google Maps times
3. A traffic-aware risk analysis (per time slot: rush, off-peak, night, weekend)

Stations are automatically discovered from the input Excel file (same file
used by busomloop_optimizer.py), so this script works with any input file
without hardcoded station lists.

Usage:
    python google_maps_distances.py --input casus.xlsx --key API_KEY
    python google_maps_distances.py --input casus.xlsx --traffic   # fetch 6 time-slot matrices
    python google_maps_distances.py --input casus.xlsx --verify    # check addresses first
    python google_maps_distances.py --input casus.xlsx             # uses key from .env
    python google_maps_distances.py --from-cache deadhead_matrix.json --validate output.xlsx

API key can be provided via --key or in a .env file:
    GOOGLE_MAPS_API_KEY=AIza...
"""

from __future__ import annotations

import argparse
import datetime
import json
import os
import sys
import time
from pathlib import Path

import requests

API_URL = "https://maps.googleapis.com/maps/api/distancematrix/json"
GEOCODE_URL = "https://maps.googleapis.com/maps/api/geocode/json"

# Time slot definitions for traffic-aware matrix fetching.
# Ranges are in minutes from midnight.
TIME_SLOTS = {
    "nacht":        {"range": (0, 360),    "fetch_hour": 2,  "label": "Nacht (00:00-06:00)"},
    "ochtendspits": {"range": (360, 570),  "fetch_hour": 8,  "label": "Ochtendspits (06:00-09:30)"},
    "dal":          {"range": (570, 930),  "fetch_hour": 12, "label": "Daluren (09:30-15:30)"},
    "middagspits":  {"range": (930, 1110), "fetch_hour": 17, "label": "Middagspits (15:30-18:30)"},
    "avond":        {"range": (1110, 1440),"fetch_hour": 21, "label": "Avond (18:30-00:00)"},
    "weekend":      {"range": (0, 1440),   "fetch_hour": 12, "label": "Weekend (hele dag)"},
}


def load_dotenv():
    """Load .env file from the script's directory into os.environ."""
    env_path = Path(__file__).parent / ".env"
    if not env_path.exists():
        return
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                key, _, value = line.partition("=")
                key = key.strip()
                value = value.strip().strip("'\"")
                os.environ.setdefault(key, value)


def build_station_addresses(registry: dict, halts: dict = None) -> dict:
    """Build Google Maps searchable addresses from the station registry.

    Uses halt/stop info (halteplaats) from the input data when available,
    giving much more precise locations than just "Station X".

    For example:
      - "Veenendaal-De Klomp" + halt "OV busstation, Dokter Hoolboomweg"
        -> "Veenendaal-De Klomp OV busstation, Dokter Hoolboomweg"
      - "Breukelen" + halt "OV busstation, Stationsweg"
        -> "Breukelen OV busstation, Stationsweg"

    registry: {canonical_key: display_name} from build_station_registry().
    halts:    {canonical_key: set of halt names} from get_station_halts().
    Returns:  {canonical_key: Google Maps search query string}
    """
    halts = halts or {}
    addresses = {}
    for canonical, display in registry.items():
        halt_set = halts.get(canonical, set())
        if halt_set:
            # Pick the most descriptive halt (prefer ones with street/location info)
            halt = _pick_best_halt(halt_set)
            addresses[canonical] = f"{display} {halt}, Nederland"
        else:
            # Fallback: generic station search
            addresses[canonical] = f"Station {display}, Nederland"
    return addresses


def _pick_best_halt(halt_set: set) -> str:
    """Pick the most useful halt name for Google Maps geocoding.

    Prefers halts with street names or specific location info.
    Filters out direction-specific suffixes ("richting X") since those
    are the same physical location.
    """
    import re

    # Normalize: strip direction info ("richting ...")
    cleaned = set()
    for h in halt_set:
        # Remove "richting <anything>" at the end
        base = re.sub(r'\s+richting\s+.*$', '', h, flags=re.IGNORECASE).strip()
        cleaned.add(base)

    # Prefer halts with comma (usually have street name: "OV busstation, Stationsweg")
    with_street = [h for h in cleaned if ',' in h]
    if with_street:
        # Pick the longest one (most descriptive)
        return max(with_street, key=len)

    # Otherwise pick the longest halt name
    return max(cleaned, key=len) if cleaned else ""


def normalize_display_name(name: str) -> str:
    """Convert a display station name to canonical location (lowercase)."""
    return name.strip().lower()


def fetch_distance_matrix(api_key: str, locations: list[str],
                          addresses: dict, batch_size: int = 10,
                          departure_time: int = None) -> dict:
    """Fetch full NxN distance matrix from Google Maps.

    locations: list of canonical location keys.
    addresses: dict {canonical_key: Google Maps address string}.
    departure_time: optional Unix timestamp for traffic-aware results.
                    When set, uses duration_in_traffic instead of duration.

    Returns dict: {(origin, dest): {"distance_m": int, "duration_s": int,
                                     "duration_min": float, "distance_km": float}}
    """
    addr_list = []
    for loc in locations:
        addr = addresses.get(loc)
        if not addr:
            print(f"  WAARSCHUWING: geen adres voor station '{loc}', wordt overgeslagen")
            continue
        addr_list.append(addr)

    n = len(locations)
    matrix = {}

    for i_start in range(0, n, batch_size):
        i_end = min(i_start + batch_size, n)
        origins = "|".join(addr_list[i_start:i_end])
        origin_locs = locations[i_start:i_end]

        for j_start in range(0, n, batch_size):
            j_end = min(j_start + batch_size, n)
            destinations = "|".join(addr_list[j_start:j_end])
            dest_locs = locations[j_start:j_end]

            params = {
                "origins": origins,
                "destinations": destinations,
                "key": api_key,
                "mode": "driving",
                "language": "nl",
            }
            if departure_time is not None:
                params["departure_time"] = departure_time

            print(f"  Fetching {len(origin_locs)}x{len(dest_locs)} "
                  f"({origin_locs[0]}..{origin_locs[-1]} -> "
                  f"{dest_locs[0]}..{dest_locs[-1]}) ...")

            try:
                resp = requests.get(API_URL, params=params, timeout=30)
                resp.raise_for_status()
            except requests.RequestException as e:
                print(f"  FOUT: API-verzoek mislukt: {e}")
                print(f"  Doorgaan met volgende batch...")
                for o in origin_locs:
                    for d in dest_locs:
                        matrix[(o, d)] = {
                            "distance_m": None, "duration_s": None,
                            "duration_min": None, "distance_km": None,
                        }
                continue

            data = resp.json()

            if data["status"] != "OK":
                print(f"  FOUT: API gaf status {data['status']}")
                print(f"  {data.get('error_message', '')}")
                if data["status"] == "REQUEST_DENIED":
                    print("  Controleer of de API-sleutel geldig is en Distance Matrix API "
                          "is ingeschakeld.")
                    sys.exit(1)
                # For other errors, continue with None values
                for o in origin_locs:
                    for d in dest_locs:
                        matrix[(o, d)] = {
                            "distance_m": None, "duration_s": None,
                            "duration_min": None, "distance_km": None,
                        }
                continue

            for ri, row in enumerate(data["rows"]):
                for ci, element in enumerate(row["elements"]):
                    o = origin_locs[ri]
                    d = dest_locs[ci]
                    if element["status"] == "OK":
                        # Prefer duration_in_traffic when available (traffic-aware)
                        if "duration_in_traffic" in element:
                            dur_s = element["duration_in_traffic"]["value"]
                        else:
                            dur_s = element["duration"]["value"]
                        matrix[(o, d)] = {
                            "distance_m": element["distance"]["value"],
                            "duration_s": dur_s,
                            "duration_min": round(dur_s / 60, 1),
                            "distance_km": round(element["distance"]["value"] / 1000, 1),
                        }
                    else:
                        print(f"  WAARSCHUWING: Geen route {o} -> {d}: {element['status']}")
                        matrix[(o, d)] = {
                            "distance_m": None, "duration_s": None,
                            "duration_min": None, "distance_km": None,
                        }

            time.sleep(0.2)

    return matrix


def save_deadhead_json(matrix: dict, locations: list[str], output_file: str):
    """Save deadhead matrix as JSON for the optimizer to import.

    Format: {origin: {dest: {"min": duration_min, "km": distance_km}}}
    For backward compatibility, if only duration is needed, optimizer can
    access deadhead[o][d]["min"] or check if value is a dict.
    """
    deadhead = {}
    for o in locations:
        deadhead[o] = {}
        for d in locations:
            entry = matrix.get((o, d))
            if entry and entry["duration_min"] is not None:
                deadhead[o][d] = {
                    "min": entry["duration_min"],
                    "km": entry.get("distance_km"),
                }
    with open(output_file, "w") as f:
        json.dump(deadhead, f, indent=2, ensure_ascii=False)
    print(f"Deadhead JSON saved to {output_file} (includes distance_km)")


def save_deadhead_json_from_nested(nested_dict: dict, output_file: str):
    """Save a nested {origin: {dest: minutes}} dict as deadhead JSON."""
    with open(output_file, "w") as f:
        json.dump(nested_dict, f, indent=2, ensure_ascii=False)
    print(f"Deadhead JSON saved to {output_file}")


def load_matrix_from_cache(cache_file: str) -> dict:
    """Load distance matrix from cached JSON.

    Handles both old format (just minutes) and new format (dict with min/km).
    """
    with open(cache_file) as f:
        cached = json.load(f)
    matrix = {}
    for o, dests in cached.items():
        for d, val in dests.items():
            if isinstance(val, (int, float)):
                # Old format: just the duration_min value
                matrix[(o, d)] = {"duration_min": val, "distance_km": None,
                                  "duration_s": None, "distance_m": None}
            elif isinstance(val, dict) and "min" in val:
                # New format: {"min": duration_min, "km": distance_km}
                matrix[(o, d)] = {"duration_min": val["min"],
                                  "distance_km": val.get("km"),
                                  "duration_s": None, "distance_m": None}
            else:
                # Full format from API
                matrix[(o, d)] = val
    return matrix


# ---------------------------------------------------------------------------
# Trip validation from optimizer output
# ---------------------------------------------------------------------------

def load_trips_from_output(output_file: str) -> list[dict]:
    """Read trips from optimizer output 'Overzicht Ritsamenhang' sheet.

    Returns list of dicts: {bus_id, origin, dest, origin_loc, dest_loc,
                            departure, arrival, duration_min, service, direction}
    """
    import openpyxl

    wb = openpyxl.load_workbook(output_file, data_only=True)

    # Find the ritsamenhang sheet
    ws = None
    for name in wb.sheetnames:
        if "ritsamenhang" in name.lower():
            ws = wb[name]
            break

    if not ws:
        print(f"Kon 'Ritsamenhang' sheet niet vinden in {output_file}")
        wb.close()
        return []

    # Find header row (contains "Bus ID", "Van", "Naar", etc.)
    header_map = {}
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        val = ws.cell(row_idx, 1).value
        if val and str(val).strip().lower() == "bus id":
            header_row = row_idx
            for c in range(1, ws.max_column + 1):
                h = ws.cell(row_idx, c).value
                if h:
                    header_map[str(h).strip().lower()] = c
            break

    if not header_row:
        print("Kon header-rij niet vinden in Ritsamenhang sheet")
        wb.close()
        return []

    # Map columns
    col_bus = header_map.get("bus id", 1)
    col_type = header_map.get("bustype", 2)
    col_service = header_map.get("busdienst", 5)
    col_direction = header_map.get("richting", 6)
    col_origin = header_map.get("van", 7)
    col_dest = header_map.get("naar", 8)
    col_dep = header_map.get("vertrek", 9)
    col_arr = header_map.get("aankomst", 10)
    col_dur = header_map.get("duur (min)", 11)

    trips = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        bus_id = ws.cell(row_idx, col_bus).value
        if not bus_id:
            continue

        origin = ws.cell(row_idx, col_origin).value or ""
        dest = ws.cell(row_idx, col_dest).value or ""
        dep = ws.cell(row_idx, col_dep).value
        arr = ws.cell(row_idx, col_arr).value
        dur = ws.cell(row_idx, col_dur).value
        service = ws.cell(row_idx, col_service).value or ""
        direction = ws.cell(row_idx, col_direction).value or ""
        bus_type = ws.cell(row_idx, col_type).value or ""

        # Convert times to minutes
        dep_min = None
        arr_min = None
        if isinstance(dep, datetime.time):
            dep_min = dep.hour * 60 + dep.minute
        if isinstance(arr, datetime.time):
            arr_min = arr.hour * 60 + arr.minute

        duration = dur if dur else (arr_min - dep_min if dep_min is not None and arr_min is not None else None)

        trips.append({
            "bus_id": str(bus_id),
            "bus_type": bus_type,
            "origin": origin,
            "dest": dest,
            "origin_loc": normalize_display_name(origin),
            "dest_loc": normalize_display_name(dest),
            "departure": dep_min,
            "arrival": arr_min,
            "duration_min": duration,
            "service": service,
            "direction": direction,
        })

    wb.close()
    return trips


def validate_trips(matrix: dict, trips: list[dict]) -> list[dict]:
    """Compare scheduled trip durations against Google Maps driving times.

    Returns list of route validations with buffer analysis.
    """
    # Group trips by unique route (origin_loc, dest_loc)
    route_trips = {}  # (origin_loc, dest_loc) -> list of trips
    for t in trips:
        if t["duration_min"] is None or t["duration_min"] <= 0:
            continue
        key = (t["origin_loc"], t["dest_loc"])
        if key not in route_trips:
            route_trips[key] = []
        route_trips[key].append(t)

    results = []
    for (o_loc, d_loc), route_list in sorted(route_trips.items()):
        gmaps = matrix.get((o_loc, d_loc))
        gmaps_min = gmaps["duration_min"] if gmaps and gmaps["duration_min"] else None

        durations = [t["duration_min"] for t in route_list]
        min_dur = min(durations)
        max_dur = max(durations)
        avg_dur = sum(durations) / len(durations)

        buffer_min = min_dur - gmaps_min if gmaps_min else None
        buffer_pct = (buffer_min / gmaps_min * 100) if gmaps_min and buffer_min is not None else None

        results.append({
            "origin": route_list[0]["origin"],
            "dest": route_list[0]["dest"],
            "origin_loc": o_loc,
            "dest_loc": d_loc,
            "num_trips": len(route_list),
            "min_scheduled": min_dur,
            "max_scheduled": max_dur,
            "avg_scheduled": round(avg_dur, 1),
            "gmaps_min": gmaps_min,
            "buffer_min": round(buffer_min, 1) if buffer_min is not None else None,
            "buffer_pct": round(buffer_pct, 1) if buffer_pct is not None else None,
            "services": list(set(t["service"] for t in route_list)),
        })

    results.sort(key=lambda r: r["buffer_min"] if r["buffer_min"] is not None else 999)
    return results


def print_validation_report(validations: list[dict]):
    """Print trip validation report to console."""
    print("\n" + "=" * 95)
    print("TRIP VALIDATION: Scheduled Duration vs Google Maps Driving Time")
    print("=" * 95)
    print(f"\n{'Route':<40} {'#':>3} {'Sched':>6} {'GMaps':>6} {'Buffer':>7} {'%':>6}  Status")
    print("-" * 95)

    critical = []
    warning = []
    ok = []

    for v in validations:
        route = f"{v['origin']} -> {v['dest']}"
        if len(route) > 39:
            route = route[:36] + "..."

        sched = f"{v['min_scheduled']:.0f}m"
        gmaps = f"{v['gmaps_min']:.0f}m" if v['gmaps_min'] else "?"
        buf = f"{v['buffer_min']:+.0f}m" if v['buffer_min'] is not None else "?"
        pct = f"{v['buffer_pct']:+.0f}%" if v['buffer_pct'] is not None else "?"

        if v['buffer_min'] is not None:
            if v['buffer_min'] < 0:
                status = "!! UNREALISTIC"
                critical.append(v)
            elif v['buffer_pct'] is not None and v['buffer_pct'] < 15:
                status = "!  TIGHT"
                warning.append(v)
            else:
                status = "OK"
                ok.append(v)
        else:
            status = "?  NO DATA"

        print(f"{route:<40} {v['num_trips']:>3} {sched:>6} {gmaps:>6} {buf:>7} {pct:>6}  {status}")

    # Summary
    print(f"\n--- Summary ---")
    print(f"Total unique routes: {len(validations)}")
    print(f"  OK (buffer >= 15%):    {len(ok)}")
    print(f"  TIGHT (buffer < 15%):  {len(warning)}")
    print(f"  UNREALISTIC (< GMaps): {len(critical)}")

    if critical:
        print(f"\n!! UNREALISTIC routes (scheduled time < Google Maps estimate):")
        for v in critical:
            print(f"   {v['origin']} -> {v['dest']}: "
                  f"scheduled {v['min_scheduled']:.0f}min vs GMaps {v['gmaps_min']:.0f}min "
                  f"({v['buffer_min']:+.0f}min / {v['buffer_pct']:+.0f}%)")
            print(f"      Services: {', '.join(v['services'])}")

    if warning:
        print(f"\n!  TIGHT routes (buffer < 15% of driving time):")
        for v in warning:
            print(f"   {v['origin']} -> {v['dest']}: "
                  f"scheduled {v['min_scheduled']:.0f}min vs GMaps {v['gmaps_min']:.0f}min "
                  f"({v['buffer_min']:+.0f}min / {v['buffer_pct']:+.0f}%)")


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel_output(matrix: dict, locations: list[str],
                       validations: list[dict], output_file: str):
    """Write results to Excel with multiple sheets."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # --- Sheet 1: Distance Matrix (km) ---
    ws1 = wb.active
    ws1.title = "Afstand (km)"
    _write_matrix_sheet(ws1, matrix, locations, "distance_km", "Afstand in km")

    # --- Sheet 2: Duration Matrix (min) ---
    ws2 = wb.create_sheet("Rijtijd (min)")
    _write_matrix_sheet(ws2, matrix, locations, "duration_min", "Rijtijd in minuten")

    # --- Sheet 3: Deadhead matrix ---
    ws3 = wb.create_sheet("Deadhead matrix")
    _write_deadhead_sheet(ws3, matrix, locations)

    # --- Sheet 4: Trip validation ---
    if validations:
        ws4 = wb.create_sheet("Ritvalidatie")
        _write_validation_sheet(ws4, validations)

    # --- Sheet 5: JSON data ---
    ws5 = wb.create_sheet("JSON data")
    json_data = {}
    for (o, d), vals in matrix.items():
        if o not in json_data:
            json_data[o] = {}
        json_data[o][d] = vals
    ws5.cell(1, 1, "Deadhead matrix als JSON (voor import in optimizer)")
    ws5.cell(2, 1, json.dumps(json_data, indent=2, ensure_ascii=False))

    wb.save(output_file)
    print(f"\nOutput saved to {output_file}")


def _write_matrix_sheet(ws, matrix, locations, value_key, title):
    """Write an NxN matrix sheet."""
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")

    ws.cell(1, 1, title).font = Font(bold=True, size=12)

    for j, loc in enumerate(locations):
        cell = ws.cell(3, j + 2, loc)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        ws.column_dimensions[get_column_letter(j + 2)].width = 6

    for i, o_loc in enumerate(locations):
        ws.cell(i + 4, 1, o_loc).font = Font(bold=True, size=10)
        for j, d_loc in enumerate(locations):
            val = matrix.get((o_loc, d_loc), {}).get(value_key)
            if val is not None:
                ws.cell(i + 4, j + 2, val)
            ws.cell(i + 4, j + 2).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 22


def _write_deadhead_sheet(ws, matrix, locations):
    """Write deadhead matrix in a format the optimizer can consume."""
    from openpyxl.styles import Font

    ws.cell(1, 1, "Deadhead matrix").font = Font(bold=True, size=12)
    ws.cell(2, 1, "Rijtijd in minuten tussen stations (voor lege ritten)")

    ws.cell(4, 1, "Van \\ Naar")
    for j, loc in enumerate(locations):
        ws.cell(4, j + 2, loc)

    for i, o_loc in enumerate(locations):
        ws.cell(i + 5, 1, o_loc)
        for j, d_loc in enumerate(locations):
            val = matrix.get((o_loc, d_loc), {}).get("duration_min")
            if val is not None:
                ws.cell(i + 5, j + 2, val)


def _write_validation_sheet(ws, validations):
    """Write trip validation results."""
    from openpyxl.styles import Font, Alignment, PatternFill

    ws.cell(1, 1, "Ritvalidatie: Geplande rijtijd vs Google Maps").font = Font(bold=True, size=12)

    headers = ["Van", "Naar", "Aantal ritten", "Min. gepland (min)",
               "Max. gepland (min)", "Gem. gepland (min)", "Google Maps (min)",
               "Buffer (min)", "Buffer (%)", "Status", "Busdiensten"]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, v in enumerate(validations, 4):
        ws.cell(r, 1, v["origin"])
        ws.cell(r, 2, v["dest"])
        ws.cell(r, 3, v["num_trips"])
        ws.cell(r, 4, v["min_scheduled"])
        ws.cell(r, 5, v["max_scheduled"])
        ws.cell(r, 6, v["avg_scheduled"])
        ws.cell(r, 7, v["gmaps_min"])
        ws.cell(r, 8, v["buffer_min"])
        ws.cell(r, 9, v["buffer_pct"])

        if v["buffer_min"] is not None:
            if v["buffer_min"] < 0:
                status = "ONREALISTISCH"
                fill = red_fill
            elif v["buffer_pct"] is not None and v["buffer_pct"] < 15:
                status = "KRAP"
                fill = orange_fill
            else:
                status = "OK"
                fill = green_fill
        else:
            status = "Geen data"
            fill = None

        ws.cell(r, 10, status)
        if fill:
            for c in range(1, 12):
                ws.cell(r, c).fill = fill

        ws.cell(r, 11, ", ".join(v["services"]))

    # Column widths
    widths = [25, 25, 12, 16, 16, 16, 16, 12, 12, 16, 40]
    for c, w in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c)].width = w


def print_route_analysis(matrix: dict):
    """Print analysis of driving times between all station pairs."""
    print("\n" + "=" * 80)
    print("DEADHEAD DRIVING TIMES BETWEEN STATIONS")
    print("=" * 80)
    print(f"\n{'Van':<22} {'Naar':<22} {'Afstand':>8} {'Rijtijd':>8}")
    print("-" * 62)

    entries = []
    for (o, d), vals in sorted(matrix.items()):
        if o == d:
            continue
        if vals.get("duration_min") is not None:
            entries.append((o, d, vals.get("distance_km", "?"), vals["duration_min"]))

    entries.sort(key=lambda x: x[3], reverse=True)

    for o, d, km, mins in entries:
        km_str = f"{km:.1f} km" if isinstance(km, (int, float)) else "?     "
        print(f"{o:<22} {d:<22} {km_str:>8} {mins:>6.1f} min")

    print(f"\nTotal unique routes (excl. self): {len(entries)}")

    long_routes = [e for e in entries if e[3] > 30]
    if long_routes:
        print(f"\nRoutes with deadhead > 30 min ({len(long_routes)}):")
        for o, d, km, mins in long_routes:
            km_str = f"{km:.0f} km" if isinstance(km, (int, float)) else "?"
            print(f"  {o} -> {d}: {mins:.0f} min ({km_str})")


# ---------------------------------------------------------------------------
# Traffic-aware matrix fetching
# ---------------------------------------------------------------------------

def extract_dates_from_input(input_file: str) -> tuple[str, str | None]:
    """Pick one weekday and one weekend day from the input Excel's trip dates.

    Returns (weekday_date_str, weekend_date_str_or_None).
    The weekday with the most trips is chosen; for weekend the first Saturday.
    """
    from busomloop_optimizer import parse_all_sheets
    all_trips, _, _ = parse_all_sheets(input_file)
    weekday_counts = {}  # date_str -> count
    weekend_dates = []
    for t in all_trips:
        ds = t.date_str  # e.g. "do 11-06-2026"
        prefix = ds.split()[0].lower() if ds else ""
        if prefix in ("za", "zo"):
            if ds not in weekend_dates:
                weekend_dates.append(ds)
        else:
            weekday_counts[ds] = weekday_counts.get(ds, 0) + 1

    # Pick weekday with most trips
    best_weekday = max(weekday_counts, key=weekday_counts.get) if weekday_counts else None
    best_weekend = weekend_dates[0] if weekend_dates else None
    return best_weekday, best_weekend


def _parse_date_str(date_str: str) -> datetime.date:
    """Parse a date string like 'do 11-06-2026' to a datetime.date."""
    parts = date_str.strip().split()
    date_part = parts[-1]  # '11-06-2026'
    day, month, year = date_part.split("-")
    return datetime.date(int(year), int(month), int(day))


def _make_departure_timestamp(date_str: str, hour: int) -> int:
    """Create a Unix timestamp for a given date and hour (local time)."""
    d = _parse_date_str(date_str)
    dt = datetime.datetime(d.year, d.month, d.day, hour, 0, 0)
    # Use UTC offset for Netherlands (CET = +1, CEST = +2)
    # Approximate: summer months (April-October) use CEST (+2), rest CET (+1)
    if 4 <= d.month <= 10:
        utc_offset = 2
    else:
        utc_offset = 1
    utc_dt = dt - datetime.timedelta(hours=utc_offset)
    epoch = datetime.datetime(1970, 1, 1)
    return int((utc_dt - epoch).total_seconds())


def fetch_traffic_aware_matrix(api_key: str, locations: list[str],
                                addresses: dict, weekday_date: str,
                                weekend_date: str = None,
                                batch_size: int = 10) -> dict:
    """Fetch distance matrices for all time slots.

    Returns dict: {"time_slots": {slot_name: {origin: {dest: minutes}}},
                   "baseline": {origin: {dest: minutes}}}
    """
    result = {"time_slots": {}, "baseline": {}}

    # Weekday slots
    weekday_slots = ["nacht", "ochtendspits", "dal", "middagspits", "avond"]
    for slot_name in weekday_slots:
        slot_info = TIME_SLOTS[slot_name]
        ts = _make_departure_timestamp(weekday_date, slot_info["fetch_hour"])
        print(f"\n--- Tijdslot: {slot_info['label']} (departure_time={ts}) ---")
        matrix = fetch_distance_matrix(api_key, locations, addresses,
                                       batch_size=batch_size,
                                       departure_time=ts)
        # Convert to nested dict format
        slot_data = {}
        for o in locations:
            slot_data[o] = {}
            for d in locations:
                entry = matrix.get((o, d))
                if entry and entry["duration_min"] is not None:
                    slot_data[o][d] = entry["duration_min"]
        result["time_slots"][slot_name] = slot_data

    # Weekend slot
    if weekend_date:
        slot_info = TIME_SLOTS["weekend"]
        ts = _make_departure_timestamp(weekend_date, slot_info["fetch_hour"])
        print(f"\n--- Tijdslot: {slot_info['label']} (departure_time={ts}) ---")
        matrix = fetch_distance_matrix(api_key, locations, addresses,
                                       batch_size=batch_size,
                                       departure_time=ts)
        slot_data = {}
        for o in locations:
            slot_data[o] = {}
            for d in locations:
                entry = matrix.get((o, d))
                if entry and entry["duration_min"] is not None:
                    slot_data[o][d] = entry["duration_min"]
        result["time_slots"]["weekend"] = slot_data
    else:
        print("\nGeen weekenddag gevonden in de input, weekend-slot wordt overgeslagen.")

    # Baseline = fetch without departure_time (no traffic)
    print(f"\n--- Baseline (geen verkeer) ---")
    matrix = fetch_distance_matrix(api_key, locations, addresses,
                                   batch_size=batch_size,
                                   departure_time=None)
    for o in locations:
        result["baseline"][o] = {}
        for d in locations:
            entry = matrix.get((o, d))
            if entry and entry["duration_min"] is not None:
                result["baseline"][o][d] = entry["duration_min"]

    return result


def save_traffic_aware_json(traffic_data: dict, output_file: str):
    """Save traffic-aware matrix data to JSON."""
    with open(output_file, "w") as f:
        json.dump(traffic_data, f, indent=2, ensure_ascii=False)
    n_slots = len(traffic_data.get("time_slots", {}))
    print(f"Traffic-aware JSON saved to {output_file} ({n_slots} tijdsloten + baseline)")


def load_matrix_from_cache_traffic(cache_file: str) -> dict:
    """Load traffic-aware or legacy matrix from cached JSON.

    If the file has "time_slots" key, returns the full traffic-aware structure.
    Otherwise, wraps legacy flat format as baseline-only.
    """
    with open(cache_file) as f:
        data = json.load(f)

    if "time_slots" in data:
        return data

    # Legacy flat format: {origin: {dest: minutes}}
    return {"time_slots": {}, "baseline": data}


def extract_stations_from_input(input_file: str) -> tuple[dict, dict]:
    """Extract stations and halt info from input Excel using the optimizer's parser.

    Returns: (registry, halts)
      registry: {canonical_key: display_name}
      halts:    {canonical_key: set of halt names}
    """
    from busomloop_optimizer import parse_all_sheets, build_station_registry, get_station_halts
    print(f"Stations extraheren uit {input_file}...")
    all_trips, reserves, _ = parse_all_sheets(input_file)
    registry = build_station_registry(all_trips, reserves)
    halts = get_station_halts()
    print(f"  {len(registry)} stations gevonden: {', '.join(sorted(registry.values()))}")
    return registry, halts


def verify_addresses(api_key: str, addresses: dict):
    """Geocode each address and show what Google Maps resolves it to.

    This lets the user verify that search queries point to the right
    locations before spending API credits on the full NxN distance matrix.
    """
    print(f"\n{'='*90}")
    print("ADRES VERIFICATIE: Google Maps Geocoding")
    print(f"{'='*90}")
    print(f"\n{'Station':<25} {'Opgelost adres':<50} {'Lat/Lng'}")
    print("-" * 90)

    issues = []
    for canonical, address in sorted(addresses.items()):
        params = {
            "address": address,
            "key": api_key,
            "language": "nl",
            "region": "nl",
        }
        try:
            resp = requests.get(GEOCODE_URL, params=params, timeout=15)
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as e:
            print(f"  {canonical:<25} FOUT: {e}")
            issues.append((canonical, address, f"API fout: {e}"))
            continue

        if data["status"] != "OK" or not data.get("results"):
            print(f"  {canonical:<25} NIET GEVONDEN ({data['status']})")
            issues.append((canonical, address, f"Niet gevonden: {data['status']}"))
            continue

        result = data["results"][0]
        resolved = result["formatted_address"]
        loc = result["geometry"]["location"]
        lat, lng = loc["lat"], loc["lng"]

        # Check if resolved address is in Netherlands (lat ~50.5-53.5, lng ~3.3-7.2)
        in_nl = 50.5 <= lat <= 53.7 and 3.3 <= lng <= 7.3
        flag = "" if in_nl else " !! BUITEN NL"
        if not in_nl:
            issues.append((canonical, address, f"Buiten Nederland: {resolved}"))

        resolved_short = resolved[:49] if len(resolved) > 49 else resolved
        print(f"  {canonical:<25} {resolved_short:<50} {lat:.4f}, {lng:.4f}{flag}")

        time.sleep(0.1)  # Rate limit

    print()
    if issues:
        print(f"WAARSCHUWINGEN ({len(issues)}):")
        for canonical, query, issue in issues:
            print(f"  {canonical}: {issue}")
            print(f"    Zoekopdracht was: {query}")
        print(f"\nControleer bovenstaande adressen voordat je de volledige matrix ophaalt.")
        print(f"Gebruik --key zonder --verify om de matrix op te halen.")
    else:
        print("Alle adressen succesvol geverifieerd.")
        print(f"Gebruik --key zonder --verify om de volledige {len(addresses)}x{len(addresses)} matrix op te halen.")


def main():
    parser = argparse.ArgumentParser(
        description="Fetch Google Maps distances between bus stations. "
                    "Stations are auto-discovered from the input Excel file.")
    parser.add_argument("--input", "-i", default=None,
                        help="Input Excel file (same as busomloop_optimizer.py). "
                             "Required when using --key to fetch new distances.")
    parser.add_argument("--key", default=None,
                        help="Google Maps API key (of stel GOOGLE_MAPS_API_KEY in .env in)")
    parser.add_argument("--output", default="afstanden_stations.xlsx",
                        help="Output Excel file (default: afstanden_stations.xlsx)")
    parser.add_argument("--json-output", default="deadhead_matrix.json",
                        help="Output JSON file for deadhead matrix")
    parser.add_argument("--from-cache", default=None,
                        help="Load matrix from cached JSON instead of API call")
    parser.add_argument("--validate", default=None,
                        help="Optimizer output .xlsx file for trip validation")
    parser.add_argument("--verify", action="store_true",
                        help="Alleen adressen verifiëren via Geocoding API (geen matrix ophalen). "
                             "Laat zien wat Google Maps per station oplost, zodat je kunt "
                             "controleren of de locaties kloppen voordat je de matrix ophaalt.")
    parser.add_argument("--traffic", action="store_true",
                        help="Haal 6 tijdslot-matrices op (nacht, ochtendspits, dal, "
                             "middagspits, avond, weekend) voor risico-analyse. "
                             "Slaat op als traffic_matrix.json.")
    parser.add_argument("--traffic-json", default="traffic_matrix.json",
                        help="Output JSON file for traffic-aware matrix (default: traffic_matrix.json)")
    args = parser.parse_args()

    # Load API key from .env if not provided on command line
    load_dotenv()
    if not args.key:
        args.key = os.environ.get("GOOGLE_MAPS_API_KEY")

    # Determine station list
    if args.input:
        registry, halts = extract_stations_from_input(args.input)
        addresses = build_station_addresses(registry, halts)
        locations = sorted(registry.keys())
    elif args.from_cache:
        # When loading from cache without input file, derive locations from cache
        print("Geen invoerbestand opgegeven, stations worden uit cache afgeleid...")
        with open(args.from_cache) as f:
            cached = json.load(f)
        locations = sorted(cached.keys())
        addresses = {}  # Not needed for cache-only mode
        registry = {loc: loc for loc in locations}
    else:
        print("FOUT: Geef --input EXCEL_BESTAND op (nodig om stations te ontdekken)")
        print("      Of gebruik --from-cache CACHE_BESTAND om eerder opgehaalde data te laden")
        sys.exit(1)

    print(f"\nStations: {len(locations)}")
    for loc in locations:
        addr = addresses.get(loc, "(uit cache)")
        print(f"  {loc}: {addr}")

    # Verify-only mode: geocode addresses and show results, then exit
    if args.verify:
        if not args.key:
            print("\nFOUT: --verify vereist een API key (--key of GOOGLE_MAPS_API_KEY in .env)")
            sys.exit(1)
        if not args.input:
            print("\nFOUT: --verify vereist --input EXCEL_BESTAND")
            sys.exit(1)
        verify_addresses(args.key, addresses)
        return

    # Traffic-aware mode: fetch 6 time-slot matrices + baseline
    if args.traffic:
        if not args.key:
            print("\nFOUT: --traffic vereist een API key (--key of GOOGLE_MAPS_API_KEY in .env)")
            sys.exit(1)
        if not args.input:
            print("\nFOUT: --traffic vereist --input EXCEL_BESTAND")
            sys.exit(1)
        print(f"\nDatums bepalen uit invoerbestand...")
        weekday_date, weekend_date = extract_dates_from_input(args.input)
        print(f"  Weekdag: {weekday_date}")
        print(f"  Weekend: {weekend_date or '(geen)'}")
        n_slots = 5 + (1 if weekend_date else 0) + 1  # weekday slots + weekend + baseline
        n_elements = len(locations) ** 2
        print(f"\nFetching {n_slots} matrices x {len(locations)}x{len(locations)} = "
              f"{n_slots * n_elements} API-elementen...")
        traffic_data = fetch_traffic_aware_matrix(
            args.key, locations, addresses, weekday_date, weekend_date)
        save_traffic_aware_json(traffic_data, args.traffic_json)
        # Also save baseline as standard deadhead matrix
        save_deadhead_json_from_nested(traffic_data["baseline"], args.json_output)
        print(f"\nTraffic-aware matrix opgeslagen. Gebruik --deadhead {args.json_output} "
              f"en --traffic-matrix {args.traffic_json} in de optimizer.")
        return

    # Load or fetch matrix
    if args.from_cache:
        print(f"\nLoading cached matrix from {args.from_cache}...")
        matrix = load_matrix_from_cache(args.from_cache)
        locations = sorted(set(o for o, d in matrix.keys()))

    elif args.key:
        if not args.input:
            print("FOUT: --input is vereist bij --key (nodig om stations te ontdekken)")
            sys.exit(1)
        print(f"\nFetching {len(locations)}x{len(locations)} distance matrix "
              f"({len(locations)**2} elements)...")
        matrix = fetch_distance_matrix(args.key, locations, addresses)
        print(f"Received {len(matrix)} route entries")
        save_deadhead_json(matrix, locations, args.json_output)
    else:
        print("\nFOUT: Geef --key API_KEY of --from-cache BESTAND op")
        print("      (of stel GOOGLE_MAPS_API_KEY in .env in)")
        sys.exit(1)

    # Trip validation
    validations = []
    if args.validate:
        print(f"\nLoading trips from {args.validate}...")
        trips = load_trips_from_output(args.validate)
        print(f"Loaded {len(trips)} trips")
        if trips:
            validations = validate_trips(matrix, trips)
            print_validation_report(validations)

    # Write Excel output
    write_excel_output(matrix, locations, validations, args.output)

    # Print route analysis
    print_route_analysis(matrix)


if __name__ == "__main__":
    main()
