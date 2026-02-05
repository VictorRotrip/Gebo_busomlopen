#!/usr/bin/env python3
"""
ze_charging_planner.py

Version 6: ZE Assignment & Charging Strategy for NS TVV Tender (K3 requirement)

This module assigns Zero Emission (ZE) touringcars to suitable bus rotations and
generates a charging strategy. NS requires minimum 5 ZE touringcars in the tender.

Usage:
  # Standalone (reads version 5 output Excel):
  python ze_charging_planner.py --roster busomloop_v5.xlsx --output ze_plan.xlsx

  # With charging station data:
  python ze_charging_planner.py --roster busomloop_v5.xlsx --tanklocaties tanklocaties.json

  # Specify minimum ZE count (default: 5):
  python ze_charging_planner.py --roster busomloop_v5.xlsx --min-ze 7

Input files:
  - Roster Excel from version 5 (with Busomloop sheet)
  - financieel_input.xlsx — ZE range per bus type (Buskosten sheet)
  - tanklocaties.json — charging stations per bus station (optional)

Output:
  - Excel with sheets: "ZE Inzet", "Laadstrategie", "Samenvatting"

Requirements:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    sys.exit("Error: openpyxl not installed. Run: pip install openpyxl")


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class ChargingStation:
    """A charging station near a bus station."""
    name: str
    operator: str
    lat: float
    lon: float
    distance_km: float
    max_power_kw: float
    num_points: int
    connectors: list
    category: str  # ultra_fast, fast, semi_fast, slow


@dataclass
class ZEFeasibility:
    """ZE feasibility assessment for a rotation."""
    rotation_id: str
    bus_type: str
    total_km: float
    ze_range_km: float
    is_feasible: bool
    needs_charging: bool
    buffer_km: float  # remaining range after rotation
    charging_opportunities: list  # [(station, idle_window_min, chargers), ...]
    recommended_charging: list  # planned charging stops
    reason: str  # why feasible/not feasible


@dataclass
class Rotation:
    """Simplified rotation data from the roster."""
    rotation_id: str
    bus_type: str
    date_str: str
    trips: list  # [(origin, dest, dep_min, arr_min, km), ...]
    total_km: float
    total_duration_min: int
    idle_windows: list  # [(station, start_min, end_min, duration_min), ...]
    start_time: int
    end_time: int


# ---------------------------------------------------------------------------
# Configuration loading
# ---------------------------------------------------------------------------

def load_ze_config(financieel_xlsx: str = "financieel_input.xlsx") -> dict:
    """Load ZE configuration from financieel_input.xlsx Buskosten sheet.

    Returns dict with:
        - ze_range_km: {bus_type: range_km}
        - ze_consumption_kwh_per_100km: {bus_type: kwh}
        - charging_time_min_per_100km: estimated charging time
    """
    config = {
        "ze_range_km": {},
        "ze_consumption_kwh_per_100km": {},
    }

    path = Path(financieel_xlsx)
    if not path.exists():
        print(f"Warning: {financieel_xlsx} not found, using defaults")
        # Defaults based on typical ZE touringcars
        config["ze_range_km"] = {
            "Touringcar": 300,
            "Dubbeldekker": 250,
            "Lagevloerbus": 280,
            "Midi bus": 350,
            "Taxibus": 400,
        }
        config["ze_consumption_kwh_per_100km"] = {
            "Touringcar": 130,
            "Dubbeldekker": 180,
            "Lagevloerbus": 150,
            "Midi bus": 100,
            "Taxibus": 50,
        }
        return config

    wb = openpyxl.load_workbook(path, data_only=True)
    if "Buskosten" not in wb.sheetnames:
        print(f"Warning: 'Buskosten' sheet not found in {financieel_xlsx}, using defaults")
        wb.close()
        return load_ze_config("__nonexistent__")  # Trigger defaults

    ws = wb["Buskosten"]

    # Parse the sheet to find ZE range and consumption values
    # Expected format: variable name in col A, value in col B
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        var_name = str(row[0]).lower()
        value = row[1]

        # ZE range (actieradius)
        if "actieradius" in var_name and "_ze" in var_name and value:
            if "dubbeldekker" in var_name:
                config["ze_range_km"]["Dubbeldekker"] = float(value)
            elif "touringcar" in var_name:
                config["ze_range_km"]["Touringcar"] = float(value)
            elif "lagevloer" in var_name:
                config["ze_range_km"]["Lagevloerbus"] = float(value)
            elif "midi" in var_name:
                config["ze_range_km"]["Midi bus"] = float(value)
            elif "taxi" in var_name:
                config["ze_range_km"]["Taxibus"] = float(value)

        # ZE consumption (kWh per 100km)
        if "verbruik" in var_name and "_ze" in var_name and "kwh" in var_name and value:
            if "dubbeldekker" in var_name:
                config["ze_consumption_kwh_per_100km"]["Dubbeldekker"] = float(value)
            elif "touringcar" in var_name:
                config["ze_consumption_kwh_per_100km"]["Touringcar"] = float(value)
            elif "lagevloer" in var_name:
                config["ze_consumption_kwh_per_100km"]["Lagevloerbus"] = float(value)
            elif "midi" in var_name:
                config["ze_consumption_kwh_per_100km"]["Midi bus"] = float(value)
            elif "taxi" in var_name:
                config["ze_consumption_kwh_per_100km"]["Taxibus"] = float(value)

    wb.close()

    # Fill in any missing values with defaults
    defaults_range = {"Touringcar": 300, "Dubbeldekker": 250, "Lagevloerbus": 280,
                      "Midi bus": 350, "Taxibus": 400}
    defaults_consumption = {"Touringcar": 130, "Dubbeldekker": 180, "Lagevloerbus": 150,
                           "Midi bus": 100, "Taxibus": 50}

    for bt in defaults_range:
        if bt not in config["ze_range_km"]:
            config["ze_range_km"][bt] = defaults_range[bt]
        if bt not in config["ze_consumption_kwh_per_100km"]:
            config["ze_consumption_kwh_per_100km"][bt] = defaults_consumption[bt]

    return config


def load_charging_stations(tanklocaties_json: str = "tanklocaties.json") -> dict:
    """Load charging station data from tanklocaties.json.

    Returns: {station_name: [ChargingStation, ...]}
    """
    path = Path(tanklocaties_json)
    if not path.exists():
        print(f"Warning: {tanklocaties_json} not found, charging analysis will be limited")
        return {}

    with open(path, "r") as f:
        data = json.load(f)

    stations_by_location = {}

    for station_name, station_data in data.get("stations", {}).items():
        chargers = []
        for cs in station_data.get("charging_stations", []):
            chargers.append(ChargingStation(
                name=cs.get("name", "Unknown"),
                operator=cs.get("operator", ""),
                lat=cs.get("lat", 0),
                lon=cs.get("lon", 0),
                distance_km=cs.get("distance_km", 0),
                max_power_kw=cs.get("max_power_kw", 0),
                num_points=cs.get("num_points", 1),
                connectors=cs.get("connectors", []),
                category=cs.get("category", "slow"),
            ))

        # Sort by power (prefer fast chargers) then by distance
        chargers.sort(key=lambda c: (-c.max_power_kw, c.distance_km))
        stations_by_location[station_name] = chargers

    return stations_by_location


# ---------------------------------------------------------------------------
# Roster parsing
# ---------------------------------------------------------------------------

def parse_roster_from_excel(roster_xlsx: str) -> list[Rotation]:
    """Parse bus rotations from a version 5 output Excel file.

    Looks for 'Omloop' or 'Busomloop' sheets and extracts rotation data.
    """
    wb = openpyxl.load_workbook(roster_xlsx, data_only=True)

    # Find rotation sheets (named "Omloop TC do", "Busomloop", etc.)
    rotation_sheets = []
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Omloop") or sheet_name.startswith("Busomloop"):
            rotation_sheets.append(sheet_name)

    if not rotation_sheets:
        wb.close()
        sys.exit(f"Error: No rotation sheets found in {roster_xlsx}")

    print(f"  Found sheets: {', '.join(rotation_sheets)}")

    rotations = []

    # Parse each rotation sheet
    for sheet_name in rotation_sheets:
        ws = wb[sheet_name]

        # Extract bus type and date from sheet name (e.g., "Omloop TC do" -> Touringcar, do)
        parts = sheet_name.replace("Omloop ", "").split()
        bus_type_code = parts[0] if parts else ""
        date_code = parts[1] if len(parts) > 1 else ""

        bus_type_map = {
            "DD": "Dubbeldekker",
            "TC": "Touringcar",
            "LV": "Lagevloerbus",
            "Taxi": "Taxibus",
            "MI": "Midi bus",
        }
        bus_type = bus_type_map.get(bus_type_code, bus_type_code)

        # Parse rotations from this sheet
        current_rotation_id = None
        trip_count = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or not any(row):
                continue

            first_cell = str(row[0] or "").strip()

            # Detect bus ID patterns (e.g., "TC-DO-001", "DD-VR-002", "1", "2")
            if first_cell:
                # Check if it looks like a bus ID or number
                is_bus_id = False
                if "-" in first_cell and len(first_cell) <= 15:
                    is_bus_id = True
                elif first_cell.isdigit() and int(first_cell) <= 100:
                    # Simple number ID
                    is_bus_id = True

                if is_bus_id:
                    # Save previous rotation
                    if current_rotation_id:
                        estimated_km = max(trip_count * 25, 50)  # ~25km per trip
                        rotations.append(Rotation(
                            rotation_id=f"{bus_type_code}-{date_code.upper()}-{current_rotation_id}",
                            bus_type=bus_type,
                            date_str=date_code,
                            trips=[],
                            total_km=estimated_km,
                            total_duration_min=trip_count * 45,
                            idle_windows=[],
                            start_time=0,
                            end_time=0,
                        ))

                    current_rotation_id = first_cell
                    trip_count = 0
                    continue

            # Count rows with data as trips
            if current_rotation_id and any(row[1:5]):
                trip_count += 1

        # Save last rotation from this sheet
        if current_rotation_id:
            estimated_km = max(trip_count * 25, 50)
            rotations.append(Rotation(
                rotation_id=f"{bus_type_code}-{date_code.upper()}-{current_rotation_id}",
                bus_type=bus_type,
                date_str=date_code,
                trips=[],
                total_km=estimated_km,
                total_duration_min=trip_count * 45,
                idle_windows=[],
                start_time=0,
                end_time=0,
            ))

    wb.close()
    return rotations


def _extract_rotations_simplified(roster_xlsx: str) -> list[Rotation]:
    """Simplified rotation extraction - looks for bus IDs and counts trips."""
    wb = openpyxl.load_workbook(roster_xlsx, data_only=True)
    rotations = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("Busomloop"):
            continue

        ws = wb[sheet_name]
        current_rotation_id = None
        current_bus_type = None
        trip_count = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row:
                continue

            first_cell = str(row[0] or "").strip()

            # Bus type header
            if first_cell in ["Dubbeldekker", "Touringcar", "Lagevloerbus", "Midi bus", "Taxibus"]:
                current_bus_type = first_cell
                continue

            # Bus ID pattern (e.g., "TC-DO-001", "DD-VR-002")
            if first_cell and len(first_cell) <= 15:
                parts = first_cell.split("-")
                if len(parts) >= 2 and len(parts[0]) == 2:
                    # Save previous rotation
                    if current_rotation_id:
                        # Estimate km based on trip count (rough estimate: 30km per trip)
                        estimated_km = trip_count * 30
                        rotations.append(Rotation(
                            rotation_id=current_rotation_id,
                            bus_type=current_bus_type or "Unknown",
                            date_str="",
                            trips=[],
                            total_km=estimated_km,
                            total_duration_min=trip_count * 45,  # ~45 min per trip
                            idle_windows=[],
                            start_time=0,
                            end_time=0,
                        ))

                    current_rotation_id = first_cell
                    trip_count = 0

                    # Infer bus type from ID prefix
                    prefix = parts[0].upper()
                    type_map = {"DD": "Dubbeldekker", "TC": "Touringcar",
                               "LV": "Lagevloerbus", "MI": "Midi bus", "TA": "Taxibus"}
                    if prefix in type_map:
                        current_bus_type = type_map[prefix]
                    continue

            # Count non-empty rows as potential trips
            if current_rotation_id and any(row[1:5]):
                trip_count += 1

        # Save last rotation
        if current_rotation_id:
            estimated_km = trip_count * 30
            rotations.append(Rotation(
                rotation_id=current_rotation_id,
                bus_type=current_bus_type or "Unknown",
                date_str="",
                trips=[],
                total_km=estimated_km,
                total_duration_min=trip_count * 45,
                idle_windows=[],
                start_time=0,
                end_time=0,
            ))

    wb.close()
    return rotations


# ---------------------------------------------------------------------------
# ZE Feasibility Analysis
# ---------------------------------------------------------------------------

def analyze_ze_feasibility(rotation: Rotation, ze_config: dict,
                           charging_stations: dict) -> ZEFeasibility:
    """Analyze whether a rotation can be done with a ZE bus."""

    bus_type = rotation.bus_type
    ze_range = ze_config["ze_range_km"].get(bus_type, 300)
    total_km = rotation.total_km

    # Simple feasibility: total km <= range
    buffer_km = ze_range - total_km
    is_feasible_without_charging = buffer_km >= 0

    # Check charging opportunities during idle windows
    charging_opportunities = []
    for idle in rotation.idle_windows:
        station, start_min, end_min, duration_min = idle
        # Minimum 30 minutes idle for useful charging
        if duration_min >= 30:
            chargers = charging_stations.get(station, [])
            fast_chargers = [c for c in chargers if c.max_power_kw >= 50]
            if fast_chargers:
                charging_opportunities.append((station, duration_min, fast_chargers))

    # Calculate if charging can extend range sufficiently
    needs_charging = not is_feasible_without_charging
    is_feasible_with_charging = False
    recommended_charging = []

    if needs_charging and charging_opportunities:
        # Estimate km that can be recovered with charging
        # Assume 50kW charger can add ~50km per 30 minutes for a touringcar
        consumption = ze_config["ze_consumption_kwh_per_100km"].get(bus_type, 130)

        total_recoverable_km = 0
        for station, duration_min, chargers in charging_opportunities:
            best_charger = chargers[0]  # Already sorted by power
            # kWh charged = power * time
            kwh_charged = best_charger.max_power_kw * (duration_min / 60) * 0.8  # 80% efficiency
            km_recovered = (kwh_charged / consumption) * 100
            total_recoverable_km += km_recovered

            if total_km <= ze_range + total_recoverable_km:
                recommended_charging.append({
                    "station": station,
                    "duration_min": duration_min,
                    "charger": best_charger.name,
                    "power_kw": best_charger.max_power_kw,
                    "km_recovered": round(km_recovered, 1),
                })
                is_feasible_with_charging = True
                break

    is_feasible = is_feasible_without_charging or is_feasible_with_charging

    # Determine reason
    if is_feasible_without_charging:
        reason = f"Bereik voldoende: {total_km:.0f} km < {ze_range:.0f} km (buffer: {buffer_km:.0f} km)"
    elif is_feasible_with_charging:
        reason = f"Haalbaar met opladen tijdens wachttijd"
    else:
        reason = f"Niet haalbaar: {total_km:.0f} km > {ze_range:.0f} km, onvoldoende laadmogelijkheden"

    return ZEFeasibility(
        rotation_id=rotation.rotation_id,
        bus_type=bus_type,
        total_km=total_km,
        ze_range_km=ze_range,
        is_feasible=is_feasible,
        needs_charging=needs_charging and is_feasible,
        buffer_km=buffer_km,
        charging_opportunities=charging_opportunities,
        recommended_charging=recommended_charging,
        reason=reason,
    )


def assign_ze_buses(rotations: list[Rotation], min_ze_count: int,
                    ze_config: dict, charging_stations: dict,
                    target_bus_type: str = "Touringcar") -> dict:
    """Assign ZE to rotations, ensuring minimum count for target bus type.

    Returns: {rotation_id: ZEFeasibility}
    """
    # Filter to target bus type
    target_rotations = [r for r in rotations if r.bus_type == target_bus_type]

    if len(target_rotations) < min_ze_count:
        print(f"Warning: Only {len(target_rotations)} {target_bus_type} rotations, "
              f"but {min_ze_count} ZE required")

    # Analyze feasibility for all target rotations
    feasibility_results = []
    for rotation in target_rotations:
        feas = analyze_ze_feasibility(rotation, ze_config, charging_stations)
        feasibility_results.append(feas)

    # Sort by suitability for ZE:
    # 1. Feasible without charging first
    # 2. Then by buffer (more buffer = safer choice)
    # 3. Then by total km (shorter = better)
    def ze_score(f: ZEFeasibility) -> tuple:
        return (
            0 if f.is_feasible and not f.needs_charging else 1,  # No charging needed is best
            0 if f.is_feasible else 1,  # Feasible is better
            -f.buffer_km,  # More buffer is better (negative for sort)
            f.total_km,  # Shorter is better
        )

    feasibility_results.sort(key=ze_score)

    # Assign ZE to top N feasible rotations
    ze_assignments = {}
    assigned_count = 0

    for feas in feasibility_results:
        if feas.is_feasible and assigned_count < min_ze_count:
            ze_assignments[feas.rotation_id] = feas
            assigned_count += 1

    # If not enough feasible, note the gap
    if assigned_count < min_ze_count:
        print(f"Warning: Could only assign {assigned_count} ZE buses, "
              f"need {min_ze_count} (insufficient feasible rotations)")

    # Also include non-assigned feasibility results for reporting
    for feas in feasibility_results:
        if feas.rotation_id not in ze_assignments:
            ze_assignments[feas.rotation_id] = feas

    return ze_assignments


# ---------------------------------------------------------------------------
# Excel Output
# ---------------------------------------------------------------------------

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ZE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def generate_ze_output(rotations: list[Rotation], ze_assignments: dict,
                       charging_stations: dict, output_xlsx: str,
                       min_ze_count: int = 5):
    """Generate Excel output for tender with ZE assignment and charging strategy."""

    wb = openpyxl.Workbook()

    # --- Sheet 1: ZE Inzet (ZE Assignment) ---
    ws1 = wb.active
    ws1.title = "ZE Inzet"

    headers1 = ["Bus ID", "Bustype", "Totaal km", "ZE Bereik km", "Buffer km",
                "ZE Toegewezen", "Laden Nodig", "Reden"]

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Get all Touringcar rotations with their feasibility
    touringcar_feas = [(r, ze_assignments.get(r.rotation_id))
                       for r in rotations if r.bus_type == "Touringcar"]

    # Sort: assigned ZE first, then by feasibility
    def sort_key(item):
        r, f = item
        if f is None:
            return (2, 0, r.total_km)
        assigned = f.rotation_id in [fid for fid, feas in ze_assignments.items()
                                      if feas.is_feasible and not feas.needs_charging][:min_ze_count]
        return (0 if assigned else 1, 0 if f.is_feasible else 1, r.total_km)

    touringcar_feas.sort(key=sort_key)

    row = 2
    ze_assigned_count = 0
    for rotation, feas in touringcar_feas:
        if feas is None:
            continue

        is_assigned = feas.is_feasible and ze_assigned_count < min_ze_count
        if is_assigned:
            ze_assigned_count += 1

        ws1.cell(row=row, column=1, value=rotation.rotation_id).border = THIN_BORDER
        ws1.cell(row=row, column=2, value=rotation.bus_type).border = THIN_BORDER
        ws1.cell(row=row, column=3, value=round(feas.total_km, 1)).border = THIN_BORDER
        ws1.cell(row=row, column=4, value=feas.ze_range_km).border = THIN_BORDER
        ws1.cell(row=row, column=5, value=round(feas.buffer_km, 1)).border = THIN_BORDER

        ze_cell = ws1.cell(row=row, column=6, value="JA" if is_assigned else "Nee")
        ze_cell.border = THIN_BORDER
        if is_assigned:
            ze_cell.fill = ZE_FILL
            ze_cell.font = Font(bold=True)

        charging_cell = ws1.cell(row=row, column=7,
                                  value="Ja" if feas.needs_charging else "Nee")
        charging_cell.border = THIN_BORDER
        if feas.needs_charging:
            charging_cell.fill = WARNING_FILL

        ws1.cell(row=row, column=8, value=feas.reason).border = THIN_BORDER

        row += 1

    # Auto-width columns
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    ws1.column_dimensions["H"].width = 50

    # --- Sheet 2: Laadstrategie (Charging Strategy) ---
    ws2 = wb.create_sheet("Laadstrategie")

    headers2 = ["Bus ID", "Station", "Laadduur (min)", "Lader", "Vermogen (kW)",
                "Geschatte km opgeladen", "Opmerkingen"]

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row = 2
    ze_assigned_count = 0
    for rotation, feas in touringcar_feas:
        if feas is None or not feas.is_feasible:
            continue
        if ze_assigned_count >= min_ze_count:
            break
        ze_assigned_count += 1

        if feas.needs_charging and feas.recommended_charging:
            for charge in feas.recommended_charging:
                ws2.cell(row=row, column=1, value=rotation.rotation_id).border = THIN_BORDER
                ws2.cell(row=row, column=2, value=charge["station"]).border = THIN_BORDER
                ws2.cell(row=row, column=3, value=charge["duration_min"]).border = THIN_BORDER
                ws2.cell(row=row, column=4, value=charge["charger"]).border = THIN_BORDER
                ws2.cell(row=row, column=5, value=charge["power_kw"]).border = THIN_BORDER
                ws2.cell(row=row, column=6, value=charge["km_recovered"]).border = THIN_BORDER
                ws2.cell(row=row, column=7, value="Laden tijdens wachttijd").border = THIN_BORDER
                row += 1
        else:
            # No charging needed
            ws2.cell(row=row, column=1, value=rotation.rotation_id).border = THIN_BORDER
            ws2.cell(row=row, column=2, value="-").border = THIN_BORDER
            ws2.cell(row=row, column=3, value="-").border = THIN_BORDER
            ws2.cell(row=row, column=4, value="-").border = THIN_BORDER
            ws2.cell(row=row, column=5, value="-").border = THIN_BORDER
            ws2.cell(row=row, column=6, value="-").border = THIN_BORDER
            ws2.cell(row=row, column=7, value="Geen tussentijds laden nodig (bereik voldoende)").border = THIN_BORDER
            row += 1

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    ws2.column_dimensions["G"].width = 40

    # --- Sheet 3: Samenvatting (Summary) ---
    ws3 = wb.create_sheet("Samenvatting")

    # Count statistics
    total_touringcar = len([r for r in rotations if r.bus_type == "Touringcar"])
    ze_feasible = len([f for f in ze_assignments.values() if f.is_feasible])
    ze_no_charging = len([f for f in ze_assignments.values()
                          if f.is_feasible and not f.needs_charging])
    ze_with_charging = len([f for f in ze_assignments.values()
                            if f.is_feasible and f.needs_charging])

    summary_data = [
        ("ZE TOURINGCAR INZET - SAMENVATTING", ""),
        ("", ""),
        ("NS Vereiste", f"Minimaal {min_ze_count} ZE touringcars"),
        ("", ""),
        ("Totaal aantal touringcar-omlopen", total_touringcar),
        ("Waarvan ZE-geschikt", ze_feasible),
        ("  - Zonder tussentijds laden", ze_no_charging),
        ("  - Met tussentijds laden", ze_with_charging),
        ("", ""),
        ("ZE touringcars toegewezen", min(min_ze_count, ze_feasible)),
        ("Voldoet aan NS vereiste", "JA" if ze_feasible >= min_ze_count else "NEE"),
        ("", ""),
        ("TOELICHTING", ""),
        ("", ""),
        ("De ZE touringcars zijn toegewezen aan omlopen die:", ""),
        ("1. Binnen het elektrische bereik vallen, OF", ""),
        ("2. Voldoende wachttijd hebben om tussentijds op te laden", ""),
        ("", ""),
        ("Selectiecriteria (in volgorde van prioriteit):", ""),
        ("1. Omlopen die geen tussentijds laden nodig hebben", ""),
        ("2. Omlopen met grotere marge (buffer) op bereik", ""),
        ("3. Kortere omlopen (minder km)", ""),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell1 = ws3.cell(row=row_idx, column=1, value=label)
        cell2 = ws3.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell1.font = Font(bold=True, size=14)
        elif label in ["NS Vereiste", "ZE touringcars toegewezen", "Voldoet aan NS vereiste",
                       "TOELICHTING"]:
            cell1.font = Font(bold=True)

        if label == "Voldoet aan NS vereiste":
            if value == "JA":
                cell2.fill = ZE_FILL
            else:
                cell2.fill = ERROR_FILL
            cell2.font = Font(bold=True)

    ws3.column_dimensions["A"].width = 45
    ws3.column_dimensions["B"].width = 35

    # Save
    wb.save(output_xlsx)
    print(f"ZE plan saved to {output_xlsx}")

    return {
        "total_touringcar": total_touringcar,
        "ze_feasible": ze_feasible,
        "ze_assigned": min(min_ze_count, ze_feasible),
        "meets_requirement": ze_feasible >= min_ze_count,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Version 6: Assign ZE touringcars and generate charging strategy for NS tender"
    )
    parser.add_argument("--roster", "-r", required=True,
                        help="Input roster Excel file (version 5 output)")
    parser.add_argument("--output", "-o", default="ze_laadstrategie.xlsx",
                        help="Output Excel file (default: ze_laadstrategie.xlsx)")
    parser.add_argument("--tanklocaties", default="tanklocaties.json",
                        help="Charging station data JSON (from fetch_tanklocaties.py)")
    parser.add_argument("--financieel", default="financieel_input.xlsx",
                        help="Financial input Excel with ZE range data")
    parser.add_argument("--min-ze", type=int, default=5,
                        help="Minimum number of ZE touringcars required (default: 5)")

    args = parser.parse_args()

    print("=" * 60)
    print("ZE Touringcar Toewijzing & Laadstrategie (Versie 6)")
    print(f"NS Vereiste: minimaal {args.min_ze} ZE touringcars")
    print("=" * 60)

    # Load configuration
    print(f"\n[1/4] Laden ZE configuratie uit {args.financieel}...")
    ze_config = load_ze_config(args.financieel)
    print(f"  ZE bereik Touringcar: {ze_config['ze_range_km'].get('Touringcar', 'N/A')} km")

    # Load charging stations
    print(f"\n[2/4] Laden laadstations uit {args.tanklocaties}...")
    charging_stations = load_charging_stations(args.tanklocaties)
    total_chargers = sum(len(v) for v in charging_stations.values())
    print(f"  {len(charging_stations)} locaties met {total_chargers} laadpunten")

    # Parse roster
    print(f"\n[3/4] Parsen busomloop uit {args.roster}...")
    rotations = parse_roster_from_excel(args.roster)
    touringcar_count = len([r for r in rotations if r.bus_type == "Touringcar"])
    print(f"  {len(rotations)} omlopen gevonden ({touringcar_count} touringcars)")

    # Assign ZE
    print(f"\n[4/4] Toewijzen ZE touringcars...")
    ze_assignments = assign_ze_buses(
        rotations, args.min_ze, ze_config, charging_stations, "Touringcar"
    )

    # Generate output
    stats = generate_ze_output(
        rotations, ze_assignments, charging_stations, args.output, args.min_ze
    )

    # Summary
    print("\n" + "=" * 60)
    print("RESULTAAT")
    print("=" * 60)
    print(f"  Totaal touringcar omlopen: {stats['total_touringcar']}")
    print(f"  ZE-geschikt: {stats['ze_feasible']}")
    print(f"  ZE toegewezen: {stats['ze_assigned']}")
    print(f"  Voldoet aan NS vereiste ({args.min_ze} ZE): {'JA ✓' if stats['meets_requirement'] else 'NEE ✗'}")
    print(f"\nOutput: {args.output}")


if __name__ == "__main__":
    main()
