#!/usr/bin/env python3
"""
fetch_fuel_charging_prices.py

Fetches the latest fuel and electricity prices from public APIs and updates
the 'Brandstofprijzen' sheet in additional_inputs.xlsx.

Data sources:
  - CBS OData (table 80416ned): daily Dutch diesel B7 pump prices (free, no auth)
  - EnergyZero API: hourly Dutch electricity prices EPEX spot (free, no auth)
  - Fieten Olie: HVO100 + B7 advisory prices (web scrape)

Usage:
  python fetch_fuel_charging_prices.py                     # update all prices
  python fetch_fuel_charging_prices.py --diesel-only       # only diesel from CBS
  python fetch_fuel_charging_prices.py --electricity-only  # only electricity
  python fetch_fuel_charging_prices.py --hvo-only          # only HVO100/B7 advisory
  python fetch_fuel_charging_prices.py --dry-run           # fetch but don't write

Requirements:
  pip install openpyxl requests
"""

import argparse
import json
import re
import sys
from datetime import datetime, date, timedelta
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Error: openpyxl not installed. Run: pip install openpyxl")

try:
    import requests
except ImportError:
    sys.exit("Error: requests not installed. Run: pip install requests")


EXCEL_PATH = Path(__file__).parent / "additional_inputs.xlsx"
SHEET_NAME = "Brandstofprijzen"


# ---------------------------------------------------------------------------
# CBS OData: Diesel B7 pump price
# ---------------------------------------------------------------------------

def fetch_cbs_diesel_price() -> dict:
    """
    Fetch the latest diesel B7 pump price from CBS OData table 81567NED.
    This table has weekly fuel prices (pompprijzen motorbrandstoffen).
    Returns dict with price (EUR/L) and date.
    """
    print("[CBS] Fetching diesel B7 pump price from table 81567NED...")

    current_year = datetime.now().year
    min_year = current_year - 1  # Accept data from last year too

    # CBS OData table 81567NED: "Pompprijzen motorbrandstoffen; brandstofsoort, per dag"
    # Note: CBS OData $orderby doesn't work reliably, so we filter for recent periods
    # and sort in Python. Use $filter with startswith to get only recent data.

    # Try fetching recent data first (current year)
    for year in [current_year, current_year - 1]:
        year_prefix = str(year)
        url = (
            "https://opendata.cbs.nl/ODataApi/odata/81567NED/TypedDataSet"
            f"?$filter=startswith(Perioden,'{year_prefix}')"
            "&$format=json"
        )

        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            records = data.get("value", [])
            if records:
                print(f"  [CBS] Found {len(records)} records for {year}")
                break
        except requests.RequestException as e:
            print(f"  [CBS] Failed to fetch {year} data: {e}")
            records = []
            continue

    if not records:
        # Fallback: try without filter (might get old data)
        print("  [CBS] Trying fallback without year filter...")
        url = (
            "https://opendata.cbs.nl/ODataApi/odata/81567NED/TypedDataSet"
            "?$top=500"
            "&$format=json"
        )
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            records = data.get("value", [])
        except requests.RequestException as e:
            print(f"  [CBS] Fallback also failed: {e}")
            return {"price": None, "date": None, "error": str(e)}

    if not records:
        return {"price": None, "date": None, "error": "No records returned"}

    # Debug: show available columns
    first_record = records[0]
    all_cols = list(first_record.keys())
    print(f"  [CBS] All columns: {all_cols}")

    # Find price columns - CBS uses various naming patterns
    # Look for columns containing price-related terms
    price_columns = []
    for key in all_cols:
        key_lower = key.lower()
        # Skip metadata columns
        if key in ['ID', 'Perioden', 'BrandstofSoorten']:
            continue
        # Look for price/pompprijs columns
        if any(term in key_lower for term in ['prijs', 'price', 'eur', 'pomppr']):
            price_columns.append(key)
        # Also check for numeric-looking column names ending in _1, _2, etc.
        elif key_lower.endswith(('_1', '_2', '_3')) and key not in price_columns:
            price_columns.append(key)

    if not price_columns:
        # Just try all non-metadata columns
        price_columns = [k for k in all_cols if k not in ['ID', 'Perioden', 'BrandstofSoorten']]

    print(f"  [CBS] Price columns to try: {price_columns}")

    # Sort records by period (descending) - we need to do this in Python since CBS $orderby doesn't work
    def get_period_sortkey(rec):
        p = rec.get("Perioden", "")
        # Convert to sortable format: "2025MM01" -> "202501", "2025JJ00" -> "202500"
        return p.replace("MM", "").replace("JJ", "").replace("KW", "")

    records_sorted = sorted(records, key=get_period_sortkey, reverse=True)
    print(f"  [CBS] Latest periods: {[r.get('Perioden', '?') for r in records_sorted[:5]]}")

    # Find diesel records (BrandstofSoorten might indicate fuel type)
    # Look for records where BrandstofSoorten contains "Diesel" or similar
    diesel_records = []
    for record in records_sorted:
        brandstof = str(record.get("BrandstofSoorten", "")).lower()
        if "diesel" in brandstof or not record.get("BrandstofSoorten"):
            diesel_records.append(record)

    if not diesel_records:
        diesel_records = records_sorted  # Use all if no diesel-specific found

    # Find the latest record with a valid price
    for record in diesel_records:
        period = record.get("Perioden", "")

        # Parse period year
        try:
            period_year = int(str(period)[:4])
        except (ValueError, IndexError):
            continue

        # Skip old data
        if period_year < min_year:
            continue

        # Try each price column
        for col in price_columns:
            price_val = record.get(col)
            if price_val is not None:
                try:
                    price_float = float(price_val)
                    # CBS prices are in EUR/L, should be between 1.00 and 3.00
                    if 0.5 < price_float < 5.0:
                        price = round(price_float, 4)
                        brandstof = record.get("BrandstofSoorten", "?")
                        print(f"  [CBS] Diesel B7: EUR {price}/L (period: {period}, col: {col}, type: {brandstof})")
                        return {"price": price, "date": period}
                except (ValueError, TypeError):
                    continue

    print(f"  [CBS] No recent diesel price found (looking for {min_year}+)")
    return {"price": None, "date": None, "error": f"No diesel price found for {min_year}+"}


# ---------------------------------------------------------------------------
# EnergyZero: Electricity price
# ---------------------------------------------------------------------------

def fetch_energyzero_electricity() -> dict:
    """
    Fetch current Dutch electricity price from EnergyZero API.
    Returns dict with current price, 24h average, and timestamp.
    """
    print("[EnergyZero] Fetching electricity prices...")

    today = date.today()
    tomorrow = today + timedelta(days=1)

    # EnergyZero expects UTC timestamps
    from_date = f"{today.isoformat()}T00:00:00.000Z"
    till_date = f"{tomorrow.isoformat()}T00:00:00.000Z"

    url = (
        f"https://api.energyzero.nl/v1/energyprices"
        f"?fromDate={from_date}"
        f"&tillDate={till_date}"
        f"&interval=4"
        f"&usageType=1"
        f"&inclBtw=true"
    )

    try:
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as e:
        print(f"  [EnergyZero] Failed: {e}")
        return {"current": None, "average": None, "error": str(e)}

    prices_data = data.get("Prices", [])
    if not prices_data:
        return {"current": None, "average": None, "error": "No prices returned"}

    # Extract all price values
    price_values = []
    for p in prices_data:
        val = p.get("price")
        if val is not None:
            price_values.append(float(val))

    if not price_values:
        return {"current": None, "average": None, "error": "No valid price values"}

    # Find current hour's price
    now_hour = datetime.utcnow().hour
    current_price = None
    if now_hour < len(price_values):
        current_price = round(price_values[now_hour], 4)

    avg_price = round(sum(price_values) / len(price_values), 4)

    print(f"  [EnergyZero] Current: EUR {current_price}/kWh, "
          f"24h avg: EUR {avg_price}/kWh ({len(price_values)} hours)")

    return {
        "current": current_price,
        "average": avg_price,
        "date": today.isoformat(),
    }


# ---------------------------------------------------------------------------
# Fieten Olie: HVO100 + B7 advisory prices (web scrape)
# ---------------------------------------------------------------------------

def fetch_fieten_prices() -> dict:
    """
    Attempt to fetch HVO100 and Diesel B7 advisory prices from Fieten Olie.
    Returns dict with hvo100 and b7 prices in EUR/L.
    """
    print("[Fieten] Fetching advisory prices from fieten.info...")

    url = "https://www.fieten.info/adviesprijzen/"

    try:
        resp = requests.get(url, timeout=30, headers={
            "User-Agent": "Mozilla/5.0 (compatible; BusOmloop/1.0; fuel-price-fetcher)"
        })
        resp.raise_for_status()
        html = resp.text
    except requests.RequestException as e:
        print(f"  [Fieten] Failed to fetch page: {e}")
        return {"hvo100": None, "b7": None, "error": str(e)}

    # Try to extract prices from the HTML
    # Fieten shows prices per 100L; we need to find both Diesel and HVO100 values.
    # The page structure may vary, so we try multiple patterns.

    hvo100_price = None
    b7_price = None

    # Pattern: look for price values near "HVO" and "Diesel" text
    # Prices are typically shown as "XXX,XX" (Dutch decimal separator) per 100L

    # Try to find structured price data
    # Pattern for prices like "167,69" or "180,91"
    price_pattern = r'(\d{2,3}[,\.]\d{2})'

    # Look for HVO100 price
    hvo_sections = re.findall(
        r'(?:HVO\s*100?|hvo\s*100?).*?(' + price_pattern.replace('(', '(?:') + r')',
        html, re.IGNORECASE | re.DOTALL
    )

    # Look for Diesel price
    diesel_sections = re.findall(
        r'(?:Diesel|diesel)(?:\s*(?:B7)?)?.*?(' + price_pattern.replace('(', '(?:') + r')',
        html, re.IGNORECASE | re.DOTALL
    )

    # More general approach: find all price-like patterns and their context
    all_prices = []
    for match in re.finditer(r'(\d{2,3})[,.](\d{2})', html):
        start = max(0, match.start() - 200)
        context = html[start:match.end() + 50].lower()
        price_str = f"{match.group(1)}.{match.group(2)}"
        price_val = float(price_str)

        # Only consider reasonable fuel prices (per 100L: 100-300, per L: 1.00-3.00)
        if 100 <= price_val <= 300:
            # This is likely per 100L
            all_prices.append({
                "value_per_100l": price_val,
                "value_per_l": round(price_val / 100, 4),
                "context": context,
                "is_hvo": "hvo" in context,
                "is_diesel": "diesel" in context and "hvo" not in context,
            })

    # Find best matches
    for p in all_prices:
        if p["is_hvo"] and hvo100_price is None:
            hvo100_price = p["value_per_l"]
        if p["is_diesel"] and b7_price is None:
            b7_price = p["value_per_l"]

    if hvo100_price:
        print(f"  [Fieten] HVO100: EUR {hvo100_price}/L")
    else:
        print("  [Fieten] Could not extract HVO100 price from page")

    if b7_price:
        print(f"  [Fieten] Diesel B7: EUR {b7_price}/L")
    else:
        print("  [Fieten] Could not extract Diesel B7 price from page")

    return {
        "hvo100": hvo100_price,
        "b7": b7_price,
        "date": date.today().isoformat(),
        "note": "Prices scraped from fieten.info/adviesprijzen/; verify manually"
                if (hvo100_price or b7_price)
                else "Could not scrape prices; update manually",
    }


# ---------------------------------------------------------------------------
# HVO100 incentive calculation (contract formula)
# ---------------------------------------------------------------------------

def calculate_hvo_incentive(fieten_hvo, fieten_b7,
                            pk_hvo=None, pk_b7=None,
                            bp_hvo=None, bp_b7=None) -> dict:
    """
    Calculate HVO100 incentive per the NS contract formula.
    Uses average price difference across available sources.
    """
    diffs = []

    if fieten_hvo is not None and fieten_b7 is not None:
        diffs.append(fieten_hvo - fieten_b7)
    if pk_hvo is not None and pk_b7 is not None:
        diffs.append(pk_hvo - pk_b7)
    if bp_hvo is not None and bp_b7 is not None:
        diffs.append(bp_hvo - bp_b7)

    if not diffs:
        return {"diff": None, "incentive": None, "sources_used": 0}

    avg_diff = sum(diffs) / len(diffs)

    if avg_diff > 0:
        price_diff_component = min(avg_diff, 0.35)
        stimulans = 0.05
        incentive = min(price_diff_component + stimulans, 0.40)
    else:
        incentive = 0.0

    return {
        "diff": round(avg_diff, 4),
        "incentive": round(incentive, 4),
        "sources_used": len(diffs),
    }


# ---------------------------------------------------------------------------
# Update Excel
# ---------------------------------------------------------------------------

def update_excel(diesel_data, electricity_data, fieten_data, hvo_calc,
                 excel_path=EXCEL_PATH):
    """Write fetched prices into the Brandstofprijzen sheet."""
    if not excel_path.exists():
        sys.exit(f"Error: {excel_path} not found.")

    wb = openpyxl.load_workbook(excel_path)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Error: sheet '{SHEET_NAME}' not found in {excel_path}")

    ws = wb[SHEET_NAME]
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Build a lookup: variable name -> row number
    var_rows = {}
    for row in range(1, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if name:
            var_rows[name] = row

    def set_value(var_name, value, updated_str=None):
        if var_name in var_rows:
            r = var_rows[var_name]
            ws.cell(row=r, column=2, value=value)
            if updated_str:
                ws.cell(row=r, column=5, value=updated_str)

    # CBS Diesel
    if diesel_data and diesel_data.get("price"):
        set_value("diesel_b7_pompprijs_eur_per_liter",
                  diesel_data["price"],
                  f"{now_str} (CBS: {diesel_data.get('date', '?')})")

    # EnergyZero Electricity
    if electricity_data:
        if electricity_data.get("current") is not None:
            set_value("elektriciteit_prijs_eur_per_kwh",
                      electricity_data["current"],
                      f"{now_str} (uur {datetime.utcnow().hour}:00 UTC)")
        if electricity_data.get("average") is not None:
            set_value("elektriciteit_prijs_gemiddeld_24h_eur_per_kwh",
                      electricity_data["average"],
                      f"{now_str} (dag-gem.)")

    # Fieten Olie
    if fieten_data:
        if fieten_data.get("hvo100") is not None:
            set_value("hvo100_adviesprijs_fieten_eur_per_liter",
                      fieten_data["hvo100"], f"{now_str} (fieten.info)")
        if fieten_data.get("b7") is not None:
            set_value("diesel_b7_adviesprijs_fieten_eur_per_liter",
                      fieten_data["b7"], f"{now_str} (fieten.info)")

    # HVO incentive calculation
    if hvo_calc:
        if hvo_calc.get("diff") is not None:
            set_value("hvo100_b7_verschil_gemiddeld_eur_per_liter",
                      hvo_calc["diff"],
                      f"{now_str} ({hvo_calc['sources_used']} bronnen)")
        if hvo_calc.get("incentive") is not None:
            set_value("hvo100_vergoeding_per_liter",
                      hvo_calc["incentive"],
                      f"{now_str} (contractformule)")

    wb.save(excel_path)
    print(f"\nUpdated {excel_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Fetch latest fuel/electricity prices and update additional_inputs.xlsx"
    )
    parser.add_argument("--diesel-only", action="store_true",
                        help="Only fetch diesel B7 from CBS")
    parser.add_argument("--electricity-only", action="store_true",
                        help="Only fetch electricity from EnergyZero")
    parser.add_argument("--hvo-only", action="store_true",
                        help="Only fetch HVO100/B7 from Fieten Olie")
    parser.add_argument("--dry-run", action="store_true",
                        help="Fetch prices but don't write to Excel")
    parser.add_argument("--excel", type=str, default=str(EXCEL_PATH),
                        help=f"Path to Excel file (default: {EXCEL_PATH})")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    fetch_all = not (args.diesel_only or args.electricity_only or args.hvo_only)

    print("=" * 60)
    print("Fuel & Electricity Price Fetcher")
    print(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)
    print()

    diesel_data = None
    electricity_data = None
    fieten_data = None
    hvo_calc = None

    # Fetch diesel
    if fetch_all or args.diesel_only:
        diesel_data = fetch_cbs_diesel_price()
        print()

    # Fetch electricity
    if fetch_all or args.electricity_only:
        electricity_data = fetch_energyzero_electricity()
        print()

    # Fetch HVO100 / B7 advisory prices
    if fetch_all or args.hvo_only:
        fieten_data = fetch_fieten_prices()

        # Calculate incentive with available data (Fieten Olie is sufficient for internal use)
        # Note: NS contract requires average of 3 sources (Fieten, PK Energy, BP) for official claims
        hvo_calc = calculate_hvo_incentive(
            fieten_hvo=fieten_data.get("hvo100") if fieten_data else None,
            fieten_b7=fieten_data.get("b7") if fieten_data else None,
            pk_hvo=None, pk_b7=None,  # Optional: PK Energy
            bp_hvo=None, bp_b7=None,  # Optional: BP Nederland
        )

        if hvo_calc.get("diff") is not None:
            print(f"\n  NS HVO100 vergoeding (duurzaamheidsstimulans):")
            print(f"    HVO100 is hernieuwbare diesel (100% plantaardige olie).")
            print(f"    NS vergoedt het prijsverschil + stimulans aan vervoerders.")
            print(f"    ")
            print(f"    HVO100 prijs:        EUR {fieten_data.get('hvo100', '?')}/L")
            print(f"    Diesel B7 prijs:     EUR {fieten_data.get('b7', '?')}/L")
            print(f"    Prijsverschil:       EUR {hvo_calc['diff']}/L")
            print(f"    + Stimulans:         EUR 0.05/L")
            print(f"    ─────────────────────────────")
            print(f"    NS vergoeding:       EUR {hvo_calc['incentive']}/L (max €0.40)")
            print(f"    ")
            print(f"    → Met HVO100 kost brandstof effectief:")
            effective_cost = round(fieten_data.get('hvo100', 0) - hvo_calc['incentive'], 4)
            print(f"      {fieten_data.get('hvo100', '?')} - {hvo_calc['incentive']} = EUR {effective_cost}/L")
            diesel_cost = fieten_data.get('b7', 0)
            if diesel_cost and effective_cost:
                savings = round(diesel_cost - effective_cost, 4)
                print(f"      Dat is EUR {savings}/L {'goedkoper' if savings > 0 else 'duurder'} dan diesel!")
        print()

    # Write to Excel
    if not args.dry_run:
        update_excel(diesel_data, electricity_data, fieten_data, hvo_calc,
                     excel_path=excel_path)
    else:
        print("[DRY RUN] Skipping Excel update.")

    # Summary
    print()
    print("Summary:")
    print("-" * 40)
    if diesel_data and diesel_data.get("price"):
        print(f"  Diesel B7 (CBS):         EUR {diesel_data['price']}/L")
    else:
        print(f"  Diesel B7 (CBS):         not available")

    if electricity_data and electricity_data.get("current") is not None:
        print(f"  Electricity (current):   EUR {electricity_data['current']}/kWh")
        print(f"  Electricity (24h avg):   EUR {electricity_data['average']}/kWh")
    else:
        print(f"  Electricity:             not available")

    if fieten_data and fieten_data.get("hvo100"):
        print(f"  HVO100 (Fieten):         EUR {fieten_data['hvo100']}/L")
    else:
        print(f"  HVO100 (Fieten):         not available")

    if fieten_data and fieten_data.get("b7"):
        print(f"  Diesel B7 (Fieten):      EUR {fieten_data['b7']}/L")
    else:
        print(f"  Diesel B7 (Fieten):      not available")

    if hvo_calc and hvo_calc.get("incentive") is not None:
        print(f"  HVO100 incentive:        EUR {hvo_calc['incentive']}/L")

    print()
    print("Prijzen bijgewerkt. De financiele calculator gebruikt deze prijzen automatisch.")


if __name__ == "__main__":
    main()
