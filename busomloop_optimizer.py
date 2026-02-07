"""
Busomloop Optimizer - NS Trein Vervangend Vervoer (TVV)
=======================================================
Leest het invoer-Excel bestand (Bijlage J) in en genereert:
  1. Busomloop per bus (Transvision-stijl)
  2. Overzicht van ritsamenhang
  3. Berekeningen en optimalisatie-details

Gebruik:
    python busomloop_optimizer.py <invoer.xlsx> [--output <uitvoer.xlsx>]

Keertijden worden automatisch bepaald uit de data (kleinste gap per bustype,
minimum 2 minuten). Handmatig overschrijven kan met:
    --keer-dd 15  --keer-tc 8  --keer-lvb 12  --keer-midi 10  --keer-taxi 5
"""

from __future__ import annotations

import argparse
import datetime
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Financial calculator for version 7+
try:
    from financial_calculator import (
        load_financial_config, calculate_total_financials,
        calculate_rotation_financials, FinancialConfig
    )
    FINANCIAL_CALCULATOR_AVAILABLE = True
except ImportError:
    FINANCIAL_CALCULATOR_AVAILABLE = False


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class Trip:
    trip_id: str
    bus_nr: int
    service: str          # Sheet / busdienst name
    date_str: str         # e.g. "do 11-06-2026"
    date_label: str       # e.g. "donderdag 11 juni"
    direction: str        # "heen" or "terug"
    bus_type: str         # Dubbeldekker, Touringcar, Taxibus
    snel_stop: str        # snelbus / stopbus
    pattern: str
    multiplicity: int     # Aantal bussen
    origin_code: str
    origin_name: str
    origin_halt: str
    dest_code: str
    dest_name: str
    dest_halt: str
    departure: int        # minutes from midnight
    arrival: int          # minutes from midnight
    stops: list           # list of (station_code, station_name, halt_code, halt_name, time_minutes)
    copy_nr: int = 1      # which copy (1..multiplicity)
    is_reserve: bool = False  # phantom trip for reserve bus duty

    @property
    def duration(self) -> int:
        return self.arrival - self.departure

    def time_str(self, minutes: int) -> str:
        h, m = divmod(minutes % 1440, 60)
        return f"{h:02d}:{m:02d}"

    @property
    def dep_str(self) -> str:
        return self.time_str(self.departure)

    @property
    def arr_str(self) -> str:
        return self.time_str(self.arrival)


@dataclass
class BusRotation:
    bus_id: str
    bus_type: str
    date_str: str
    trips: list = field(default_factory=list)
    _cached_idle_minutes: int = field(default=None, repr=False)

    @property
    def start_time(self) -> int:
        return self.trips[0].departure if self.trips else 0

    @property
    def end_time(self) -> int:
        return self.trips[-1].arrival if self.trips else 0

    @property
    def total_ride_minutes(self) -> int:
        return sum(t.arrival - t.departure for t in self.trips if not t.is_reserve)

    @property
    def total_reserve_minutes(self) -> int:
        return sum(t.arrival - t.departure for t in self.trips if t.is_reserve)

    @property
    def total_idle_minutes(self) -> int:
        # Return cached value if set (for multiday rotations with cross-day gaps)
        if self._cached_idle_minutes is not None:
            return self._cached_idle_minutes
        idle = 0
        for i in range(1, len(self.trips)):
            gap = self.trips[i].departure - self.trips[i - 1].arrival
            idle += gap
        return idle

    @total_idle_minutes.setter
    def total_idle_minutes(self, value: int):
        self._cached_idle_minutes = value

    @property
    def total_dienst_minutes(self) -> int:
        return self.end_time - self.start_time if self.trips else 0

    @property
    def real_trips(self) -> list:
        return [t for t in self.trips if not t.is_reserve]

    @property
    def reserve_trip_list(self) -> list:
        return [t for t in self.trips if t.is_reserve]


@dataclass
class ReserveBus:
    station: str
    count: int
    day: str
    start: int       # minutes from midnight
    end: int         # minutes from midnight
    remark: str = ""


# ---------------------------------------------------------------------------
# ZE (Zero Emission) Data Classes - Version 6
# ---------------------------------------------------------------------------

@dataclass
class ChargingStation:
    """A charging station near a bus station."""
    name: str
    operator: str
    lat: float
    lon: float
    distance_km: float
    max_power_kw: float
    num_points: int
    connectors: list
    category: str  # ultra_fast, fast, semi_fast, slow
    drive_time_min: float = None  # Actual driving time from Google Maps (if available)
    drive_distance_km: float = None  # Actual driving distance from Google Maps (if available)


@dataclass
class ZEFeasibility:
    """ZE feasibility assessment for a rotation."""
    rotation_id: str
    bus_type: str
    total_km: float
    ze_range_km: float
    is_feasible: bool
    needs_charging: bool
    buffer_km: float  # remaining range after rotation
    charging_opportunities: list  # [(station, idle_window_min, chargers), ...]
    recommended_charging: list  # planned charging stops
    reason: str  # why feasible/not feasible


# ---------------------------------------------------------------------------
# Time helpers
# ---------------------------------------------------------------------------

def time_to_minutes(t) -> Optional[int]:
    """Convert datetime.time or decimal fraction to minutes from midnight."""
    if t is None:
        return None
    if isinstance(t, datetime.time):
        return t.hour * 60 + t.minute
    if isinstance(t, (int, float)):
        # Could be decimal day fraction
        total = round(t * 24 * 60)
        return total
    return None


def minutes_to_str(m: int) -> str:
    if m is None:
        return ""
    h, mi = divmod(m % 1440, 60)
    return f"{h:02d}:{mi:02d}"


def minutes_to_time(m: int) -> datetime.time:
    h, mi = divmod(m % 1440, 60)
    return datetime.time(h, mi)


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------

def parse_reserve_buses(wb) -> list:
    """Parse reserve bus info from Voorblad."""
    ws = wb["Voorblad"]
    reserves = []
    in_reserve = False
    for row in ws.iter_rows(min_row=1, max_row=100, max_col=15, values_only=False):
        cell_vals = {c.column: c.value for c in row}
        if cell_vals.get(1) == "Reservebussen":
            in_reserve = True
            continue
        if in_reserve and cell_vals.get(2) == "Station":
            continue  # header row
        if in_reserve and cell_vals.get(2) and cell_vals.get(4):
            station = cell_vals.get(2, "")
            count = int(cell_vals.get(4, 0))
            day = cell_vals.get(5, "")
            start_t = time_to_minutes(cell_vals.get(6))
            end_t = time_to_minutes(cell_vals.get(8))
            remark = cell_vals.get(10, "") or ""
            if start_t is not None and end_t is not None:
                reserves.append(ReserveBus(station, count, day, start_t, end_t, remark))
    return reserves


BUS_TYPE_VALUES = {"dubbeldekker", "touringcar", "taxibus", "lagevloerbus",
                    "gelede bus", "midi bus", "midibus"}


def normalize_bus_type(raw: str) -> str:
    """Normalize bus type names to canonical form matching the pricing categories."""
    low = raw.strip().lower()
    if "dubbeldek" in low:
        return "Dubbeldekker"
    if "touring" in low:
        return "Touringcar"
    if "lagevloer" in low or "gelede" in low:
        return "Lagevloerbus"
    if "midi" in low:
        return "Midi bus"
    if "taxi" in low:
        return "Taxibus"
    # Return with title case if unknown
    return raw.strip() if raw.strip() else "Onbekend"


def parse_direction_block(ws, start_row, max_col=100):
    """
    Parse a 'Busbewegingen in heen/terugrichting' block.
    Returns list of dicts with trip info per bus column.
    """
    bus_numbers = {}
    patterns = {}
    bus_types = {}
    snel_stop = {}
    aantal = {}
    looptijd = {}
    stops = []

    for row in ws.iter_rows(min_row=start_row, max_row=start_row + 80,
                            max_col=max_col, values_only=False):
        cell_map = {c.column: c.value for c in row}
        label = str(cell_map.get(5, "") or "").strip()
        col1 = cell_map.get(1, "")

        data_vals = {col: val for col, val in cell_map.items()
                     if col >= 6 and val is not None}

        if label == "Busnummer":
            bus_numbers = data_vals
        elif label == "Patroon":
            # Detect what this "Patroon" row actually contains
            first_val = next(iter(data_vals.values()), "") if data_vals else ""
            if isinstance(first_val, (int, float)) and not isinstance(first_val, datetime.time):
                # Integer values = these are bus numbers
                bus_numbers = data_vals
            elif str(first_val).strip().lower() in BUS_TYPE_VALUES:
                bus_types = data_vals
            elif isinstance(first_val, datetime.time):
                # Time values = this is actually Looptijd
                looptijd = {col: time_to_minutes(val) for col, val in data_vals.items()}
            else:
                patterns = data_vals
        elif label == "Type bus":
            bus_types = data_vals
        elif "Snelbus" in label or "Stopbus" in label:
            snel_stop = data_vals
        elif label == "Aantal bussen":
            aantal = data_vals
        elif "Looptijd" in label:
            looptijd = {col: time_to_minutes(val) for col, val in data_vals.items()}
        elif cell_map.get(2) and cell_map.get(3):
            # Station row - must have station code in col 2 and name in col 3
            station_code = cell_map.get(2, "")
            station_name = cell_map.get(3, "")
            halt_code = cell_map.get(4, "")
            halt_name = cell_map.get(5, "")
            rijtijd = cell_map.get(1, "")
            times = {}
            for col, val in data_vals.items():
                t = time_to_minutes(val)
                if t is not None:
                    times[col] = t
            stops.append({
                "station_code": str(station_code),
                "station_name": str(station_name),
                "halt_code": str(halt_code) if halt_code else "",
                "halt_name": str(halt_name) if halt_name else "",
                "rijtijd": rijtijd,
                "times": times,
            })
        elif col1 and str(col1).startswith("Uurpatronen"):
            break
        elif col1 and str(col1).startswith("Busbewegingen"):
            break

    return bus_numbers, patterns, bus_types, snel_stop, aantal, looptijd, stops


def parse_sheet(wb, sheet_name) -> list:
    """Parse a single service sheet and return list of Trip objects."""
    ws = wb[sheet_name]
    trips = []

    # Read busdienst and datum
    service = ""
    date_str = ""
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=5, values_only=False):
        cell_map = {c.column: c.value for c in row}
        if cell_map.get(1) == "Busdienst":
            service = str(cell_map.get(2, ""))
        elif cell_map.get(1) == "Datum":
            date_str = str(cell_map.get(2, ""))

    # Find direction blocks
    max_col = ws.max_column or 20
    if max_col > 100:
        max_col = 100

    heen_start = None
    terug_start = None
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row or 100,
                     max_col=5, values_only=False), start=1
    ):
        cell_map = {c.column: c.value for c in row}
        val = cell_map.get(1, "")
        if val and "Busbewegingen in heenrichting" in str(val):
            heen_start = row_idx + 1
        elif val and "Busbewegingen in terugrichting" in str(val):
            terug_start = row_idx + 1

    trip_counter = 0

    for direction, start_row in [("heen", heen_start), ("terug", terug_start)]:
        if start_row is None:
            continue
        bus_numbers, patterns, bus_types_map, snel_stop_map, aantal_map, looptijd_map, stops = \
            parse_direction_block(ws, start_row, max_col)

        if not bus_numbers or not stops:
            continue

        # For each bus column, build a trip
        for col, bus_nr in sorted(bus_numbers.items()):
            # Check this column has times in stops
            col_times = []
            for s in stops:
                t = s["times"].get(col)
                if t is not None:
                    col_times.append((s, t))

            if len(col_times) < 2:
                continue

            origin = col_times[0]
            dest = col_times[-1]
            all_stops = [(s["station_code"], s["station_name"],
                          s["halt_code"], s["halt_name"], t)
                         for s, t in col_times]

            multiplicity = int(aantal_map.get(col, 1))
            bus_type = normalize_bus_type(str(bus_types_map.get(col, "Onbekend")))
            pattern = str(patterns.get(col, ""))
            snel = str(snel_stop_map.get(col, ""))

            # Handle times that cross midnight
            dep_time = origin[1]
            arr_time = dest[1]
            if arr_time < dep_time:
                arr_time += 1440  # next day

            for copy in range(1, multiplicity + 1):
                trip_counter += 1
                trip_id = f"{sheet_name}_{direction}_{bus_nr}_{copy}"

                t = Trip(
                    trip_id=trip_id,
                    bus_nr=bus_nr,
                    service=service,
                    date_str=date_str,
                    date_label=sheet_name,
                    direction=direction,
                    bus_type=bus_type,
                    snel_stop=snel,
                    pattern=pattern,
                    multiplicity=multiplicity,
                    origin_code=origin[0]["station_code"],
                    origin_name=origin[0]["station_name"],
                    origin_halt=origin[0]["halt_name"],
                    dest_code=dest[0]["station_code"],
                    dest_name=dest[0]["station_name"],
                    dest_halt=dest[0]["halt_name"],
                    departure=dep_time,
                    arrival=arr_time,
                    stops=all_stops,
                    copy_nr=copy,
                )
                trips.append(t)

    return trips


def parse_all_sheets(input_file: str):
    """Parse all service sheets from the input Excel."""
    wb = openpyxl.load_workbook(input_file, data_only=True)
    all_trips = []
    reserves = parse_reserve_buses(wb)

    for sheet_name in wb.sheetnames:
        if sheet_name == "Voorblad":
            continue
        try:
            sheet_trips = parse_sheet(wb, sheet_name)
            all_trips.extend(sheet_trips)
        except Exception as e:
            print(f"  Waarschuwing: Fout bij parsen van '{sheet_name}': {e}")

    return all_trips, reserves, wb.sheetnames


# ---------------------------------------------------------------------------
# Optimizer - Greedy best-fit bus chaining
# ---------------------------------------------------------------------------

# Default minimum turnaround times per bus type (minutes)
MIN_TURNAROUND_DEFAULTS = {
    "Dubbeldekker": 8,
    "Touringcar": 6,
    "Lagevloerbus": 5,
    "Midi bus": 4,
    "Taxibus": 3,
}
MIN_TURNAROUND_FALLBACK = 6  # fallback for unknown bus types
MIN_TURNAROUND_FLOOR = 2     # absolute minimum turnaround (minutes)


def detect_turnaround_times(trips: list, within_service_only: bool = False) -> dict:
    """
    Auto-detect minimum turnaround time per bus type from the trip data.

    If within_service_only=True, only considers trips from the same service
    (= same Excel tab). This gives a conservative baseline turnaround since
    it avoids accidental short gaps between unrelated services.

    If within_service_only=False, considers all trips at the same location
    regardless of service (the original behavior).

    Returns dict {bus_type: minutes}, with a floor of MIN_TURNAROUND_FLOOR.
    """
    # Group arrivals and departures
    arrivals = {}
    departures = {}

    for t in trips:
        dest_loc = normalize_location(t.dest_code)
        orig_loc = normalize_location(t.origin_code)
        if within_service_only:
            arr_key = (t.bus_type, t.date_str, dest_loc, t.service)
            dep_key = (t.bus_type, t.date_str, orig_loc, t.service)
        else:
            arr_key = (t.bus_type, t.date_str, dest_loc)
            dep_key = (t.bus_type, t.date_str, orig_loc)
        arrivals.setdefault(arr_key, []).append(t.arrival)
        departures.setdefault(dep_key, []).append(t.departure)

    # For each bus type, find minimum gap between any arrival and subsequent departure
    min_gap_per_type = {}

    for arr_key, arr_times in arrivals.items():
        bus_type = arr_key[0]
        dep_key = arr_key  # same key structure
        if dep_key not in departures:
            continue

        dep_times = sorted(departures[dep_key])
        for arr_t in arr_times:
            for dep_t in dep_times:
                gap = dep_t - arr_t
                if gap >= MIN_TURNAROUND_FLOOR:
                    if bus_type not in min_gap_per_type or gap < min_gap_per_type[bus_type]:
                        min_gap_per_type[bus_type] = gap
                    break

    result = {}
    for bus_type in set(t.bus_type for t in trips):
        if bus_type in min_gap_per_type:
            result[bus_type] = min_gap_per_type[bus_type]
        else:
            result[bus_type] = MIN_TURNAROUND_FALLBACK

    return result


def detect_turnaround_per_service(trips: list) -> dict:
    """
    Detect the minimum turnaround time per service (= per Excel tab).
    Returns dict {service_name: (bus_type, min_gap_minutes or None, dates_list, n_trips, directions)}.
    """
    by_service = {}
    for t in trips:
        by_service.setdefault(t.service, []).append(t)

    result = {}
    for service, svc_trips in by_service.items():
        bus_type = svc_trips[0].bus_type if svc_trips else "Onbekend"
        dates = sorted(set(t.date_str for t in svc_trips))
        n_trips = len(svc_trips)
        directions = sorted(set(t.direction for t in svc_trips))

        # Group by (date, location) to avoid cross-day comparisons
        arrivals = {}   # (date, loc) -> [arrival_minutes]
        departures = {} # (date, loc) -> [departure_minutes]
        for t in svc_trips:
            dest_loc = normalize_location(t.dest_code)
            orig_loc = normalize_location(t.origin_code)
            arrivals.setdefault((t.date_str, dest_loc), []).append(t.arrival)
            departures.setdefault((t.date_str, orig_loc), []).append(t.departure)

        min_gap = None
        for (date, loc), arr_times in arrivals.items():
            dep_times = departures.get((date, loc), [])
            dep_sorted = sorted(dep_times)
            for arr_t in arr_times:
                for dep_t in dep_sorted:
                    gap = dep_t - arr_t
                    if gap >= MIN_TURNAROUND_FLOOR:
                        if min_gap is None or gap < min_gap:
                            min_gap = gap
                        break

        # min_gap is None when no turnaround exists (e.g. one-way only)
        result[service] = (bus_type, min_gap, dates, n_trips, directions)

    return result


def normalize_location(code: str) -> str:
    """Normalize station codes for matching (e.g. same city = same location).

    Uses the dynamically built STATION_REGISTRY if available, otherwise
    falls back to the code itself (lowercased).
    """
    code = code.strip().lower()
    # Look up in dynamic registry first
    if code in _STATION_CODE_TO_CANONICAL:
        return _STATION_CODE_TO_CANONICAL[code]
    return code


def normalize_reserve_station(station_name: str) -> str:
    """Normalize a reserve bus station name to match normalize_location output.

    Uses the dynamically built STATION_REGISTRY if available.
    """
    name = station_name.strip().lower()
    if name in _STATION_NAME_TO_CANONICAL:
        return _STATION_NAME_TO_CANONICAL[name]
    # Fallback: clean up to a slug-like form
    return _name_to_canonical(name)


def _name_to_canonical(name: str) -> str:
    """Convert a station name to a canonical key (lowercase, underscored)."""
    # "Driebergen-Zeist" -> "driebergen-zeist"
    # "Utrecht Centraal" -> "utrecht centraal"
    return name.strip().lower()


# ---------------------------------------------------------------------------
# Station registry - built dynamically from input data
# ---------------------------------------------------------------------------
# Maps: station_code (lowercase) -> canonical key
_STATION_CODE_TO_CANONICAL: dict = {}
# Maps: station_name (lowercase) -> canonical key
_STATION_NAME_TO_CANONICAL: dict = {}
# Maps: canonical key -> display name (for output/Google Maps)
_CANONICAL_TO_DISPLAY: dict = {}
# Maps: canonical key -> set of halt names (for Google Maps address building)
_CANONICAL_TO_HALTS: dict = {}


def build_station_registry(all_trips: list, reserves: list = None):
    """Build the station registry from parsed trip data.

    This populates the module-level lookup dicts so that normalize_location()
    and normalize_reserve_station() work correctly.

    Must be called after parse_all_sheets() and before optimize_rotations().
    """
    _STATION_CODE_TO_CANONICAL.clear()
    _STATION_NAME_TO_CANONICAL.clear()
    _CANONICAL_TO_DISPLAY.clear()
    _CANONICAL_TO_HALTS.clear()

    def _register(code: str, name: str, halt: str = ""):
        if not code or not name:
            return
        code_lower = code.strip().lower()
        name_clean = name.strip()
        canonical = _name_to_canonical(name_clean)
        _STATION_CODE_TO_CANONICAL[code_lower] = canonical
        _STATION_NAME_TO_CANONICAL[canonical] = canonical
        _CANONICAL_TO_DISPLAY[canonical] = name_clean
        if halt and halt.strip():
            _CANONICAL_TO_HALTS.setdefault(canonical, set()).add(halt.strip())

    # Collect station code -> name -> halt mappings from all trip stops
    for t in all_trips:
        _register(t.origin_code, t.origin_name, t.origin_halt)
        _register(t.dest_code, t.dest_name, t.dest_halt)

        # Also register intermediate stops if available
        if hasattr(t, 'stops') and t.stops:
            for stop in t.stops:
                s_code = stop[0] if len(stop) > 0 else ""
                s_name = stop[1] if len(stop) > 1 else ""
                s_halt = stop[3] if len(stop) > 3 else ""
                _register(s_code, s_name, s_halt)

    # Register reserve bus station names (no halt info available)
    if reserves:
        for rb in reserves:
            if rb.station:
                name_clean = rb.station.strip()
                canonical = _name_to_canonical(name_clean)
                _STATION_NAME_TO_CANONICAL[canonical] = canonical
                if canonical not in _CANONICAL_TO_DISPLAY:
                    _CANONICAL_TO_DISPLAY[canonical] = name_clean

    return dict(_CANONICAL_TO_DISPLAY)  # return for external use


def get_station_registry() -> dict:
    """Return the current station registry: {canonical_key: display_name}."""
    return dict(_CANONICAL_TO_DISPLAY)


def get_station_halts() -> dict:
    """Return halt info per station: {canonical_key: set of halt names}."""
    return {k: set(v) for k, v in _CANONICAL_TO_HALTS.items()}


# ---------------------------------------------------------------------------
# ZE (Zero Emission) Configuration - Version 6
# ---------------------------------------------------------------------------

# Default fuel configuration (used if additional_inputs.xlsx not available)
FUEL_DEFAULTS = {
    # Diesel range per bus type (km) - from tank capacity / consumption
    "diesel_range_km": {
        "Touringcar": 1562,      # 500L tank / 32L per 100km
        "Dubbeldekker": 889,     # 400L tank / 45L per 100km
        "Lagevloerbus": 789,     # 300L tank / 38L per 100km
        "Midi bus": 800,         # 200L tank / 25L per 100km
        "Taxibus": 667,          # 80L tank / 12L per 100km
    },
    # Diesel consumption per bus type (L/100km)
    "diesel_consumption_l_per_100km": {
        "Touringcar": 32,
        "Dubbeldekker": 45,
        "Lagevloerbus": 38,
        "Midi bus": 25,
        "Taxibus": 12,
    },
    # Refuel time in minutes (including buffer)
    "refuel_time_min": 20,       # 15 min tank + 5 min buffer
    # Average speed to fuel station (km/h)
    "speed_to_station_kmh": 30,
}

# Default ZE configuration (used if additional_inputs.xlsx not available)
ZE_DEFAULTS = {
    "ze_range_km": {
        "Touringcar": 300,
        "Dubbeldekker": 250,
        "Lagevloerbus": 280,
        "Midi bus": 350,
        "Taxibus": 400,
    },
    "ze_consumption_kwh_per_100km": {
        "Touringcar": 130,
        "Dubbeldekker": 180,
        "Lagevloerbus": 150,
        "Midi bus": 100,
        "Taxibus": 50,
    },
    # Estimated average speed for km calculation (km/h)
    "avg_speed_kmh": {
        "Touringcar": 50,
        "Dubbeldekker": 45,
        "Lagevloerbus": 40,
        "Midi bus": 45,
        "Taxibus": 50,
    },
}


def load_ze_config(inputs_xlsx: str = "additional_inputs.xlsx") -> dict:
    """Load ZE configuration from additional_inputs.xlsx Buskosten sheet.

    Also supports the older financieel_input.xlsx format for backward compatibility.

    Returns dict with:
        - ze_range_km: {bus_type: range_km}
        - ze_consumption_kwh_per_100km: {bus_type: kwh}
        - avg_speed_kmh: {bus_type: km/h for estimating distances}
    """
    import json
    config = {
        "ze_range_km": dict(ZE_DEFAULTS["ze_range_km"]),
        "ze_consumption_kwh_per_100km": dict(ZE_DEFAULTS["ze_consumption_kwh_per_100km"]),
        "avg_speed_kmh": dict(ZE_DEFAULTS["avg_speed_kmh"]),
    }

    path = Path(inputs_xlsx)
    if not path.exists():
        print(f"  ZE config: {inputs_xlsx} niet gevonden, standaardwaarden gebruikt")
        return config

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Buskosten" not in wb.sheetnames:
            print(f"  ZE config: 'Buskosten' blad niet gevonden, standaardwaarden gebruikt")
            wb.close()
            return config

        ws = wb["Buskosten"]

        # Parse the sheet to find ZE range and consumption values
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            var_name = str(row[0]).lower()
            value = row[1]

            # ZE range (actieradius)
            if "actieradius" in var_name and "_ze" in var_name and value:
                if "dubbeldekker" in var_name:
                    config["ze_range_km"]["Dubbeldekker"] = float(value)
                elif "touringcar" in var_name:
                    config["ze_range_km"]["Touringcar"] = float(value)
                elif "lagevloer" in var_name:
                    config["ze_range_km"]["Lagevloerbus"] = float(value)
                elif "midi" in var_name:
                    config["ze_range_km"]["Midi bus"] = float(value)
                elif "taxi" in var_name:
                    config["ze_range_km"]["Taxibus"] = float(value)

            # ZE consumption (kWh per 100km)
            if "verbruik" in var_name and "_ze" in var_name and "kwh" in var_name and value:
                if "dubbeldekker" in var_name:
                    config["ze_consumption_kwh_per_100km"]["Dubbeldekker"] = float(value)
                elif "touringcar" in var_name:
                    config["ze_consumption_kwh_per_100km"]["Touringcar"] = float(value)
                elif "lagevloer" in var_name:
                    config["ze_consumption_kwh_per_100km"]["Lagevloerbus"] = float(value)
                elif "midi" in var_name:
                    config["ze_consumption_kwh_per_100km"]["Midi bus"] = float(value)
                elif "taxi" in var_name:
                    config["ze_consumption_kwh_per_100km"]["Taxibus"] = float(value)

        wb.close()
        print(f"  ZE config geladen: bereik Touringcar = {config['ze_range_km']['Touringcar']} km")
    except Exception as e:
        print(f"  ZE config laden mislukt ({e}), standaardwaarden gebruikt")

    return config


def load_fuel_config(inputs_xlsx: str = "additional_inputs.xlsx") -> dict:
    """Load fuel configuration from additional_inputs.xlsx Buskosten sheet.

    Returns dict with:
        - diesel_range_km: {bus_type: range_km}
        - diesel_consumption_l_per_100km: {bus_type: liters}
        - refuel_time_min: total refuel time including buffer
        - speed_to_station_kmh: average speed driving to fuel station
        - avg_speed_kmh: {bus_type: km/h for estimating trip distances}
    """
    config = {
        "diesel_range_km": dict(FUEL_DEFAULTS["diesel_range_km"]),
        "diesel_consumption_l_per_100km": dict(FUEL_DEFAULTS["diesel_consumption_l_per_100km"]),
        "refuel_time_min": FUEL_DEFAULTS["refuel_time_min"],
        "speed_to_station_kmh": FUEL_DEFAULTS["speed_to_station_kmh"],
        "avg_speed_kmh": dict(ZE_DEFAULTS["avg_speed_kmh"]),
    }

    path = Path(inputs_xlsx)
    if not path.exists():
        print(f"  Fuel config: {inputs_xlsx} niet gevonden, standaardwaarden gebruikt")
        return config

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Buskosten" not in wb.sheetnames:
            print(f"  Fuel config: 'Buskosten' blad niet gevonden, standaardwaarden gebruikt")
            wb.close()
            return config

        ws = wb["Buskosten"]

        # Parse the sheet to find diesel range and other values
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            var_name = str(row[0]).lower()
            value = row[1]

            # Diesel range (actieradius)
            if "actieradius" in var_name and "_diesel" in var_name and value:
                if "dubbeldekker" in var_name:
                    config["diesel_range_km"]["Dubbeldekker"] = float(value)
                elif "touringcar" in var_name:
                    config["diesel_range_km"]["Touringcar"] = float(value)
                elif "lagevloer" in var_name:
                    config["diesel_range_km"]["Lagevloerbus"] = float(value)
                elif "midi" in var_name:
                    config["diesel_range_km"]["Midi bus"] = float(value)
                elif "taxi" in var_name:
                    config["diesel_range_km"]["Taxibus"] = float(value)

            # Diesel consumption (L per 100km)
            if "verbruik" in var_name and "_diesel" in var_name and "l_per" in var_name and value:
                if "dubbeldekker" in var_name:
                    config["diesel_consumption_l_per_100km"]["Dubbeldekker"] = float(value)
                elif "touringcar" in var_name:
                    config["diesel_consumption_l_per_100km"]["Touringcar"] = float(value)
                elif "lagevloer" in var_name:
                    config["diesel_consumption_l_per_100km"]["Lagevloerbus"] = float(value)
                elif "midi" in var_name:
                    config["diesel_consumption_l_per_100km"]["Midi bus"] = float(value)
                elif "taxi" in var_name:
                    config["diesel_consumption_l_per_100km"]["Taxibus"] = float(value)

            # Refuel time
            if "tanktijd_diesel" in var_name and value:
                config["refuel_time_min"] = float(value)
            if "tanktijd_buffer" in var_name and value:
                config["refuel_time_min"] += float(value)

            # Speed to station
            if "snelheid_naar_tankstation" in var_name and value:
                config["speed_to_station_kmh"] = float(value)

            # Average speed per bus type (for km estimation)
            if "avg_snelheid" in var_name and "_kmh" in var_name and value:
                if "dubbeldekker" in var_name:
                    config["avg_speed_kmh"]["Dubbeldekker"] = float(value)
                elif "touringcar" in var_name:
                    config["avg_speed_kmh"]["Touringcar"] = float(value)
                elif "lagevloer" in var_name:
                    config["avg_speed_kmh"]["Lagevloerbus"] = float(value)
                elif "midi" in var_name:
                    config["avg_speed_kmh"]["Midi bus"] = float(value)
                elif "taxi" in var_name:
                    config["avg_speed_kmh"]["Taxibus"] = float(value)

        wb.close()
        print(f"  Fuel config geladen: bereik Touringcar = {config['diesel_range_km']['Touringcar']} km, "
              f"snelheid = {config['avg_speed_kmh']['Touringcar']} km/h")
    except Exception as e:
        print(f"  Fuel config laden mislukt ({e}), standaardwaarden gebruikt")

    return config


def load_fuel_stations(tanklocaties_json: str = "tanklocaties.json") -> dict:
    """Load fuel station data from tanklocaties.json.

    Returns: {station_name: [{"name": str, "distance_km": float, "fuel_types": list,
                              "drive_time_min": float, "drive_distance_km": float}, ...]}
    """
    import json
    path = Path(tanklocaties_json)
    if not path.exists():
        print(f"  Tankstations: {tanklocaties_json} niet gevonden")
        return {}

    try:
        with open(path, "r") as f:
            data = json.load(f)

        stations_by_location = {}
        has_drive_times = data.get("metadata", {}).get("has_drive_times", False)

        for station_name, station_data in data.get("stations", {}).items():
            fuel_stations = []
            for fs in station_data.get("fuel_stations", []):
                fuel_stations.append({
                    "name": fs.get("name", "Unknown"),
                    "distance_km": fs.get("distance_km", 0),
                    "fuel_types": fs.get("fuel_types", ["diesel"]),
                    "drive_time_min": fs.get("drive_time_min"),
                    "drive_distance_km": fs.get("drive_distance_km"),
                })

            # Sort by distance (prefer closest)
            fuel_stations.sort(key=lambda s: s["distance_km"])
            stations_by_location[station_name] = fuel_stations

        total_stations = sum(len(v) for v in stations_by_location.values())
        drive_info = " (met rijtijden)" if has_drive_times else ""
        print(f"  Tankstations geladen: {len(stations_by_location)} locaties, {total_stations} tankpunten{drive_info}")
        return stations_by_location
    except Exception as e:
        print(f"  Tankstations laden mislukt: {e}")
        return {}


def load_charging_stations(tanklocaties_json: str = "tanklocaties.json") -> dict:
    """Load charging station data from tanklocaties.json.

    Returns: {station_name: [ChargingStation, ...]}
    """
    import json
    path = Path(tanklocaties_json)
    if not path.exists():
        print(f"  Laadstations: {tanklocaties_json} niet gevonden")
        return {}

    try:
        with open(path, "r") as f:
            data = json.load(f)

        stations_by_location = {}

        for station_name, station_data in data.get("stations", {}).items():
            chargers = []
            for cs in station_data.get("charging_stations", []):
                chargers.append(ChargingStation(
                    name=cs.get("name", "Unknown"),
                    operator=cs.get("operator", ""),
                    lat=cs.get("lat", 0),
                    lon=cs.get("lon", 0),
                    distance_km=cs.get("distance_km", 0),
                    max_power_kw=cs.get("max_power_kw", 0),
                    num_points=cs.get("num_points", 1),
                    connectors=cs.get("connectors", []),
                    category=cs.get("category", "slow"),
                    drive_time_min=cs.get("drive_time_min"),
                    drive_distance_km=cs.get("drive_distance_km"),
                ))

            # Sort by power (prefer fast chargers) then by distance
            chargers.sort(key=lambda c: (-c.max_power_kw, c.distance_km))
            stations_by_location[station_name] = chargers

        total_chargers = sum(len(v) for v in stations_by_location.values())
        print(f"  Laadstations geladen: {len(stations_by_location)} locaties, {total_chargers} laadpunten")
        return stations_by_location
    except Exception as e:
        print(f"  Laadstations laden mislukt: {e}")
        return {}


def calculate_gmaps_avg_speed(deadhead_matrix: dict = None,
                               deadhead_km_matrix: dict = None) -> float | None:
    """Calculate actual average speed from Google Maps deadhead data.

    Uses all available distance/duration pairs from the deadhead matrices
    to compute the weighted average speed: sum(km) / sum(hours).

    This gives a more accurate average speed for the specific geographic area
    covered by the routes, accounting for actual road conditions.

    Args:
        deadhead_matrix: {origin: {dest: duration_min}} from Google Maps
        deadhead_km_matrix: {origin: {dest: distance_km}} from Google Maps

    Returns:
        Average speed in km/h, or None if no valid data available.
    """
    if not deadhead_matrix or not deadhead_km_matrix:
        return None

    total_km = 0.0
    total_hours = 0.0

    for origin, dests_time in deadhead_matrix.items():
        if origin not in deadhead_km_matrix:
            continue
        dests_km = deadhead_km_matrix[origin]

        for dest, duration_min in dests_time.items():
            if dest not in dests_km or origin == dest:
                continue
            distance_km = dests_km[dest]

            # Skip invalid entries
            if not duration_min or duration_min <= 0:
                continue
            if not distance_km or distance_km <= 0:
                continue

            total_km += distance_km
            total_hours += duration_min / 60

    if total_hours <= 0:
        return None

    avg_speed = total_km / total_hours
    return round(avg_speed, 1)


def load_bus_speed_factors(inputs_xlsx: str = "additional_inputs.xlsx") -> dict:
    """Load bus speed factors from additional_inputs.xlsx.

    These factors convert Google Maps car speed to bus speed (buses are slower).

    Returns dict: {bus_type: factor} where factor is typically 0.85-0.95
    """
    # Default factors (used if Excel doesn't have them)
    factors = {
        "Touringcar": 0.95,
        "Dubbeldekker": 0.90,
        "Lagevloerbus": 0.85,
        "Midi bus": 0.92,
        "Taxibus": 0.95,
    }

    path = Path(inputs_xlsx)
    if not path.exists():
        return factors

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        if "Buskosten" not in wb.sheetnames:
            wb.close()
            return factors

        ws = wb["Buskosten"]

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if not row or not row[0]:
                continue
            var_name = str(row[0]).lower()
            value = row[1]

            if "snelheidsfactor" in var_name and value:
                if "dubbeldekker" in var_name:
                    factors["Dubbeldekker"] = float(value)
                elif "touringcar" in var_name:
                    factors["Touringcar"] = float(value)
                elif "lagevloer" in var_name:
                    factors["Lagevloerbus"] = float(value)
                elif "midi" in var_name:
                    factors["Midi bus"] = float(value)
                elif "taxi" in var_name:
                    factors["Taxibus"] = float(value)

        wb.close()
    except Exception as e:
        print(f"  Snelheidsfactoren laden mislukt ({e}), standaardwaarden gebruikt")

    return factors


def update_config_with_gmaps_speed(config: dict, deadhead_matrix: dict = None,
                                    deadhead_km_matrix: dict = None,
                                    inputs_xlsx: str = "additional_inputs.xlsx") -> dict:
    """Update fuel/ZE config with calculated average speed from Google Maps.

    If Google Maps distance data is available, calculates the actual average
    driving speed for the area and uses it for all bus types (as the geographic
    conditions are the same). Falls back to configured values if no data.

    Args:
        config: Fuel or ZE config dict with avg_speed_kmh key
        deadhead_matrix: {origin: {dest: duration_min}}
        deadhead_km_matrix: {origin: {dest: distance_km}}
        inputs_xlsx: Path to additional_inputs.xlsx for speed factors

    Returns:
        Updated config dict with potentially adjusted avg_speed_kmh values.
    """
    gmaps_speed = calculate_gmaps_avg_speed(deadhead_matrix, deadhead_km_matrix)

    if gmaps_speed is None:
        return config

    print(f"  Google Maps gemiddelde snelheid berekend: {gmaps_speed} km/h")

    # Load bus speed factors from Excel (buses are slower than cars)
    bus_speed_factor = load_bus_speed_factors(inputs_xlsx)

    for bus_type in config["avg_speed_kmh"]:
        factor = bus_speed_factor.get(bus_type, 0.90)
        adjusted_speed = round(gmaps_speed * factor, 1)
        config["avg_speed_kmh"][bus_type] = adjusted_speed

    print(f"  Snelheden aangepast op basis van Google Maps data (factoren uit {inputs_xlsx}):")
    for bus_type, speed in sorted(config["avg_speed_kmh"].items()):
        factor = bus_speed_factor.get(bus_type, 0.90)
        print(f"    {bus_type}: {speed} km/h (factor {factor})")

    return config


def estimate_rotation_km(rotation: BusRotation, ze_config: dict) -> float:
    """Estimate total km for a rotation based on ride time and average speed."""
    bus_type = rotation.bus_type
    avg_speed = ze_config["avg_speed_kmh"].get(bus_type, 45)

    # Use total ride minutes (excluding waiting time)
    ride_minutes = rotation.total_ride_minutes
    if ride_minutes <= 0:
        return 0.0

    # km = speed (km/h) × time (h)
    return avg_speed * (ride_minutes / 60)


def get_idle_windows(rotation: BusRotation, min_idle_min: int = 30) -> list:
    """Get idle windows for a rotation where charging could happen.

    Returns: [(station_code, start_min, end_min, duration_min), ...]
    """
    windows = []
    trips = rotation.trips

    for i in range(1, len(trips)):
        prev_trip = trips[i - 1]
        next_trip = trips[i]

        gap = next_trip.departure - prev_trip.arrival
        if gap >= min_idle_min:
            # The bus is waiting at the destination of prev_trip
            station = normalize_location(prev_trip.dest_code)
            windows.append((station, prev_trip.arrival, next_trip.departure, gap))

    return windows


def analyze_ze_feasibility(rotation: BusRotation, ze_config: dict,
                           charging_stations: dict) -> ZEFeasibility:
    """Analyze whether a rotation can be done with a ZE bus."""
    bus_type = rotation.bus_type
    ze_range = ze_config["ze_range_km"].get(bus_type, 300)
    total_km = estimate_rotation_km(rotation, ze_config)

    # Simple feasibility: total km <= range
    buffer_km = ze_range - total_km
    is_feasible_without_charging = buffer_km >= 0

    # Check charging opportunities during idle windows
    idle_windows = get_idle_windows(rotation)
    charging_opportunities = []

    for station, start_min, end_min, duration_min in idle_windows:
        # Find matching charging station (try exact match and normalized)
        chargers = charging_stations.get(station, [])
        if not chargers:
            # Try normalized version
            for loc_name in charging_stations:
                if normalize_location(loc_name) == station:
                    chargers = charging_stations[loc_name]
                    break

        fast_chargers = [c for c in chargers if c.max_power_kw >= 50]
        if fast_chargers:
            charging_opportunities.append((station, duration_min, fast_chargers))

    # Calculate if charging can extend range sufficiently
    needs_charging = not is_feasible_without_charging
    is_feasible_with_charging = False
    recommended_charging = []

    if needs_charging and charging_opportunities:
        consumption = ze_config["ze_consumption_kwh_per_100km"].get(bus_type, 130)

        total_recoverable_km = 0
        for station, duration_min, chargers in charging_opportunities:
            best_charger = chargers[0]  # Already sorted by power

            # Account for driving time to/from charger
            # Use Google Maps drive time if available, else estimate from distance
            if best_charger.drive_time_min:
                drive_time_one_way = best_charger.drive_time_min
                extra_km_one_way = best_charger.drive_distance_km or best_charger.distance_km
            else:
                # Assume 30 km/h average speed in urban area
                drive_time_one_way = (best_charger.distance_km / 30) * 60  # minutes
                extra_km_one_way = best_charger.distance_km
            drive_time_total = 2 * drive_time_one_way  # round trip
            extra_km = 2 * extra_km_one_way  # round trip adds to total km

            # Actual time available for charging
            actual_charge_time = duration_min - drive_time_total
            if actual_charge_time < 10:  # Need at least 10 min to make it worthwhile
                continue

            # kWh charged = power * time * efficiency
            kwh_charged = best_charger.max_power_kw * (actual_charge_time / 60) * 0.8
            km_recovered = (kwh_charged / consumption) * 100

            # Net km benefit = km recovered - extra km driven to charger
            net_km_benefit = km_recovered - extra_km
            if net_km_benefit <= 0:
                continue

            total_recoverable_km += net_km_benefit

            if total_km <= ze_range + total_recoverable_km:
                recommended_charging.append({
                    "station": station,
                    "duration_min": duration_min,
                    "drive_time_min": round(drive_time_total, 0),
                    "actual_charge_min": round(actual_charge_time, 0),
                    "charger": best_charger.name,
                    "charger_distance_km": best_charger.distance_km,
                    "power_kw": best_charger.max_power_kw,
                    "km_recovered": round(km_recovered, 1),
                    "extra_km_driven": round(extra_km, 1),
                    "net_km_benefit": round(net_km_benefit, 1),
                })
                is_feasible_with_charging = True
                break

    is_feasible = is_feasible_without_charging or is_feasible_with_charging

    # Determine reason
    if is_feasible_without_charging:
        reason = f"Bereik voldoende: {total_km:.0f} km < {ze_range:.0f} km (buffer: {buffer_km:.0f} km)"
    elif is_feasible_with_charging:
        reason = f"Haalbaar met opladen tijdens wachttijd"
    else:
        reason = f"Niet haalbaar: {total_km:.0f} km > {ze_range:.0f} km"

    return ZEFeasibility(
        rotation_id=rotation.bus_id,
        bus_type=bus_type,
        total_km=total_km,
        ze_range_km=ze_range,
        is_feasible=is_feasible,
        needs_charging=needs_charging and is_feasible,
        buffer_km=buffer_km,
        charging_opportunities=charging_opportunities,
        recommended_charging=recommended_charging,
        reason=reason,
    )


def assign_ze_buses(rotations: list, min_ze_count: int, ze_config: dict,
                    charging_stations: dict,
                    target_bus_type: str = "Touringcar") -> dict:
    """Assign ZE to rotations, ensuring minimum count for target bus type.

    Returns: {rotation_id: ZEFeasibility}
    """
    # Filter to target bus type
    target_rotations = [r for r in rotations if r.bus_type == target_bus_type]

    if len(target_rotations) < min_ze_count:
        print(f"  Waarschuwing: slechts {len(target_rotations)} {target_bus_type} omlopen, "
              f"maar {min_ze_count} ZE vereist")

    # Analyze feasibility for all target rotations
    feasibility_results = []
    for rotation in target_rotations:
        feas = analyze_ze_feasibility(rotation, ze_config, charging_stations)
        feasibility_results.append(feas)

    # Sort by suitability for ZE:
    # 1. Feasible without charging first
    # 2. Then by buffer (more buffer = safer choice)
    # 3. Then by total km (shorter = better)
    def ze_score(f: ZEFeasibility) -> tuple:
        return (
            0 if f.is_feasible and not f.needs_charging else 1,
            0 if f.is_feasible else 1,
            -f.buffer_km,
            f.total_km,
        )

    feasibility_results.sort(key=ze_score)

    # Assign ZE to top N feasible rotations
    ze_assignments = {}
    assigned_count = 0

    for feas in feasibility_results:
        if feas.is_feasible and assigned_count < min_ze_count:
            ze_assignments[feas.rotation_id] = feas
            assigned_count += 1

    if assigned_count < min_ze_count:
        print(f"  Waarschuwing: slechts {assigned_count} ZE toewijsbaar, "
              f"{min_ze_count} vereist")

    # Include all feasibility results for reporting
    all_results = {f.rotation_id: f for f in feasibility_results}

    return ze_assignments, all_results, assigned_count


# ---------------------------------------------------------------------------
# Fuel Constraint Validation - Version 6 (Integrated into optimization)
# ---------------------------------------------------------------------------

@dataclass
class FuelStop:
    """A planned fuel stop during a rotation."""
    station_name: str
    fuel_station_name: str
    fuel_station_distance_km: float
    idle_start_min: int
    idle_end_min: int
    idle_duration_min: int
    drive_time_min: float
    km_before_stop: float


@dataclass
class FuelValidationResult:
    """Result of fuel constraint validation for a rotation."""
    rotation_id: str
    bus_type: str
    total_km: float
    fuel_range_km: float
    is_feasible: bool
    needs_refuel: bool
    fuel_stops: list  # List of FuelStop
    split_points: list  # List of trip indices where chain should be split
    reason: str


def estimate_trip_km(trip: Trip, fuel_config: dict) -> float:
    """Estimate km for a single trip based on duration and average speed."""
    avg_speed = fuel_config["avg_speed_kmh"].get(trip.bus_type, 45)
    duration_hours = trip.duration / 60
    return avg_speed * duration_hours


def validate_fuel_feasibility(rotation: BusRotation, fuel_config: dict,
                               fuel_stations: dict,
                               deadhead_matrix: dict = None,
                               deadhead_km_matrix: dict = None) -> FuelValidationResult:
    """Validate fuel feasibility for a rotation.

    Checks if the cumulative km exceeds fuel range, and if so, whether
    refueling opportunities exist during idle windows.

    Args:
        rotation: The bus rotation to validate
        fuel_config: Fuel configuration (range, speeds, etc.)
        fuel_stations: Fuel stations per location
        deadhead_matrix: Optional {origin: {dest: minutes}} for deadhead time
        deadhead_km_matrix: Optional {origin: {dest: km}} from Google Maps
            - If available, uses actual km from Google Maps (more accurate)
            - If not, estimates km from duration × avg speed

    Returns FuelValidationResult with:
    - is_feasible: True if can complete rotation (with or without refueling)
    - needs_refuel: True if refueling is needed
    - fuel_stops: List of planned fuel stops
    - split_points: If not feasible, indices where chain should be split
    """
    bus_type = rotation.bus_type
    fuel_range = fuel_config["diesel_range_km"].get(bus_type, 1000)
    refuel_time = fuel_config["refuel_time_min"]
    speed_to_station = fuel_config["speed_to_station_kmh"]

    trips = rotation.trips
    fuel_stops = []
    split_points = []

    cumulative_km = 0.0
    remaining_range = fuel_range

    for i, trip in enumerate(trips):
        trip_km = estimate_trip_km(trip, fuel_config)

        # Add deadhead km from previous trip (if any)
        deadhead_km = 0.0
        if i > 0:
            prev_dest = normalize_location(trips[i - 1].dest_code)
            curr_orig = normalize_location(trip.origin_code)
            if prev_dest != curr_orig:
                # Prefer Google Maps distance_km if available
                if deadhead_km_matrix:
                    deadhead_km = deadhead_km_matrix.get(prev_dest, {}).get(curr_orig, 0) or 0

                # Fall back to estimating from time if no km data
                if deadhead_km == 0 and deadhead_matrix:
                    dh_time = deadhead_matrix.get(prev_dest, {}).get(curr_orig, 0)
                    if dh_time > 0:
                        # Estimate deadhead km from time (assume avg speed to station)
                        deadhead_km = (dh_time / 60) * speed_to_station

                trip_km += deadhead_km

        # Check if adding this trip exceeds remaining range
        if trip_km > remaining_range and i > 0:
            # Need to check for refueling opportunity BEFORE this trip
            prev_trip = trips[i - 1]
            idle_gap = trip.departure - prev_trip.arrival

            # Find nearest fuel station at the location
            station_loc = normalize_location(prev_trip.dest_code)
            stations = fuel_stations.get(station_loc, [])

            # Try normalized match if exact match fails
            if not stations:
                for loc_name in fuel_stations:
                    if normalize_location(loc_name) == station_loc:
                        stations = fuel_stations[loc_name]
                        break

            can_refuel = False
            if stations:
                nearest = stations[0]
                # Use Google Maps drive time if available, else estimate from distance
                if nearest.get("drive_time_min"):
                    drive_time_one_way = nearest["drive_time_min"]
                else:
                    drive_time_one_way = (nearest["distance_km"] / speed_to_station) * 60
                drive_time_total = 2 * drive_time_one_way  # round trip
                total_time_needed = refuel_time + drive_time_total

                if idle_gap >= total_time_needed:
                    can_refuel = True
                    fuel_stops.append(FuelStop(
                        station_name=station_loc,
                        fuel_station_name=nearest["name"],
                        fuel_station_distance_km=nearest.get("drive_distance_km") or nearest["distance_km"],
                        idle_start_min=prev_trip.arrival,
                        idle_end_min=trip.departure,
                        idle_duration_min=idle_gap,
                        drive_time_min=drive_time_total,
                        km_before_stop=cumulative_km,
                    ))
                    # Reset range after refueling
                    remaining_range = fuel_range
                    cumulative_km = 0.0

            if not can_refuel:
                # Cannot refuel, mark split point
                split_points.append(i)
                # Reset for new bus
                remaining_range = fuel_range
                cumulative_km = 0.0

        # Add trip km to cumulative
        cumulative_km += trip_km
        remaining_range -= trip_km

    total_km = sum(estimate_trip_km(t, fuel_config) for t in trips)
    is_feasible = len(split_points) == 0
    needs_refuel = len(fuel_stops) > 0

    if is_feasible and not needs_refuel:
        reason = f"Bereik voldoende: {total_km:.0f} km < {fuel_range:.0f} km"
    elif is_feasible and needs_refuel:
        reason = f"Haalbaar met {len(fuel_stops)} tankstop(s)"
    else:
        reason = f"Niet haalbaar: {len(split_points)} splits nodig"

    return FuelValidationResult(
        rotation_id=rotation.bus_id,
        bus_type=bus_type,
        total_km=total_km,
        fuel_range_km=fuel_range,
        is_feasible=is_feasible,
        needs_refuel=needs_refuel,
        fuel_stops=fuel_stops,
        split_points=split_points,
        reason=reason,
    )


def _calculate_rotation_km(rotation: BusRotation, fuel_config: dict,
                           deadhead_matrix: dict = None,
                           deadhead_km_matrix: dict = None) -> float:
    """Calculate total km for a rotation including deadhead."""
    total_km = 0.0
    speed_to_station = fuel_config.get("speed_to_station_kmh", 30)

    for i, trip in enumerate(rotation.trips):
        trip_km = estimate_trip_km(trip, fuel_config)

        # Add deadhead km from previous trip
        if i > 0:
            prev_dest = normalize_location(rotation.trips[i - 1].dest_code)
            curr_orig = normalize_location(trip.origin_code)
            if prev_dest != curr_orig:
                deadhead_km = 0.0
                if deadhead_km_matrix:
                    deadhead_km = deadhead_km_matrix.get(prev_dest, {}).get(curr_orig, 0) or 0
                if deadhead_km == 0 and deadhead_matrix:
                    dh_time = deadhead_matrix.get(prev_dest, {}).get(curr_orig, 0)
                    if dh_time > 0:
                        deadhead_km = (dh_time / 60) * speed_to_station
                trip_km += deadhead_km

        total_km += trip_km

    return total_km


def _can_insert_trip_with_fuel(rotation: BusRotation, trip: Trip, insert_pos: str,
                                fuel_config: dict, fuel_stations: dict,
                                deadhead_matrix: dict = None,
                                deadhead_km_matrix: dict = None) -> tuple:
    """Check if a trip can be inserted into a rotation respecting fuel constraints.

    Args:
        rotation: Existing rotation
        trip: Trip to insert
        insert_pos: "start" or "end"
        fuel_config, fuel_stations, deadhead_matrix, deadhead_km_matrix: Fuel config

    Returns: (can_insert, needs_refuel, refuel_info)
    """
    bus_type = rotation.bus_type
    fuel_range = fuel_config["diesel_range_km"].get(bus_type, 1000)
    refuel_time = fuel_config["refuel_time_min"]
    speed_to_station = fuel_config.get("speed_to_station_kmh", 30)

    # Calculate current rotation km
    current_km = _calculate_rotation_km(rotation, fuel_config, deadhead_matrix, deadhead_km_matrix)

    # Calculate additional km from the new trip
    trip_km = estimate_trip_km(trip, fuel_config)

    # Calculate deadhead km for connection
    deadhead_km = 0.0
    if insert_pos == "end" and rotation.trips:
        last_dest = normalize_location(rotation.trips[-1].dest_code)
        trip_orig = normalize_location(trip.origin_code)
        if last_dest != trip_orig:
            if deadhead_km_matrix:
                deadhead_km = deadhead_km_matrix.get(last_dest, {}).get(trip_orig, 0) or 0
            if deadhead_km == 0 and deadhead_matrix:
                dh_time = deadhead_matrix.get(last_dest, {}).get(trip_orig, 0)
                if dh_time > 0:
                    deadhead_km = (dh_time / 60) * speed_to_station
    elif insert_pos == "start" and rotation.trips:
        trip_dest = normalize_location(trip.dest_code)
        first_orig = normalize_location(rotation.trips[0].origin_code)
        if trip_dest != first_orig:
            if deadhead_km_matrix:
                deadhead_km = deadhead_km_matrix.get(trip_dest, {}).get(first_orig, 0) or 0
            if deadhead_km == 0 and deadhead_matrix:
                dh_time = deadhead_matrix.get(trip_dest, {}).get(first_orig, 0)
                if dh_time > 0:
                    deadhead_km = (dh_time / 60) * speed_to_station

    total_new_km = current_km + trip_km + deadhead_km

    if total_new_km <= fuel_range:
        return True, False, None

    # Need refueling - check if there's an opportunity
    if insert_pos == "end" and rotation.trips:
        last_trip = rotation.trips[-1]
        idle_gap = trip.departure - last_trip.arrival

        station_loc = normalize_location(last_trip.dest_code)
        stations = fuel_stations.get(station_loc, [])

        if not stations:
            for loc_name in fuel_stations:
                if normalize_location(loc_name) == station_loc:
                    stations = fuel_stations[loc_name]
                    break

        if stations:
            nearest = stations[0]
            if nearest.get("drive_time_min"):
                drive_time_one_way = nearest["drive_time_min"]
            else:
                drive_time_one_way = (nearest["distance_km"] / speed_to_station) * 60
            drive_time_total = 2 * drive_time_one_way
            total_time_needed = refuel_time + drive_time_total

            if idle_gap >= total_time_needed:
                return True, True, {
                    "station": station_loc,
                    "fuel_station": nearest["name"],
                    "drive_time": drive_time_total,
                    "idle_gap": idle_gap
                }

    return False, False, None


def apply_fuel_constraints(rotations: list, fuel_config: dict,
                           fuel_stations: dict,
                           deadhead_matrix: dict = None,
                           deadhead_km_matrix: dict = None,
                           turnaround_map: dict = None,
                           algorithm: str = "mincost") -> tuple:
    """Apply fuel constraints to rotations with iterative re-optimization.

    When fuel constraints cause splits, the remaining trips are re-optimized
    using the selected algorithm instead of just being assigned greedily.

    Algorithm:
    1. Validate all rotations for fuel feasibility
    2. Keep feasible rotations and trips up to first split point
    3. Collect remaining trips after split points
    4. Re-run optimization on remaining trips using selected algorithm
    5. Repeat until no more splits needed

    Args:
        rotations: List of bus rotations
        fuel_config: Fuel configuration (range, speeds, etc.)
        fuel_stations: Fuel stations per location
        deadhead_matrix: Optional {origin: {dest: minutes}} for deadhead time
        deadhead_km_matrix: Optional {origin: {dest: km}} from Google Maps
        turnaround_map: Optional {bus_type: min_turnaround_minutes}
        algorithm: Algorithm to use for re-optimization ("greedy" or "mincost")

    Returns: (new_rotations, validation_results, split_count)
    """
    if turnaround_map is None:
        turnaround_map = dict(MIN_TURNAROUND_DEFAULTS)

    all_feasible_rotations = []
    validation_results = {}
    total_split_count = 0
    iteration = 0
    max_iterations = 10  # Safety limit

    # Current rotations to process
    current_rotations = list(rotations)

    while current_rotations and iteration < max_iterations:
        iteration += 1
        remaining_trips = []  # Trips that need re-optimization

        for rotation in current_rotations:
            result = validate_fuel_feasibility(
                rotation, fuel_config, fuel_stations, deadhead_matrix, deadhead_km_matrix
            )
            validation_results[rotation.bus_id] = result

            if result.is_feasible:
                # No splits needed, keep rotation as-is
                all_feasible_rotations.append(rotation)
            else:
                # Need to split
                total_split_count += 1
                trips = rotation.trips
                first_split = result.split_points[0] if result.split_points else len(trips)

                # Keep trips before first split as a rotation
                first_segment = trips[:first_split]
                if first_segment:
                    new_rot = BusRotation(
                        bus_id=f"{rotation.bus_id}_s{iteration}",
                        bus_type=rotation.bus_type,
                        date_str=rotation.date_str,
                        trips=list(first_segment),
                    )
                    all_feasible_rotations.append(new_rot)

                # Collect remaining trips for re-optimization
                remaining_segment = trips[first_split:]
                remaining_trips.extend(remaining_segment)

        if not remaining_trips:
            break  # All rotations are feasible

        # Re-optimize remaining trips using selected algorithm
        # Group by date and bus type
        from collections import defaultdict
        groups = defaultdict(list)
        for trip in remaining_trips:
            key = (trip.date_str, trip.bus_type)
            groups[key].append(trip)

        # Select algorithm function
        algo_func = _optimize_greedy if algorithm == "greedy" else _optimize_mincost

        # Run selected algorithm on each group and convert to rotations
        new_rotations = []
        for (date_str, bus_type), group_trips in groups.items():
            if not group_trips:
                continue

            # Use selected algorithm to re-optimize these trips
            chains = algo_func(
                group_trips, turnaround_map,
                service_constraint=False,
                deadhead_matrix=deadhead_matrix,
                trip_turnaround_overrides=None
            )

            # Convert chains to rotations
            sorted_trips = sorted(group_trips, key=lambda t: (t.departure, t.arrival))
            for chain_idx, chain in enumerate(chains):
                chain_trips = [sorted_trips[i] for i in chain]
                rot = BusRotation(
                    bus_id=f"fuel_i{iteration}_{date_str}_{bus_type[:2]}_{chain_idx+1:03d}",
                    bus_type=bus_type,
                    date_str=date_str,
                    trips=chain_trips,
                )
                new_rotations.append(rot)

        if iteration > 1:
            print(f"    Fuel re-optimization iteration {iteration}: "
                  f"{len(remaining_trips)} trips → {len(new_rotations)} buses")

        # These new rotations become the input for next iteration
        current_rotations = new_rotations

    # Add any remaining rotations from final iteration
    for rot in current_rotations:
        result = validate_fuel_feasibility(
            rot, fuel_config, fuel_stations, deadhead_matrix, deadhead_km_matrix
        )
        if result.is_feasible:
            all_feasible_rotations.append(rot)
        else:
            # Still infeasible - split and keep what we can
            trips = rot.trips
            first_split = result.split_points[0] if result.split_points else len(trips)
            if first_split > 0:
                new_rot = BusRotation(
                    bus_id=f"{rot.bus_id}_final",
                    bus_type=rot.bus_type,
                    date_str=rot.date_str,
                    trips=list(trips[:first_split]),
                )
                all_feasible_rotations.append(new_rot)
            # Remaining trips become single-trip rotations
            for i, trip in enumerate(trips[first_split:]):
                single_rot = BusRotation(
                    bus_id=f"{rot.bus_id}_single_{i+1}",
                    bus_type=rot.bus_type,
                    date_str=rot.date_str,
                    trips=[trip],
                )
                all_feasible_rotations.append(single_rot)
            print(f"    Warning: rotation {rot.bus_id} still infeasible after {iteration} iterations, created single-trip buses")

    # Recalculate idle times for all rotations
    for rot in all_feasible_rotations:
        if len(rot.trips) > 1:
            total_idle = 0
            for i in range(len(rot.trips) - 1):
                gap = rot.trips[i + 1].departure - rot.trips[i].arrival
                if deadhead_matrix:
                    dh = deadhead_matrix.get(
                        normalize_location(rot.trips[i].dest_code), {}
                    ).get(normalize_location(rot.trips[i + 1].origin_code), 0)
                    gap = max(0, gap - dh)
                total_idle += max(0, gap)
            rot.total_idle_minutes = total_idle

    return all_feasible_rotations, validation_results, total_split_count


def match_reserve_day(reserve_day: str, trip_dates: list) -> str:
    """Match a reserve day string to a trip date_str.
    E.g. 'donderdag 11 juni' -> 'do 11-06-2026'
    """
    day_lower = reserve_day.strip().lower()
    # Extract day number
    parts = day_lower.split()
    for trip_date in trip_dates:
        # trip_date format: "do 11-06-2026"
        td_parts = trip_date.split()
        day_num = td_parts[1].split("-")[0]  # "11"
        # Check if day number appears in reserve day and weekday prefix matches
        day_map = {
            "maandag": "ma", "dinsdag": "di", "woensdag": "wo",
            "donderdag": "do", "vrijdag": "vr", "zaterdag": "za", "zondag": "zo"
        }
        for full, short in day_map.items():
            if full in day_lower and td_parts[0] == short and day_num in parts:
                return trip_date
    return ""


def analyze_reserve_coverage(rotations: list, reserves: list, trip_dates: list) -> list:
    """
    Analyze which bus rotations cover reserve bus requirements.

    A rotation covers a reserve requirement if:
    - It has a gap (idle period) at the reserve station
    - The gap fully covers the reserve time window
    - The rotation's bus is at that station during the entire reserve window

    Returns list of dicts with coverage info per reserve requirement.
    """
    results = []

    for rb in reserves:
        date_str = match_reserve_day(rb.day, trip_dates)
        res_loc = normalize_reserve_station(rb.station)

        # Find rotations that are idle at this station during the reserve window
        covering_buses = []

        for rot in rotations:
            if rot.date_str != date_str:
                continue

            # Check each gap between consecutive trips
            for i in range(len(rot.trips) - 1):
                prev_trip = rot.trips[i]
                next_trip = rot.trips[i + 1]

                # Bus is at prev_trip's destination after prev_trip ends
                bus_loc = normalize_location(prev_trip.dest_code)
                idle_start = prev_trip.arrival
                idle_end = next_trip.departure

                # Does this gap cover the reserve window?
                if (bus_loc == res_loc and
                    idle_start <= rb.start and
                    idle_end >= rb.end):
                    covering_buses.append({
                        "bus_id": rot.bus_id,
                        "bus_type": rot.bus_type,
                        "idle_start": idle_start,
                        "idle_end": idle_end,
                        "prev_trip": f"{prev_trip.origin_name}→{prev_trip.dest_name}",
                        "next_trip": f"{next_trip.origin_name}→{next_trip.dest_name}",
                    })

            # Also check: bus arrives at last trip and has no more trips
            if rot.trips:
                last = rot.trips[-1]
                bus_loc = normalize_location(last.dest_code)
                if (bus_loc == res_loc and last.arrival <= rb.start):
                    covering_buses.append({
                        "bus_id": rot.bus_id,
                        "bus_type": rot.bus_type,
                        "idle_start": last.arrival,
                        "idle_end": 1440,  # end of day
                        "prev_trip": f"{last.origin_name}→{last.dest_name}",
                        "next_trip": "(einde dienst)",
                    })

            # Also check: bus starts first trip from reserve station
            if rot.trips:
                first = rot.trips[0]
                bus_loc = normalize_location(first.origin_code)
                if (bus_loc == res_loc and first.departure >= rb.end):
                    covering_buses.append({
                        "bus_id": rot.bus_id,
                        "bus_type": rot.bus_type,
                        "idle_start": 0,
                        "idle_end": first.departure,
                        "prev_trip": "(start dienst)",
                        "next_trip": f"{first.origin_name}→{first.dest_name}",
                    })

        results.append({
            "reserve": rb,
            "date_str": date_str,
            "location": res_loc,
            "required": rb.count,
            "covered": len(covering_buses),
            "covering_buses": covering_buses,
            "shortfall": max(0, rb.count - len(covering_buses)),
        })

    return results


def optimize_reserve_idle_matching(rotations: list, reserves: list, trip_dates: list) -> list:
    """
    Optimally allocate idle bus time to cover reserve requirements using
    bipartite matching (Hopcroft-Karp).  Maximises the number of reserve
    slots covered by buses that are already idle at the right station.

    Returns a list of dicts (same format as analyze_reserve_coverage) but
    with the optimal assignment.
    """
    # Build idle slots: (rotation, location, idle_start, idle_end, date_str)
    idle_slots = []
    for rot in rotations:
        # Gaps between consecutive trips
        for i in range(len(rot.trips) - 1):
            prev_t = rot.trips[i]
            next_t = rot.trips[i + 1]
            loc = normalize_location(prev_t.dest_code)
            idle_slots.append((rot, loc, prev_t.arrival, next_t.departure))
        # After last trip
        if rot.trips:
            last = rot.trips[-1]
            idle_slots.append((rot, normalize_location(last.dest_code), last.arrival, 1440))
        # Before first trip
        if rot.trips:
            first = rot.trips[0]
            idle_slots.append((rot, normalize_location(first.origin_code), 0, first.departure))

    # Expand reserves into individual slots (one per count)
    reserve_slots = []  # (ReserveBus, copy_idx, date_str, normalized_location)
    for rb in reserves:
        date_str = match_reserve_day(rb.day, trip_dates)
        res_loc = normalize_reserve_station(rb.station)
        for i in range(rb.count):
            reserve_slots.append((rb, i, date_str, res_loc))

    n_idle = len(idle_slots)
    n_res = len(reserve_slots)

    # Build bipartite adjacency: idle slot i → reserve slot j
    adj = [[] for _ in range(n_idle)]
    for i, (rot, loc, istart, iend) in enumerate(idle_slots):
        for j, (rb, _, date_str, res_loc) in enumerate(reserve_slots):
            if rot.date_str == date_str and loc == res_loc and istart <= rb.start and iend >= rb.end:
                adj[i].append(j)

    match_l, match_r = _hopcroft_karp(adj, n_idle, n_res)

    # Build results grouped by original ReserveBus
    results = []
    slot_idx = 0
    for rb in reserves:
        date_str = match_reserve_day(rb.day, trip_dates)
        res_loc = normalize_reserve_station(rb.station)
        covered = 0
        covering_buses = []
        for _ in range(rb.count):
            if slot_idx < n_res and match_r[slot_idx] != -1:
                idle_idx = match_r[slot_idx]
                rot = idle_slots[idle_idx][0]
                covering_buses.append({"bus_id": rot.bus_id, "bus_type": rot.bus_type})
                covered += 1
            slot_idx += 1
        results.append({
            "reserve": rb,
            "date_str": date_str,
            "location": res_loc,
            "required": rb.count,
            "covered": covered,
            "covering_buses": covering_buses,
            "shortfall": max(0, rb.count - covered),
        })
    return results


def assign_reserves_to_bus_types(reserves: list, all_trips: list) -> list:
    """
    Assign each reserve requirement to the bus type most common at that
    station on that day.

    Returns list of (ReserveBus, bus_type, date_str) tuples.
    """
    trip_dates = sorted(set(t.date_str for t in all_trips))

    # Count trips per (normalized_station, date, bus_type)
    station_type_count = {}
    for t in all_trips:
        for loc_code in [t.origin_code, t.dest_code]:
            loc = normalize_location(loc_code)
            key = (loc, t.date_str, t.bus_type)
            station_type_count[key] = station_type_count.get(key, 0) + 1

    assignments = []
    for rb in reserves:
        date_str = match_reserve_day(rb.day, trip_dates)
        if not date_str:
            continue
        res_loc = normalize_reserve_station(rb.station)

        # Find bus type with most trips at this station on this day
        best_type = None
        best_count = 0
        for (loc, d, bt), count in station_type_count.items():
            if loc == res_loc and d == date_str and count > best_count:
                best_count = count
                best_type = bt

        if best_type:
            assignments.append((rb, best_type, date_str))

    return assignments


def create_reserve_trips(reserves: list, all_trips: list) -> list:
    """
    Create phantom Trip objects for reserve bus requirements.

    Each reserve of count N at station S produces N phantom trips that
    occupy a bus at S for the reserve time window.  The bus type is
    assigned to the most common type at that station/day.
    """
    assignments = assign_reserves_to_bus_types(reserves, all_trips)
    trip_dates = sorted(set(t.date_str for t in all_trips))

    # Build station code / name lookup from real trips
    loc_info = {}  # normalized_loc → (code, name)
    for t in all_trips:
        for code, name in [(t.origin_code, t.origin_name), (t.dest_code, t.dest_name)]:
            nloc = normalize_location(code)
            if nloc not in loc_info:
                loc_info[nloc] = (code, name)

    reserve_trips = []
    for rb, bus_type, date_str in assignments:
        res_loc = normalize_reserve_station(rb.station)
        code, name = loc_info.get(res_loc, ("", rb.station))

        for i in range(rb.count):
            trip_id = f"RES_{rb.station.replace(' ', '')}_{date_str.replace(' ', '')}_{i + 1}"
            reserve_trips.append(Trip(
                trip_id=trip_id,
                bus_nr=0,
                service=f"Reserve {rb.station}",
                date_str=date_str,
                date_label=f"Reserve {rb.day}",
                direction="reserve",
                bus_type=bus_type,
                snel_stop="",
                pattern="",
                multiplicity=1,
                origin_code=code,
                origin_name=name,
                origin_halt="",
                dest_code=code,
                dest_name=name,
                dest_halt="",
                departure=rb.start,
                arrival=rb.end,
                stops=[],
                is_reserve=True,
            ))

    return reserve_trips


# ---------------------------------------------------------------------------
# Halt capacity check
# ---------------------------------------------------------------------------

def check_halt_capacity(rotations: list, halt_capacity: dict) -> list:
    """Check if bus rotations exceed halt capacity limits.

    For each station with a capacity limit, count the maximum number of buses
    present simultaneously. A bus is "present" at a station from its arrival
    until its next departure (i.e. during idle time between trips).

    Args:
        rotations: list of BusRotation objects
        halt_capacity: {station_name: max_buses} e.g. {"Utrecht Centraal": 6}

    Returns:
        list of dicts with violations:
        [{"station": str, "date": str, "time": str, "count": int, "capacity": int}]
    """
    if not halt_capacity:
        return []

    # Collect all presence intervals: (station, date, start_min, end_min)
    # A bus is present at trip.dest_name from trip.arrival until next_trip.departure
    intervals = []
    for rot in rotations:
        for i, trip in enumerate(rot.trips):
            if i < len(rot.trips) - 1:
                next_trip = rot.trips[i + 1]
                # Bus waits at dest station from arrival until next departure
                intervals.append((trip.dest_name, rot.date_str,
                                  trip.arrival, next_trip.departure))

    # For each capped station, find the peak concurrent bus count
    violations = []
    for station, capacity in halt_capacity.items():
        # Filter intervals for this station (fuzzy match: check if station
        # name is contained in the interval station name or vice versa)
        station_intervals = [
            (date, start, end) for (stn, date, start, end) in intervals
            if station.lower() in stn.lower() or stn.lower() in station.lower()
        ]
        if not station_intervals:
            continue

        # Group by date
        by_date = {}
        for date, start, end in station_intervals:
            by_date.setdefault(date, []).append((start, end))

        for date, ivs in by_date.items():
            # Sweep-line: collect all start/end events
            events = []
            for start, end in ivs:
                if start < end:  # valid interval
                    events.append((start, +1))
                    events.append((end, -1))
            events.sort(key=lambda e: (e[0], e[1]))

            concurrent = 0
            peak = 0
            peak_time = 0
            for time_min, delta in events:
                concurrent += delta
                if concurrent > peak:
                    peak = concurrent
                    peak_time = time_min

            if peak > capacity:
                h, m = divmod(peak_time, 60)
                violations.append({
                    "station": station,
                    "date": date,
                    "time": f"{h:02d}:{m:02d}",
                    "count": peak,
                    "capacity": capacity,
                })

    return violations


# ---------------------------------------------------------------------------
# Traffic-aware risk analysis
# ---------------------------------------------------------------------------

def get_time_slot(departure_minutes: int, is_weekend: bool = False) -> str:
    """Determine the time slot for a trip based on departure time.

    departure_minutes: minutes from midnight.
    Returns the time slot name (e.g. 'ochtendspits', 'dal', 'weekend').
    """
    if is_weekend:
        return "weekend"
    from google_maps_distances import TIME_SLOTS
    for slot_name in ["nacht", "ochtendspits", "dal", "middagspits", "avond"]:
        lo, hi = TIME_SLOTS[slot_name]["range"]
        if lo <= departure_minutes < hi:
            return slot_name
    return "avond"  # fallback for exactly midnight


def is_weekend_date(date_str: str) -> bool:
    """Check if a date string like 'za 13-06-2026' is a weekend day."""
    if not date_str:
        return False
    prefix = date_str.strip().split()[0].lower()
    return prefix in ("za", "zo")


def compute_trip_turnaround_overrides(
    trips: list,
    traffic_data: dict,
    base_turnaround_map: dict,
) -> tuple[dict, list]:
    """Compute per-trip turnaround time overrides based on traffic risk.

    For each trip, compares the scheduled duration against the Google Maps
    traffic-aware driving time for the trip's time slot. If the Google Maps
    time exceeds the scheduled time (negative buffer), the excess is added
    to the turnaround time for that trip.

    Args:
        trips: list of Trip objects
        traffic_data: {"time_slots": {slot: {orig: {dest: min}}}, "baseline": {...}}
        base_turnaround_map: {bus_type: minutes} default turnaround

    Returns:
        (overrides, report)
        overrides: {trip_id: adjusted_turnaround_minutes}
        report: list of dicts with per-trip risk details for Excel output
    """
    time_slots = traffic_data.get("time_slots", {})
    baseline = traffic_data.get("baseline", {})
    overrides = {}
    report = []

    for trip in trips:
        if trip.is_reserve:
            continue

        is_we = is_weekend_date(trip.date_str)
        slot = get_time_slot(trip.departure, is_we)
        base_turn = base_turnaround_map.get(trip.bus_type, 8)

        origin_loc = normalize_location(trip.origin_code)
        dest_loc = normalize_location(trip.dest_code)
        scheduled_dur = trip.duration

        # Get traffic-aware driving time for this slot
        slot_matrix = time_slots.get(slot, {})
        traffic_min = slot_matrix.get(origin_loc, {}).get(dest_loc)
        baseline_min = baseline.get(origin_loc, {}).get(dest_loc)

        if traffic_min is not None and scheduled_dur > 0:
            buffer = scheduled_dur - traffic_min
            if buffer < 0:
                # Negative buffer: trip takes longer than scheduled under traffic
                # Add the deficit to turnaround time
                extra = abs(buffer)
                adjusted = max(base_turn + extra, 2)
                overrides[trip.trip_id] = round(adjusted, 1)
            else:
                adjusted = max(base_turn, 2)
                extra = 0.0
        else:
            buffer = None
            adjusted = base_turn
            extra = 0.0
            traffic_min = None

        report.append({
            "trip_id": trip.trip_id,
            "service": trip.service,
            "direction": trip.direction,
            "origin": trip.origin_name,
            "dest": trip.dest_name,
            "departure": trip.departure,
            "arrival": trip.arrival,
            "scheduled_min": scheduled_dur,
            "time_slot": slot,
            "traffic_min": round(traffic_min, 1) if traffic_min is not None else None,
            "baseline_min": round(baseline_min, 1) if baseline_min is not None else None,
            "buffer_min": round(buffer, 1) if buffer is not None else None,
            "base_turnaround": base_turn,
            "extra_turnaround": round(extra, 1),
            "adjusted_turnaround": round(adjusted, 1),
            "risk": "HOOG" if (buffer is not None and buffer < 0) else
                    "MATIG" if (buffer is not None and buffer < 5) else "OK",
        })

    return overrides, report


def can_connect(prev_trip: Trip, next_trip: Trip, turnaround_map: dict,
                service_constraint: bool = False,
                deadhead_matrix: dict = None,
                trip_turnaround_overrides: dict = None) -> tuple:
    """Check if a bus finishing prev_trip can start next_trip.

    If service_constraint=True, two real (non-reserve) trips must belong to
    the same service.  Reserve trips can bridge different services.

    deadhead_matrix: optional dict {origin: {dest: minutes}} for repositioning.
    If provided, allows connections where dest != origin if the bus can drive
    there in time (deadhead).

    trip_turnaround_overrides: optional dict {trip_id: minutes} for per-trip
    turnaround time overrides (from traffic risk analysis). If present, uses
    the override for prev_trip instead of the bus-type default.

    Returns (connectable: bool, deadhead_time: float).
    deadhead_time is 0 if same location, or the driving time in minutes if
    the bus needs to reposition. Returns (False, 0) if not connectable.
    """
    # Must be same bus type
    if prev_trip.bus_type != next_trip.bus_type:
        return False, 0
    # Must be same date
    if prev_trip.date_str != next_trip.date_str:
        return False, 0
    # Service constraint: real-to-real must be same service
    if service_constraint:
        if not prev_trip.is_reserve and not next_trip.is_reserve:
            if prev_trip.service != next_trip.service:
                return False, 0

    dest_loc = normalize_location(prev_trip.dest_code)
    orig_loc = normalize_location(next_trip.origin_code)
    deadhead_time = 0.0

    if dest_loc != orig_loc:
        # Different locations: check if deadhead is possible
        if deadhead_matrix is None:
            return False, 0
        dh = deadhead_matrix.get(dest_loc, {}).get(orig_loc)
        if dh is None:
            return False, 0
        deadhead_time = dh

    # Timing: reserve trips need 0 turnaround (bus just stays at station)
    if prev_trip.is_reserve or next_trip.is_reserve:
        min_turnaround = 0
    else:
        # Check per-trip override first, then fall back to bus-type default
        if trip_turnaround_overrides and prev_trip.trip_id in trip_turnaround_overrides:
            min_turnaround = trip_turnaround_overrides[prev_trip.trip_id]
        else:
            min_turnaround = turnaround_map.get(prev_trip.bus_type, MIN_TURNAROUND_FALLBACK)
        # Absolute minimum: never go below 2 minutes for real trips
        min_turnaround = max(min_turnaround, 2)

    gap = next_trip.departure - prev_trip.arrival
    # Gap must accommodate both deadhead driving and turnaround time
    if gap < deadhead_time + min_turnaround:
        return False, 0
    return True, deadhead_time


def _parse_date_to_ordinal(date_str: str) -> int:
    """Convert date string to ordinal for cross-day calculations.

    Example: "do 11-06-2026" -> ordinal number
    """
    from datetime import datetime
    if not date_str:
        return 0
    try:
        # Format: "do 11-06-2026" or "donderdag 11 juni"
        parts = date_str.split()
        if len(parts) >= 2 and '-' in parts[1]:
            # Format: "do 11-06-2026"
            date_part = parts[1]
            dt = datetime.strptime(date_part, "%d-%m-%Y")
            return dt.toordinal()
        else:
            # Try other formats
            return 0
    except:
        return 0


def can_connect_multiday(prev_trip, next_trip, turnaround_map, service_constraint=False,
                         deadhead_matrix=None, trip_turnaround_overrides=None):
    """
    Check if two trips can be connected by the same bus, allowing cross-day connections.

    Same as can_connect, but allows trips on consecutive days to be chained.
    Returns (bool, deadhead_time) tuple.
    """
    # Must be same bus type
    if prev_trip.bus_type != next_trip.bus_type:
        return False, 0

    # Calculate absolute times (minutes from start of first day)
    prev_day = _parse_date_to_ordinal(prev_trip.date_str)
    next_day = _parse_date_to_ordinal(next_trip.date_str)

    if prev_day == 0 or next_day == 0:
        # Can't parse dates, fall back to same-day check
        if prev_trip.date_str != next_trip.date_str:
            return False, 0
        day_offset = 0
    else:
        day_offset = next_day - prev_day
        if day_offset < 0:
            # next_trip is before prev_trip in calendar
            return False, 0
        if day_offset > 1:
            # More than 1 day gap - don't chain (could be made configurable)
            return False, 0

    # Service constraint: real-to-real must be same service (within same day)
    if service_constraint and day_offset == 0:
        if not prev_trip.is_reserve and not next_trip.is_reserve:
            if prev_trip.service != next_trip.service:
                return False, 0

    dest_loc = normalize_location(prev_trip.dest_code)
    orig_loc = normalize_location(next_trip.origin_code)
    deadhead_time = 0.0

    if dest_loc != orig_loc:
        # Different locations: check if deadhead is possible
        if deadhead_matrix is None:
            return False, 0
        dh = deadhead_matrix.get(dest_loc, {}).get(orig_loc)
        if dh is None:
            return False, 0
        deadhead_time = dh

    # Timing: reserve trips need 0 turnaround
    if prev_trip.is_reserve or next_trip.is_reserve:
        min_turnaround = 0
    else:
        if trip_turnaround_overrides and prev_trip.trip_id in trip_turnaround_overrides:
            min_turnaround = trip_turnaround_overrides[prev_trip.trip_id]
        else:
            min_turnaround = turnaround_map.get(prev_trip.bus_type, MIN_TURNAROUND_FALLBACK)
        min_turnaround = max(min_turnaround, 2)

    # Calculate gap including day offset
    # day_offset of 1 means next_trip is on the next day
    gap = (day_offset * 1440) + next_trip.departure - prev_trip.arrival

    if gap < deadhead_time + min_turnaround:
        return False, 0
    return True, deadhead_time


def _group_trips_multiday(trips, turnaround_map):
    """Group trips by bus_type only (for cross-day optimization)."""
    if turnaround_map is None:
        turnaround_map = dict(MIN_TURNAROUND_DEFAULTS)
    groups = {}
    for t in trips:
        key = t.bus_type
        groups.setdefault(key, []).append(t)
    return groups, turnaround_map


def _group_trips(trips, turnaround_map):
    """Common setup: group trips by (date, bus_type), build compatibility edges."""
    if turnaround_map is None:
        turnaround_map = dict(MIN_TURNAROUND_DEFAULTS)
    groups = {}
    for t in trips:
        key = (t.date_str, t.bus_type)
        groups.setdefault(key, []).append(t)
    return groups, turnaround_map


def _build_rotations(group_trips, date_str, bus_type, chains, rotation_counter):
    """Convert list of trip-index chains into BusRotation objects."""
    rotations = []
    for chain in chains:
        rotation_counter += 1
        bus_id = f"{bus_type[:2].upper()}-{date_str.split()[0].upper()}-{rotation_counter:03d}"
        rotations.append(BusRotation(
            bus_id=bus_id,
            bus_type=bus_type,
            date_str=date_str,
            trips=[group_trips[i] for i in chain],
        ))
    return rotations, rotation_counter


# ---------------------------------------------------------------------------
# Algorithm 1: Greedy best-fit
# ---------------------------------------------------------------------------

def _optimize_greedy(group_trips, turnaround_map, service_constraint=False,
                     deadhead_matrix=None, trip_turnaround_overrides=None,
                     multiday=False):
    """Greedy best-fit: assign each trip to the bus with shortest idle time.

    If multiday=True, allows chaining trips across consecutive days.
    """
    # Sort by absolute time for multiday, or just by time for single-day
    if multiday:
        # Sort by date ordinal then by time
        group_trips.sort(key=lambda t: (_parse_date_to_ordinal(t.date_str), t.departure, t.arrival))
    else:
        group_trips.sort(key=lambda t: (t.departure, t.arrival))

    buses = []  # list of lists of trip indices
    connect_func = can_connect_multiday if multiday else can_connect

    for idx, trip in enumerate(group_trips):
        best_bus = None
        best_gap = float('inf')

        for bus in buses:
            last = group_trips[bus[-1]]
            ok, _dh = connect_func(last, trip, turnaround_map, service_constraint,
                                   deadhead_matrix, trip_turnaround_overrides)
            if ok:
                # Calculate gap (including day offset for multiday)
                if multiday:
                    day_offset = _parse_date_to_ordinal(trip.date_str) - _parse_date_to_ordinal(last.date_str)
                    gap = (day_offset * 1440) + trip.departure - last.arrival
                else:
                    gap = trip.departure - last.arrival
                if gap < best_gap:
                    best_gap = gap
                    best_bus = bus

        if best_bus is not None:
            best_bus.append(idx)
        else:
            buses.append([idx])

    return buses


# ---------------------------------------------------------------------------
# Bipartite matching helpers (used by min-cost and reserve idle matching)
# ---------------------------------------------------------------------------

def _hopcroft_karp(adj, n_left, n_right):
    """
    Hopcroft-Karp algorithm for maximum bipartite matching.
    adj[u] = list of right-side nodes that left-side node u can match to.
    Returns (match_l, match_r) where match_l[u] = matched right node or -1.
    """
    from collections import deque

    match_l = [-1] * n_left
    match_r = [-1] * n_right

    def bfs():
        dist = [0] * n_left
        queue = deque()
        for u in range(n_left):
            if match_l[u] == -1:
                dist[u] = 0
                queue.append(u)
            else:
                dist[u] = float('inf')
        found = False
        while queue:
            u = queue.popleft()
            for v in adj[u]:
                w = match_r[v]
                if w == -1:
                    found = True
                elif dist[w] == float('inf'):
                    dist[w] = dist[u] + 1
                    queue.append(w)
        return found, dist

    def dfs(u, dist):
        for v in adj[u]:
            w = match_r[v]
            if w == -1 or (dist[w] == dist[u] + 1 and dfs(w, dist)):
                match_l[u] = v
                match_r[v] = u
                return True
        dist[u] = float('inf')
        return False

    while True:
        found, dist = bfs()
        if not found:
            break
        for u in range(n_left):
            if match_l[u] == -1:
                dfs(u, dist)

    return match_l, match_r


def _matching_to_chains(n, match_l):
    """Convert a matching into chains of trip indices."""
    # match_l[i] = j means trip i is followed by trip j
    matched_targets = set(v for v in match_l if v != -1)
    chains = []
    for i in range(n):
        if i not in matched_targets:
            # i is the start of a chain
            chain = [i]
            current = i
            while match_l[current] != -1:
                current = match_l[current]
                chain.append(current)
            chains.append(chain)
    return chains


# ---------------------------------------------------------------------------
# Euro-based edge cost calculation for Version 8
# ---------------------------------------------------------------------------

def calculate_euro_edge_cost(trip_i, trip_j, deadhead_min: int, deadhead_km: float,
                              financial_config, fuel_config: dict) -> float:
    """
    Calculate euro-based cost for connecting trip_i to trip_j.

    This is used in Version 8 cost-optimized chaining to prefer connections
    that minimize actual operational costs rather than just time.

    Cost components:
    1. Driver idle cost: paid time with no revenue
    2. Deadhead fuel cost: fuel burned driving empty
    3. ORT penalty: extra cost if connection time falls in unsocial hours

    Args:
        trip_i: First trip (ending)
        trip_j: Second trip (starting)
        deadhead_min: Deadhead driving time in minutes
        deadhead_km: Deadhead distance in km (from matrix)
        financial_config: FinancialConfig object with rates
        fuel_config: Dict with fuel consumption and prices

    Returns:
        Cost in euros (lower is better)
    """
    cost = 0.0

    # 1. Idle time cost (driver paid but no revenue)
    idle_minutes = trip_j.departure - trip_i.arrival
    driver_hourly_rate = getattr(financial_config, 'base_hourly_wage', 22.0)  # Default €22/hour
    idle_cost = (idle_minutes / 60.0) * driver_hourly_rate
    cost += idle_cost

    # 2. Deadhead fuel cost
    if deadhead_km > 0:
        # Get bus type consumption (L/100km for diesel)
        bus_type = trip_i.bus_type
        consumption_map = fuel_config.get('consumption', {})
        consumption_per_100km = consumption_map.get(bus_type, 30)  # Default 30 L/100km
        consumption_per_km = consumption_per_100km / 100.0

        # Fuel price
        fuel_price = fuel_config.get('diesel_price', 1.70)  # Default €1.70/L

        deadhead_fuel_cost = deadhead_km * consumption_per_km * fuel_price
        cost += deadhead_fuel_cost

    # 3. ORT penalty estimate (for idle/deadhead in unsocial hours)
    # ORT windows: weekdays 19:00-07:30, all Saturday, all Sunday
    # We check if the connection time (arrival to departure) overlaps ORT windows

    # Get connection time window
    connection_start = trip_i.arrival  # Minutes from midnight
    connection_end = trip_j.departure

    # Check for evening ORT (19:00 = 1140 min, 07:30 = 450 min next day)
    evening_start = 19 * 60  # 1140
    morning_end = 7 * 60 + 30  # 450

    ort_minutes = 0

    # Simplified ORT check (weekday evening 19:00-24:00)
    if connection_start < 24 * 60:  # Same day
        if connection_end > evening_start:
            # Some time in evening ORT window
            ort_start = max(connection_start, evening_start)
            ort_end = min(connection_end, 24 * 60)
            ort_minutes += max(0, ort_end - ort_start)

        # Early morning (00:00-07:30)
        if connection_start < morning_end:
            ort_start = max(connection_start, 0)
            ort_end = min(connection_end, morning_end)
            ort_minutes += max(0, ort_end - ort_start)

    # ORT surcharge (average €5.50/hour for evening, higher for night)
    if ort_minutes > 0:
        ort_hourly_rate = 5.50  # Average ORT surcharge
        ort_cost = (ort_minutes / 60.0) * ort_hourly_rate
        cost += ort_cost

    return cost


# ---------------------------------------------------------------------------
# Algorithm 2: Minimum-cost maximum matching (successive shortest path)
#   Minimizes buses first, then minimizes total idle time.
# ---------------------------------------------------------------------------

def _optimize_mincost(group_trips, turnaround_map, service_constraint=False,
                      deadhead_matrix=None, trip_turnaround_overrides=None,
                      euro_cost_mode=False, financial_config=None, fuel_config=None,
                      distance_matrix=None, multiday=False):
    """
    Min-cost max matching via successive shortest path (SPFA).
    Minimizes number of buses (primary) and total deadhead+idle time (secondary).

    Args:
        euro_cost_mode: If True, use euro-based costs instead of time-based (Version 8)
        financial_config: FinancialConfig for euro costs (required if euro_cost_mode=True)
        fuel_config: Fuel config dict for euro costs (required if euro_cost_mode=True)
        distance_matrix: Dict with distances in km for fuel cost calculation
        multiday: If True, allows chaining trips across consecutive days
    """
    from collections import deque

    # Sort by absolute time for multiday, or just by time for single-day
    if multiday:
        group_trips.sort(key=lambda t: (_parse_date_to_ordinal(t.date_str), t.departure, t.arrival))
    else:
        group_trips.sort(key=lambda t: (t.departure, t.arrival))
    n = len(group_trips)

    # Build adjacency with costs
    # Default: Cost = deadhead time (penalizes empty driving) + idle time
    # Euro mode: Cost = driver idle cost + deadhead fuel cost + ORT penalty
    connect_func = can_connect_multiday if multiday else can_connect
    adj = [[] for _ in range(n)]
    cost_map = {}
    for i in range(n):
        for j in range(i + 1, n):
            ok, dh = connect_func(group_trips[i], group_trips[j], turnaround_map,
                                  service_constraint, deadhead_matrix,
                                  trip_turnaround_overrides)
            if ok:
                if euro_cost_mode and financial_config and fuel_config:
                    # Version 8: Euro-based cost
                    # Get deadhead km from distance matrix if available
                    dh_km = 0.0
                    if distance_matrix and dh > 0:
                        from_st = group_trips[i].dest_code.lower()
                        to_st = group_trips[j].origin_code.lower()
                        if from_st in distance_matrix and to_st in distance_matrix.get(from_st, {}):
                            dh_km = distance_matrix[from_st].get(to_st, 0)
                        elif to_st in distance_matrix and from_st in distance_matrix.get(to_st, {}):
                            dh_km = distance_matrix[to_st].get(from_st, 0)

                    cost = calculate_euro_edge_cost(
                        group_trips[i], group_trips[j],
                        deadhead_min=dh,
                        deadhead_km=dh_km,
                        financial_config=financial_config,
                        fuel_config=fuel_config
                    )
                else:
                    # Default: Time-based cost
                    if multiday:
                        day_offset = _parse_date_to_ordinal(group_trips[j].date_str) - _parse_date_to_ordinal(group_trips[i].date_str)
                        idle = (day_offset * 1440) + group_trips[j].departure - group_trips[i].arrival
                    else:
                        idle = group_trips[j].departure - group_trips[i].arrival
                    # Weight deadhead more heavily: it costs fuel and driver time
                    cost = dh * 2 + idle if dh > 0 else idle

                adj[i].append(j)
                cost_map[(i, j)] = cost

    # Successive shortest path: find augmenting paths in order of increasing cost
    # Model as flow network with residual graph
    # match_l[i] = j means left i matched to right j
    match_l = [-1] * n
    match_r = [-1] * n

    def spfa_augment():
        """Find minimum-cost augmenting path using SPFA."""
        dist = [float('inf')] * n  # distance to right-side node j
        prev_l = [-1] * n  # previous left node on path to right j
        in_queue = [False] * n
        queue = deque()

        # Start from all unmatched left nodes
        # dist_left[u] = 0 for unmatched u
        dist_left = [float('inf')] * n
        for u in range(n):
            if match_l[u] == -1:
                dist_left[u] = 0
                # Relax edges from u
                for v in adj[u]:
                    c = cost_map[(u, v)]
                    if c < dist[v]:
                        dist[v] = c
                        prev_l[v] = u
                        if not in_queue[v]:
                            queue.append(v)
                            in_queue[v] = True

        # SPFA: relax through alternating paths
        while queue:
            v = queue.popleft()
            in_queue[v] = False

            # v is a right-side node, matched to w = match_r[v]
            w = match_r[v]
            if w == -1:
                continue  # free node, potential augmenting path end

            # Relax from w (go through the matched edge, then to new right nodes)
            new_dist_w = dist[v]  # cost to reach w through v's matched edge (0 cost)
            if new_dist_w < dist_left[w]:
                dist_left[w] = new_dist_w
                for v2 in adj[w]:
                    c = new_dist_w + cost_map[(w, v2)]
                    if c < dist[v2]:
                        dist[v2] = c
                        prev_l[v2] = w
                        if not in_queue[v2]:
                            queue.append(v2)
                            in_queue[v2] = True

        # Find the free right-side node with minimum distance
        best_v = -1
        best_d = float('inf')
        for v in range(n):
            if match_r[v] == -1 and dist[v] < best_d:
                best_d = dist[v]
                best_v = v

        if best_v == -1:
            return False  # no augmenting path

        # Trace back and augment
        v = best_v
        while v != -1:
            u = prev_l[v]
            old_v = match_l[u]
            match_l[u] = v
            match_r[v] = u
            v = old_v

        return True

    # Find all augmenting paths
    while spfa_augment():
        pass

    return _matching_to_chains(n, match_l)


# ---------------------------------------------------------------------------
# Algorithm 3: Profit-maximizing optimization (Version 8)
#   Explores different bus counts to find maximum profit
# ---------------------------------------------------------------------------

def _optimize_profit_maximizing(group_trips, turnaround_map, service_constraint=False,
                                 deadhead_matrix=None, trip_turnaround_overrides=None,
                                 financial_config=None, fuel_config=None,
                                 distance_matrix=None, max_extra_buses_pct=30,
                                 algorithm="mincost"):
    """
    Profit-maximizing optimization that explores different bus counts.

    Instead of minimizing buses, this algorithm finds the number of buses
    that maximizes profit by balancing:
    - Garage costs (more buses = more garage travel)
    - ORT costs (fewer buses = longer shifts = more ORT)
    - Overtime costs (fewer buses = more overtime)
    - Deadhead costs (depends on chaining)

    Args:
        group_trips: List of trips to chain
        turnaround_map: Minimum turnaround times per bus type
        service_constraint: Whether to enforce same-service connections
        deadhead_matrix: Optional deadhead time matrix
        trip_turnaround_overrides: Per-trip turnaround overrides
        financial_config: FinancialConfig object for profit calculation
        fuel_config: Fuel config dict
        distance_matrix: Distances in km for fuel calculations
        max_extra_buses_pct: Maximum % extra buses to try beyond minimum
        algorithm: Which algorithm to use for initial chaining ("greedy" or "mincost")

    Returns:
        Tuple of (chains, profit_info) where profit_info contains the analysis
    """
    from collections import deque

    if not financial_config:
        # Fallback to selected algorithm if no financial config
        algo_func = _optimize_greedy if algorithm == "greedy" else _optimize_mincost
        return algo_func(group_trips, turnaround_map, service_constraint,
                         deadhead_matrix, trip_turnaround_overrides), None

    group_trips_sorted = sorted(group_trips, key=lambda t: (t.departure, t.arrival))
    n = len(group_trips_sorted)

    if n == 0:
        return [], {'best_buses': 0, 'best_profit': 0, 'explored': []}

    # Get initial minimum-buses solution using selected algorithm
    if algorithm == "greedy":
        # Use greedy algorithm for initial chains
        min_buses_chains = _optimize_greedy(group_trips_sorted, turnaround_map,
                                            service_constraint, deadhead_matrix,
                                            trip_turnaround_overrides)
    else:
        # Use euro-cost based mincost matching for better profit optimization
        # Build connection graph with euro costs
        adj = [[] for _ in range(n)]
        cost_map = {}  # (i, j) -> euro cost of connection

        for i in range(n):
            for j in range(i + 1, n):
                ok, dh = can_connect(group_trips_sorted[i], group_trips_sorted[j],
                                     turnaround_map, service_constraint, deadhead_matrix,
                                     trip_turnaround_overrides)
                if ok:
                    # Calculate euro cost for this connection
                    dh_km = 0.0
                    if distance_matrix and dh > 0:
                        from_st = group_trips_sorted[i].dest_code.lower()
                        to_st = group_trips_sorted[j].origin_code.lower()
                        if from_st in distance_matrix and to_st in distance_matrix.get(from_st, {}):
                            dh_km = distance_matrix[from_st].get(to_st, 0)
                        elif to_st in distance_matrix and from_st in distance_matrix.get(to_st, {}):
                            dh_km = distance_matrix[to_st].get(from_st, 0)

                    cost = calculate_euro_edge_cost(
                        group_trips_sorted[i], group_trips_sorted[j],
                        deadhead_min=dh, deadhead_km=dh_km,
                        financial_config=financial_config,
                        fuel_config=fuel_config
                    )
                    adj[i].append(j)
                    cost_map[(i, j)] = cost

        # Find minimum buses solution via max matching with euro costs
        match_l = [-1] * n
        match_r = [-1] * n

        def spfa_augment():
            """Find minimum-cost augmenting path using SPFA."""
            dist = [float('inf')] * n
            prev_l = [-1] * n
            in_queue = [False] * n
            queue = deque()

            dist_left = [float('inf')] * n
            for u in range(n):
                if match_l[u] == -1:
                    dist_left[u] = 0
                    for v in adj[u]:
                        c = cost_map[(u, v)]
                        if c < dist[v]:
                            dist[v] = c
                            prev_l[v] = u
                            if not in_queue[v]:
                                queue.append(v)
                                in_queue[v] = True

            while queue:
                v = queue.popleft()
                in_queue[v] = False

                w = match_r[v]
                if w == -1:
                    continue

                new_dist_w = dist[v]
                if new_dist_w < dist_left[w]:
                    dist_left[w] = new_dist_w
                    for v2 in adj[w]:
                        c = new_dist_w + cost_map[(w, v2)]
                        if c < dist[v2]:
                            dist[v2] = c
                            prev_l[v2] = w
                            if not in_queue[v2]:
                                queue.append(v2)
                                in_queue[v2] = True

            best_v = -1
            best_d = float('inf')
            for v in range(n):
                if match_r[v] == -1 and dist[v] < best_d:
                    best_d = dist[v]
                    best_v = v

            if best_v == -1:
                return False

            v = best_v
            while v != -1:
                u = prev_l[v]
                old_v = match_l[u]
                match_l[u] = v
                match_r[v] = u
                v = old_v

            return True

        # Find max matching
        while spfa_augment():
            pass

        min_buses_chains = _matching_to_chains(n, match_l)

    min_buses = len(min_buses_chains)

    # Calculate max buses to try
    max_extra = max(1, int(min_buses * max_extra_buses_pct / 100))
    max_buses = min(n, min_buses + max_extra)

    # Helper: Calculate profit for a given chaining
    def calculate_chain_profit(chains):
        """Calculate total profit for chains using full financial model."""
        from financial_calculator import calculate_rotation_financials

        total_revenue = 0.0
        total_cost = 0.0

        for chain in chains:
            trips = [group_trips_sorted[i] for i in chain]

            # Build a simple rotation object for financial calculation
            class SimpleRotation:
                def __init__(self, trips):
                    self.trips = trips
                    self.bus_type = trips[0].bus_type if trips else 'Touringcar'
                    self.date_str = trips[0].date_str if trips else ''
                    self.bus_id = 'temp'
                    self.total_km = None
                    self.deadhead_km = 0

                    # Calculate deadhead km
                    for i in range(1, len(trips)):
                        prev_trip = trips[i-1]
                        curr_trip = trips[i]
                        from_st = prev_trip.dest_code.lower() if hasattr(prev_trip, 'dest_code') else ''
                        to_st = curr_trip.origin_code.lower() if hasattr(curr_trip, 'origin_code') else ''
                        if from_st and to_st and from_st != to_st:
                            if distance_matrix:
                                if from_st in distance_matrix and to_st in distance_matrix.get(from_st, {}):
                                    self.deadhead_km += distance_matrix[from_st].get(to_st, 0)
                                elif to_st in distance_matrix and from_st in distance_matrix.get(to_st, {}):
                                    self.deadhead_km += distance_matrix[to_st].get(from_st, 0)

            rot = SimpleRotation(trips)
            fin = calculate_rotation_financials(rot, financial_config, fuel_type="diesel")

            total_revenue += fin.revenue
            total_cost += fin.driver_cost.total_cost + fin.fuel_cost + fin.garage_fuel_cost

        return total_revenue - total_cost

    # Explore different bus counts
    explored = []
    best_chains = min_buses_chains
    best_profit = calculate_chain_profit(min_buses_chains)
    best_buses = min_buses

    explored.append({
        'buses': min_buses,
        'profit': best_profit,
        'chains': min_buses_chains
    })

    # For more buses, we need to split chains
    # Strategy: Start from min-cost chains, then greedily split at best points

    for target_buses in range(min_buses + 1, max_buses + 1):
        # Start from previous solution and split one chain
        # Find the split that results in highest profit

        current_chains = [list(c) for c in best_chains]  # Copy

        # Try splitting each chain at each possible point
        best_split_profit = float('-inf')
        best_split_chains = None

        for chain_idx, chain in enumerate(current_chains):
            if len(chain) <= 1:
                continue  # Can't split a single-trip chain

            for split_pos in range(1, len(chain)):
                # Split chain into two
                chain1 = chain[:split_pos]
                chain2 = chain[split_pos:]

                # Create new chain set
                new_chains = [c for i, c in enumerate(current_chains) if i != chain_idx]
                new_chains.append(chain1)
                new_chains.append(chain2)

                # Calculate profit
                split_profit = calculate_chain_profit(new_chains)

                if split_profit > best_split_profit:
                    best_split_profit = split_profit
                    best_split_chains = new_chains

        if best_split_chains is None:
            break  # No more splits possible

        explored.append({
            'buses': target_buses,
            'profit': best_split_profit,
            'chains': best_split_chains
        })

        if best_split_profit > best_profit:
            best_profit = best_split_profit
            best_chains = best_split_chains
            best_buses = target_buses

        # Use this as starting point for next iteration
        current_chains = best_split_chains

    profit_info = {
        'best_buses': best_buses,
        'best_profit': best_profit,
        'min_buses': min_buses,
        'explored': explored
    }

    return best_chains, profit_info


# ---------------------------------------------------------------------------
# Main dispatcher
# ---------------------------------------------------------------------------

ALGORITHMS = {
    "greedy": ("Greedy best-fit", _optimize_greedy),
    "mincost": ("Min-cost maximum matching (SPFA)", _optimize_mincost),
}


def optimize_rotations(trips: list, turnaround_map: dict = None,
                       algorithm: str = "greedy",
                       per_service: bool = False,
                       service_constraint: bool = False,
                       deadhead_matrix: dict = None,
                       trip_turnaround_overrides: dict = None) -> list:
    """
    Optimize bus rotations using the specified algorithm.

    If per_service=True, only chains trips within the same service (Excel tab).
    If per_service=False, chains across all services (cross-tab optimization).
    If service_constraint=True (only when per_service=False), real-to-real trip
    connections must be within the same service; reserve trips can bridge services.
    deadhead_matrix: optional {origin: {dest: minutes}} for repositioning trips.
    trip_turnaround_overrides: optional {trip_id: minutes} for per-trip turnaround.
    """
    groups, turnaround_map = _group_trips(trips, turnaround_map)
    algo_name, algo_func = ALGORITHMS[algorithm]

    # If per_service, further split groups by service
    if per_service:
        new_groups = {}
        for (date_str, bus_type), group_trips in groups.items():
            by_svc = {}
            for t in group_trips:
                by_svc.setdefault(t.service, []).append(t)
            for svc, svc_trips in by_svc.items():
                new_groups[(date_str, bus_type, svc)] = svc_trips
        all_rotations = []
        rotation_counter = 0
        for key, group_trips in sorted(new_groups.items()):
            date_str, bus_type = key[0], key[1]
            chains = algo_func(group_trips, turnaround_map,
                               service_constraint=False,
                               deadhead_matrix=deadhead_matrix,
                               trip_turnaround_overrides=trip_turnaround_overrides)
            rotations, rotation_counter = _build_rotations(
                group_trips, date_str, bus_type, chains, rotation_counter
            )
            all_rotations.extend(rotations)
        return all_rotations

    all_rotations = []
    rotation_counter = 0

    for (date_str, bus_type), group_trips in sorted(groups.items()):
        chains = algo_func(group_trips, turnaround_map,
                           service_constraint=service_constraint,
                           deadhead_matrix=deadhead_matrix,
                           trip_turnaround_overrides=trip_turnaround_overrides)
        rotations, rotation_counter = _build_rotations(
            group_trips, date_str, bus_type, chains, rotation_counter
        )
        all_rotations.extend(rotations)

    return all_rotations


# ---------------------------------------------------------------------------
# Excel Output Generator
# ---------------------------------------------------------------------------

# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
BUS_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
RESERVE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TIME_FORMAT = "HH:MM"


def apply_header_style(ws, row, col_start, col_end, fill=None, font=None):
    if fill is None:
        fill = HEADER_FILL
    if font is None:
        font = HEADER_FONT_WHITE
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_omloop_sheet(wb_out, rotations: list, reserves: list,
                       deadhead_matrix: dict = None):
    """
    Tab 1: Busomloop - Transvision style per bus.
    Groups by date + bus_type, shows each bus's trip sequence.
    """
    # Group rotations by date
    by_date = {}
    for r in rotations:
        by_date.setdefault(r.date_str, []).append(r)

    for date_str, date_rotations in sorted(by_date.items()):
        # Group by bus type within date
        by_type = {}
        for r in date_rotations:
            by_type.setdefault(r.bus_type, []).append(r)

        for bus_type, type_rotations in sorted(by_type.items()):
            # Create sheet per date+type
            type_abbrev = {"Dubbeldekker": "DD", "Touringcar": "TC", "Lagevloerbus": "LVB", "Midi bus": "Midi", "Taxibus": "Taxi"}.get(bus_type, bus_type[:4])
            day_abbrev = date_str.split()[0] if date_str else "dag"
            sheet_name = f"Omloop {type_abbrev} {day_abbrev}"
            # Ensure unique sheet name (max 31 chars)
            sheet_name = sheet_name[:31]
            ws = wb_out.create_sheet(title=sheet_name)

            # Sort rotations by first departure
            type_rotations.sort(key=lambda r: r.start_time)

            # Layout: 3 buses per block of columns, like the example
            buses_per_row = 3
            cols_per_bus = 7  # Van, Naar, Van(t), Tot(t), Duur, Hold, spacer

            row = 1
            # Title
            ws.cell(row=row, column=1, value=f"Busomlopen {bus_type} - {date_str}")
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=14)
            row += 1

            # Summary line
            total_buses = len(type_rotations)
            ws.cell(row=row, column=1, value=f"Totaal bussen: {total_buses}")
            ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=11)
            row += 2

            # Helper: build expanded row list for a bus (trips + deadhead rows)
            DEADHEAD_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

            def _expand_bus_rows(bus):
                """Return list of row dicts: either a trip or a deadhead entry."""
                rows = []
                for t_idx, t in enumerate(bus.trips):
                    # Check if deadhead is needed before this trip
                    if t_idx > 0 and deadhead_matrix:
                        prev_t = bus.trips[t_idx - 1]
                        prev_dest = normalize_location(prev_t.dest_code)
                        this_orig = normalize_location(t.origin_code)
                        if prev_dest != this_orig:
                            dh_min = deadhead_matrix.get(prev_dest, {}).get(this_orig)
                            rows.append({
                                "type": "deadhead",
                                "from_name": prev_t.dest_name,
                                "to_name": t.origin_name,
                                "dh_minutes": dh_min,
                            })
                    rows.append({"type": "trip", "trip": t, "trip_idx": t_idx})
                return rows

            # Process in blocks of buses_per_row
            for block_start in range(0, len(type_rotations), buses_per_row):
                block = type_rotations[block_start:block_start + buses_per_row]
                block_rows = [_expand_bus_rows(bus) for bus in block]
                max_rows = max(len(br) for br in block_rows)

                # Bus headers
                for i, bus in enumerate(block):
                    base_col = 1 + i * cols_per_bus
                    ws.cell(row=row, column=base_col, value=bus.bus_id)
                    ws.cell(row=row, column=base_col).font = Font(bold=True, size=11, color="FFFFFF")
                    ws.cell(row=row, column=base_col).fill = BUS_HEADER_FILL
                    ws.cell(row=row, column=base_col + 1, value=bus.bus_type)
                    ws.cell(row=row, column=base_col + 1).font = Font(bold=True, size=11, color="FFFFFF")
                    ws.cell(row=row, column=base_col + 1).fill = BUS_HEADER_FILL
                    # Fill remaining header cols
                    for cc in range(base_col + 2, base_col + cols_per_bus - 1):
                        ws.cell(row=row, column=cc).fill = BUS_HEADER_FILL

                    # Dienst info
                    dienst_str = (f"{minutes_to_str(bus.start_time)} - {minutes_to_str(bus.end_time)} "
                                  f"({bus.total_dienst_minutes // 60}u{bus.total_dienst_minutes % 60:02d})")
                    ws.cell(row=row, column=base_col + 2, value=dienst_str)
                    ws.cell(row=row, column=base_col + 2).font = Font(bold=True, size=9, color="FFFFFF")
                    ws.cell(row=row, column=base_col + 2).fill = BUS_HEADER_FILL

                row += 1

                # Column headers
                headers = ["Van", "Naar", "Vertrek", "Aankomst", "Duur", "Wacht"]
                for i, bus in enumerate(block):
                    base_col = 1 + i * cols_per_bus
                    for j, h in enumerate(headers):
                        cell = ws.cell(row=row, column=base_col + j, value=h)
                        cell.font = HEADER_FONT
                        cell.fill = SUBHEADER_FILL
                        cell.border = THIN_BORDER
                        cell.alignment = Alignment(horizontal="center")
                row += 1
                trip_start_row = row

                # Trip + deadhead rows
                for row_idx in range(max_rows):
                    for i, bus in enumerate(block):
                        base_col = 1 + i * cols_per_bus
                        if row_idx < len(block_rows[i]):
                            entry = block_rows[i][row_idx]

                            if entry["type"] == "deadhead":
                                # Deadhead repositioning row
                                dh_min = entry["dh_minutes"]
                                dh_str = f"{round(dh_min)} min" if dh_min is not None else "?"
                                ws.cell(row=row, column=base_col, value=entry["from_name"])
                                ws.cell(row=row, column=base_col + 1, value=entry["to_name"])
                                ws.cell(row=row, column=base_col + 4, value=dh_str)
                                for cc in range(base_col, base_col + 6):
                                    ws.cell(row=row, column=cc).border = THIN_BORDER
                                    ws.cell(row=row, column=cc).alignment = Alignment(horizontal="center")
                                    ws.cell(row=row, column=cc).fill = DEADHEAD_FILL
                                    ws.cell(row=row, column=cc).font = Font(italic=True, size=9)
                            else:
                                # Normal trip row
                                t = entry["trip"]
                                t_idx = entry["trip_idx"]
                                if t.is_reserve:
                                    ws.cell(row=row, column=base_col, value="RESERVE")
                                else:
                                    ws.cell(row=row, column=base_col, value=t.origin_name)
                                ws.cell(row=row, column=base_col + 1, value=t.dest_name)
                                ws.cell(row=row, column=base_col + 2, value=minutes_to_time(t.departure))
                                ws.cell(row=row, column=base_col + 2).number_format = "HH:MM"
                                ws.cell(row=row, column=base_col + 3, value=minutes_to_time(t.arrival))
                                ws.cell(row=row, column=base_col + 3).number_format = "HH:MM"
                                dur = t.arrival - t.departure
                                ws.cell(row=row, column=base_col + 4, value=f"{dur // 60}:{dur % 60:02d}")

                                # Hold/wait time until next trip
                                if t_idx < len(bus.trips) - 1:
                                    next_t = bus.trips[t_idx + 1]
                                    hold = next_t.departure - t.arrival
                                    ws.cell(row=row, column=base_col + 5, value=f"{hold // 60}:{hold % 60:02d}")

                                # Apply borders + reserve highlight
                                for cc in range(base_col, base_col + 6):
                                    ws.cell(row=row, column=cc).border = THIN_BORDER
                                    ws.cell(row=row, column=cc).alignment = Alignment(horizontal="center")
                                    if t.is_reserve:
                                        ws.cell(row=row, column=cc).fill = RESERVE_FILL
                    row += 1

                # Subtotals for this block – placed directly under each bus's last trip
                for i, bus in enumerate(block):
                    sub_row = trip_start_row + len(block_rows[i])
                    base_col = 1 + i * cols_per_bus
                    ws.cell(row=sub_row, column=base_col, value="Ritten:")
                    ws.cell(row=sub_row, column=base_col).font = Font(bold=True, size=9)
                    ws.cell(row=sub_row, column=base_col + 1, value=len(bus.trips))
                    ws.cell(row=sub_row, column=base_col + 2, value="Rijtijd:")
                    ws.cell(row=sub_row, column=base_col + 2).font = Font(bold=True, size=9)
                    ride = bus.total_ride_minutes
                    ws.cell(row=sub_row, column=base_col + 3, value=f"{ride // 60}:{ride % 60:02d}")
                    ws.cell(row=sub_row, column=base_col + 4, value="Wacht:")
                    ws.cell(row=sub_row, column=base_col + 4).font = Font(bold=True, size=9)
                    idle = int(bus.total_idle_minutes)
                    ws.cell(row=sub_row, column=base_col + 5, value=f"{idle // 60}:{idle % 60:02d}")
                row += 2

            # Reserve buses section for this date
            # Match reserves by checking day name substring in reserve's day field
            day_map = {"do": "donderdag", "vr": "vrijdag", "za": "zaterdag",
                       "zo": "zondag", "ma": "maandag"}
            full_day = day_map.get(day_abbrev, day_abbrev)
            date_reserves = [r for r in reserves
                            if full_day in r.day.lower() or day_abbrev in r.day.lower()[:2]]
            if date_reserves:
                ws.cell(row=row, column=1, value="Reservebussen")
                ws.cell(row=row, column=1).font = Font(bold=True, size=12)
                row += 1
                headers_r = ["Station", "Aantal", "Van", "Tot", "Opmerking"]
                for j, h in enumerate(headers_r):
                    cell = ws.cell(row=row, column=1 + j, value=h)
                    cell.font = HEADER_FONT
                    cell.fill = RESERVE_FILL
                    cell.border = THIN_BORDER
                row += 1
                for rb in date_reserves:
                    ws.cell(row=row, column=1, value=rb.station)
                    ws.cell(row=row, column=2, value=rb.count)
                    ws.cell(row=row, column=3, value=minutes_to_time(rb.start))
                    ws.cell(row=row, column=3).number_format = "HH:MM"
                    ws.cell(row=row, column=4, value=minutes_to_time(rb.end))
                    ws.cell(row=row, column=4).number_format = "HH:MM"
                    ws.cell(row=row, column=5, value=rb.remark)
                    ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
                    for cc in range(1, 6):
                        ws.cell(row=row, column=cc).border = THIN_BORDER
                    row += 1

            # Column widths
            for c in range(1, 22):
                ws.column_dimensions[get_column_letter(c)].width = 16


def write_overzicht_sheet(wb_out, rotations: list, all_trips: list):
    """
    Tab 2: Overzicht - Overview of how trips connect per bus.
    Single sheet showing all trip chains.
    """
    ws = wb_out.create_sheet(title="Overzicht Ritsamenhang")

    row = 1
    ws.cell(row=row, column=1, value="Overzicht Ritsamenhang - Alle Busomlopen")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2

    # Headers
    headers = [
        "Bus ID", "Bustype", "Datum", "Rit #", "Busdienst", "Richting",
        "Van", "Naar", "Vertrek", "Aankomst", "Duur (min)",
        "Wachttijd (min)", "Busnr", "Snel/Stop"
    ]
    for j, h in enumerate(headers):
        cell = ws.cell(row=row, column=1 + j, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Sort rotations by date, then bus type, then start time
    sorted_rotations = sorted(rotations, key=lambda r: (r.date_str, r.bus_type, r.start_time))

    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for r_idx, rot in enumerate(sorted_rotations):
        use_fill = alt_fill if r_idx % 2 == 0 else None
        for t_idx, trip in enumerate(rot.trips):
            # Wachttijd
            wait = ""
            if t_idx < len(rot.trips) - 1:
                next_trip = rot.trips[t_idx + 1]
                wait = next_trip.departure - trip.arrival

            values = [
                rot.bus_id, rot.bus_type, rot.date_str,
                t_idx + 1, trip.service, trip.direction,
                trip.origin_name, trip.dest_name,
                minutes_to_time(trip.departure),
                minutes_to_time(trip.arrival),
                trip.arrival - trip.departure,
                wait if isinstance(wait, int) else "",
                trip.bus_nr, trip.snel_stop,
            ]
            for j, v in enumerate(values):
                cell = ws.cell(row=row, column=1 + j, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                if trip.is_reserve:
                    cell.fill = RESERVE_FILL
                elif use_fill:
                    cell.fill = use_fill
                if j in (8, 9):
                    cell.number_format = "HH:MM"
            row += 1

    # Column widths
    widths = [16, 14, 18, 8, 20, 10, 22, 22, 10, 10, 12, 14, 12, 12]
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(1 + j)].width = w

    # Freeze panes
    ws.freeze_panes = "A4"


def write_berekeningen_sheet(wb_out, rotations: list, all_trips: list, reserves: list,
                             turnaround_map: dict = None, algorithm: str = "greedy",
                             output_mode: int = 1):
    """
    Tab 3: Berekeningen - Calculations and KPIs.
    """
    if turnaround_map is None:
        turnaround_map = dict(MIN_TURNAROUND_DEFAULTS)
    ws = wb_out.create_sheet(title="Berekeningen")

    row = 1
    ws.cell(row=row, column=1, value="Berekeningen Busomloop Optimalisatie")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2

    # --- Section 1: Summary per date + bus type ---
    ws.cell(row=row, column=1, value="1. Samenvatting per datum en bustype")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    sum_headers = [
        "Datum", "Bustype", "Aantal bussen", "Totaal ritten",
        "Totaal rijtijd (uur)", "Totaal wachttijd (uur)",
        "Totaal diensttijd (uur)", "Gem. ritten/bus",
        "Gem. diensttijd/bus (uur)", "Benutting (%)"
    ]
    for j, h in enumerate(sum_headers):
        cell = ws.cell(row=row, column=1 + j, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    # Group rotations
    groups = {}
    for r in rotations:
        key = (r.date_str, r.bus_type)
        groups.setdefault(key, []).append(r)

    grand_buses = 0
    grand_trips = 0
    grand_ride = 0
    grand_idle = 0
    grand_dienst = 0

    for (date_str, bus_type), rots in sorted(groups.items()):
        n_buses = len(rots)
        n_trips = sum(len(r.trips) for r in rots)
        ride_min = sum(r.total_ride_minutes for r in rots)
        idle_min = sum(r.total_idle_minutes for r in rots)
        dienst_min = sum(r.total_dienst_minutes for r in rots)
        benutting = (ride_min / dienst_min * 100) if dienst_min > 0 else 0

        grand_buses += n_buses
        grand_trips += n_trips
        grand_ride += ride_min
        grand_idle += idle_min
        grand_dienst += dienst_min

        values = [
            date_str, bus_type, n_buses, n_trips,
            round(ride_min / 60, 1), round(idle_min / 60, 1),
            round(dienst_min / 60, 1),
            round(n_trips / n_buses, 1) if n_buses > 0 else 0,
            round(dienst_min / n_buses / 60, 1) if n_buses > 0 else 0,
            round(benutting, 1),
        ]
        for j, v in enumerate(values):
            cell = ws.cell(row=row, column=1 + j, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
        row += 1

    # Totals row
    benutting_total = (grand_ride / grand_dienst * 100) if grand_dienst > 0 else 0
    totals = [
        "TOTAAL", "", grand_buses, grand_trips,
        round(grand_ride / 60, 1), round(grand_idle / 60, 1),
        round(grand_dienst / 60, 1),
        round(grand_trips / grand_buses, 1) if grand_buses > 0 else 0,
        round(grand_dienst / grand_buses / 60, 1) if grand_buses > 0 else 0,
        round(benutting_total, 1),
    ]
    for j, v in enumerate(totals):
        cell = ws.cell(row=row, column=1 + j, value=v)
        cell.border = THIN_BORDER
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    row += 3

    # --- Section 2: Per-bus detail ---
    ws.cell(row=row, column=1, value="2. Detail per bus")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    detail_headers = [
        "Bus ID", "Bustype", "Datum", "Eerste rit",
        "Laatste rit", "Dienststart", "Diensteinde",
        "Diensttijd (uur)", "Rijtijd (uur)", "Wachttijd (uur)",
        "Aantal ritten", "Benutting (%)"
    ]
    for j, h in enumerate(detail_headers):
        cell = ws.cell(row=row, column=1 + j, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    sorted_rots = sorted(rotations, key=lambda r: (r.date_str, r.bus_type, r.start_time))
    for rot in sorted_rots:
        if not rot.trips:
            continue
        first = rot.trips[0]
        last = rot.trips[-1]
        benutting = (rot.total_ride_minutes / rot.total_dienst_minutes * 100) if rot.total_dienst_minutes > 0 else 0
        values = [
            rot.bus_id, rot.bus_type, rot.date_str,
            f"{first.origin_name} -> {first.dest_name}",
            f"{last.origin_name} -> {last.dest_name}",
            minutes_to_time(rot.start_time),
            minutes_to_time(rot.end_time),
            round(rot.total_dienst_minutes / 60, 1),
            round(rot.total_ride_minutes / 60, 1),
            round(rot.total_idle_minutes / 60, 1),
            len(rot.trips),
            round(benutting, 1),
        ]
        for j, v in enumerate(values):
            cell = ws.cell(row=row, column=1 + j, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if j in (5, 6):
                cell.number_format = "HH:MM"
        row += 1

    row += 2

    # --- Section 3: Reservebussen analyse ---
    trip_dates = sorted(set(t.date_str for t in all_trips))
    real_trip_count = sum(1 for t in all_trips if not t.is_reserve)

    if output_mode in (1, 2):
        # Modes 1 & 2: post-hoc coverage analysis
        mode_label = {
            1: "3. Reservebussen - Dekkingsanalyse (wachttijd, greedy)",
            2: "3. Reservebussen - Optimale toewijzing (wachttijd, matching)",
        }[output_mode]
        ws.cell(row=row, column=1, value=mode_label)
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        if output_mode == 1:
            coverage = analyze_reserve_coverage(rotations, reserves, trip_dates)
        else:
            coverage = optimize_reserve_idle_matching(rotations, reserves, trip_dates)

        res_headers = ["Station", "Dag", "Van", "Tot", "Nodig", "Gedekt door omloop",
                       "Extra nodig", "Opmerking", "Dekkende bus(sen)"]
        for j, h in enumerate(res_headers):
            cell = ws.cell(row=row, column=1 + j, value=h)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

        total_reserve = 0
        total_covered = 0
        total_extra = 0
        for c in coverage:
            rb = c["reserve"]
            covered = min(c["covered"], c["required"])
            extra = c["shortfall"]
            total_reserve += rb.count
            total_covered += covered
            total_extra += extra

            bus_names = ", ".join(b["bus_id"] for b in c["covering_buses"][:c["required"]])

            values = [rb.station, rb.day, minutes_to_time(rb.start), minutes_to_time(rb.end),
                      rb.count, covered, extra, rb.remark, bus_names]
            for j, v in enumerate(values):
                cell = ws.cell(row=row, column=1 + j, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                if j in (2, 3):
                    cell.number_format = "HH:MM"
                if j == 6 and extra > 0:
                    cell.font = Font(bold=True, color="FF0000")
                elif j == 6 and extra == 0:
                    cell.font = Font(bold=True, color="008000")
            row += 1

        # Totals
        row += 1
        summary_items = [
            ("Totaal reservebussen nodig:", total_reserve),
            ("Gedekt door bestaande busomlopen:", total_covered),
            ("Extra bussen nodig voor reserve:", total_extra),
            ("Totaal vloot (omloop + extra reserve):", len(rotations) + total_extra),
        ]
        for label, val in summary_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

    else:
        # Modes 3 & 4: reserves are phantom trips in the rotations
        ws.cell(row=row, column=1, value="3. Reservebussen - Ingepland in busomlopen")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Count reserve coverage from the rotations themselves
        reserve_in_rot = []
        for rot in rotations:
            for t in rot.trips:
                if t.is_reserve:
                    reserve_in_rot.append((rot.bus_id, rot.bus_type, t))

        # Summarise per reserve requirement
        res_headers = ["Station", "Dag", "Van", "Tot", "Nodig",
                       "Ingepland", "Extra nodig", "Opmerking", "Bus(sen)"]
        for j, h in enumerate(res_headers):
            cell = ws.cell(row=row, column=1 + j, value=h)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

        total_reserve = 0
        total_planned = 0
        total_extra = 0
        for rb in reserves:
            date_str = match_reserve_day(rb.day, trip_dates)
            res_loc = normalize_reserve_station(rb.station)
            # Find reserve trips in rotations matching this requirement
            matching = [(bid, bt) for bid, bt, t in reserve_in_rot
                        if t.date_str == date_str
                        and normalize_location(t.origin_code) == res_loc
                        and t.departure == rb.start and t.arrival == rb.end]
            planned = min(len(matching), rb.count)
            extra = max(0, rb.count - planned)
            total_reserve += rb.count
            total_planned += planned
            total_extra += extra
            bus_names = ", ".join(bid for bid, _ in matching[:rb.count])
            values = [rb.station, rb.day, minutes_to_time(rb.start), minutes_to_time(rb.end),
                      rb.count, planned, extra, rb.remark, bus_names]
            for j, v in enumerate(values):
                cell = ws.cell(row=row, column=1 + j, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                if j in (2, 3):
                    cell.number_format = "HH:MM"
                if j == 6 and extra > 0:
                    cell.font = Font(bold=True, color="FF0000")
                elif j == 6 and extra == 0:
                    cell.font = Font(bold=True, color="008000")
            row += 1

        n_total_buses = len(rotations)
        n_with_trips = len([r for r in rotations if r.real_trips])
        n_reserve_only = len([r for r in rotations if not r.real_trips and r.reserve_trip_list])
        row += 1
        summary_items = [
            ("Totaal reservebussen nodig:", total_reserve),
            ("Ingepland in busomlopen:", total_planned),
            ("Extra bussen nodig voor reserve:", total_extra),
            ("Bussen met ritten:", n_with_trips),
            ("Bussen alleen reserve:", n_reserve_only),
            ("Totaal bussen (optimizer):", n_total_buses),
            ("Totaal vloot (incl. extra reserve):", n_total_buses + total_extra),
        ]
        for label, val in summary_items:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=val)
            ws.cell(row=row, column=2).font = Font(bold=True)
            row += 1

    row += 2

    # --- Section 4: Optimalisatie parameters ---
    ws.cell(row=row, column=1, value="4. Optimalisatie parameters & toelichting")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    params = [
        ("Algoritme", ALGORITHMS[algorithm][0]),
        ("Minimum keertijd", ", ".join(f"{bt}: {mins} min" for bt, mins in turnaround_map.items())),
        ("Doel", "Minimaliseer aantal bussen, daarna minimaliseer wachttijd"),
        ("Locatie-matching", "Bus eindlocatie moet gelijk zijn aan volgende rit startlocatie"),
        ("Bustype-constraint", "Bussen worden alleen ingezet op ritten met hetzelfde bustype"),
        ("Datum-constraint", "Bussen worden per dag apart geoptimaliseerd"),
        ("Multiplicity", "Ritten met 'Aantal bussen' > 1 worden uitgesplitst in aparte ritten"),
    ]
    for label, val in params:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=val)
        row += 1

    row += 2

    # --- Section 5: Niet-toegewezen ritten check ---
    ws.cell(row=row, column=1, value="5. Controle")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    total_input = real_trip_count
    total_assigned = sum(len(r.real_trips) for r in rotations)
    ws.cell(row=row, column=1, value="Totaal ritten in invoer:")
    ws.cell(row=row, column=2, value=total_input)
    row += 1
    ws.cell(row=row, column=1, value="Totaal ritten toegewezen:")
    ws.cell(row=row, column=2, value=total_assigned)
    row += 1
    ws.cell(row=row, column=1, value="Niet-toegewezen ritten:")
    ws.cell(row=row, column=2, value=total_input - total_assigned)
    if total_input != total_assigned:
        ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
    row += 1

    row += 2

    # --- Section 6: Algorithm examples ---
    row = _write_algo_examples(ws, row)

    # Column widths (don't override col A width set by _write_algo_examples)
    widths = [None, 16, 18, 36, 36, 14, 14, 16, 14, 16, 14, 14]
    for j, w in enumerate(widths):
        if w is not None and j < 12:
            ws.column_dimensions[get_column_letter(1 + j)].width = w


def write_businzet_sheet(wb_out, rotations: list, all_trips: list, reserves: list):
    """
    Tab 4: Overzicht Businzet - Matrix of services x dates with bus counts.
    Similar to 'Overzicht businzet.xlsx' but auto-generated with details.
    """
    ws = wb_out.create_sheet(title="Overzicht Businzet")

    row = 1
    ws.cell(row=row, column=1, value="Overzicht Businzet")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2

    # --- Section 1: Service x Date matrix ---
    ws.cell(row=row, column=1, value="1. Busdiensten per datum")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    # Collect unique dates and services from trips
    dates = sorted(set(t.date_str for t in all_trips))
    services = sorted(set(t.service for t in all_trips))

    # Build lookup: (service, date) -> {bus_type, trip_count, bus_count}
    from collections import defaultdict
    svc_date = defaultdict(lambda: {"trips": 0, "bus_types": set()})
    for t in all_trips:
        key = (t.service, t.date_str)
        svc_date[key]["trips"] += 1
        svc_date[key]["bus_types"].add(t.bus_type)

    # Count buses per service+date from rotations
    rot_svc_date = defaultdict(set)
    for rot in rotations:
        for trip in rot.trips:
            rot_svc_date[(trip.service, rot.date_str)].add(rot.bus_id)

    # Header row
    ws.cell(row=row, column=1, value="Busdienst")
    ws.cell(row=row, column=1).font = HEADER_FONT_WHITE
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).border = THIN_BORDER
    for j, date in enumerate(dates):
        col_base = 2 + j * 3
        for offset, header in enumerate(["Ritten", "Bussen", "Type"]):
            cell = ws.cell(row=row, column=col_base + offset, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
    # Date labels above
    date_row = row - 1
    for j, date in enumerate(dates):
        col_base = 2 + j * 3
        cell = ws.cell(row=date_row, column=col_base, value=date)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=date_row, start_column=col_base,
                       end_row=date_row, end_column=col_base + 2)
    row += 1

    type_abbrev = {"Dubbeldekker": "DD", "Touringcar": "TC", "Lagevloerbus": "LVB", "Midi bus": "Midi", "Taxibus": "Taxi"}
    alt_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for s_idx, service in enumerate(services):
        use_fill = alt_fill if s_idx % 2 == 0 else None
        cell = ws.cell(row=row, column=1, value=service)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER
        if use_fill:
            cell.fill = use_fill

        for j, date in enumerate(dates):
            col_base = 2 + j * 3
            key = (service, date)
            info = svc_date.get(key)
            if info and info["trips"] > 0:
                n_trips = info["trips"]
                n_buses = len(rot_svc_date.get(key, set()))
                types_str = ", ".join(type_abbrev.get(bt, bt) for bt in sorted(info["bus_types"]))

                ws.cell(row=row, column=col_base, value=n_trips)
                ws.cell(row=row, column=col_base + 1, value=n_buses)
                ws.cell(row=row, column=col_base + 2, value=types_str)
            else:
                ws.cell(row=row, column=col_base, value="-")
                ws.cell(row=row, column=col_base + 1, value="-")
                ws.cell(row=row, column=col_base + 2, value="-")

            for offset in range(3):
                c = ws.cell(row=row, column=col_base + offset)
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center")
                if use_fill:
                    c.fill = use_fill
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    for j, date in enumerate(dates):
        col_base = 2 + j * 3
        date_trips = sum(1 for t in all_trips if t.date_str == date)
        date_buses = len(set(r.bus_id for r in rotations if r.date_str == date))
        ws.cell(row=row, column=col_base, value=date_trips)
        ws.cell(row=row, column=col_base + 1, value=date_buses)
        for offset in range(3):
            c = ws.cell(row=row, column=col_base + offset)
            c.border = THIN_BORDER
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")
    row += 3

    # --- Section 2: Buses per date + type summary ---
    ws.cell(row=row, column=1, value="2. Bussen per datum en bustype")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    bus_types_all = sorted(set(t.bus_type for t in all_trips))
    headers2 = ["Datum"] + [type_abbrev.get(bt, bt) for bt in bus_types_all] + ["Totaal", "Reserve"]
    for j, h in enumerate(headers2):
        cell = ws.cell(row=row, column=1 + j, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
    row += 1

    for date in dates:
        ws.cell(row=row, column=1, value=date)
        ws.cell(row=row, column=1).border = THIN_BORDER
        total_date = 0
        for bt_idx, bt in enumerate(bus_types_all):
            n = len([r for r in rotations if r.date_str == date and r.bus_type == bt])
            ws.cell(row=row, column=2 + bt_idx, value=n)
            ws.cell(row=row, column=2 + bt_idx).border = THIN_BORDER
            ws.cell(row=row, column=2 + bt_idx).alignment = Alignment(horizontal="center")
            total_date += n
        ws.cell(row=row, column=2 + len(bus_types_all), value=total_date)
        ws.cell(row=row, column=2 + len(bus_types_all)).border = THIN_BORDER
        ws.cell(row=row, column=2 + len(bus_types_all)).font = Font(bold=True)
        ws.cell(row=row, column=2 + len(bus_types_all)).alignment = Alignment(horizontal="center")

        # Reserve count for this date
        day_map = {"do": "donderdag", "vr": "vrijdag", "za": "zaterdag",
                   "zo": "zondag", "ma": "maandag"}
        day_abbrev = date.split()[0] if date else ""
        full_day = day_map.get(day_abbrev, day_abbrev)
        res_count = sum(r.count for r in reserves
                       if full_day in r.day.lower() or day_abbrev in r.day.lower()[:2])
        ws.cell(row=row, column=3 + len(bus_types_all), value=res_count)
        ws.cell(row=row, column=3 + len(bus_types_all)).border = THIN_BORDER
        ws.cell(row=row, column=3 + len(bus_types_all)).alignment = Alignment(horizontal="center")
        row += 1

    # Grand total
    ws.cell(row=row, column=1, value="TOTAAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    grand = 0
    for bt_idx, bt in enumerate(bus_types_all):
        n = len([r for r in rotations if r.bus_type == bt])
        ws.cell(row=row, column=2 + bt_idx, value=n)
        ws.cell(row=row, column=2 + bt_idx).font = Font(bold=True)
        ws.cell(row=row, column=2 + bt_idx).border = THIN_BORDER
        ws.cell(row=row, column=2 + bt_idx).alignment = Alignment(horizontal="center")
        grand += n
    ws.cell(row=row, column=2 + len(bus_types_all), value=grand)
    ws.cell(row=row, column=2 + len(bus_types_all)).font = Font(bold=True)
    ws.cell(row=row, column=2 + len(bus_types_all)).border = THIN_BORDER
    ws.cell(row=row, column=2 + len(bus_types_all)).alignment = Alignment(horizontal="center")
    res_total = sum(r.count for r in reserves)
    ws.cell(row=row, column=3 + len(bus_types_all), value=res_total)
    ws.cell(row=row, column=3 + len(bus_types_all)).font = Font(bold=True)
    ws.cell(row=row, column=3 + len(bus_types_all)).border = THIN_BORDER
    ws.cell(row=row, column=3 + len(bus_types_all)).alignment = Alignment(horizontal="center")
    row += 3

    # --- Section 3: Diensttijden per busdienst ---
    ws.cell(row=row, column=1, value="3. Diensttijden per busdienst")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    headers3 = ["Busdienst", "Datum", "Bustype", "Eerste vertrek", "Laatste aankomst",
                "Diensttijd (uur)", "Ritten", "Bussen ingezet"]
    for j, h in enumerate(headers3):
        cell = ws.cell(row=row, column=1 + j, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    row += 1

    for s_idx, service in enumerate(services):
        use_fill = alt_fill if s_idx % 2 == 0 else None
        for date in dates:
            key = (service, date)
            info = svc_date.get(key)
            if not info or info["trips"] == 0:
                continue
            svc_trips = [t for t in all_trips if t.service == service and t.date_str == date]
            if not svc_trips:
                continue
            first_dep = min(t.departure for t in svc_trips)
            last_arr = max(t.arrival for t in svc_trips)
            dienst_min = last_arr - first_dep
            n_buses = len(rot_svc_date.get(key, set()))
            types_str = ", ".join(type_abbrev.get(bt, bt) for bt in sorted(info["bus_types"]))

            values = [
                service, date, types_str,
                minutes_to_time(first_dep), minutes_to_time(last_arr),
                round(dienst_min / 60, 1), info["trips"], n_buses,
            ]
            for j, v in enumerate(values):
                cell = ws.cell(row=row, column=1 + j, value=v)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                if use_fill:
                    cell.fill = use_fill
                if j in (3, 4):
                    cell.number_format = "HH:MM"
            row += 1

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 20
    for c in range(2, 2 + len(dates) * 3 + 5):
        ws.column_dimensions[get_column_letter(c)].width = 14


def _write_algo_examples(ws, row):
    """Write simple, accessible algorithm examples to a worksheet."""
    ws.cell(row=row, column=1, value="6. Hoe werken de algoritmes? (voorbeelden)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 2

    # Example scenario
    ws.cell(row=row, column=1, value="Stel: 4 ritten op dezelfde dag, allemaal Touringcar, minimale keertijd = 8 minuten")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 1

    example_headers = ["Rit", "Van", "Naar", "Vertrek", "Aankomst"]
    for j, h in enumerate(example_headers):
        c = ws.cell(row=row, column=1 + j, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")
        c.border = THIN_BORDER
    row += 1

    example_trips = [
        ("Rit 1", "Utrecht", "Ede", "06:00", "06:42"),
        ("Rit 2", "Ede", "Utrecht", "06:50", "07:32"),
        ("Rit 3", "Utrecht", "Ede", "07:00", "07:42"),
        ("Rit 4", "Ede", "Utrecht", "07:50", "08:32"),
    ]
    for vals in example_trips:
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=1 + j, value=v)
            c.border = THIN_BORDER
        row += 1
    row += 1

    # --- GREEDY ---
    ws.cell(row=row, column=1, value="A) Greedy best-fit (\"pak de eerste de beste\")")
    ws.cell(row=row, column=1).font = Font(bold=True, color="1F4E79")
    row += 1
    greedy_lines = [
        "Het algoritme loopt de ritten af op vertrektijd en koppelt elke rit aan de bus",
        "met de kleinste wachttijd tussen zijn vorige aankomst en het nieuwe potentiële vertrek.",
        "",
        "Stap 1: Rit 1 (Ut→Ed 06:00-06:42) → geen bus beschikbaar → Bus A",
        "Stap 2: Rit 2 (Ed→Ut 06:50-07:32) → Bus A staat in Ede, wacht 8 min → Bus A",
        "Stap 3: Rit 3 (Ut→Ed 07:00-07:42) → Bus A is onderweg → geen bus → Bus B",
        "Stap 4: Rit 4 (Ed→Ut 07:50-08:32) → Bus A in Ut (wacht 18 min), Bus B in Ede (wacht 8 min)",
        "        → Bus B (kleinste wachttijd) → Bus B",
        "",
        "Resultaat: 2 bussen. Bus A: Rit 1→2 | Bus B: Rit 3→4",
        "Voordeel: Snel, werkt goed in de praktijk. Nadeel: vindt niet altijd het absolute minimum.",
    ]
    for line in greedy_lines:
        ws.cell(row=row, column=1, value=line)
        row += 1
    row += 1

    # --- MINCOST ---
    ws.cell(row=row, column=1, value="B) Min-cost matching (\"optimaal bussen + minimale wachttijd\")")
    ws.cell(row=row, column=1).font = Font(bold=True, color="1F4E79")
    row += 1
    mincost_lines = [
        "Bouwt een netwerk van alle mogelijke koppelingen tussen ritten en vindt",
        "de verdeling met het minimum aantal bussen EN de minste totale wachttijd.",
        "",
        "Mogelijke koppelingen (aankomstlocatie = vertreklocatie, genoeg keertijd):",
        "  Rit 1 (aankomst Ede 06:42) → Rit 2 (vertrek Ede 06:50): gap 8 min ✓",
        "  Rit 1 (aankomst Ede 06:42) → Rit 4 (vertrek Ede 07:50): gap 68 min ✓",
        "  Rit 3 (aankomst Ede 07:42) → Rit 4 (vertrek Ede 07:50): gap 8 min ✓",
        "",
        "Mogelijke oplossingen met 2 bussen:",
        "  Optie 1: Bus A: Rit 1→2 (wacht 8 min) | Bus B: Rit 3→4 (wacht 8 min)  → totaal 16 min wacht",
        "  Optie 2: Bus A: Rit 1→4 (wacht 68 min) | Bus B: Rit 3 | Bus C: Rit 2   → 3 bussen, slechter",
        "",
        "Min-cost kiest Optie 1: 2 bussen, 16 minuten totale wachttijd.",
        "",
        "NB: Zonder deadhead (huidig model) geeft greedy altijd hetzelfde resultaat.",
        "Min-cost is noodzakelijk als deadhead/repositionering wordt toegevoegd.",
    ]
    for line in mincost_lines:
        ws.cell(row=row, column=1, value=line)
        row += 1

    ws.column_dimensions["A"].width = 90
    return row


def write_sensitivity_sheet(wb_out, all_trips: list, base_turnaround_map: dict,
                            algorithm: str = "greedy"):
    """
    Tab: Sensitiviteitsanalyse - shows impact of different turnaround times.
    For each bus type present in the data, varies turnaround time and shows bus count.
    Always uses greedy algorithm for speed (mincost would be too slow for ~40+ runs).
    """
    # Always use greedy for sensitivity: mincost is too slow for many iterations
    sens_algo = "greedy"
    ws = wb_out.create_sheet(title="Sensitiviteit")
    row = 1
    ws.cell(row=row, column=1, value="Sensitiviteitsanalyse Keertijden")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2

    ws.cell(row=row, column=1, value="Wat als de minimale keertijd anders is? Hoeveel bussen zijn er dan nodig?")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 1
    ws.cell(row=row, column=1, value=f"Algoritme: {ALGORITHMS[sens_algo][0]} (sensitiviteit altijd greedy)")
    row += 2

    # Get bus types present in real trips only
    active_types = sorted(set(t.bus_type for t in all_trips if not t.is_reserve))
    # Test values from 2 to max(base+4, 15)
    max_test = max(max(base_turnaround_map.get(bt, 8) for bt in active_types) + 4, 15)
    test_values = list(range(2, max_test + 1))

    # --- Section 1: Total buses per turnaround time variation ---
    ws.cell(row=row, column=1, value="1. Totaal bussen per keertijd-variatie (per bustype)")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    for bus_type in active_types:
        ws.cell(row=row, column=1, value=f"Bustype: {bus_type}")
        ws.cell(row=row, column=1).font = Font(bold=True)
        base_val = base_turnaround_map.get(bus_type, 8)
        ws.cell(row=row, column=2, value=f"(baseline: {base_val} min)")
        row += 1

        # Headers
        headers = ["Keertijd (min)", "Bussen nodig", "Verschil t.o.v. baseline", "Benutting (%)"]
        for j, h in enumerate(headers):
            c = ws.cell(row=row, column=1 + j, value=h)
            c.font = HEADER_FONT_WHITE
            c.fill = HEADER_FILL
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center")
        row += 1

        # Compute for each test value
        bt_trips = [t for t in all_trips if t.bus_type == bus_type]
        baseline_buses = None

        for tv in test_values:
            test_map = dict(base_turnaround_map)
            test_map[bus_type] = tv
            rots = optimize_rotations(all_trips, test_map, algorithm=sens_algo)
            bt_rots = [r for r in rots if r.bus_type == bus_type and r.real_trips]
            n_buses = len(bt_rots)
            ride = sum(r.total_ride_minutes for r in bt_rots)
            dienst = sum(r.total_dienst_minutes for r in bt_rots)
            benutting = (ride / dienst * 100) if dienst > 0 else 0

            if tv == base_val:
                baseline_buses = n_buses

            diff = n_buses - baseline_buses if baseline_buses is not None else 0
            diff_str = f"{diff:+d}" if baseline_buses is not None else ""

            is_base = (tv == base_val)
            vals = [f"{tv} min", n_buses, diff_str, round(benutting, 1)]
            for j, v in enumerate(vals):
                c = ws.cell(row=row, column=1 + j, value=v)
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center")
                if is_base:
                    c.font = Font(bold=True)
                    c.fill = PatternFill("solid", fgColor="E2EFDA")
            row += 1

        row += 2

    # --- Section 2: Detail of short turnarounds per route ---
    ws.cell(row=row, column=1, value="2. Overzicht korte keertijden per traject")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1, value="Welke trajecten hebben korte keertijden? Zijn kortere keertijden realistisch?")
    ws.cell(row=row, column=1).font = Font(italic=True)
    row += 1

    route_headers = ["Dienst", "Bustype", "Min. keertijd (min)", "Locatie", "Voorbeeld aankomst", "Voorbeeld vertrek"]
    for j, h in enumerate(route_headers):
        c = ws.cell(row=row, column=1 + j, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")
    row += 1

    # Find min gap per service with example
    by_service = {}
    for t in all_trips:
        by_service.setdefault(t.service, []).append(t)

    service_gaps = []
    for service, svc_trips in sorted(by_service.items()):
        bus_type = svc_trips[0].bus_type
        arrivals_by_loc = {}
        departures_by_loc = {}
        for t in svc_trips:
            dest_loc = normalize_location(t.dest_code)
            orig_loc = normalize_location(t.origin_code)
            arrivals_by_loc.setdefault(dest_loc, []).append(t)
            departures_by_loc.setdefault(orig_loc, []).append(t)

        best_gap = None
        best_example = None
        for loc, arrs in arrivals_by_loc.items():
            deps = departures_by_loc.get(loc, [])
            deps_sorted = sorted(deps, key=lambda x: x.departure)
            for a in arrs:
                for d in deps_sorted:
                    g = d.departure - a.arrival
                    if g >= MIN_TURNAROUND_FLOOR:
                        if best_gap is None or g < best_gap:
                            best_gap = g
                            best_example = (loc, a, d)
                        break

        if best_gap is not None and best_example is not None:
            loc, a_trip, d_trip = best_example
            service_gaps.append((best_gap, service, bus_type, loc, a_trip, d_trip))

    service_gaps.sort(key=lambda x: x[0])
    for gap, service, bus_type, loc, a_trip, d_trip in service_gaps:
        vals = [
            service, bus_type, gap, loc,
            f"{minutes_to_str(a_trip.arrival)} ({a_trip.origin_name}→{a_trip.dest_name})",
            f"{minutes_to_str(d_trip.departure)} ({d_trip.origin_name}→{d_trip.dest_name})",
        ]
        for j, v in enumerate(vals):
            c = ws.cell(row=row, column=1 + j, value=v)
            c.border = THIN_BORDER
            c.alignment = Alignment(horizontal="center" if j in (2,) else "left")
        row += 1

    # Column widths
    for j, w in enumerate([28, 16, 18, 18, 40, 40]):
        ws.column_dimensions[get_column_letter(1 + j)].width = w


def write_risk_analysis_sheet(wb_out, risk_report: list):
    """Write a risk analysis sheet showing per-trip traffic risk and turnaround adjustments."""
    ws = wb_out.create_sheet(title="Risico-analyse")

    row = 1
    ws.cell(row=row, column=1, value="Risico-analyse: Verkeersinvloed op keertijden")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2

    # Explanation
    ws.cell(row=row, column=1, value="Toelichting kolommen:")
    ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
    row += 1
    ws.cell(row=row, column=1, value="• NS Gepland (min): Reistijd zoals gepland in de NS dienstregeling")
    row += 1
    ws.cell(row=row, column=1, value="• Rijtijd met verkeer (min): Werkelijke rijtijd volgens Google Maps met verkeer voor dat tijdslot")
    row += 1
    ws.cell(row=row, column=1, value="• Rijtijd zonder verkeer (min): Werkelijke rijtijd volgens Google Maps zonder verkeer (baseline)")
    row += 1
    ws.cell(row=row, column=1, value="• Marge t.o.v. verkeer (min): NS Gepland - Rijtijd met verkeer (negatief = te krap gepland)")
    row += 2

    # Summary
    n_total = len(risk_report)
    n_high = sum(1 for r in risk_report if r["risk"] == "HOOG")
    n_medium = sum(1 for r in risk_report if r["risk"] == "MATIG")
    n_ok = sum(1 for r in risk_report if r["risk"] == "OK")

    ws.cell(row=row, column=1, value="Samenvatting:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value=f"Totaal ritten geanalyseerd: {n_total}")
    row += 1
    ws.cell(row=row, column=1, value=f"Hoog risico (marge < 0 min, NS planning korter dan verkeer): {n_high}")
    ws.cell(row=row, column=1).font = Font(color="FF0000")
    row += 1
    ws.cell(row=row, column=1, value=f"Matig risico (marge < 5 min): {n_medium}")
    ws.cell(row=row, column=1).font = Font(color="FF8C00")
    row += 1
    ws.cell(row=row, column=1, value=f"OK (marge >= 5 min): {n_ok}")
    ws.cell(row=row, column=1).font = Font(color="008000")
    row += 2

    # Detail table
    headers = [
        "Rit ID", "Busdienst", "Richting", "Van", "Naar",
        "Vertrek", "Aankomst", "NS Gepland (min)", "Tijdslot",
        "Rijtijd met verkeer (min)", "Rijtijd zonder verkeer (min)", "Marge t.o.v. verkeer (min)",
        "Basis keertijd", "Extra keertijd", "Aangepaste keertijd", "Risico",
    ]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    row += 1

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # Sort by risk: HOOG first, then MATIG, then OK
    risk_order = {"HOOG": 0, "MATIG": 1, "OK": 2}
    sorted_report = sorted(risk_report,
                          key=lambda r: (risk_order.get(r["risk"], 9), -(r["buffer_min"] or 999)))

    def _time_str(minutes):
        if minutes is None:
            return ""
        h, m = divmod(int(minutes) % 1440, 60)
        return f"{h:02d}:{m:02d}"

    for r in sorted_report:
        vals = [
            r["trip_id"], r["service"], r["direction"], r["origin"], r["dest"],
            _time_str(r["departure"]), _time_str(r["arrival"]), r["scheduled_min"],
            r["time_slot"], r["traffic_min"], r["baseline_min"], r["buffer_min"],
            r["base_turnaround"], r["extra_turnaround"], r["adjusted_turnaround"],
            r["risk"],
        ]
        if r["risk"] == "HOOG":
            fill = red_fill
        elif r["risk"] == "MATIG":
            fill = orange_fill
        else:
            fill = green_fill

        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            cell.fill = fill
            if c >= 6:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # Column widths
    widths = [18, 25, 10, 22, 22, 8, 8, 12, 16, 12, 12, 12, 14, 14, 16, 10]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ---------------------------------------------------------------------------
# ZE Output Sheets - Version 6
# ---------------------------------------------------------------------------

# ZE-specific style constants
ZE_HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ZE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ZE_WARNING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
ZE_ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def write_ze_inzet_sheet(wb, ze_assignments: dict, all_results: dict,
                          rotations: list, min_ze_count: int):
    """Write ZE Inzet sheet showing ZE touringcar assignments."""
    ws = wb.create_sheet("ZE Inzet")

    headers = ["Bus ID", "Bustype", "Datum", "Totaal km", "ZE Bereik km",
               "Buffer km", "ZE Toegewezen", "Laden Nodig", "Reden"]

    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = ZE_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # Get touringcar rotations with feasibility
    touringcar_rots = [r for r in rotations if r.bus_type == "Touringcar"]

    # Sort: assigned first, then by feasibility, then by km
    def sort_key(r):
        feas = all_results.get(r.bus_id)
        if feas is None:
            return (2, 0, 0)
        is_assigned = r.bus_id in ze_assignments
        return (0 if is_assigned else 1, 0 if feas.is_feasible else 1, feas.total_km)

    touringcar_rots.sort(key=sort_key)

    row = 2
    for rotation in touringcar_rots:
        feas = all_results.get(rotation.bus_id)
        if feas is None:
            continue

        is_assigned = rotation.bus_id in ze_assignments

        ws.cell(row=row, column=1, value=rotation.bus_id).border = THIN_BORDER
        ws.cell(row=row, column=2, value=rotation.bus_type).border = THIN_BORDER
        ws.cell(row=row, column=3, value=rotation.date_str).border = THIN_BORDER
        ws.cell(row=row, column=4, value=round(feas.total_km, 1)).border = THIN_BORDER
        ws.cell(row=row, column=5, value=feas.ze_range_km).border = THIN_BORDER
        ws.cell(row=row, column=6, value=round(feas.buffer_km, 1)).border = THIN_BORDER

        ze_cell = ws.cell(row=row, column=7, value="JA" if is_assigned else "Nee")
        ze_cell.border = THIN_BORDER
        if is_assigned:
            ze_cell.fill = ZE_FILL
            ze_cell.font = Font(bold=True)

        charging_cell = ws.cell(row=row, column=8,
                                value="Ja" if feas.needs_charging else "Nee")
        charging_cell.border = THIN_BORDER
        if feas.needs_charging:
            charging_cell.fill = ZE_WARNING_FILL

        ws.cell(row=row, column=9, value=feas.reason).border = THIN_BORDER

        row += 1

    # Column widths
    col_widths = [15, 12, 15, 12, 14, 12, 14, 12, 55]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_ze_laadstrategie_sheet(wb, ze_assignments: dict, rotations: list):
    """Write Laadstrategie sheet showing charging plan for ZE buses."""
    ws = wb.create_sheet("Laadstrategie")

    headers = ["Bus ID", "Station", "Laadduur (min)", "Lader", "Vermogen (kW)",
               "Geschatte km opgeladen", "Opmerkingen"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = ZE_HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for rot_id, feas in sorted(ze_assignments.items()):
        if feas.needs_charging and feas.recommended_charging:
            for charge in feas.recommended_charging:
                ws.cell(row=row, column=1, value=rot_id).border = THIN_BORDER
                ws.cell(row=row, column=2, value=charge["station"]).border = THIN_BORDER
                ws.cell(row=row, column=3, value=charge["duration_min"]).border = THIN_BORDER
                ws.cell(row=row, column=4, value=charge["charger"]).border = THIN_BORDER
                ws.cell(row=row, column=5, value=charge["power_kw"]).border = THIN_BORDER
                ws.cell(row=row, column=6, value=charge["km_recovered"]).border = THIN_BORDER
                ws.cell(row=row, column=7, value="Laden tijdens wachttijd").border = THIN_BORDER
                row += 1
        else:
            ws.cell(row=row, column=1, value=rot_id).border = THIN_BORDER
            ws.cell(row=row, column=2, value="-").border = THIN_BORDER
            ws.cell(row=row, column=3, value="-").border = THIN_BORDER
            ws.cell(row=row, column=4, value="-").border = THIN_BORDER
            ws.cell(row=row, column=5, value="-").border = THIN_BORDER
            ws.cell(row=row, column=6, value="-").border = THIN_BORDER
            ws.cell(row=row, column=7, value="Geen tussentijds laden nodig").border = THIN_BORDER
            row += 1

    col_widths = [15, 20, 14, 25, 14, 20, 40]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_ze_samenvatting_sheet(wb, ze_assignments: dict, all_results: dict,
                                 rotations: list, min_ze_count: int, assigned_count: int):
    """Write ZE Samenvatting (summary) sheet."""
    ws = wb.create_sheet("ZE Samenvatting")

    total_touringcar = len([r for r in rotations if r.bus_type == "Touringcar"])
    ze_feasible = len([f for f in all_results.values() if f.is_feasible])
    ze_no_charging = len([f for f in all_results.values()
                           if f.is_feasible and not f.needs_charging])
    ze_with_charging = len([f for f in all_results.values()
                             if f.is_feasible and f.needs_charging])
    meets_requirement = assigned_count >= min_ze_count

    summary_data = [
        ("ZE TOURINGCAR INZET - SAMENVATTING", ""),
        ("", ""),
        ("NS Vereiste", f"Minimaal {min_ze_count} ZE touringcars"),
        ("", ""),
        ("Totaal aantal touringcar-omlopen", total_touringcar),
        ("Waarvan ZE-geschikt", ze_feasible),
        ("  - Zonder tussentijds laden", ze_no_charging),
        ("  - Met tussentijds laden", ze_with_charging),
        ("", ""),
        ("ZE touringcars toegewezen", assigned_count),
        ("Voldoet aan NS vereiste", "JA" if meets_requirement else "NEE"),
        ("", ""),
        ("TOELICHTING", ""),
        ("", ""),
        ("De ZE touringcars zijn toegewezen aan omlopen die:", ""),
        ("1. Binnen het elektrische bereik vallen, OF", ""),
        ("2. Voldoende wachttijd hebben om tussentijds op te laden", ""),
        ("", ""),
        ("Selectiecriteria (in volgorde van prioriteit):", ""),
        ("1. Omlopen die geen tussentijds laden nodig hebben", ""),
        ("2. Omlopen met grotere marge (buffer) op bereik", ""),
        ("3. Kortere omlopen (minder km)", ""),
    ]

    for row_idx, (label, value) in enumerate(summary_data, 1):
        cell1 = ws.cell(row=row_idx, column=1, value=label)
        cell2 = ws.cell(row=row_idx, column=2, value=value)

        if row_idx == 1:
            cell1.font = Font(bold=True, size=14)
        elif label in ["NS Vereiste", "ZE touringcars toegewezen",
                       "Voldoet aan NS vereiste", "TOELICHTING"]:
            cell1.font = Font(bold=True)

        if label == "Voldoet aan NS vereiste":
            if value == "JA":
                cell2.fill = ZE_FILL
            else:
                cell2.fill = ZE_ERROR_FILL
            cell2.font = Font(bold=True)

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 35


def write_fuel_analysis_sheet(wb, fuel_results: dict, fuel_stations: dict, fuel_config: dict):
    """Write fuel analysis sheet showing diesel range validation per rotation.

    Args:
        wb: openpyxl workbook
        fuel_results: Dict of {rotation_id: FuelValidationResult} objects
        fuel_stations: Dict of fuel stations per location
        fuel_config: Fuel configuration with ranges and speeds
    """
    ws = wb.create_sheet("Brandstof Analyse")

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # Headers
    headers = [
        "Bus ID", "Bustype", "Totale km", "Bereik (km)", "% Bereik",
        "Status", "Tankstops", "Dichtstbijzijnde Tank", "Afstand Tank (km)",
        "Rijtijd Tank (min)", "Toelichting"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Status colors
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    row = 2
    for result in fuel_results.values():
        # Calculate percentage of range used
        pct_range = (result.total_km / result.fuel_range_km * 100) if result.fuel_range_km > 0 else 0

        # Determine status
        if result.is_feasible and not result.needs_refuel:
            status = "OK"
            status_fill = green_fill
        elif result.is_feasible and result.needs_refuel:
            status = f"Tankstop nodig"
            status_fill = yellow_fill
        else:
            status = "Gesplitst"
            status_fill = red_fill

        # Find nearest fuel station info from the first fuel stop if available
        nearest_name = "-"
        nearest_dist = "-"
        nearest_time = "-"
        if result.fuel_stops:
            stop = result.fuel_stops[0]
            nearest_name = stop.fuel_station_name
            nearest_dist = f"{stop.fuel_station_distance_km:.1f}"
            nearest_time = f"{stop.drive_time_min:.0f}"

        ws.cell(row=row, column=1, value=result.rotation_id)
        ws.cell(row=row, column=2, value=result.bus_type)
        ws.cell(row=row, column=3, value=round(result.total_km, 1))
        ws.cell(row=row, column=4, value=round(result.fuel_range_km, 0))
        ws.cell(row=row, column=5, value=f"{pct_range:.0f}%")
        status_cell = ws.cell(row=row, column=6, value=status)
        status_cell.fill = status_fill
        ws.cell(row=row, column=7, value=len(result.fuel_stops))
        ws.cell(row=row, column=8, value=nearest_name)
        ws.cell(row=row, column=9, value=nearest_dist)
        ws.cell(row=row, column=10, value=nearest_time)
        ws.cell(row=row, column=11, value=result.reason)

        row += 1

    # Summary section
    row += 2
    ws.cell(row=row, column=1, value="SAMENVATTING").font = Font(bold=True)
    row += 1

    total_rotations = len(fuel_results)
    ok_count = sum(1 for r in fuel_results.values() if r.is_feasible and not r.needs_refuel)
    refuel_count = sum(1 for r in fuel_results.values() if r.is_feasible and r.needs_refuel)
    split_count = sum(1 for r in fuel_results.values() if not r.is_feasible)

    ws.cell(row=row, column=1, value="Totaal omlopen:")
    ws.cell(row=row, column=2, value=total_rotations)
    row += 1
    ws.cell(row=row, column=1, value="Binnen bereik (geen tankstop):")
    ws.cell(row=row, column=2, value=ok_count)
    row += 1
    ws.cell(row=row, column=1, value="Tankstop nodig:")
    ws.cell(row=row, column=2, value=refuel_count)
    row += 1
    ws.cell(row=row, column=1, value="Gesplitst (onhaalbaar):")
    ws.cell(row=row, column=2, value=split_count)

    # Fuel range info per bus type
    row += 2
    ws.cell(row=row, column=1, value="BEREIK PER BUSTYPE").font = Font(bold=True)
    row += 1
    for bus_type, range_km in fuel_config.get("diesel_range_km", {}).items():
        ws.cell(row=row, column=1, value=bus_type)
        ws.cell(row=row, column=2, value=f"{range_km:.0f} km")
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 30
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 40


def write_financial_sheet(wb, financials: dict):
    """Write financial overview sheet showing revenue, costs, and profit per rotation.

    Args:
        wb: openpyxl workbook
        financials: Dict from calculate_total_financials() with 'rotations' and 'totals'
    """
    ws = wb.create_sheet("Financieel Overzicht")

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    money_format = '€ #,##0.00'
    pct_format = '0.0%'

    # Title
    ws.cell(row=1, column=1, value="Financieel Overzicht - Versie 7")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Summary section
    row = 3
    ws.cell(row=row, column=1, value="SAMENVATTING").font = Font(bold=True, size=12)
    row += 1

    totals = financials['totals']
    summary_data = [
        ("Totale omzet (rijtijd × tarief)", totals['total_revenue']),
        ("Totale chauffeurskosten", totals['total_driver_cost']),
        ("Totale brandstofkosten", totals['total_fuel_cost']),
        ("Bruto winst", totals['total_gross_profit']),
        ("ZE bonus", totals['total_ze_bonus']),
        ("HVO bonus", totals['total_hvo_bonus']),
        ("Netto winst", totals['total_net_profit']),
    ]

    for label, value in summary_data:
        ws.cell(row=row, column=1, value=label)
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = money_format
        if "winst" in label.lower():
            if value >= 0:
                cell.font = Font(color="008000", bold=True)  # Green for profit
            else:
                cell.font = Font(color="FF0000", bold=True)  # Red for loss
        row += 1

    # Additional stats
    row += 1
    ws.cell(row=row, column=1, value="Totale rij-uren (betaald door NS)")
    ws.cell(row=row, column=2, value=totals['total_driving_hours'])
    ws.cell(row=row, column=2).number_format = '0.0'
    row += 1
    ws.cell(row=row, column=1, value="Totale dienst-uren (chauffeur)")
    ws.cell(row=row, column=2, value=totals['total_shift_hours'])
    ws.cell(row=row, column=2).number_format = '0.0'
    row += 1
    ws.cell(row=row, column=1, value="Totale ORT uren")
    ws.cell(row=row, column=2, value=totals['total_ort_hours'])
    ws.cell(row=row, column=2).number_format = '0.0'
    row += 1
    ws.cell(row=row, column=1, value="Totale ORT kosten")
    ws.cell(row=row, column=2, value=totals['total_ort_amount'])
    ws.cell(row=row, column=2).number_format = money_format
    row += 1
    ws.cell(row=row, column=1, value="Totale km")
    ws.cell(row=row, column=2, value=totals['total_km'])
    ws.cell(row=row, column=2).number_format = '#,##0'

    # Profit margin
    row += 2
    if totals['total_revenue'] > 0:
        margin = totals['total_net_profit'] / totals['total_revenue']
        ws.cell(row=row, column=1, value="Winstmarge")
        cell = ws.cell(row=row, column=2, value=margin)
        cell.number_format = pct_format
        if margin >= 0:
            cell.font = Font(color="008000", bold=True)
        else:
            cell.font = Font(color="FF0000", bold=True)

    # Detail table header
    row += 3
    ws.cell(row=row, column=1, value="DETAIL PER OMLOOP").font = Font(bold=True, size=12)
    row += 1

    headers = [
        "Bus ID", "Bustype", "Datum", "Rijtijd (min)", "Dienst (uur)",
        "Omzet", "Chauffeur", "ORT", "Brandstof", "Bruto winst", "Netto winst"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Color fills for profit/loss
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Data rows
    for fin in financials['rotations']:
        ws.cell(row=row, column=1, value=fin.rotation_id)
        ws.cell(row=row, column=2, value=fin.bus_type)
        ws.cell(row=row, column=3, value=fin.date_str)
        ws.cell(row=row, column=4, value=fin.driving_minutes)
        ws.cell(row=row, column=5, value=round(fin.driver_cost.shift_duration_hours, 1))

        ws.cell(row=row, column=6, value=fin.revenue).number_format = money_format
        ws.cell(row=row, column=7, value=fin.driver_cost.total_cost).number_format = money_format
        ws.cell(row=row, column=8, value=fin.driver_cost.ort_amount).number_format = money_format
        ws.cell(row=row, column=9, value=fin.fuel_cost).number_format = money_format

        gross_cell = ws.cell(row=row, column=10, value=fin.gross_profit)
        gross_cell.number_format = money_format
        gross_cell.fill = green_fill if fin.gross_profit >= 0 else red_fill

        net_cell = ws.cell(row=row, column=11, value=fin.net_profit)
        net_cell.number_format = money_format
        net_cell.fill = green_fill if fin.net_profit >= 0 else red_fill

        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12


def write_financial_comparison_sheet(wb, all_financials: dict):
    """Write comparison sheet showing financials for all version permutations.

    Args:
        wb: openpyxl workbook
        all_financials: Dict mapping permutation name to financials dict
                       e.g. {"basis": {...}, "deadhead": {...}, "multidag": {...}, ...}
    """
    ws = wb.create_sheet("Vergelijking Versies")

    # Header styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    money_format = '€ #,##0.00'
    pct_format = '0.0%'
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    best_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # Title
    ws.cell(row=1, column=1, value="Financiële Vergelijking - Alle Versies")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Permutation name mapping for display
    display_names = {
        "basis": "Basis",
        "basis_brandstof": "Basis + Brandstof",
        "basis_risico": "Basis + Risico",
        "basis_risico_brandstof": "Basis + Risico + Brandstof",
        "deadhead": "Deadhead",
        "deadhead_brandstof": "Deadhead + Brandstof",
        "deadhead_risico": "Deadhead + Risico",
        "deadhead_risico_brandstof": "Deadhead + Risico + Brandstof",
        "multidag": "Multidag",
        "multidag_brandstof": "Multidag + Brandstof",
        "multidag_risico": "Multidag + Risico",
        "multidag_risico_brandstof": "Multidag + Risico + Brandstof",
        "deadhead_multidag": "Deadhead + Multidag",
        "deadhead_multidag_brandstof": "Deadhead + Multidag + Brandstof",
        "deadhead_multidag_risico": "Deadhead + Multidag + Risico",
        "deadhead_multidag_risico_brandstof": "Alle opties",
    }

    # Header row with permutation names
    row = 3
    ws.cell(row=row, column=1, value="Kenmerk").font = Font(bold=True)
    for col, (key, _) in enumerate(all_financials.items(), 2):
        cell = ws.cell(row=row, column=col, value=display_names.get(key, key))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Metrics to compare
    row += 1
    metrics = [
        ("Aantal bussen", lambda t: t.get('bus_count', 0), '0'),
        ("Totale rij-uren", lambda t: t.get('total_driving_hours', 0), '0.0'),
        ("Totale dienst-uren", lambda t: t.get('total_shift_hours', 0), '0.0'),
        ("Totale ORT uren", lambda t: t.get('total_ort_hours', 0), '0.0'),
        ("", None, None),  # Spacer
        ("Totale omzet", lambda t: t.get('total_revenue', 0), money_format),
        ("Chauffeurkosten", lambda t: t.get('total_driver_cost', 0), money_format),
        ("ORT kosten", lambda t: t.get('total_ort_amount', 0), money_format),
        ("Brandstofkosten (ritten)", lambda t: t.get('total_fuel_cost', 0), money_format),
        ("Brandstofkosten (garage)", lambda t: t.get('total_garage_fuel_cost', 0), money_format),
        ("", None, None),  # Spacer
        ("Bruto winst", lambda t: t.get('total_gross_profit', 0), money_format),
        ("ZE bonus", lambda t: t.get('total_ze_bonus', 0), money_format),
        ("HVO bonus", lambda t: t.get('total_hvo_bonus', 0), money_format),
        ("Netto winst", lambda t: t.get('total_net_profit', 0), money_format),
        ("Winstmarge", lambda t: t.get('total_net_profit', 0) / t.get('total_revenue', 1) if t.get('total_revenue', 0) > 0 else 0, pct_format),
    ]

    # Find best net profit for highlighting
    best_profit = max(
        (fin['totals'].get('total_net_profit', 0) for fin in all_financials.values()),
        default=0
    )

    for label, getter, fmt in metrics:
        row += 1
        if getter is None:
            continue  # Spacer row

        ws.cell(row=row, column=1, value=label).font = Font(bold=True) if "winst" in label.lower() else None

        for col, (key, fin) in enumerate(all_financials.items(), 2):
            totals = fin['totals']
            value = getter(totals)
            cell = ws.cell(row=row, column=col, value=value)
            if fmt:
                cell.number_format = fmt

            # Highlight best net profit
            if label == "Netto winst" and value == best_profit and best_profit > 0:
                cell.fill = best_fill
                cell.font = Font(bold=True, color="006400")
            elif "winst" in label.lower() and value >= 0:
                cell.fill = green_fill

    # Add difference from basis row
    if "basis" in all_financials and len(all_financials) > 1:
        row += 2
        ws.cell(row=row, column=1, value="Verschil t.o.v. basis").font = Font(bold=True, size=11)
        basis_profit = all_financials["basis"]['totals'].get('total_net_profit', 0)

        row += 1
        ws.cell(row=row, column=1, value="Winst verschil (EUR)")
        for col, (key, fin) in enumerate(all_financials.items(), 2):
            if key == "basis":
                ws.cell(row=row, column=col, value="-")
            else:
                diff = fin['totals'].get('total_net_profit', 0) - basis_profit
                cell = ws.cell(row=row, column=col, value=diff)
                cell.number_format = '+€ #,##0.00;-€ #,##0.00'
                if diff > 0:
                    cell.font = Font(color="006400", bold=True)
                elif diff < 0:
                    cell.font = Font(color="FF0000")

        row += 1
        ws.cell(row=row, column=1, value="Winst verschil (%)")
        for col, (key, fin) in enumerate(all_financials.items(), 2):
            if key == "basis":
                ws.cell(row=row, column=col, value="-")
            elif basis_profit != 0:
                pct_diff = (fin['totals'].get('total_net_profit', 0) - basis_profit) / abs(basis_profit)
                cell = ws.cell(row=row, column=col, value=pct_diff)
                cell.number_format = '+0.0%;-0.0%'
                if pct_diff > 0:
                    cell.font = Font(color="006400", bold=True)
                elif pct_diff < 0:
                    cell.font = Font(color="FF0000")

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col in range(2, len(all_financials) + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22


def write_cost_calculation_sheet(wb, financial_config):
    """Write explanation sheet showing how costs are calculated step-by-step.

    This helps human roster planners understand and verify the cost calculations.
    """
    ws = wb.create_sheet("Kostenberekening Uitleg")

    # Styling
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    section_font = Font(bold=True, size=12)
    formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    row = 1
    ws.cell(row=row, column=1, value="Kostenberekening - Stap-voor-stap Uitleg")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    row += 2

    # Section 1: Overview
    ws.cell(row=row, column=1, value="1. OVERZICHT FORMULE").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Netto Winst = Omzet - Chauffeurskosten - Brandstofkosten + Bonussen")
    ws.cell(row=row, column=1).fill = formula_fill
    row += 2

    # Section 2: Revenue
    ws.cell(row=row, column=1, value="2. OMZET (Revenue)").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Formule:")
    ws.cell(row=row, column=2, value="Rijtijd (uren) × Uurtarief per bustype")
    row += 1
    ws.cell(row=row, column=1, value="Let op:")
    ws.cell(row=row, column=2, value="Alleen actieve rijtijd wordt betaald, NIET wachttijd of deadhead")
    row += 2

    ws.cell(row=row, column=1, value="Uurtarieven per bustype:")
    row += 1
    rates = getattr(financial_config, 'rates', {})
    for bus_type, rate in rates.items():
        ws.cell(row=row, column=2, value=bus_type)
        ws.cell(row=row, column=3, value=rate)
        ws.cell(row=row, column=3).number_format = '€ #,##0.00'
        row += 1
    row += 1

    # Section 3: Driver costs
    ws.cell(row=row, column=1, value="3. CHAUFFEURSKOSTEN").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Formule:")
    ws.cell(row=row, column=2, value="(Betaalde uren × Basis uurloon + ORT) × Werkgever opslag")
    row += 1

    ws.cell(row=row, column=1, value="Basis uurloon:")
    ws.cell(row=row, column=2, value=getattr(financial_config, 'basis_uurloon', 18.50))
    ws.cell(row=row, column=2).number_format = '€ #,##0.00'
    row += 1

    ws.cell(row=row, column=1, value="Werkgever opslag:")
    ws.cell(row=row, column=2, value=f"{getattr(financial_config, 'werkgever_opslag', 1.35):.0%}")
    row += 2

    # Pauzestaffel
    ws.cell(row=row, column=1, value="3a. Pauzestaffel (onbetaalde pauze per dienstduur):")
    row += 1
    pauzestaffel = getattr(financial_config, 'pauzestaffel', [])
    for max_hours, deduction_min in pauzestaffel:
        ws.cell(row=row, column=2, value=f"Dienst ≤ {max_hours} uur:")
        ws.cell(row=row, column=3, value=f"{deduction_min} min pauze")
        row += 1
    ws.cell(row=row, column=2, value="Betaalde uren = Dienst uren - Pauze minuten/60")
    row += 2

    # ORT
    ws.cell(row=row, column=1, value="3b. ORT (Onregelmatigheidstoeslag):")
    row += 1
    ws.cell(row=row, column=2, value="Doordeweeks avond (19:00-07:30):")
    ws.cell(row=row, column=3, value=getattr(financial_config, 'ort_weekday_rate', 4.80))
    ws.cell(row=row, column=3).number_format = '€ #,##0.00'
    ws.cell(row=row, column=4, value="per uur")
    row += 1
    ws.cell(row=row, column=2, value="Zaterdag (hele dag):")
    ws.cell(row=row, column=3, value=getattr(financial_config, 'ort_saturday_rate', 4.80))
    ws.cell(row=row, column=3).number_format = '€ #,##0.00'
    ws.cell(row=row, column=4, value="per uur")
    row += 1
    ws.cell(row=row, column=2, value="Zondag (hele dag):")
    ws.cell(row=row, column=3, value=getattr(financial_config, 'ort_sunday_rate', 6.68))
    ws.cell(row=row, column=3).number_format = '€ #,##0.00'
    ws.cell(row=row, column=4, value="per uur")
    row += 2

    # Section 4: Fuel costs
    ws.cell(row=row, column=1, value="4. BRANDSTOFKOSTEN").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Formule:")
    ws.cell(row=row, column=2, value="Totale km × Verbruik (L/100km) / 100 × Dieselprijs (€/L)")
    row += 1
    ws.cell(row=row, column=1, value="Dieselprijs:")
    ws.cell(row=row, column=2, value=getattr(financial_config, 'diesel_price', 1.65))
    ws.cell(row=row, column=2).number_format = '€ #,##0.00'
    ws.cell(row=row, column=3, value="per liter")
    row += 2

    ws.cell(row=row, column=1, value="Verbruik per bustype (L/100km):")
    row += 1
    fuel_consumption = getattr(financial_config, 'fuel_consumption', {})
    for bus_type, consumption in fuel_consumption.items():
        ws.cell(row=row, column=2, value=bus_type)
        ws.cell(row=row, column=3, value=consumption)
        ws.cell(row=row, column=3).number_format = '0.0'
        ws.cell(row=row, column=4, value="L/100km")
        row += 1
    row += 1

    # Section 5: Garage travel
    ws.cell(row=row, column=1, value="5. GARAGE/REMISE REISKOSTEN").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Per bus: 2× enkele reis (heen + terug)")
    row += 1
    ws.cell(row=row, column=2, value="Enkele afstand:")
    ws.cell(row=row, column=3, value=getattr(financial_config, 'garage_afstand_enkel_km', 50))
    ws.cell(row=row, column=4, value="km")
    row += 1
    ws.cell(row=row, column=2, value="Enkele reistijd:")
    ws.cell(row=row, column=3, value=getattr(financial_config, 'garage_reistijd_enkel_min', 60))
    ws.cell(row=row, column=4, value="min")
    row += 1
    ws.cell(row=row, column=2, value="Telt mee in diensttijd:")
    ws.cell(row=row, column=3, value="Ja" if getattr(financial_config, 'garage_include_in_shift', True) else "Nee")
    row += 2

    # Section 6: Bonuses
    ws.cell(row=row, column=1, value="6. DUURZAAMHEIDSBONUSSEN").font = section_font
    row += 1
    ws.cell(row=row, column=2, value="ZE bonus:")
    ws.cell(row=row, column=3, value=getattr(financial_config, 'ze_bonus_per_km', 0.12))
    ws.cell(row=row, column=3).number_format = '€ #,##0.00'
    ws.cell(row=row, column=4, value="per km (Zero Emissie)")
    row += 1
    ws.cell(row=row, column=2, value="HVO bonus:")
    ws.cell(row=row, column=3, value=getattr(financial_config, 'hvo_bonus_per_liter', 0.05))
    ws.cell(row=row, column=3).number_format = '€ #,##0.00'
    ws.cell(row=row, column=4, value="per liter HVO100")
    row += 2

    # Section 7: Example calculation
    ws.cell(row=row, column=1, value="7. VOORBEELDBEREKENING").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Bus met 6 uur rijtijd (Touringcar), 8 uur dienst, 200 km:")
    row += 1
    example_rate = rates.get("Touringcar", 80.45)
    ws.cell(row=row, column=2, value="Omzet:")
    ws.cell(row=row, column=3, value=f"6 uur × €{example_rate:.2f} = €{6*example_rate:.2f}")
    row += 1

    basis_loon = getattr(financial_config, 'basis_uurloon', 18.50)
    werkgever = getattr(financial_config, 'werkgever_opslag', 1.35)
    paid_hours = 8 - 0.5  # 30 min break for 8-hour shift
    driver_cost = paid_hours * basis_loon * werkgever
    ws.cell(row=row, column=2, value="Chauffeurskosten:")
    ws.cell(row=row, column=3, value=f"{paid_hours} uur × €{basis_loon:.2f} × {werkgever:.2f} = €{driver_cost:.2f}")
    row += 1

    diesel_price = getattr(financial_config, 'diesel_price', 1.65)
    consumption = fuel_consumption.get("Touringcar", 32)
    fuel_cost = 200 * consumption / 100 * diesel_price
    ws.cell(row=row, column=2, value="Brandstof:")
    ws.cell(row=row, column=3, value=f"200 km × {consumption}L/100km × €{diesel_price:.2f} = €{fuel_cost:.2f}")
    row += 1

    net_profit = 6*example_rate - driver_cost - fuel_cost
    ws.cell(row=row, column=2, value="Netto winst:")
    cell = ws.cell(row=row, column=3, value=f"€{6*example_rate:.2f} - €{driver_cost:.2f} - €{fuel_cost:.2f} = €{net_profit:.2f}")
    cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15


def generate_ze_output(rotations: list, output_file: str, ze_config: dict,
                        charging_stations: dict, min_ze_count: int = 5,
                        append_to_existing: bool = False) -> dict:
    """Generate Version 6 ZE output: ZE touringcar assignment and charging strategy.

    Args:
        append_to_existing: If True, loads existing file and adds ZE sheets to it

    Returns dict with stats: total_touringcar, ze_feasible, assigned_count, meets_requirement
    """
    # Analyze and assign ZE buses
    ze_assignments, all_results, assigned_count = assign_ze_buses(
        rotations, min_ze_count, ze_config, charging_stations, "Touringcar"
    )

    # Create or load workbook
    if append_to_existing and Path(output_file).exists():
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        # Write Uitleg sheet first (Version 6) - only for new workbooks
        write_uitleg_sheet(wb, 6)

    # Write ZE sheets
    write_ze_inzet_sheet(wb, ze_assignments, all_results, rotations, min_ze_count)
    write_ze_laadstrategie_sheet(wb, ze_assignments, rotations)
    write_ze_samenvatting_sheet(wb, ze_assignments, all_results, rotations,
                                 min_ze_count, assigned_count)

    wb.save(output_file)

    # Calculate stats
    total_touringcar = len([r for r in rotations if r.bus_type == "Touringcar"])
    ze_feasible = len([f for f in all_results.values() if f.is_feasible])

    return {
        "total_touringcar": total_touringcar,
        "ze_feasible": ze_feasible,
        "assigned_count": assigned_count,
        "meets_requirement": assigned_count >= min_ze_count,
    }


# ---------------------------------------------------------------------------
# Version explanation texts for Uitleg sheet
# ---------------------------------------------------------------------------

VERSION_EXPLANATIONS = {
    1: {
        "titel": "Versie 1: Per Dienst (Baseline)",
        "beschrijving": [
            "Deze versie optimaliseert elke busdienst (Excel-tab) APART.",
            "Ritten van verschillende diensten worden NIET aan elkaar gekoppeld.",
            "Reservebussen worden als aparte voertuigen geteld (niet geïntegreerd).",
        ],
        "hoe_het_werkt": [
            "1. Per busdienst worden ritten chronologisch gesorteerd",
            "2. Ritten worden gekoppeld als: zelfde bustype, zelfde locatie, voldoende keertijd",
            "3. Resultaat: minimaal aantal bussen per dienst",
        ],
        "verschil_vorige": [],  # No previous version
        "use_case": "Baseline vergelijking, contractuele per-dienst vereisten",
    },
    2: {
        "titel": "Versie 2: Per Dienst + Reserve Matching",
        "beschrijving": [
            "Zelfde ritketening als Versie 1, maar met OPTIMALE RESERVE TOEWIJZING.",
            "Reservediensten worden toegewezen aan bussen tijdens hun wachttijd.",
            "Bipartite matching zorgt voor optimale dekking van reserve-eisen.",
        ],
        "hoe_het_werkt": [
            "1. Eerst: zelfde ritketening als Versie 1",
            "2. Dan: voor elke bus, check of wachttijden overlappen met reserve-eisen",
            "3. Bipartite matching wijst reserves optimaal toe aan beschikbare bussen",
            "4. Ongedekte reserves vereisen extra standalone bussen",
        ],
        "verschil_vorige": [
            "NIEUW t.o.v. Versie 1:",
            "• Reservediensten worden slim toegewezen aan bestaande bussen",
            "• Minder totaal bussen nodig doordat wachttijd benut wordt voor reserve",
            "• Bipartite matching garandeert optimale reserve-dekking",
        ],
        "use_case": "Efficiëntere inzet door reserve-integratie",
    },
    3: {
        "titel": "Versie 3: Gecombineerd + Reserves + Sensitiviteit",
        "beschrijving": [
            "Cross-dienst optimalisatie: ritten van VERSCHILLENDE diensten worden gekoppeld.",
            "Reservebussen als 'phantom trips' geïntegreerd in omlopen.",
            "Sensitiviteitsanalyse toont impact van keertijd-wijzigingen.",
        ],
        "hoe_het_werkt": [
            "1. ALLE ritten van alle diensten worden samengevoegd",
            "2. Optimalisatie over de gehele set (niet per dienst)",
            "3. Reservediensten worden als phantom-ritten toegevoegd",
            "4. Sensitiviteitsanalyse berekent: wat als keertijd +/- X minuten?",
        ],
        "verschil_vorige": [
            "NIEUW t.o.v. Versie 2:",
            "• Cross-dienst ketening: een bus kan ritten van meerdere diensten doen",
            "• Significante busreductie door slimmere ketening",
            "• Reserves als phantom-trips (niet apart toegewezen)",
            "• Sensitiviteitsanalyse sheet toegevoegd",
        ],
        "use_case": "Maximale efficiëntie zonder locatie-verplaatsing",
    },
    4: {
        "titel": "Versie 4: Gecombineerd + Risico-Gebaseerde Keertijd",
        "beschrijving": [
            "Zelfde als Versie 3, maar met VERKEERSAFHANKELIJKE keertijden.",
            "Google Maps verkeerdata bepaalt risico per rit.",
            "Keertijden worden verhoogd voor hoog-risico verbindingen.",
        ],
        "hoe_het_werkt": [
            "1. Eerst: laad traffic_matrix.json met tijdslot-specifieke reistijden",
            "2. Per rit: vergelijk geplande duur met Google Maps verwachting",
            "3. Als Google Maps > gepland: markeer als risico (HOOG/MATIG/OK)",
            "4. Verhoog keertijd voor ritten met hoog risico",
            "5. Optimaliseer met aangepaste keertijden",
        ],
        "verschil_vorige": [
            "NIEUW t.o.v. Versie 3:",
            "• Verkeersbewuste keertijden (spits vs dal vs weekend)",
            "• Risico-analyse sheet toont per-rit risicoscore",
            "• Robuustere planning die rekening houdt met verkeersvertragingen",
            "• Vereist: --traffic-matrix met Google Maps data",
        ],
        "use_case": "Robuuste planning met verkeersrisico-analyse",
    },
    5: {
        "titel": "Versie 5: Gecombineerd + Deadhead Repositionering",
        "beschrijving": [
            "Maximale flexibiliteit: bussen kunnen LEEG naar andere locaties rijden.",
            "Cross-locatie verbindingen mogelijk via deadhead matrix.",
            "Kostenfunctie weegt deadhead-tijd 2x zwaarder dan wachttijd.",
        ],
        "hoe_het_werkt": [
            "1. Laad deadhead matrix: reistijden tussen alle stations",
            "2. Een bus op locatie A kan nu een rit pakken die start op locatie B",
            "3. Voorwaarde: voldoende tijd om leeg van A naar B te rijden + keertijd",
            "4. Kostenfunctie: deadhead × 2 + wachttijd (straft lege ritten af)",
            "5. Min-cost matching vindt optimale balans",
        ],
        "verschil_vorige": [
            "NIEUW t.o.v. Versie 4:",
            "• Cross-locatie verbindingen: bus rijdt leeg naar andere halte",
            "• Potentieel minder bussen, maar meer lege kilometers",
            "• Deadhead rijen zichtbaar in het busomloopschema",
            "• Vereist: --deadhead met station-naar-station reistijden",
        ],
        "use_case": "Maximale optimalisatie, trade-off bussen vs lege km",
    },
    6: {
        "titel": "Versie 6: Brandstof/Laad Constraints",
        "beschrijving": [
            "Integreert BRANDSTOF en LAAD beperkingen in de optimalisatie.",
            "Cumulatieve km per bus wordt gevalideerd tegen actieradius.",
            "Ketens worden gesplitst als bereik overschreden zonder tankgelegenheid.",
        ],
        "hoe_het_werkt": [
            "1. Laad brandstofconfiguratie: actieradius per bustype",
            "2. Per keten: bereken cumulatieve km (rit-km + deadhead-km)",
            "3. Bij elke rit-overgang: check of km < resterende actieradius",
            "4. Als overschreden: zoek tankgelegenheid in wachttijd",
            "5. Als geen tankgelegenheid: splits keten (extra bus nodig)",
            "6. ZE-analyse: wijs minimaal 5 ZE Touringcars toe (NS K3 eis)",
        ],
        "verschil_vorige": [
            "NIEUW t.o.v. Versie 5:",
            "• Brandstofbereik validatie (diesel actieradius per bustype)",
            "• Automatische keten-splitsing bij bereik-overschrijding",
            "• Tankstop planning tijdens wachttijden",
            "• ZE (Zero Emission) haalbaarheidsanalyse",
            "• Laadstrategie voor elektrische bussen",
            "• Vereist: --fuel-constraints en --tanklocaties",
        ],
        "use_case": "Realistische brandstof-logistiek en ZE planning",
    },
}


def write_uitleg_sheet(wb, version: int):
    """Write explanation sheet as first tab in the workbook."""
    if version not in VERSION_EXPLANATIONS:
        return

    info = VERSION_EXPLANATIONS[version]
    ws = wb.create_sheet(title="Uitleg", index=0)

    # Styles
    from openpyxl.styles import Font, Alignment, PatternFill

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    row = 1

    # Title
    ws.cell(row, 1, info["titel"]).font = title_font
    row += 2

    # Beschrijving
    ws.cell(row, 1, "Beschrijving:").font = header_font
    row += 1
    for line in info["beschrijving"]:
        ws.cell(row, 1, line)
        row += 1
    row += 1

    # Hoe het werkt
    ws.cell(row, 1, "Hoe het werkt:").font = header_font
    row += 1
    for line in info["hoe_het_werkt"]:
        ws.cell(row, 1, line)
        row += 1
    row += 1

    # Verschil met vorige versie
    if info["verschil_vorige"]:
        ws.cell(row, 1, "Verschil met vorige versie:").font = header_font
        ws.cell(row, 1).fill = header_fill
        row += 1
        for line in info["verschil_vorige"]:
            ws.cell(row, 1, line)
            row += 1
        row += 1

    # Use case
    ws.cell(row, 1, "Toepassing:").font = header_font
    row += 1
    ws.cell(row, 1, info["use_case"])

    # Set column width
    ws.column_dimensions["A"].width = 80


def generate_output(rotations: list, all_trips: list, reserves: list, output_file: str,
                    turnaround_map: dict = None, algorithm: str = "greedy",
                    include_sensitivity: bool = False, output_mode: int = 1,
                    risk_report: list = None, deadhead_matrix: dict = None,
                    version: int = 1):
    """Generate the complete output Excel workbook.

    output_mode:
        1 = baseline per dienst (no reserves)
        2 = per dienst + optimal idle reserve matching
        3 = per dienst + reserve phantom trips
        4 = gecombineerd + reserve phantom trips + sensitivity
    risk_report: optional list of dicts from compute_trip_turnaround_overrides.
    deadhead_matrix: optional {origin: {dest: minutes}} for showing repositioning trips.
    version: version number (1-6) for the Uitleg sheet explanation.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Tab 0: Uitleg (explanation of this version)
    write_uitleg_sheet(wb, version)

    # Tab 1: Busomloop (Transvision-stijl)
    write_omloop_sheet(wb, rotations, reserves, deadhead_matrix=deadhead_matrix)

    # Tab 2: Overzicht ritsamenhang
    write_overzicht_sheet(wb, rotations, all_trips)

    # Tab 3: Berekeningen
    write_berekeningen_sheet(wb, rotations, all_trips, reserves, turnaround_map, algorithm,
                             output_mode=output_mode)

    # Tab 4: Overzicht Businzet
    write_businzet_sheet(wb, rotations, all_trips, reserves)

    # Tab 5 (optional): Sensitiviteitsanalyse
    if include_sensitivity:
        write_sensitivity_sheet(wb, all_trips, turnaround_map, algorithm)

    # Tab 6 (optional): Risico-analyse
    if risk_report:
        write_risk_analysis_sheet(wb, risk_report)

    wb.save(output_file)
    return output_file


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Busomloop Optimizer - Genereert optimale busomlopen uit NS TVV dienstregeling"
    )
    parser.add_argument(
        "input_file",
        help="Invoer Excel bestand (Bijlage J casus busdiensten)",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Uitvoer Excel bestand (standaard: busomloop_output.xlsx)",
    )
    parser.add_argument(
        "--algoritme", "-a",
        choices=list(ALGORITHMS.keys()) + ["all"],
        default="all",
        help="Optimalisatie-algoritme: greedy (snel, optimaal zonder deadhead), "
             "mincost (optimaal min. bussen + min. wachttijd, ook met deadhead), "
             "all (beide). Standaard: all",
    )
    parser.add_argument(
        "--snel",
        action="store_true",
        help="Snelle modus: outputs 1-4 alleen met greedy, alleen de deadhead-"
             "output (5) met alle algoritmes. Scheelt rekentijd.",
    )
    parser.add_argument(
        "--keertijd",
        type=int,
        default=None,
        help="Keertijd in minuten voor ALLE bustypes. "
             "Overschrijft de standaardwaarden.",
    )
    parser.add_argument(
        "--data-keertijd",
        action="store_true",
        help="Gebruik de minimale keertijd uit de dienstregeling (kleinste gat "
             "tussen twee opeenvolgende ritten per tabblad) in plaats van de "
             "standaardwaarden. Let op: dit kan onrealistisch lage waarden geven.",
    )
    parser.add_argument(
        "--keer-dd",
        type=int,
        default=None,
        help=f"Keertijd dubbeldekker in minuten (standaard: {MIN_TURNAROUND_DEFAULTS['Dubbeldekker']})",
    )
    parser.add_argument(
        "--keer-tc",
        type=int,
        default=None,
        help=f"Keertijd touringcar in minuten (standaard: {MIN_TURNAROUND_DEFAULTS['Touringcar']})",
    )
    parser.add_argument(
        "--keer-lvb",
        type=int,
        default=None,
        help=f"Keertijd lagevloerbus/gelede bus in minuten (standaard: {MIN_TURNAROUND_DEFAULTS['Lagevloerbus']})",
    )
    parser.add_argument(
        "--keer-midi",
        type=int,
        default=None,
        help=f"Keertijd midi bus in minuten (standaard: {MIN_TURNAROUND_DEFAULTS['Midi bus']})",
    )
    parser.add_argument(
        "--keer-taxi",
        type=int,
        default=None,
        help=f"Keertijd taxibus in minuten (standaard: {MIN_TURNAROUND_DEFAULTS['Taxibus']})",
    )
    parser.add_argument(
        "--deadhead",
        default=None,
        help="JSON bestand met deadhead matrix (rijtijden tussen stations). "
             "Gegenereerd door google_maps_distances.py. "
             "Als opgegeven, mogen bussen lege ritten maken tussen stations.",
    )
    parser.add_argument(
        "--traffic-matrix",
        default=None,
        help="JSON bestand met traffic-aware matrices (per tijdslot). "
             "Gegenereerd door google_maps_distances.py --traffic. "
             "Wordt gebruikt voor risico-analyse in output 4.",
    )
    parser.add_argument(
        "--capaciteit",
        default=None,
        help="JSON bestand met haltecapaciteiten (max gelijktijdige bussen). "
             'Formaat: {"Utrecht Centraal": 6, "Ede-Wageningen": 5}. '
             "Na optimalisatie wordt gecontroleerd of de capaciteit overschreden wordt.",
    )
    # Version 6: ZE (Zero Emission) arguments
    parser.add_argument(
        "--ze",
        action="store_true",
        help="Genereer Output 6: ZE touringcar toewijzing en laadstrategie (voor NS tender K3).",
    )
    parser.add_argument(
        "--min-ze",
        type=int,
        default=5,
        help="Minimum aantal ZE touringcars (NS vereiste). Standaard: 5",
    )
    parser.add_argument(
        "--tanklocaties",
        default="tanklocaties.json",
        help="JSON bestand met laadstations per busstation. "
             "Gegenereerd door fetch_tanklocaties.py. Standaard: tanklocaties.json",
    )
    parser.add_argument(
        "--inputs",
        default="additional_inputs.xlsx",
        help="Excel bestand met busspecificaties, tarieven, tankinhoud, etc. "
             "Gegenereerd door create_additional_inputs.py. Standaard: additional_inputs.xlsx",
    )
    parser.add_argument(
        "--fuel-constraints",
        action="store_true",
        help="Pas brandstofbeperkingen toe: controleer actieradius per bustype en "
             "splits omlopen als brandstofbereik overschreden wordt zonder tankmogelijkheid.",
    )
    # Version 6: Multi-day cross-day optimization
    parser.add_argument(
        "--multiday",
        action="store_true",
        help="Genereer Output 6: Meerdaagse optimalisatie. Combineert ritten over "
             "opeenvolgende dagen zodat dezelfde bus meerdere dagen kan rijden "
             "(met verschillende chauffeurs).",
    )
    # Version 8: Garage travel (depot → start, end → depot)
    # Version 9: Financial analysis
    parser.add_argument(
        "--financieel",
        action="store_true",
        help="Genereer Output 8 (garage-reistijden) en Output 9 (financieel overzicht) "
             "met omzet, kosten en winst per omloop (bouwt voort op versie 7).",
    )
    # Version 10: Profit maximization
    parser.add_argument(
        "--kosten-optimalisatie",
        action="store_true",
        help="Genereer Output 10: Winstmaximalisatie. Onderzoekt verschillende aantallen "
             "bussen om de meest winstgevende configuratie te vinden. Weegt garagekosten, "
             "ORT, brandstof en dienstlengte tegen elkaar af.",
    )
    args = parser.parse_args()

    if args.output is None:
        args.output = "busomloop_output"

    # Strip .xlsx if user provided it (we'll add suffixes)
    output_base = args.output.replace(".xlsx", "")

    algos = list(ALGORITHMS.keys()) if args.algoritme == "all" else [args.algoritme]

    # Load deadhead matrix if provided
    import json
    deadhead_matrix = None
    deadhead_km_matrix = None  # Distance in km (from Google Maps)
    if args.deadhead:
        dh_path = Path(args.deadhead)
        if not dh_path.exists():
            print(f"WAARSCHUWING: Deadhead bestand '{args.deadhead}' niet gevonden, "
                  "wordt overgeslagen (alleen directe verbindingen)")
        else:
            try:
                with open(dh_path) as f:
                    raw_data = json.load(f)

                # Parse both old format (just minutes) and new format (dict with min/km)
                deadhead_matrix = {}
                deadhead_km_matrix = {}
                has_distance_data = False

                for origin, dests in raw_data.items():
                    deadhead_matrix[origin] = {}
                    deadhead_km_matrix[origin] = {}
                    for dest, val in dests.items():
                        if isinstance(val, (int, float)):
                            # Old format: just the duration_min value
                            deadhead_matrix[origin][dest] = val
                        elif isinstance(val, dict):
                            # New format: {"min": duration_min, "km": distance_km}
                            deadhead_matrix[origin][dest] = val.get("min", val.get("duration_min", 0))
                            if val.get("km") is not None:
                                deadhead_km_matrix[origin][dest] = val["km"]
                                has_distance_data = True

                dh_locs = len(deadhead_matrix)
                if has_distance_data:
                    print(f"  Deadhead matrix met afstandsdata (km) geladen")
                else:
                    deadhead_km_matrix = None  # No distance data available
            except (json.JSONDecodeError, IOError) as e:
                print(f"WAARSCHUWING: Deadhead bestand kon niet geladen worden: {e}")
                print("  Wordt overgeslagen (alleen directe verbindingen)")
    if deadhead_matrix:
        dh_locs = len(deadhead_matrix)
    else:
        dh_locs = 0

    # Load traffic-aware matrix for risk analysis (output 4)
    traffic_data = None
    if args.traffic_matrix:
        tm_path = Path(args.traffic_matrix)
        if not tm_path.exists():
            print(f"WAARSCHUWING: Traffic matrix '{args.traffic_matrix}' niet gevonden, "
                  "risico-analyse wordt overgeslagen")
        else:
            try:
                from google_maps_distances import load_matrix_from_cache_traffic
                traffic_data = load_matrix_from_cache_traffic(str(tm_path))
                n_slots = len(traffic_data.get("time_slots", {}))
                print(f"Traffic matrix geladen: {n_slots} tijdsloten + baseline")

                # Extract distances_km from traffic matrix if available
                # (distances don't change with traffic, so stored separately)
                if traffic_data.get("distances_km"):
                    traffic_km = traffic_data["distances_km"]
                    # Merge with deadhead_km_matrix (traffic distances take precedence)
                    if deadhead_km_matrix is None:
                        deadhead_km_matrix = {}
                    for origin, dests in traffic_km.items():
                        if origin not in deadhead_km_matrix:
                            deadhead_km_matrix[origin] = {}
                        for dest, km in dests.items():
                            if km is not None:
                                deadhead_km_matrix[origin][dest] = km
                    print(f"  Afstanden (km) uit traffic matrix geëxtraheerd")
            except Exception as e:
                print(f"WAARSCHUWING: Traffic matrix kon niet geladen worden: {e}")

    # Load halt capacity limits if provided
    halt_capacity = None
    if args.capaciteit:
        cap_path = Path(args.capaciteit)
        if not cap_path.exists():
            print(f"WAARSCHUWING: Capaciteitsbestand '{args.capaciteit}' niet gevonden, "
                  "capaciteitscheck wordt overgeslagen")
        else:
            try:
                with open(cap_path) as f:
                    halt_capacity = json.load(f)
                print(f"Haltecapaciteiten geladen: {len(halt_capacity)} haltes")
            except (json.JSONDecodeError, IOError) as e:
                print(f"WAARSCHUWING: Capaciteitsbestand kon niet geladen worden: {e}")

    # Load ZE configuration if --ze is enabled
    ze_config = None
    charging_stations = None
    if args.ze:
        print("ZE configuratie laden...")
        ze_config = load_ze_config(args.inputs)
        charging_stations = load_charging_stations(args.tanklocaties)

        # Update avg_speed with calculated values from Google Maps if available
        if deadhead_km_matrix:
            ze_config = update_config_with_gmaps_speed(
                ze_config, deadhead_matrix, deadhead_km_matrix, args.inputs
            )

    # Load fuel configuration if --fuel-constraints is enabled
    fuel_config = None
    fuel_stations = None
    if args.fuel_constraints:
        print("Brandstofconfiguratie laden...")
        fuel_config = load_fuel_config(args.inputs)
        fuel_stations = load_fuel_stations(args.tanklocaties)

        # Update avg_speed with calculated values from Google Maps if available
        if deadhead_km_matrix:
            fuel_config = update_config_with_gmaps_speed(
                fuel_config, deadhead_matrix, deadhead_km_matrix, args.inputs
            )

        # Show range per bus type
        print("  Actieradius per bustype:")
        for bt, range_km in fuel_config["diesel_range_km"].items():
            print(f"    {bt}: {range_km:.0f} km")

    # Load financial configuration if --financieel is enabled
    financial_config = None
    if args.financieel:
        if not FINANCIAL_CALCULATOR_AVAILABLE:
            print("WAARSCHUWING: financial_calculator.py niet gevonden, financieel overzicht overgeslagen")
        else:
            print("Financiele configuratie laden...")
            financial_config = load_financial_config(args.inputs)
            print(f"  Tarieven geladen: {len(financial_config.rates)} bustypes")
            print(f"  Diesel prijs: {financial_config.diesel_price:.2f} EUR/liter")

    # --snel mode: only useful when deadhead is provided + multiple algos
    snel_mode = args.snel and len(algos) > 1 and deadhead_matrix is not None

    n_outputs = 3
    has_traffic = traffic_data and traffic_data.get("time_slots")
    if has_traffic:
        n_outputs = 4
    if deadhead_matrix:
        n_outputs = 5 if has_traffic else 4
    # Version 6: multi-day cross-day optimization (only when --multiday enabled)
    # Works with or without deadhead - without deadhead, bus must end at same station where it starts next day
    if getattr(args, 'multiday', False):
        n_outputs = 6
    # Version 7: fuel/charging constraints (only when --fuel-constraints enabled)
    if args.fuel_constraints and fuel_config and fuel_stations:
        n_outputs = 7
    # Version 8: garage travel (depot → start, end → depot) - when --financieel enabled
    # Version 9: financial analysis (all permutations) - when --financieel enabled
    if args.financieel and financial_config:
        n_outputs = 9
    # Version 10: profit maximization (only when --kosten-optimalisatie enabled)
    # Create a basic fuel_config from financial_config if not already loaded
    if getattr(args, 'kosten_optimalisatie', False) and financial_config:
        if not fuel_config:
            # Create minimal fuel_config from financial_config for profit calculations
            fuel_config = {
                'consumption': financial_config.fuel_consumption,
                'diesel_price': financial_config.diesel_price,
                'hvo_price': financial_config.hvo_price,
            }
        n_outputs = 10
    if snel_mode:
        # outputs 1-(n_outputs-1) greedy only + last output all algos
        n_basis = n_outputs - 1 if deadhead_matrix else n_outputs
        n_files = n_basis * 1 + (1 if deadhead_matrix else 0) * len(algos)
    else:
        n_files = len(algos) * n_outputs

    print(f"Busomloop Optimizer")
    print(f"{'='*60}")
    print(f"Invoer:        {args.input_file}")
    if snel_mode:
        print(f"Algoritme(s):  greedy (outputs 1-{n_basis}), {', '.join(algos)} (output {n_outputs})")
    else:
        print(f"Algoritme(s):  {', '.join(algos)}")
    if deadhead_matrix:
        print(f"Deadhead:      {args.deadhead} ({dh_locs} locaties)")
    else:
        print(f"Deadhead:      niet opgegeven (alleen directe verbindingen)")
    if traffic_data:
        print(f"Traffic:       {args.traffic_matrix} ({n_slots} tijdsloten)")
    if snel_mode and deadhead_matrix and len(algos) > 1:
        print(f"Uitvoer:       {n_files} bestanden ({n_basis} outputs x 1 greedy + 1 output x {len(algos)} algoritmes)")
    else:
        print(f"Uitvoer:       {n_files} bestanden ({n_outputs} outputs x {len(algos)} algoritme{'s' if len(algos) > 1 else ''})")
    print()

    # ===== PARSE =====
    print("Stap 1: Invoer parsen...")
    all_trips, reserves, sheet_names = parse_all_sheets(args.input_file)
    print(f"  {len(sheet_names) - 1} dienstbladen gevonden")
    print(f"  {len(all_trips)} ritten geparsed (inclusief multipliciteit)")
    print(f"  {len(reserves)} reservebus-regels gevonden")

    # Build dynamic station registry from parsed data
    station_reg = build_station_registry(all_trips, reserves)
    print(f"  {len(station_reg)} unieke stations geregistreerd: "
          + ", ".join(sorted(station_reg.values())))

    # Check deadhead coverage: warn about trip endpoint stations missing from deadhead matrix
    if deadhead_matrix:
        dh_keys = set(deadhead_matrix.keys())
        endpoint_locs = set()
        for t in all_trips:
            endpoint_locs.add(normalize_location(t.origin_code))
            endpoint_locs.add(normalize_location(t.dest_code))
        missing = endpoint_locs - dh_keys
        if missing:
            print(f"  WAARSCHUWING: {len(missing)} ritstation(s) ontbreken in deadhead matrix: "
                  + ", ".join(sorted(missing)))
            print(f"  Lege ritten van/naar deze stations zijn niet mogelijk.")

    by_type = {}
    for t in all_trips:
        by_type.setdefault(t.bus_type, []).append(t)
    for bt, trips in sorted(by_type.items()):
        print(f"    {bt}: {len(trips)} ritten")
    print()

    # ===== DETERMINE TURNAROUND TIMES =====
    print("Stap 2: Keertijden bepalen...")

    # Start with defaults for all known bus types
    baseline_turnaround = dict(MIN_TURNAROUND_DEFAULTS)

    # Optionally detect from data (--data-keertijd)
    detected_from_data = {}
    if args.data_keertijd:
        detected_from_data = detect_turnaround_times(all_trips, within_service_only=True)
        # Data-detected values override defaults
        for bt, val in detected_from_data.items():
            baseline_turnaround[bt] = val

    # Apply global --keertijd override (all types)
    if args.keertijd is not None:
        for bt in set(baseline_turnaround.keys()):
            baseline_turnaround[bt] = args.keertijd

    # Apply per-type CLI overrides (highest priority)
    cli_overrides = {
        "Dubbeldekker": args.keer_dd,
        "Touringcar": args.keer_tc,
        "Lagevloerbus": args.keer_lvb,
        "Midi bus": args.keer_midi,
        "Taxibus": args.keer_taxi,
    }
    for bt, val in cli_overrides.items():
        if val is not None:
            baseline_turnaround[bt] = val

    # Ensure any bus types in data but not in defaults are covered
    used_bus_types = set(t.bus_type for t in all_trips)
    for bt in used_bus_types:
        if bt not in baseline_turnaround:
            baseline_turnaround[bt] = MIN_TURNAROUND_FALLBACK

    # Display turnaround times with source labels
    mode_label = "uit data + standaard" if args.data_keertijd else "standaardwaarden"
    print(f"  Keertijden ({mode_label}):")
    for bt, mins in sorted(baseline_turnaround.items()):
        if cli_overrides.get(bt) is not None:
            source = "handmatig"
        elif args.keertijd is not None:
            source = f"--keertijd {args.keertijd}"
        elif bt in detected_from_data:
            source = f"uit data"
        else:
            source = "standaard"
        used = "" if bt in used_bus_types else "  [niet gebruikt]"
        print(f"    {bt:20s} {mins:3d} min  ({source}){used}")

    # Show per-service detail: gap in schedule vs. turnaround we actually use
    svc_turnarounds = detect_turnaround_per_service(all_trips)
    print(f"\n  Keertijden per dienst (dienstregeling vs. optimizer):")

    # Pre-compute column widths for alignment
    items = sorted(svc_turnarounds.items(),
                   key=lambda x: (x[1][1] is None, x[1][1] or 0))
    svc_w = max(len(svc) for svc in svc_turnarounds) + 1
    bt_w = max(len(bt) for bt, *_ in svc_turnarounds.values()) + 2  # for parens
    date_strs = {svc: ", ".join(dates) for svc, (_, _, dates, _, _) in items}
    date_w = max(len(ds) for ds in date_strs.values()) + 2  # for brackets

    for svc, (bt, gap, dates, n_trips, dirs) in items:
        used_val = baseline_turnaround.get(bt, MIN_TURNAROUND_FALLBACK)
        col1 = f"{svc:<{svc_w}s}"
        col2 = f"({bt})"
        col2 = f"{col2:<{bt_w}s}"
        col3 = f"[{date_strs[svc]}]"
        col3 = f"{col3:<{date_w}s}"
        prefix = f"    {col1} {col2} {col3}"
        if gap is None:
            dir_str = " + ".join(dirs)
            reason = f"{n_trips} rit{'ten' if n_trips > 1 else ''}, alleen {dir_str}" if len(dirs) == 1 else f"{n_trips} ritten, geen keerpunt"
            print(f"{prefix}  {'n.v.t.':>7s}       optimizer gebruikt {used_val} min ({reason})")
        elif gap < used_val:
            print(f"{prefix}  {gap:3d} min       optimizer gebruikt {used_val} min (+{used_val - gap} t.o.v. dienstregeling)")
        elif gap > used_val:
            print(f"{prefix}  {gap:3d} min       optimizer gebruikt {used_val} min (marge {gap - used_val} min)")
        else:
            print(f"{prefix}  {gap:3d} min       optimizer gebruikt {used_val} min (gelijk)")
    print()

    total_reserves = sum(r.count for r in reserves)

    # Create reserve phantom trips (used by outputs 3+)
    reserve_trip_list = create_reserve_trips(reserves, all_trips)
    trips_with_reserves = all_trips + reserve_trip_list
    n_phantom = len(reserve_trip_list)
    n_unmatched = total_reserves - n_phantom

    print(f"  Reservebussen in input: {total_reserves}")
    if n_unmatched > 0:
        # Show which stations could not be matched
        trip_dates = sorted(set(t.date_str for t in all_trips))
        unmatched_stations = []
        for rb in reserves:
            date_str = match_reserve_day(rb.day, trip_dates)
            res_loc = normalize_reserve_station(rb.station)
            has_trips = any(
                t.date_str == date_str and
                (normalize_location(t.origin_code) == res_loc or
                 normalize_location(t.dest_code) == res_loc)
                for t in all_trips
            )
            if not has_trips:
                unmatched_stations.append(f"{rb.station} ({rb.day}, {rb.count}x)")
        print(f"    {n_phantom} inplanbaar (station komt voor in ritten)")
        print(f"    {n_unmatched} niet inplanbaar (geen ritten via dat station):")
        for s in unmatched_stations:
            print(f"      - {s}")
        print(f"    Deze {n_unmatched} tellen altijd mee als extra losse reservebussen.")
    else:
        print(f"    Alle {n_phantom} reserves inplanbaar als phantom trips.")
    print()

    trip_dates = sorted(set(t.date_str for t in all_trips))

    # Determine which algorithms to run
    if args.algoritme == "all":
        algo_keys = list(ALGORITHMS.keys())
    else:
        algo_keys = [args.algoritme]

    # ===================================================================
    # Per-algorithm results collector for comparison table
    # ===================================================================
    # results[algo_key] = {1: {...}, 2: {...}, 3: {...}, 4: {...}}
    all_results = {}

    for algo_idx, algo_key in enumerate(algo_keys):
        algo_name = ALGORITHMS[algo_key][0]
        algo_short = {"greedy": "greedy", "mincost": "mincost"}[algo_key]

        # --snel: skip non-greedy for outputs 1-4, only run deadhead output
        snel_skip_basis = snel_mode and algo_key != "greedy"

        if len(algo_keys) > 1:
            print(f"{'='*60}")
            if snel_skip_basis:
                print(f"Algoritme {algo_idx+1}/{len(algo_keys)}: {algo_name} (alleen deadhead output)")
            else:
                print(f"Algoritme {algo_idx+1}/{len(algo_keys)}: {algo_name}")
            print(f"{'='*60}")
        else:
            print(f"Algoritme: {algo_name}")

        algo_results = {}

        # --snel: skip outputs 1-4 for non-greedy algorithms
        if not snel_skip_basis:
            # ---------------------------------------------------------------
            # OUTPUT 1: Per dienst, geen reserves
            # ---------------------------------------------------------------
            print(f"  Output 1 - Per dienst, geen reserves...")
            rot1 = optimize_rotations(all_trips, baseline_turnaround,
                                      algorithm=algo_key, per_service=True)

            n1 = len(rot1)
            n1_idle = sum(r.total_idle_minutes for r in rot1)
            print(f"    {n1} bussen met ritten + {total_reserves} reserve = {n1 + total_reserves} totaal")
            print(f"    Totale wachttijd: {n1_idle} min ({n1_idle / 60:.1f} uur)")

            file1 = f"{output_base}_{algo_short}_1_per_dienst.xlsx"
            print(f"    Schrijven {file1}...", end=" ", flush=True)
            generate_output(rot1, all_trips, reserves, file1, baseline_turnaround, algo_key,
                            output_mode=1, version=1)
            print("OK")

            algo_results[1] = {"rotations": rot1, "buses_met_ritten": n1,
                               "reserve_bussen": total_reserves, "idle_min": n1_idle,
                               "file": file1}

            # ---------------------------------------------------------------
            # OUTPUT 2: Per dienst + optimale idle reserve matching
            # ---------------------------------------------------------------
            print(f"  Output 2 - Per dienst + optimale reserve matching...")
            file2 = f"{output_base}_{algo_short}_2_per_dienst_reservematch.xlsx"
            print(f"    Schrijven {file2}...", end=" ", flush=True)
            generate_output(rot1, all_trips, reserves, file2, baseline_turnaround, algo_key,
                            output_mode=2, version=2)
            print("OK")

            idle_cov = optimize_reserve_idle_matching(rot1, reserves, trip_dates)
            idle_covered = sum(min(c["covered"], c["required"]) for c in idle_cov)
            n2_reserve_bussen = total_reserves - idle_covered
            print(f"    {n1} bussen met ritten + {n2_reserve_bussen} reserve = {n1 + n2_reserve_bussen} totaal")
            print(f"    ({idle_covered}/{total_reserves} reserves gedekt door bestaande bussen)")
            print(f"    Totale wachttijd: {n1_idle} min ({n1_idle / 60:.1f} uur)")

            algo_results[2] = {"rotations": rot1, "buses_met_ritten": n1,
                               "reserve_bussen": n2_reserve_bussen, "idle_min": n1_idle,
                               "file": file2}

            # ---------------------------------------------------------------
            # OUTPUT 3: Gecombineerd + reserves ingepland + sensitiviteit
            # ---------------------------------------------------------------
            print(f"  Output 3 - Gecombineerd + reserves + sensitiviteit...")
            rot3 = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                      algorithm=algo_key)

            n3_with_trips = len([r for r in rot3 if r.real_trips])
            n3_reserve_only = len([r for r in rot3 if not r.real_trips and r.reserve_trip_list])
            n3_res_planned = sum(len(r.reserve_trip_list) for r in rot3)
            n3_extra = max(0, total_reserves - n3_res_planned)
            n3_reserve_bussen = n3_reserve_only + n3_extra
            n3_idle = sum(r.total_idle_minutes for r in rot3)
            print(f"    {n3_with_trips} bussen met ritten + {n3_reserve_bussen} reserve = {n3_with_trips + n3_reserve_bussen} totaal")
            print(f"    Totale wachttijd: {n3_idle} min ({n3_idle / 60:.1f} uur)")

            file3 = f"{output_base}_{algo_short}_3_gecombineerd_met_reserve.xlsx"
            print(f"    Schrijven {file3}...", end=" ", flush=True)
            generate_output(rot3, trips_with_reserves, reserves, file3, baseline_turnaround, algo_key,
                            include_sensitivity=True, output_mode=4, version=3)
            print("OK")

            algo_results[3] = {"rotations": rot3, "buses_met_ritten": n3_with_trips,
                               "reserve_bussen": n3_reserve_bussen, "idle_min": n3_idle,
                               "file": file3}

            # ---------------------------------------------------------------
            # Compute risk overrides once (shared by output 4 and 5)
            # ---------------------------------------------------------------
            risk_report = None
            trip_overrides = None
            if traffic_data and traffic_data.get("time_slots"):
                if algo_idx == 0:  # Only print once
                    print(f"  Risico-analyse: keertijden berekenen op basis van verkeerssituatie...")
                trip_overrides, risk_report = compute_trip_turnaround_overrides(
                    trips_with_reserves, traffic_data, baseline_turnaround)
                n_overrides = len(trip_overrides)
                n_high = sum(1 for r in risk_report if r["risk"] == "HOOG")
                n_medium = sum(1 for r in risk_report if r["risk"] == "MATIG")
                if algo_idx == 0:
                    print(f"    {n_overrides} ritten met verhoogde keertijd, "
                          f"{n_high} hoog risico, {n_medium} matig risico")

            # ---------------------------------------------------------------
            # OUTPUT 4: Gecombineerd + reserves + risico-keertijden (geen deadhead)
            # Only generated when traffic data is available
            # ---------------------------------------------------------------
            if trip_overrides is not None:
                print(f"  Output 4 - Gecombineerd + reserves + risico-keertijden...")
                rot4 = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                          algorithm=algo_key,
                                          trip_turnaround_overrides=trip_overrides)

                n4_with_trips = len([r for r in rot4 if r.real_trips])
                n4_reserve_only = len([r for r in rot4 if not r.real_trips and r.reserve_trip_list])
                n4_res_planned = sum(len(r.reserve_trip_list) for r in rot4)
                n4_extra = max(0, total_reserves - n4_res_planned)
                n4_reserve_bussen = n4_reserve_only + n4_extra
                n4_idle = sum(r.total_idle_minutes for r in rot4)
                print(f"    {n4_with_trips} bussen met ritten + {n4_reserve_bussen} reserve = {n4_with_trips + n4_reserve_bussen} totaal")
                print(f"    Totale wachttijd: {n4_idle} min ({n4_idle / 60:.1f} uur)")

                file4 = f"{output_base}_{algo_short}_4_gecombineerd_risico.xlsx"
                print(f"    Schrijven {file4}...", end=" ", flush=True)
                generate_output(rot4, trips_with_reserves, reserves, file4, baseline_turnaround, algo_key,
                                include_sensitivity=True, output_mode=4,
                                risk_report=risk_report, version=4)
                print("OK")

                algo_results[4] = {"rotations": rot4, "buses_met_ritten": n4_with_trips,
                                   "reserve_bussen": n4_reserve_bussen, "idle_min": n4_idle,
                                   "file": file4}

        # ---------------------------------------------------------------
        # OUTPUT 5 (or 4): Gecombineerd + reserves + deadhead + risico
        # Only generated when --deadhead is provided
        # Always runs (not skipped by --snel)
        # ---------------------------------------------------------------
        if deadhead_matrix:
            # When --snel skipped basis, we still need risk overrides for output 5
            if snel_skip_basis:
                risk_report = None
                trip_overrides = None
                if traffic_data and traffic_data.get("time_slots"):
                    trip_overrides, risk_report = compute_trip_turnaround_overrides(
                        trips_with_reserves, traffic_data, baseline_turnaround)

            out_num = 5 if has_traffic else 4
            print(f"  Output {out_num} - Gecombineerd + reserves + deadhead"
                  f"{' + risico-keertijden' if trip_overrides else ''}...")
            rot5 = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                      algorithm=algo_key,
                                      deadhead_matrix=deadhead_matrix,
                                      trip_turnaround_overrides=trip_overrides)

            n5_with_trips = len([r for r in rot5 if r.real_trips])
            n5_reserve_only = len([r for r in rot5 if not r.real_trips and r.reserve_trip_list])
            n5_res_planned = sum(len(r.reserve_trip_list) for r in rot5)
            n5_extra = max(0, total_reserves - n5_res_planned)
            n5_reserve_bussen = n5_reserve_only + n5_extra
            n5_idle = sum(r.total_idle_minutes for r in rot5)
            print(f"    {n5_with_trips} bussen met ritten + {n5_reserve_bussen} reserve = {n5_with_trips + n5_reserve_bussen} totaal")
            print(f"    Totale wachttijd: {n5_idle} min ({n5_idle / 60:.1f} uur)")

            file5 = f"{output_base}_{algo_short}_{out_num}_gecombineerd_deadhead.xlsx"
            print(f"    Schrijven {file5}...", end=" ", flush=True)
            generate_output(rot5, trips_with_reserves, reserves, file5, baseline_turnaround, algo_key,
                            include_sensitivity=True, output_mode=4,
                            risk_report=risk_report, deadhead_matrix=deadhead_matrix, version=5)
            print("OK")

            algo_results[out_num] = {"rotations": rot5, "buses_met_ritten": n5_with_trips,
                                     "reserve_bussen": n5_reserve_bussen, "idle_min": n5_idle,
                                     "file": file5}

        # ---------------------------------------------------------------
        # OUTPUT 6: Meerdaagse optimalisatie (multi-day cross-day)
        # Only generated when --multiday is enabled
        # Combines trips across consecutive days so same bus can work multiple days
        # Works with or without deadhead - without deadhead, bus must end at same station
        # ---------------------------------------------------------------
        if getattr(args, 'multiday', False):
            print(f"  Output 6 - Meerdaagse optimalisatie...")
            print(f"    Combineert ritten over opeenvolgende dagen...")
            if not deadhead_matrix:
                print(f"    (Zonder deadhead: bus moet eindigen waar volgende dag begint)")

            # Group trips by bus type only (not by date) for cross-day optimization
            multiday_groups, _ = _group_trips_multiday(all_trips, baseline_turnaround)

            rot6 = []
            rotation_counter = 0

            # Get trip_overrides if available
            trip_overrides_v6 = trip_overrides if 'trip_overrides' in dir() else None

            for bus_type, group_trips in multiday_groups.items():
                if not group_trips:
                    continue

                # Use multiday-aware algorithm
                if algo_key == "greedy":
                    chains = _optimize_greedy(group_trips, baseline_turnaround,
                                              service_constraint=False,
                                              deadhead_matrix=deadhead_matrix,
                                              trip_turnaround_overrides=trip_overrides_v6,
                                              multiday=True)
                else:
                    chains = _optimize_mincost(group_trips, baseline_turnaround,
                                               service_constraint=False,
                                               deadhead_matrix=deadhead_matrix,
                                               trip_turnaround_overrides=trip_overrides_v6,
                                               multiday=True)

                # Convert chains to rotations
                for chain in chains:
                    rotation_counter += 1
                    chain_trips = [group_trips[i] for i in chain]
                    # Get date range for bus_id
                    dates = sorted(set(t.date_str.split()[0] for t in chain_trips if t.date_str))
                    date_label = dates[0] if len(dates) == 1 else f"{dates[0]}-{dates[-1]}"
                    bus_id = f"MD-{bus_type[:2].upper()}-{date_label}-{rotation_counter:03d}"

                    rot6.append(BusRotation(
                        bus_id=bus_id,
                        bus_type=bus_type,
                        date_str=chain_trips[0].date_str if chain_trips else "",
                        trips=chain_trips,
                    ))

            # Calculate idle time for multiday rotations
            for r in rot6:
                if len(r.trips) > 1:
                    total_idle = 0
                    for i in range(len(r.trips) - 1):
                        t1, t2 = r.trips[i], r.trips[i+1]
                        day_offset = _parse_date_to_ordinal(t2.date_str) - _parse_date_to_ordinal(t1.date_str)
                        gap = (day_offset * 1440) + t2.departure - t1.arrival
                        # Subtract deadhead time if applicable
                        if deadhead_matrix:
                            dh = deadhead_matrix.get(normalize_location(t1.dest_code), {}).get(
                                normalize_location(t2.origin_code), 0)
                            gap = max(0, gap - dh)
                        total_idle += max(0, gap)
                    r.total_idle_minutes = int(total_idle)

            n6_with_trips = len([r for r in rot6 if r.real_trips])
            n6_idle = sum(r.total_idle_minutes for r in rot6)
            n6_trips_per_bus = sum(len(r.trips) for r in rot6) / max(1, len(rot6))
            n6_multiday = len([r for r in rot6 if len(set(t.date_str for t in r.trips)) > 1])

            print(f"    {len(rot6)} bussen totaal ({n6_multiday} meerdaags)")
            print(f"    Gemiddeld {n6_trips_per_bus:.1f} ritten per bus")
            print(f"    Totale wachttijd: {n6_idle} min ({n6_idle / 60:.1f} uur)")

            file6 = f"{output_base}_{algo_short}_6_meerdaags.xlsx"
            print(f"    Schrijven {file6}...", end=" ", flush=True)
            risk_report_v6 = risk_report if 'risk_report' in dir() else None
            generate_output(rot6, trips_with_reserves, reserves, file6, baseline_turnaround, algo_key,
                            include_sensitivity=True, output_mode=4,
                            risk_report=risk_report_v6, deadhead_matrix=deadhead_matrix, version=6)
            print("OK")

            algo_results[6] = {"rotations": rot6, "buses_met_ritten": n6_with_trips,
                               "reserve_bussen": 0, "idle_min": n6_idle,
                               "file": file6, "multiday_buses": n6_multiday}

        # ---------------------------------------------------------------
        # OUTPUT 7: Brandstof/laad strategie (fuel constraints + ZE)
        # Only generated when --fuel-constraints is enabled
        # ---------------------------------------------------------------
        if args.fuel_constraints and fuel_config and fuel_stations:
            print(f"  Output 7 - Brandstof/laad strategie...")

            # Apply fuel constraints - this may split rotations
            # Use best available: multiday (rot6) > deadhead (rot5) > combined (rot3)
            if 6 in algo_results:
                base_rot = algo_results[6]['rotations']
            elif 5 in algo_results:
                base_rot = algo_results[5]['rotations']
            else:
                base_rot = algo_results[3]['rotations']

            rot7_orig_count = len(base_rot)
            rot7, fuel_results_7, fuel_splits_7 = apply_fuel_constraints(
                base_rot, fuel_config, fuel_stations, deadhead_matrix, deadhead_km_matrix,
                turnaround_map=baseline_turnaround, algorithm=algo_key
            )

            if fuel_splits_7 > 0:
                print(f"    Brandstofcontrole: {fuel_splits_7} omlopen gesplitst "
                      f"({rot7_orig_count} -> {len(rot7)} bussen)")
            else:
                print(f"    Brandstofcontrole: geen splitsingen nodig ({len(rot7)} bussen)")

            n7_with_trips = len([r for r in rot7 if r.real_trips])
            n7_reserve_only = len([r for r in rot7 if not r.real_trips and r.reserve_trip_list])
            n7_res_planned = sum(len(r.reserve_trip_list) for r in rot7)
            n7_extra = max(0, total_reserves - n7_res_planned)
            n7_reserve_bussen = n7_reserve_only + n7_extra
            n7_idle = sum(r.total_idle_minutes for r in rot7)
            print(f"    {n7_with_trips} bussen met ritten + {n7_reserve_bussen} reserve = {n7_with_trips + n7_reserve_bussen} totaal")
            print(f"    Totale wachttijd: {n7_idle} min ({n7_idle / 60:.1f} uur)")

            file7 = f"{output_base}_{algo_short}_7_brandstof_strategie.xlsx"
            print(f"    Schrijven {file7}...", end=" ", flush=True)

            # Generate full output with all schedule tabs
            risk_report_v7 = risk_report if 'risk_report' in dir() else None
            generate_output(rot7, trips_with_reserves, reserves, file7, baseline_turnaround, algo_key,
                            include_sensitivity=True, output_mode=4,
                            risk_report=risk_report_v7, deadhead_matrix=deadhead_matrix, version=7)
            print("OK")

            # Add fuel analysis sheet to version 7 output
            print(f"    Toevoegen Brandstof Analyse sheet...", end=" ", flush=True)
            wb7 = openpyxl.load_workbook(file7)
            write_fuel_analysis_sheet(wb7, fuel_results_7, fuel_stations, fuel_config)
            wb7.save(file7)
            print("OK")

            # Add ZE analysis if enabled
            if args.ze and ze_config:
                ze_stats = generate_ze_output(
                    rot7, file7, ze_config, charging_stations, args.min_ze,
                    append_to_existing=True
                )
                print(f"    ZE analyse: {ze_stats['ze_feasible']}/{ze_stats['total_touringcar']} "
                      f"touringcars ZE-geschikt, {ze_stats['assigned_count']} toegewezen")

            algo_results[7] = {"rotations": rot7, "buses_met_ritten": n7_with_trips,
                               "reserve_bussen": n7_reserve_bussen, "idle_min": n7_idle,
                               "file": file7}

        # ---------------------------------------------------------------
        # OUTPUT 8: Garage reistijden (depot → start, end → depot)
        # Shows explicit garage travel for each rotation, using financial config
        # ---------------------------------------------------------------
        if args.financieel and financial_config:
            print(f"  Output 8 - Garage reistijden (remise → start, eind → remise)...")

            # Use best available rotations
            if 7 in algo_results:
                rot8_base = algo_results[7]['rotations']
            elif 5 in algo_results:
                rot8_base = algo_results[5]['rotations']
            else:
                rot8_base = algo_results[3]['rotations']

            # Calculate garage travel info
            garage_tijd_enkel = financial_config.garage_reistijd_enkel_min
            garage_km_enkel = financial_config.garage_afstand_enkel_km
            garage_tijd_totaal = garage_tijd_enkel * 2  # Round trip
            garage_km_totaal = garage_km_enkel * 2

            n8_buses = len(rot8_base)
            n8_with_trips = len([r for r in rot8_base if r.real_trips])
            n8_garage_km = n8_buses * garage_km_totaal
            n8_garage_min = n8_buses * garage_tijd_totaal

            print(f"    {n8_buses} bussen")
            print(f"    Garage reistijd per bus: {garage_tijd_totaal} min (2x {garage_tijd_enkel} min)")
            print(f"    Garage afstand per bus: {garage_km_totaal:.1f} km (2x {garage_km_enkel:.1f} km)")
            print(f"    Totale garage km: {n8_garage_km:.1f} km")
            print(f"    Totale garage tijd: {n8_garage_min} min ({n8_garage_min/60:.1f} uur)")

            file8 = f"{output_base}_{algo_short}_8_garage_reistijden.xlsx"
            print(f"    Schrijven {file8}...", end=" ", flush=True)

            # Generate output with garage travel info
            generate_output(rot8_base, trips_with_reserves, reserves, file8, baseline_turnaround, algo_key,
                           include_sensitivity=True, output_mode=4,
                           risk_report=risk_report if 'risk_report' in dir() else None,
                           deadhead_matrix=deadhead_matrix, version=8)

            # Add garage travel sheet
            wb8 = openpyxl.load_workbook(file8)
            ws_garage = wb8.create_sheet("Garage Reistijden")

            # Header
            ws_garage.cell(row=1, column=1, value="Garage Reistijden per Bus")
            ws_garage.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws_garage.cell(row=2, column=1, value=f"Remise → Startstation: {garage_tijd_enkel} min, {garage_km_enkel:.1f} km")
            ws_garage.cell(row=3, column=1, value=f"Eindstation → Remise: {garage_tijd_enkel} min, {garage_km_enkel:.1f} km")
            ws_garage.cell(row=4, column=1, value=f"Totaal per bus: {garage_tijd_totaal} min, {garage_km_totaal:.1f} km")

            # Table headers
            headers = ["Bus ID", "Bustype", "Datum", "Eerste station", "Start rit",
                      "Vertrek remise", "Laatste station", "Einde rit", "Aankomst remise",
                      "Totale diensttijd", "Garage km"]
            for col, h in enumerate(headers, 1):
                ws_garage.cell(row=6, column=col, value=h)
                ws_garage.cell(row=6, column=col).font = Font(bold=True)

            # Data rows
            row = 7
            for rot in rot8_base:
                if not rot.trips:
                    continue
                first_trip = rot.trips[0]
                last_trip = rot.trips[-1]

                # Calculate times
                trip_start = first_trip.departure  # Minutes from midnight
                trip_end = last_trip.arrival
                remise_vertrek = trip_start - garage_tijd_enkel
                remise_aankomst = trip_end + garage_tijd_enkel
                total_dienst = remise_aankomst - remise_vertrek

                ws_garage.cell(row=row, column=1, value=rot.bus_id)
                ws_garage.cell(row=row, column=2, value=rot.bus_type)
                ws_garage.cell(row=row, column=3, value=rot.date_str)
                ws_garage.cell(row=row, column=4, value=first_trip.origin_name)
                ws_garage.cell(row=row, column=5, value=f"{trip_start//60:02d}:{trip_start%60:02d}")
                ws_garage.cell(row=row, column=6, value=f"{remise_vertrek//60:02d}:{remise_vertrek%60:02d}")
                ws_garage.cell(row=row, column=7, value=last_trip.dest_name)
                ws_garage.cell(row=row, column=8, value=f"{trip_end//60:02d}:{trip_end%60:02d}")
                ws_garage.cell(row=row, column=9, value=f"{remise_aankomst//60:02d}:{remise_aankomst%60:02d}")
                ws_garage.cell(row=row, column=10, value=f"{total_dienst//60}:{total_dienst%60:02d}")
                ws_garage.cell(row=row, column=11, value=garage_km_totaal)
                row += 1

            # Summary row
            row += 1
            ws_garage.cell(row=row, column=1, value="TOTAAL")
            ws_garage.cell(row=row, column=1).font = Font(bold=True)
            ws_garage.cell(row=row, column=11, value=n8_garage_km)
            ws_garage.cell(row=row, column=11).font = Font(bold=True)

            # Auto-fit column widths
            for col in range(1, len(headers) + 1):
                ws_garage.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            wb8.save(file8)
            print("OK")

            algo_results[8] = {"rotations": rot8_base, "buses_met_ritten": n8_with_trips,
                               "reserve_bussen": 0, "idle_min": 0,
                               "garage_km": n8_garage_km, "garage_min": n8_garage_min,
                               "file": file8}

        # ---------------------------------------------------------------
        # OUTPUT 9: Financieel overzicht (financial analysis)
        # Generates financials for ALL version permutations with dimensions:
        #   - Deadhead: no/yes
        #   - Multidag: no/yes
        #   - Risico: no/yes (if traffic data available)
        #   - Brandstof: no/yes (if fuel constraints available)
        # ---------------------------------------------------------------
        if args.financieel and financial_config:
            print(f"  Output 9 - Financieel overzicht (alle permutaties)...")

            # Dictionary to store all permutation financials for comparison
            all_financials = {}
            all_rotations = {}
            trip_overrides_v9 = trip_overrides if 'trip_overrides' in dir() else None
            has_fuel = args.fuel_constraints and fuel_config and fuel_stations
            has_risk = trip_overrides_v9 is not None

            # Helper to process a permutation (generate financials, write file)
            def process_permutation(rot, perm_key, perm_label, file_suffix,
                                   dh_matrix=None, apply_fuel=False, fuel_results_in=None):
                """Process a single permutation: apply fuel if needed, calculate financials, write file."""
                fuel_results = fuel_results_in
                if apply_fuel and has_fuel:
                    rot_before = len(rot)
                    rot, fuel_results, _ = apply_fuel_constraints(
                        rot, fuel_config, fuel_stations, dh_matrix, deadhead_km_matrix,
                        turnaround_map=baseline_turnaround, algorithm=algo_key
                    )
                    if len(rot) > rot_before:
                        print(f"      +{len(rot) - rot_before} bussen door tankbeperkingen")

                financials = calculate_total_financials(rot, financial_config, fuel_type="diesel")
                financials['totals']['bus_count'] = len(rot)
                all_financials[perm_key] = financials
                all_rotations[perm_key] = rot
                print(f"      {len(rot)} bussen, winst: {financials['totals']['total_net_profit']:,.0f} EUR")

                # Write output file
                file_path = f"{output_base}_{algo_short}_{file_suffix}.xlsx"
                generate_output(rot, trips_with_reserves, reserves, file_path, baseline_turnaround, algo_key,
                                include_sensitivity=True, output_mode=4,
                                risk_report=risk_report if 'risk_report' in dir() else None,
                                deadhead_matrix=dh_matrix, version=9)
                wb = openpyxl.load_workbook(file_path)
                if fuel_results:
                    write_fuel_analysis_sheet(wb, fuel_results, fuel_stations, fuel_config)
                write_financial_sheet(wb, financials)
                write_cost_calculation_sheet(wb, financial_config)
                wb.save(file_path)
                return rot, financials

            # =====================================================================
            # BASIS PERMUTATIONS (no deadhead, no multiday)
            # =====================================================================
            print(f"    Basis permutaties...")

            # 9a: Basis (no fuel)
            print(f"      basis...")
            rot_basis = algo_results[3]['rotations']
            process_permutation(rot_basis, "basis", "Basis", "9a_financieel_basis",
                               dh_matrix=None, apply_fuel=False)

            # 8a_brandstof: Basis with fuel constraints
            if has_fuel:
                print(f"      basis_brandstof...")
                rot_basis_copy = algo_results[3]['rotations'][:]  # Fresh copy
                process_permutation(rot_basis_copy, "basis_brandstof", "Basis + Brandstof",
                                   "9a_financieel_basis_brandstof", dh_matrix=None, apply_fuel=True)

            # 8a_risico: Basis with risk (no fuel)
            if has_risk:
                print(f"      basis_risico...")
                rot_basis_risico = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                                       algorithm=algo_key,
                                                       trip_turnaround_overrides=trip_overrides_v9)
                process_permutation(rot_basis_risico, "basis_risico", "Basis + Risico",
                                   "9a_financieel_basis_risico", dh_matrix=None, apply_fuel=False)

                # 8a_risico_brandstof: Basis with risk AND fuel
                if has_fuel:
                    print(f"      basis_risico_brandstof...")
                    rot_basis_risico_copy = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                                                algorithm=algo_key,
                                                                trip_turnaround_overrides=trip_overrides_v9)
                    process_permutation(rot_basis_risico_copy, "basis_risico_brandstof",
                                       "Basis + Risico + Brandstof",
                                       "9a_financieel_basis_risico_brandstof", dh_matrix=None, apply_fuel=True)

            # =====================================================================
            # DEADHEAD PERMUTATIONS (with deadhead, no multiday)
            # =====================================================================
            if deadhead_matrix:
                print(f"    Deadhead permutaties...")

                # 9b: Deadhead (no risk, no fuel)
                print(f"      deadhead...")
                rot_deadhead = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                                   algorithm=algo_key,
                                                   deadhead_matrix=deadhead_matrix,
                                                   trip_turnaround_overrides=None)
                process_permutation(rot_deadhead, "deadhead", "Deadhead",
                                   "9b_financieel_deadhead", dh_matrix=deadhead_matrix, apply_fuel=False)

                # 8b_brandstof: Deadhead with fuel
                if has_fuel:
                    print(f"      deadhead_brandstof...")
                    rot_deadhead_copy = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                                            algorithm=algo_key,
                                                            deadhead_matrix=deadhead_matrix,
                                                            trip_turnaround_overrides=None)
                    process_permutation(rot_deadhead_copy, "deadhead_brandstof", "Deadhead + Brandstof",
                                       "9b_financieel_deadhead_brandstof", dh_matrix=deadhead_matrix, apply_fuel=True)

                # 8b_risico: Deadhead with risk (no fuel)
                if has_risk:
                    print(f"      deadhead_risico...")
                    rot_dh_risico = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                                        algorithm=algo_key,
                                                        deadhead_matrix=deadhead_matrix,
                                                        trip_turnaround_overrides=trip_overrides_v9)
                    process_permutation(rot_dh_risico, "deadhead_risico", "Deadhead + Risico",
                                       "9b_financieel_deadhead_risico", dh_matrix=deadhead_matrix, apply_fuel=False)

                    # 8b_risico_brandstof: Deadhead with risk AND fuel
                    if has_fuel:
                        print(f"      deadhead_risico_brandstof...")
                        rot_dh_risico_copy = optimize_rotations(trips_with_reserves, baseline_turnaround,
                                                                 algorithm=algo_key,
                                                                 deadhead_matrix=deadhead_matrix,
                                                                 trip_turnaround_overrides=trip_overrides_v9)
                        process_permutation(rot_dh_risico_copy, "deadhead_risico_brandstof",
                                           "Deadhead + Risico + Brandstof",
                                           "9b_financieel_deadhead_risico_brandstof",
                                           dh_matrix=deadhead_matrix, apply_fuel=True)

            # =====================================================================
            # MULTIDAG PERMUTATIONS (no deadhead, with multiday)
            # =====================================================================
            print(f"    Multidag permutaties...")

            # Helper to generate multiday rotations
            def generate_multiday_rotations(use_deadhead, use_risk, prefix):
                """Generate multiday rotations with given options."""
                multiday_groups, _ = _group_trips_multiday(all_trips, baseline_turnaround)
                rotations = []
                counter = 0
                dh = deadhead_matrix if use_deadhead else None
                overrides = trip_overrides_v9 if use_risk else None
                for bus_type, group_trips in multiday_groups.items():
                    if not group_trips:
                        continue
                    if algo_key == "greedy":
                        chains = _optimize_greedy(group_trips, baseline_turnaround,
                                                  service_constraint=False,
                                                  deadhead_matrix=dh,
                                                  trip_turnaround_overrides=overrides,
                                                  multiday=True)
                    else:
                        chains = _optimize_mincost(group_trips, baseline_turnaround,
                                                   service_constraint=False,
                                                   deadhead_matrix=dh,
                                                   trip_turnaround_overrides=overrides,
                                                   multiday=True)
                    for chain in chains:
                        counter += 1
                        chain_trips = [group_trips[i] for i in chain]
                        dates = sorted(set(t.date_str.split()[0] for t in chain_trips if t.date_str))
                        date_label = dates[0] if len(dates) == 1 else f"{dates[0]}-{dates[-1]}"
                        bus_id = f"{prefix}-{bus_type[:2].upper()}-{date_label}-{counter:03d}"
                        rotations.append(BusRotation(
                            bus_id=bus_id, bus_type=bus_type,
                            date_str=chain_trips[0].date_str if chain_trips else "",
                            trips=chain_trips,
                        ))
                return rotations

            # 9c: Multidag (no deadhead, no risk, no fuel)
            print(f"      multidag...")
            rot_multidag = generate_multiday_rotations(use_deadhead=False, use_risk=False, prefix="8c")
            process_permutation(rot_multidag, "multidag", "Multidag",
                               "9c_financieel_multidag", dh_matrix=None, apply_fuel=False)

            # 8c_brandstof: Multidag with fuel
            if has_fuel:
                print(f"      multidag_brandstof...")
                rot_multidag_f = generate_multiday_rotations(use_deadhead=False, use_risk=False, prefix="8cf")
                process_permutation(rot_multidag_f, "multidag_brandstof", "Multidag + Brandstof",
                                   "9c_financieel_multidag_brandstof", dh_matrix=None, apply_fuel=True)

            # 8c_risico: Multidag with risk (no fuel)
            if has_risk:
                print(f"      multidag_risico...")
                rot_multidag_r = generate_multiday_rotations(use_deadhead=False, use_risk=True, prefix="8cr")
                process_permutation(rot_multidag_r, "multidag_risico", "Multidag + Risico",
                                   "9c_financieel_multidag_risico", dh_matrix=None, apply_fuel=False)

                # 8c_risico_brandstof: Multidag with risk AND fuel
                if has_fuel:
                    print(f"      multidag_risico_brandstof...")
                    rot_multidag_rf = generate_multiday_rotations(use_deadhead=False, use_risk=True, prefix="8crf")
                    process_permutation(rot_multidag_rf, "multidag_risico_brandstof",
                                       "Multidag + Risico + Brandstof",
                                       "9c_financieel_multidag_risico_brandstof", dh_matrix=None, apply_fuel=True)

            # =====================================================================
            # DEADHEAD + MULTIDAG PERMUTATIONS
            # =====================================================================
            if deadhead_matrix:
                print(f"    Deadhead + Multidag permutaties...")

                # 9d: Deadhead + Multidag (no risk, no fuel)
                print(f"      deadhead_multidag...")
                rot_dh_md = generate_multiday_rotations(use_deadhead=True, use_risk=False, prefix="8d")
                process_permutation(rot_dh_md, "deadhead_multidag", "Deadhead + Multidag",
                                   "9d_financieel_deadhead_multidag", dh_matrix=deadhead_matrix, apply_fuel=False)

                # 8d_brandstof: Deadhead + Multidag with fuel
                if has_fuel:
                    print(f"      deadhead_multidag_brandstof...")
                    rot_dh_md_f = generate_multiday_rotations(use_deadhead=True, use_risk=False, prefix="8df")
                    process_permutation(rot_dh_md_f, "deadhead_multidag_brandstof",
                                       "Deadhead + Multidag + Brandstof",
                                       "9d_financieel_deadhead_multidag_brandstof",
                                       dh_matrix=deadhead_matrix, apply_fuel=True)

                # 8d_risico: Deadhead + Multidag with risk (no fuel)
                if has_risk:
                    print(f"      deadhead_multidag_risico...")
                    rot_dh_md_r = generate_multiday_rotations(use_deadhead=True, use_risk=True, prefix="8dr")
                    process_permutation(rot_dh_md_r, "deadhead_multidag_risico",
                                       "Deadhead + Multidag + Risico",
                                       "9d_financieel_deadhead_multidag_risico",
                                       dh_matrix=deadhead_matrix, apply_fuel=False)

                    # 8d_risico_brandstof: All features combined
                    if has_fuel:
                        print(f"      deadhead_multidag_risico_brandstof...")
                        rot_dh_md_rf = generate_multiday_rotations(use_deadhead=True, use_risk=True, prefix="8drf")
                        process_permutation(rot_dh_md_rf, "deadhead_multidag_risico_brandstof",
                                           "Deadhead + Multidag + Risico + Brandstof",
                                           "9d_financieel_deadhead_multidag_risico_brandstof",
                                           dh_matrix=deadhead_matrix, apply_fuel=True)

            # ---- Create comparison file ----
            print(f"    Vergelijkingsoverzicht...")
            file8_compare = f"{output_base}_{algo_short}_9_financieel_vergelijking.xlsx"
            wb_compare = openpyxl.Workbook()
            wb_compare.remove(wb_compare.active)
            write_uitleg_sheet(wb_compare, 8)
            write_financial_comparison_sheet(wb_compare, all_financials)
            write_cost_calculation_sheet(wb_compare, financial_config)
            wb_compare.save(file8_compare)
            print(f"    Geschreven: {file8_compare}")

            # Find best permutation
            best_key = max(all_financials.keys(),
                           key=lambda k: all_financials[k]['totals']['total_net_profit'])
            best_profit = all_financials[best_key]['totals']['total_net_profit']
            print(f"    Beste optie: {best_key} "
                  f"met {all_financials[best_key]['totals']['bus_count']} bussen, "
                  f"winst: {best_profit:,.0f} EUR")

            # Store best for version 9 comparison
            rot8 = all_rotations[best_key]
            financials = all_financials[best_key]
            n8_with_trips = len([r for r in rot8 if r.real_trips])
            n8_reserve_only = len([r for r in rot8 if not r.real_trips and r.reserve_trip_list])
            n8_res_planned = sum(len(r.reserve_trip_list) for r in rot8)
            n8_extra = max(0, total_reserves - n8_res_planned)
            n8_reserve_bussen = n8_reserve_only + n8_extra
            n8_idle = sum(r.total_idle_minutes for r in rot8)

            algo_results[9] = {"rotations": rot8, "buses_met_ritten": n8_with_trips,
                               "reserve_bussen": n8_reserve_bussen, "idle_min": n8_idle,
                               "file": file8_compare, "financials": financials,
                               "all_financials": all_financials, "best_permutation": best_key}

        # ---------------------------------------------------------------
        # OUTPUT 10: Winstmaximalisatie (Profit Maximization)
        # Can run with or without deadhead data
        # Supports --multiday for reduced garage costs
        # ---------------------------------------------------------------
        if getattr(args, 'kosten_optimalisatie', False) and financial_config and fuel_config:
            use_multiday_v10 = getattr(args, 'multiday', False)
            multiday_label = " (meerdaags)" if use_multiday_v10 else ""
            print(f"  Output 10 - Winstmaximalisatie{multiday_label}...")
            print(f"    Onderzoekt verschillende aantallen bussen voor maximale winst...")

            rot10 = []
            all_profit_info = []

            # Handle trip_overrides if not available
            trip_overrides_v10 = trip_overrides if 'trip_overrides' in dir() else None

            # Group trips - either by date or multi-day
            from collections import defaultdict
            if use_multiday_v10:
                # Multi-day: group by bus type only, allowing cross-day chaining
                multiday_groups, _ = _group_trips_multiday(all_trips, baseline_turnaround)

                for bus_type, group_trips in multiday_groups.items():
                    if not group_trips:
                        continue

                    # Run profit-maximizing optimization with multiday support
                    max_extra_pct = getattr(financial_config, 'max_extra_buses_pct', 50)

                    # For multiday, we use the standard optimizer first, then explore profit variations
                    # The _optimize_profit_maximizing doesn't directly support multiday yet,
                    # so we use the regular multiday optimizer and calculate financials
                    if algo_key == "greedy":
                        chains = _optimize_greedy(group_trips, baseline_turnaround,
                                                  service_constraint=False,
                                                  deadhead_matrix=deadhead_matrix,
                                                  trip_turnaround_overrides=trip_overrides_v10,
                                                  multiday=True)
                    else:
                        chains = _optimize_mincost(group_trips, baseline_turnaround,
                                                   service_constraint=False,
                                                   deadhead_matrix=deadhead_matrix,
                                                   trip_turnaround_overrides=trip_overrides_v10,
                                                   multiday=True)

                    # Build rotations from chains
                    for chain_idx, chain in enumerate(chains):
                        chain_trips = [group_trips[i] for i in chain]
                        dates = sorted(set(t.date_str.split()[0] for t in chain_trips if t.date_str))
                        date_label = dates[0] if len(dates) == 1 else f"{dates[0]}-{dates[-1]}"
                        rot = BusRotation(
                            bus_id=f"v10_{bus_type[:2]}_{date_label}_{chain_idx+1:03d}",
                            bus_type=chain_trips[0].bus_type,
                            date_str=chain_trips[0].date_str,
                            trips=chain_trips
                        )
                        # Calculate idle time for multiday rotations
                        if len(chain_trips) > 1:
                            total_idle = 0
                            for i in range(len(chain_trips) - 1):
                                t1, t2 = chain_trips[i], chain_trips[i+1]
                                day_offset = _parse_date_to_ordinal(t2.date_str) - _parse_date_to_ordinal(t1.date_str)
                                gap = (day_offset * 1440) + t2.departure - t1.arrival
                                if deadhead_matrix:
                                    dh = deadhead_matrix.get(normalize_location(t1.dest_code), {}).get(
                                        normalize_location(t2.origin_code), 0)
                                    gap = max(0, gap - dh)
                                total_idle += max(0, gap)
                            rot.total_idle_minutes = int(total_idle)
                        rot10.append(rot)

                    print(f"    [{bus_type}] {len(chains)} bussen (meerdaags)")
            else:
                # Single-day: group by date and bus type
                groups_by_date_type = defaultdict(list)
                for t in all_trips:
                    groups_by_date_type[(t.date_str, t.bus_type)].append(t)

                for (date_str, bus_type), group_trips in groups_by_date_type.items():
                    if not group_trips:
                        continue

                    # Run profit-maximizing optimization
                    max_extra_pct = getattr(financial_config, 'max_extra_buses_pct', 50)
                    chains, profit_info = _optimize_profit_maximizing(
                        group_trips, baseline_turnaround,
                        service_constraint=True,
                        deadhead_matrix=deadhead_matrix,
                        trip_turnaround_overrides=trip_overrides_v10,
                        financial_config=financial_config,
                        fuel_config=fuel_config,
                        distance_matrix=deadhead_km_matrix,
                        max_extra_buses_pct=max_extra_pct,
                        algorithm=algo_key
                    )

                    if profit_info:
                        all_profit_info.append({
                            'date': date_str,
                            'bus_type': bus_type,
                            'info': profit_info
                        })

                        # Report findings for this group
                        if profit_info['best_buses'] != profit_info['min_buses']:
                            print(f"    [{date_str}/{bus_type}] Optimaal: {profit_info['best_buses']} bussen "
                                  f"(min={profit_info['min_buses']}, winst +{profit_info['best_profit']:,.0f} EUR)")

                    # Build rotations from chains
                    sorted_trips = sorted(group_trips, key=lambda t: (t.departure, t.arrival))
                    for chain_idx, chain in enumerate(chains):
                        chain_trips = [sorted_trips[i] for i in chain]
                        rot = BusRotation(
                            bus_id=f"v10_{date_str}_{bus_type[:2]}_{chain_idx+1:03d}",
                            bus_type=chain_trips[0].bus_type,
                            date_str=chain_trips[0].date_str,
                            trips=chain_trips
                        )
                        rot10.append(rot)

            # Note: Version 10 doesn't include phantom reserve trips in optimization
            # Reserves are handled separately in output generation

            # Apply fuel constraints if enabled (same as Version 7)
            fuel_results_10 = None
            if args.fuel_constraints and fuel_config and fuel_stations:
                rot10_before = len(rot10)
                rot10, fuel_results_10, fuel_splits_10 = apply_fuel_constraints(
                    rot10, fuel_config, fuel_stations, deadhead_matrix, deadhead_km_matrix,
                    turnaround_map=baseline_turnaround, algorithm=algo_key
                )
                if len(rot10) > rot10_before:
                    print(f"    Brandstofvalidatie: {len(rot10) - rot10_before} extra bussen door tankbeperkingen")

            # Calculate financials
            financials10 = calculate_total_financials(rot10, financial_config, fuel_type="diesel")
            totals10 = financials10['totals']

            # Show profit exploration summary
            total_min_buses = sum(p['info']['min_buses'] for p in all_profit_info)
            total_best_buses = sum(p['info']['best_buses'] for p in all_profit_info)
            if total_best_buses != total_min_buses:
                print(f"    Totaal: {total_best_buses} bussen gekozen (vs {total_min_buses} minimum)")

            print(f"    Totale omzet: {totals10['total_revenue']:,.2f} EUR")
            total_costs = totals10['total_driver_cost'] + totals10['total_fuel_cost'] + totals10['total_garage_fuel_cost']
            print(f"    Totale kosten: {total_costs:,.2f} EUR")
            print(f"      - Chauffeurkosten: {totals10['total_driver_cost']:,.2f} EUR")
            print(f"      - Brandstofkosten (ritten): {totals10['total_fuel_cost']:,.2f} EUR")
            print(f"      - Brandstofkosten (garage): {totals10['total_garage_fuel_cost']:,.2f} EUR")
            print(f"      - Garage km totaal: {totals10['total_garage_km']:,.1f} km")
            print(f"    Netto winst: {totals10['total_net_profit']:,.2f} EUR")

            # Compare with version 9
            if 9 in algo_results and 'financials' in algo_results[9]:
                profit9 = algo_results[9]['financials']['totals']['total_net_profit']
                profit10 = totals10['total_net_profit']
                diff = profit10 - profit9
                pct = (diff/profit9*100) if profit9 else 0
                print(f"    Winstverschil t.o.v. v9: {diff:+,.2f} EUR ({pct:+.2f}%)")

            file10 = f"{output_base}_{algo_short}_10_winstmaximalisatie.xlsx"
            print(f"    Schrijven {file10}...", end=" ", flush=True)

            generate_output(rot10, trips_with_reserves, reserves, file10, baseline_turnaround, algo_key,
                            include_sensitivity=True, output_mode=4,
                            risk_report=risk_report if 'risk_report' in dir() else None,
                            deadhead_matrix=deadhead_matrix, version=10)

            # Add financial sheet and fuel analysis if applicable
            wb10 = openpyxl.load_workbook(file10)
            if fuel_results_10 and args.fuel_constraints:
                write_fuel_analysis_sheet(wb10, fuel_results_10, fuel_stations, fuel_config)
            write_financial_sheet(wb10, financials10)
            wb10.save(file10)
            print("OK")

            n10_with_trips = len([r for r in rot10 if r.real_trips])
            n10_reserve_only = len([r for r in rot10 if not r.real_trips and r.reserve_trip_list])
            n10_res_planned = sum(len(r.reserve_trip_list) for r in rot10)
            n10_extra = max(0, total_reserves - n10_res_planned)
            n10_reserve_bussen = n10_reserve_only + n10_extra
            n10_idle = sum(r.total_idle_minutes for r in rot10)

            algo_results[10] = {"rotations": rot10, "buses_met_ritten": n10_with_trips,
                               "reserve_bussen": n10_reserve_bussen, "idle_min": n10_idle,
                               "file": file10, "financials": financials10,
                               "profit_info": all_profit_info}

        all_results[algo_key] = algo_results
        print()

    # ===== FINAL COMPARISON TABLE =====
    print()
    print(f"VERGELIJKINGSTABEL")
    print(f"{'='*100}")

    output_labels = {
        1: "1. Per dienst",
        2: "2. Per dienst + reserve idle",
        3: "3. Gecombineerd + reserve ingepland",
        4: "4. Gecombineerd + reserve + risico",
        5: "5. Gecombineerd + reserve + deadhead + risico",
        6: "6. Meerdaagse optimalisatie",
        7: "7. Brandstof/laad strategie",
        8: "8. Garage reistijden",
        9: "9. Financieel overzicht",
        10: "10. Winstmaximalisatie",
    }
    # Fallback label when output 4 is deadhead without traffic data
    if deadhead_matrix and not (traffic_data and traffic_data.get("time_slots")):
        output_labels[4] = "4. Gecombineerd + reserve + deadhead"

    # Header
    cw = 14  # column width per algorithm
    lw = 50  # label width
    algo_short_names = {"greedy": "Greedy", "mincost": "Min-cost"}
    print(f"{'Output':<{lw}s}", end="")
    for ak in algo_keys:
        print(f" {algo_short_names.get(ak, ak):>{cw}s}", end="")
    print()
    print(f"{'-'*lw}", end="")
    for _ in algo_keys:
        print(f" {'':->{ cw }s}", end="")
    print()

    # Collect all output numbers across all algorithms
    all_output_nums = set()
    for ak in algo_keys:
        all_output_nums.update(all_results[ak].keys())
    output_nums = sorted(all_output_nums)

    for out_num in output_nums:
        label = output_labels.get(out_num, f"{out_num}. (onbekend)")
        print(label)

        # Helper: print a row, showing "-" for algorithms that don't have this output
        def _print_row(row_label, value_func):
            print(f"{row_label:<{lw}s}", end="")
            for ak in algo_keys:
                if out_num in all_results[ak]:
                    r = all_results[ak][out_num]
                    print(f" {value_func(r):>{cw}}", end="")
                else:
                    print(f" {'-':>{cw}s}", end="")
            print()

        _print_row("  bussen met ritten", lambda r: f"{r['buses_met_ritten']:d}")
        _print_row("  reserve bussen", lambda r: f"{r['reserve_bussen']:d}")
        _print_row("  TOTAAL VLOOT", lambda r: f"{r['buses_met_ritten'] + r['reserve_bussen']:d}")
        _print_row("  wachttijd (uur)", lambda r: f"{r['idle_min'] / 60:.1f}")
        print()

    print(f"Reservebussen totaal nodig: {total_reserves}")
    print(f"Gegenereerde bestanden: {n_files}")

    # Halt capacity check (if --capaciteit provided)
    if halt_capacity:
        print()
        print("HALTECAPACITEITSCHECK")
        print("=" * 60)
        any_violation = False
        for ak in algo_keys:
            results = all_results.get(ak, {})
            if not results:
                continue
            # Use the most complete output (highest number)
            best_out = max(results.keys())
            rots = results[best_out]["rotations"]
            violations = check_halt_capacity(rots, halt_capacity)
            algo_label = algo_short_names.get(ak, ak)
            if violations:
                any_violation = True
                for v in violations:
                    print(f"  WAARSCHUWING [{algo_label}, output {best_out}]: "
                          f"{v['station']} op {v['date']} om {v['time']}: "
                          f"{v['count']} bussen (max {v['capacity']})")
            else:
                print(f"  {algo_label} (output {best_out}): geen overschrijdingen")
        if not any_violation:
            print("  Alle haltes binnen capaciteit.")

    print()
    print("Klaar!")


if __name__ == "__main__":
    main()
