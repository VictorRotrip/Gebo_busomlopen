#!/usr/bin/env python3
"""
Financial Calculator Module for Bus Rotation Optimizer

Calculates revenue, driver costs, fuel costs, and profit for bus rotations.
All financial parameters are loaded from additional_inputs.xlsx.

Revenue = sum of trip durations × hourly rate per bus type (NS pays only for driving time)
Driver cost = full shift duration with ORT, pauzestaffel, overtime, meal allowances
Fuel cost = km driven × consumption per bus type × fuel price
Profit = Revenue - Driver cost - Fuel cost + Sustainability bonuses
"""

from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import openpyxl


@dataclass
class FinancialConfig:
    """Configuration loaded from additional_inputs.xlsx"""
    # Tarieven (hourly rates per bus type)
    rates: Dict[str, float] = field(default_factory=dict)

    # Pauzestaffel (break deduction brackets)
    # List of (max_hours, deduction_minutes) tuples
    pauzestaffel: List[Tuple[float, int]] = field(default_factory=list)

    # Driver costs
    basis_uurloon: float = 18.50
    werkgever_opslag: float = 1.35

    # ORT (unsocial hours surcharge)
    ort_weekday_start: float = 19.0  # 19:00
    ort_weekday_end: float = 7.5     # 07:30
    ort_weekday_rate: float = 4.80
    ort_saturday_rate: float = 4.80
    ort_sunday_rate: float = 6.68

    # Overtime
    overwerk_threshold_per_maand: float = 173.33
    overwerk_toeslag_rijdend_pct: float = 35.0

    # Meal allowances
    maaltijd_threshold_1: float = 11.0  # hours
    maaltijd_vergoeding_1: float = 22.50
    maaltijd_threshold_2: float = 14.0  # hours
    maaltijd_vergoeding_2: float = 36.72

    # Broken shift
    gebroken_dienst_min_onderbreking: int = 120  # minutes
    gebroken_dienst_toeslag: float = 15.00

    # Fuel consumption (L/100km for diesel)
    fuel_consumption: Dict[str, float] = field(default_factory=dict)

    # Fuel prices
    diesel_price: float = 1.65
    hvo_price: float = 1.95
    electricity_price: float = 0.35

    # Sustainability bonuses
    ze_bonus_per_km: float = 0.12
    hvo_bonus_per_liter: float = 0.05
    hvo_max_total_per_liter: float = 0.40


@dataclass
class DriverCostBreakdown:
    """Detailed breakdown of driver costs for a rotation"""
    shift_start_min: int  # Minutes from midnight
    shift_end_min: int
    shift_duration_hours: float
    paid_hours: float  # After pauzestaffel deduction
    base_wage: float
    ort_hours: float
    ort_amount: float
    meal_allowance: float
    broken_shift_allowance: float
    total_cost: float


@dataclass
class RotationFinancials:
    """Complete financial analysis for a rotation"""
    rotation_id: str
    bus_type: str
    date_str: str

    # Revenue
    driving_minutes: int
    revenue: float

    # Costs
    driver_cost: DriverCostBreakdown
    fuel_cost: float
    fuel_km: float

    # Profit
    gross_profit: float

    # Sustainability
    ze_bonus: float = 0.0
    hvo_bonus: float = 0.0
    net_profit: float = 0.0


def load_financial_config(xlsx_path: str) -> FinancialConfig:
    """Load financial configuration from additional_inputs.xlsx"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    config = FinancialConfig()

    # Helper to find value by variable name in a sheet
    def get_value(ws, var_name, default=None):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            if row and row[0] == var_name:
                return row[1] if row[1] is not None else default
        return default

    # Load Tarieven
    ws = wb['Tarieven']
    config.rates = {
        'Dubbeldekker': get_value(ws, 'rate_dubbeldekker_eur_per_hour', 116.37),
        'Touringcar': get_value(ws, 'rate_touringcar_eur_per_hour', 80.455),
        'Lagevloerbus': get_value(ws, 'rate_lagevloer_gelede_eur_per_hour', 80.445),
        'Midi bus': get_value(ws, 'rate_midibus_eur_per_hour', 74.85),
        'Taxibus': get_value(ws, 'rate_taxibus_eur_per_hour', 50.455),
    }

    # Load Chauffeurkosten
    ws = wb['Chauffeurkosten']

    # Pauzestaffel
    config.pauzestaffel = []
    for i in range(1, 7):
        max_hours = get_value(ws, f'Pauzestaffel_{i}_max_diensttijd_uren')
        if max_hours is None:
            max_hours = get_value(ws, f'Pauzestaffel_{i}_min_diensttijd_uren')
        deduction = get_value(ws, f'Pauzestaffel_{i}_pauze_min', 0)
        if max_hours is not None:
            config.pauzestaffel.append((max_hours, deduction))

    # Sort by max_hours ascending
    config.pauzestaffel.sort(key=lambda x: x[0])

    # Base wage
    config.basis_uurloon = get_value(ws, 'basis_uurloon_chauffeur_eur', 18.50)
    config.werkgever_opslag = get_value(ws, 'werkgever_opslag_factor', 1.35)

    # ORT
    config.ort_weekday_start = get_value(ws, 'ORT_weekday_start_hour', 19.0)
    config.ort_weekday_end = get_value(ws, 'ORT_weekday_end_hour', 7.5)
    config.ort_weekday_rate = get_value(ws, 'ORT_weekday_rate_eur_per_hour', 4.80)
    config.ort_saturday_rate = get_value(ws, 'ORT_saturday_rate_eur_per_hour', 4.80)
    config.ort_sunday_rate = get_value(ws, 'ORT_sunday_rate_eur_per_hour', 6.68)

    # Overtime
    config.overwerk_threshold_per_maand = get_value(ws, 'overwerk_threshold_uren_per_maand', 173.33)
    config.overwerk_toeslag_rijdend_pct = get_value(ws, 'overwerk_toeslag_rijdend_pct', 35.0)

    # Meal allowances
    config.maaltijd_threshold_1 = get_value(ws, 'maaltijd_threshold_1_uren', 11.0)
    config.maaltijd_vergoeding_1 = get_value(ws, 'maaltijd_vergoeding_1_eur', 22.50)
    config.maaltijd_threshold_2 = get_value(ws, 'maaltijd_threshold_2_uren', 14.0)
    config.maaltijd_vergoeding_2 = get_value(ws, 'maaltijd_vergoeding_2_eur', 36.72)

    # Broken shift
    config.gebroken_dienst_min_onderbreking = int(get_value(ws, 'gebroken_dienst_min_onderbreking_min', 120))
    config.gebroken_dienst_toeslag = get_value(ws, 'gebroken_dienst_toeslag_eur', 15.00)

    # Load Buskosten (fuel consumption)
    ws = wb['Buskosten']
    config.fuel_consumption = {
        'Dubbeldekker': get_value(ws, 'verbruik_dubbeldekker_diesel_l_per_100km', 45),
        'Touringcar': get_value(ws, 'verbruik_touringcar_diesel_l_per_100km', 32),
        'Lagevloerbus': get_value(ws, 'verbruik_lagevloer_diesel_l_per_100km', 38),
        'Midi bus': get_value(ws, 'verbruik_midibus_diesel_l_per_100km', 25),
        'Taxibus': get_value(ws, 'verbruik_taxibus_diesel_l_per_100km', 12),
    }

    # Load Brandstofprijzen
    ws = wb['Brandstofprijzen']
    # Try API prices first, fall back to manual
    diesel_api = get_value(ws, 'diesel_b7_pompprijs_eur_per_liter')
    diesel_manual = get_value(ws, 'diesel_b7_handmatig_eur_per_liter', 1.65)
    config.diesel_price = diesel_api if diesel_api else diesel_manual

    hvo_api = get_value(ws, 'hvo100_prijs_eur_per_liter')
    hvo_manual = get_value(ws, 'hvo100_handmatig_eur_per_liter', 1.95)
    config.hvo_price = hvo_api if hvo_api else hvo_manual

    elec_api = get_value(ws, 'elektriciteit_snelladen_eur_per_kwh')
    elec_manual = get_value(ws, 'elektriciteit_handmatig_eur_per_kwh', 0.35)
    config.electricity_price = elec_api if elec_api else elec_manual

    # Load Duurzaamheid
    ws = wb['Duurzaamheid']
    config.ze_bonus_per_km = get_value(ws, 'zero_emissie_stimulans_eur_per_km', 0.12)
    config.hvo_bonus_per_liter = get_value(ws, 'hvo_stimulans_eur_per_liter', 0.05)
    config.hvo_max_total_per_liter = get_value(ws, 'hvo_max_total_eur_per_liter', 0.40)

    wb.close()
    return config


def is_weekend(date_str: str) -> Tuple[bool, bool]:
    """Check if date is Saturday or Sunday.

    Returns: (is_saturday, is_sunday)
    """
    # Parse date from format like "do 11-06-2026" or "2026-06-11"
    try:
        if '-' in date_str and len(date_str) > 10:
            # Format: "do 11-06-2026"
            parts = date_str.split()
            date_part = parts[-1] if len(parts) > 1 else date_str
            dt = datetime.strptime(date_part, "%d-%m-%Y")
        else:
            dt = datetime.strptime(date_str, "%Y-%m-%d")

        weekday = dt.weekday()  # 0=Monday, 5=Saturday, 6=Sunday
        return (weekday == 5, weekday == 6)
    except:
        return (False, False)


def calculate_ort_hours(shift_start_min: int, shift_end_min: int,
                        date_str: str, config: FinancialConfig) -> Tuple[float, float]:
    """Calculate ORT hours and amount for a shift.

    Args:
        shift_start_min: Shift start in minutes from midnight
        shift_end_min: Shift end in minutes from midnight
        date_str: Date string to determine weekend
        config: Financial configuration

    Returns: (ort_hours, ort_amount)
    """
    is_saturday, is_sunday = is_weekend(date_str)

    # Sunday/holiday: entire shift gets ORT
    if is_sunday:
        ort_hours = (shift_end_min - shift_start_min) / 60.0
        return (ort_hours, ort_hours * config.ort_sunday_rate)

    # Saturday: entire shift gets ORT
    if is_saturday:
        ort_hours = (shift_end_min - shift_start_min) / 60.0
        return (ort_hours, ort_hours * config.ort_saturday_rate)

    # Weekday: ORT for hours in 19:00-07:30 window
    ort_start_min = int(config.ort_weekday_start * 60)  # 19:00 = 1140 min
    ort_end_min = int(config.ort_weekday_end * 60)      # 07:30 = 450 min

    ort_minutes = 0

    # Handle the ORT window (19:00 to 07:30 next day = two segments)
    # Segment 1: 19:00 (1140) to 24:00 (1440)
    # Segment 2: 00:00 (0) to 07:30 (450)

    for seg_start, seg_end in [(ort_start_min, 1440), (0, ort_end_min)]:
        # Calculate overlap between shift and this ORT segment
        overlap_start = max(shift_start_min, seg_start)
        overlap_end = min(shift_end_min, seg_end)
        if overlap_end > overlap_start:
            ort_minutes += (overlap_end - overlap_start)

    ort_hours = ort_minutes / 60.0
    return (ort_hours, ort_hours * config.ort_weekday_rate)


def calculate_pauzestaffel_deduction(shift_hours: float, config: FinancialConfig) -> int:
    """Calculate break deduction in minutes based on shift duration.

    Uses pauzestaffel brackets from CAO.
    """
    for max_hours, deduction in config.pauzestaffel:
        if shift_hours <= max_hours:
            return deduction

    # If longer than all brackets, use the last bracket's deduction
    if config.pauzestaffel:
        return config.pauzestaffel[-1][1]
    return 0


def calculate_meal_allowance(shift_hours: float, config: FinancialConfig) -> float:
    """Calculate meal allowance based on shift duration."""
    if shift_hours >= config.maaltijd_threshold_2:
        return config.maaltijd_vergoeding_2
    elif shift_hours >= config.maaltijd_threshold_1:
        return config.maaltijd_vergoeding_1
    return 0.0


def has_broken_shift(idle_gaps: List[int], config: FinancialConfig) -> bool:
    """Check if rotation has a broken shift (long gap between trips)."""
    return any(gap >= config.gebroken_dienst_min_onderbreking for gap in idle_gaps)


def calculate_driver_cost(shift_start_min: int, shift_end_min: int,
                          idle_gaps: List[int], date_str: str,
                          config: FinancialConfig) -> DriverCostBreakdown:
    """Calculate full driver cost for a rotation.

    Args:
        shift_start_min: First trip departure (minutes from midnight)
        shift_end_min: Last trip arrival (minutes from midnight)
        idle_gaps: List of idle gap durations in minutes
        date_str: Date string
        config: Financial configuration

    Returns: DriverCostBreakdown with all cost components
    """
    shift_duration_min = shift_end_min - shift_start_min
    shift_duration_hours = shift_duration_min / 60.0

    # Pauzestaffel deduction
    deduction_min = calculate_pauzestaffel_deduction(shift_duration_hours, config)
    paid_hours = max(0, (shift_duration_min - deduction_min) / 60.0)

    # Base wage (with employer costs)
    hourly_cost = config.basis_uurloon * config.werkgever_opslag
    base_wage = paid_hours * hourly_cost

    # ORT
    ort_hours, ort_amount = calculate_ort_hours(shift_start_min, shift_end_min, date_str, config)

    # Meal allowance
    meal_allowance = calculate_meal_allowance(shift_duration_hours, config)

    # Broken shift allowance
    broken_shift_allowance = config.gebroken_dienst_toeslag if has_broken_shift(idle_gaps, config) else 0.0

    total_cost = base_wage + ort_amount + meal_allowance + broken_shift_allowance

    return DriverCostBreakdown(
        shift_start_min=shift_start_min,
        shift_end_min=shift_end_min,
        shift_duration_hours=shift_duration_hours,
        paid_hours=paid_hours,
        base_wage=base_wage,
        ort_hours=ort_hours,
        ort_amount=ort_amount,
        meal_allowance=meal_allowance,
        broken_shift_allowance=broken_shift_allowance,
        total_cost=total_cost,
    )


def calculate_revenue(driving_minutes: int, bus_type: str, config: FinancialConfig) -> float:
    """Calculate revenue for a rotation.

    Revenue = total trip duration × hourly rate for bus type
    NS only pays for actual driving time, not idle time.
    """
    rate = config.rates.get(bus_type, 80.0)  # Default to Touringcar rate
    return (driving_minutes / 60.0) * rate


def calculate_fuel_cost(km_driven: float, bus_type: str,
                        fuel_type: str, config: FinancialConfig) -> float:
    """Calculate fuel cost for a rotation.

    Args:
        km_driven: Total km driven (trips + deadhead)
        bus_type: Type of bus
        fuel_type: "diesel", "hvo", or "ze"
        config: Financial configuration

    Returns: Fuel cost in EUR
    """
    if fuel_type == "ze":
        # ZE buses use electricity - different calculation
        # Assume 1.5 kWh/km average for buses
        kwh_per_km = 1.5
        return km_driven * kwh_per_km * config.electricity_price

    # Diesel or HVO
    consumption_l_per_100km = config.fuel_consumption.get(bus_type, 32)  # Default Touringcar
    liters = km_driven * consumption_l_per_100km / 100.0

    if fuel_type == "hvo":
        return liters * config.hvo_price
    else:  # diesel
        return liters * config.diesel_price


def calculate_sustainability_bonus(km_driven: float, bus_type: str,
                                   fuel_type: str, config: FinancialConfig) -> Tuple[float, float]:
    """Calculate sustainability bonuses.

    Returns: (ze_bonus, hvo_bonus)
    """
    ze_bonus = 0.0
    hvo_bonus = 0.0

    if fuel_type == "ze":
        ze_bonus = km_driven * config.ze_bonus_per_km
    elif fuel_type == "hvo":
        consumption_l_per_100km = config.fuel_consumption.get(bus_type, 32)
        liters = km_driven * consumption_l_per_100km / 100.0
        # HVO bonus = stimulans + price difference (capped)
        hvo_bonus = liters * config.hvo_bonus_per_liter

    return (ze_bonus, hvo_bonus)


def calculate_rotation_financials(rotation, config: FinancialConfig,
                                  fuel_type: str = "diesel",
                                  km_estimate_speed: float = 50.0) -> RotationFinancials:
    """Calculate complete financials for a BusRotation object.

    Args:
        rotation: BusRotation object from busomloop_optimizer
        config: Financial configuration
        fuel_type: "diesel", "hvo", or "ze"
        km_estimate_speed: Speed (km/h) to estimate km from trip duration if not available

    Returns: RotationFinancials with complete breakdown
    """
    trips = rotation.trips if hasattr(rotation, 'trips') else []
    real_trips = [t for t in trips if not getattr(t, 'is_reserve', False) and not getattr(t, 'is_deadhead', False)]

    if not real_trips:
        # Empty rotation
        return RotationFinancials(
            rotation_id=rotation.bus_id,
            bus_type=rotation.bus_type,
            date_str=rotation.date_str,
            driving_minutes=0,
            revenue=0.0,
            driver_cost=DriverCostBreakdown(0, 0, 0, 0, 0, 0, 0, 0, 0, 0),
            fuel_cost=0.0,
            fuel_km=0.0,
            gross_profit=0.0,
        )

    # Calculate driving minutes (revenue time)
    driving_minutes = sum(t.duration for t in real_trips)

    # Shift times
    shift_start_min = min(t.departure for t in real_trips)
    shift_end_min = max(t.arrival for t in real_trips)

    # Calculate idle gaps (for broken shift detection)
    sorted_trips = sorted(real_trips, key=lambda t: t.departure)
    idle_gaps = []
    for i in range(1, len(sorted_trips)):
        gap = sorted_trips[i].departure - sorted_trips[i-1].arrival
        if gap > 0:
            idle_gaps.append(gap)

    # Revenue (NS pays for driving time only)
    revenue = calculate_revenue(driving_minutes, rotation.bus_type, config)

    # Driver cost
    driver_cost = calculate_driver_cost(
        shift_start_min, shift_end_min, idle_gaps, rotation.date_str, config
    )

    # Estimate km driven
    # Use total_km if available on rotation, else estimate from driving time
    if hasattr(rotation, 'total_km') and rotation.total_km:
        fuel_km = rotation.total_km
    else:
        # Estimate: driving_minutes at km_estimate_speed
        fuel_km = (driving_minutes / 60.0) * km_estimate_speed

    # Add deadhead km if tracked
    if hasattr(rotation, 'deadhead_km'):
        fuel_km += rotation.deadhead_km

    # Fuel cost
    fuel_cost = calculate_fuel_cost(fuel_km, rotation.bus_type, fuel_type, config)

    # Gross profit (before sustainability bonuses)
    gross_profit = revenue - driver_cost.total_cost - fuel_cost

    # Sustainability bonuses
    ze_bonus, hvo_bonus = calculate_sustainability_bonus(
        fuel_km, rotation.bus_type, fuel_type, config
    )

    net_profit = gross_profit + ze_bonus + hvo_bonus

    return RotationFinancials(
        rotation_id=rotation.bus_id,
        bus_type=rotation.bus_type,
        date_str=rotation.date_str,
        driving_minutes=driving_minutes,
        revenue=revenue,
        driver_cost=driver_cost,
        fuel_cost=fuel_cost,
        fuel_km=fuel_km,
        gross_profit=gross_profit,
        ze_bonus=ze_bonus,
        hvo_bonus=hvo_bonus,
        net_profit=net_profit,
    )


def calculate_total_financials(rotations: list, config: FinancialConfig,
                               fuel_type: str = "diesel") -> Dict:
    """Calculate aggregated financials for all rotations.

    Returns dict with totals and per-rotation breakdown.
    """
    results = []
    totals = {
        'total_revenue': 0.0,
        'total_driver_cost': 0.0,
        'total_fuel_cost': 0.0,
        'total_gross_profit': 0.0,
        'total_ze_bonus': 0.0,
        'total_hvo_bonus': 0.0,
        'total_net_profit': 0.0,
        'total_driving_hours': 0.0,
        'total_shift_hours': 0.0,
        'total_km': 0.0,
        'total_ort_hours': 0.0,
        'total_ort_amount': 0.0,
    }

    for rotation in rotations:
        fin = calculate_rotation_financials(rotation, config, fuel_type)
        results.append(fin)

        totals['total_revenue'] += fin.revenue
        totals['total_driver_cost'] += fin.driver_cost.total_cost
        totals['total_fuel_cost'] += fin.fuel_cost
        totals['total_gross_profit'] += fin.gross_profit
        totals['total_ze_bonus'] += fin.ze_bonus
        totals['total_hvo_bonus'] += fin.hvo_bonus
        totals['total_net_profit'] += fin.net_profit
        totals['total_driving_hours'] += fin.driving_minutes / 60.0
        totals['total_shift_hours'] += fin.driver_cost.shift_duration_hours
        totals['total_km'] += fin.fuel_km
        totals['total_ort_hours'] += fin.driver_cost.ort_hours
        totals['total_ort_amount'] += fin.driver_cost.ort_amount

    return {
        'rotations': results,
        'totals': totals,
        'config': config,
    }


if __name__ == "__main__":
    # Test loading config
    import sys

    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "additional_inputs.xlsx"

    print(f"Loading financial config from {xlsx_path}...")
    config = load_financial_config(xlsx_path)

    print(f"\nTarieven (hourly rates):")
    for bt, rate in config.rates.items():
        print(f"  {bt}: {rate:.2f} EUR/hour")

    print(f"\nPauzestaffel (break deductions):")
    for max_h, deduct in config.pauzestaffel:
        print(f"  <= {max_h}h: -{deduct} min")

    print(f"\nDriver costs:")
    print(f"  Base wage: {config.basis_uurloon:.2f} EUR/hour")
    print(f"  Employer markup: {config.werkgever_opslag:.0%}")
    print(f"  Effective cost: {config.basis_uurloon * config.werkgever_opslag:.2f} EUR/hour")

    print(f"\nORT rates:")
    print(f"  Weekday (19:00-07:30): {config.ort_weekday_rate:.2f} EUR/hour")
    print(f"  Saturday: {config.ort_saturday_rate:.2f} EUR/hour")
    print(f"  Sunday: {config.ort_sunday_rate:.2f} EUR/hour")

    print(f"\nFuel consumption (L/100km):")
    for bt, cons in config.fuel_consumption.items():
        print(f"  {bt}: {cons} L/100km")

    print(f"\nFuel prices:")
    print(f"  Diesel B7: {config.diesel_price:.2f} EUR/liter")
    print(f"  HVO100: {config.hvo_price:.2f} EUR/liter")
    print(f"  Electricity: {config.electricity_price:.2f} EUR/kWh")

    print(f"\nSustainability bonuses:")
    print(f"  ZE: {config.ze_bonus_per_km:.2f} EUR/km")
    print(f"  HVO: {config.hvo_bonus_per_liter:.2f} EUR/liter")
